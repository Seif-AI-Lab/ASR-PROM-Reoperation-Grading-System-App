# -*- coding: utf-8 -*-
"""
Step 2 Analysis

Python script exported from the corresponding notebook.
The original notebook is the primary version in the notebooks/ folder.
"""

# %% [markdown] Cell 1
# #**Step2_ModelScreening_All8Classifiers**

# %% Cell 2
# -*- coding: utf-8 -*-
"""
Step 2 dynamic ODI classifier screening for delayed 1-year lumbar reoperation prediction
======================================================================================

This script is the Step 2 counterpart of the Step 1 classifier-screening script.
It compares eight prespecified machine-learning classifiers in one dynamic ODI
landmark cohort, using paired baseline-only and dynamic ODI-expanded feature
sets with identical patient-level train/calibration/test splits and identical
group-aware CV folds.

Input expected in /content
--------------------------
Step 2_ODI_Cohort.csv

Target
------
final_reop_step2
    1 = delayed reoperation from postoperative day 91 through day 365
    0 = no delayed reoperation through day 365
"""

import os, re, sys, json, math, time, platform, subprocess, warnings
from dataclasses import dataclass
from typing import Dict, List

import numpy as np
import pandas as pd

try:
    import lightgbm as lgb
    from lightgbm import LGBMClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lightgbm"])
    import lightgbm as lgb
    from lightgbm import LGBMClassifier

try:
    from xgboost import XGBClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "xgboost"])
    from xgboost import XGBClassifier

try:
    from catboost import CatBoostClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "catboost"])
    from catboost import CatBoostClassifier

try:
    import openpyxl  # noqa
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.model_selection import StratifiedShuffleSplit, StratifiedGroupKFold, ParameterSampler
from sklearn.metrics import average_precision_score, roc_auc_score, brier_score_loss
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.svm import LinearSVC
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier, HistGradientBoostingClassifier

warnings.filterwarnings("ignore")

# ============================================================
# 1) Configuration
# ============================================================
BASE_DIR = "/content"
INPUT_CSV = os.path.join(BASE_DIR, "Step 2_ODI_Cohort.csv")
OUTPUT_DIR = os.path.join(BASE_DIR, "Step2_ClassifierScreening_All8_outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

GROUP_COL = "PersonKey"
TARGET_COL = "final_reop_step2"
RANDOM_STATE = 20260524
TEST_FRACTION = 0.20
CALIBRATION_FRACTION_OF_REMAINING = 0.20
N_CV_FOLDS = 5
N_RANDOM_COMBINATIONS = 300
N_JOBS = -1
CALIBRATION_METHOD = "isotonic"
ECE_N_BINS = 10

# Step 2 dynamic ODI definitions. These are derived from input columns and then used as features.
RELATIVE_ODI_MCID_CUTOFF = 0.30
DAYS_BETWEEN_PROM_COL = "days_between_PROMs"
INPUT_ODI_MCID_COL = "ODI_MCID_binary"
PROM_CHANGE_RATE_COL = "ODI_change_rate"
RELATIVE_ODI_MCID_COL = "ODI_relative_MCID_binary"

BASELINE_FEATURES = [
    "finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis",
    "age","sex","race","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure",
    "connective_tissue_rheumatic_disease","diabetes_status","myocardial_infarction","renal_disease","institution_type",
    "institution_size","institution_region","asa","bmi","payer_status","alif_llif","corpectomy","discectomy",
    "foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif",
    "other_lumbar_procedure","number_operated_levels","operative_region_extent","PatTobaccoUse"]
assert len(BASELINE_FEATURES) == 35

STEP2_ODI_FEATURES = [
    "preop_ODI",
    "postop_ODI",
    "ODI_change",
    PROM_CHANGE_RATE_COL,
    RELATIVE_ODI_MCID_COL,
    "postop_ODI_day",
]
BASELINE_ONLY_FEATURES = list(BASELINE_FEATURES)
DYNAMIC_PROM_FEATURES = list(BASELINE_FEATURES) + list(STEP2_ODI_FEATURES)

EXCLUDED_FEATURES = {"reop","reoptime","final_reop","final_reop_step1","final_reop_step2","death_within_90d","death_within_180d","death_within_365d","death_after_index_surgery","death_before_or_on_index_surgery","PersonDeathDate","days_to_death_from_index_surgery","removal_hardware","any_arthroplasty","final_diagnosis_complexity","procedure_complexity_score"}
bad_features = sorted(set(DYNAMIC_PROM_FEATURES) & EXCLUDED_FEATURES)
if bad_features:
    raise ValueError(f"Excluded/leakage-prone features were accidentally included: {bad_features}")

CONTINUOUS_FEATURES = ["age","bmi","preop_ODI","postop_ODI","ODI_change",PROM_CHANGE_RATE_COL,"postop_ODI_day"]
CONTINUOUS_BASELINE_FEATURES = ["age", "bmi"]  # retained for output tables only
BINARY_FEATURES = ["finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis","sex","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure","connective_tissue_rheumatic_disease","myocardial_infarction","renal_disease","institution_type","alif_llif","corpectomy","discectomy","foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif","other_lumbar_procedure","operative_region_extent",RELATIVE_ODI_MCID_COL]
ORDINAL_FEATURES = ["diabetes_status","institution_size","asa","number_operated_levels"]
NOMINAL_FEATURES = ["race","institution_region","payer_status","PatTobaccoUse"]

MISSING_STRINGS = {""," ","na","n/a","nan","none","null",".","missing","<na>"}
BINARY_MAPS = {
    "sex": {"female":0,"f":0,"male":1,"m":1},
    "ethnicity": {"non-hispanic":0,"non hispanic":0,"hispanic":1},
    "cancer_status": {"no cancer":0,"no":0,"none":0,"any cancer":1,"yes":1,"cancer":1},
    "institution_type": {"hospital":0,"non-hospital":1,"non hospital":1,"nonhospital":1},
    "operative_region_extent": {"lumbar only":0,"extended_region_involvement":1,"extended region involvement":1,"extended":1},
}
ORDINAL_MAPS = {
    "diabetes_status": {"no":0,"none":0,"0":0,"without comp":1,"without complication":1,"without complications":1,"1":1,"with comp":2,"with complication":2,"with complications":2,"2":2},
    "institution_size": {"between 1-99 beds":0,"1-99":0,"between 100-399 beds":1,"100-399":1,">= 400 beds":2,">=400 beds":2,">=400":2,">= 400":2},
    "asa": {"1":1,"i":1,"2":2,"ii":2,"3":3,"iii":3,"4":4,"iv":4,">=4":4,">= 4":4,"5":4,"v":4},
    "number_operated_levels": {"0":0,"1":1,"2":2,"3":3,"4":4,">=4":4,">= 4":4,"5":4,"6":4,"7":4,"8":4,"9":4,"10":4},
}
PREFERRED_NOMINAL_LEVELS = {"race":["White","Black","Other"],"institution_region":["South","North East","West","Midwest"],"payer_status":["Medicare","Commercial/Private","Other","Medicaid/Public/Government"],"PatTobaccoUse":["Unknown/Not reported/Multiple","Never","Former","Current"]}

# ============================================================
# 2) Utility functions
# ============================================================
def clean_scalar(x):
    if pd.isna(x): return np.nan
    if isinstance(x,str):
        s=x.strip().replace("≥",">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x

def norm_text(x):
    x=clean_scalar(x)
    if pd.isna(x): return None
    return str(x).strip().replace("≥",">=").lower()

def to_binary_target(x):
    sx=norm_text(x)
    if sx is None: return np.nan
    if sx in {"1","1.0","yes","y","true","t"}: return 1.0
    if sx in {"0","0.0","no","n","false","f"}: return 0.0
    try:
        v=float(sx); return float(v) if v in (0.0,1.0) else np.nan
    except Exception: return np.nan

def find_existing_column(columns, candidates, what):
    lookup={c.lower():c for c in columns}
    for c in candidates:
        if c in columns: return c
        if c.lower() in lookup: return lookup[c.lower()]
    raise ValueError(f"Could not find {what}. Tried: {candidates}")

def resolve_target_column(df): return find_existing_column(df.columns.tolist(), TARGET_COL_CANDIDATES, "Step 1 target column")
def safe_average_precision(y,s): return np.nan if len(np.unique(y))<2 else float(average_precision_score(y,s))
def safe_roc_auc(y,s): return np.nan if len(np.unique(y))<2 else float(roc_auc_score(y,s))

def ece(y,p,n_bins=10):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); bins=np.linspace(0,1,n_bins+1); out=0.0
    if len(y)==0: return np.nan
    for i in range(n_bins):
        m=(p>=bins[i]) & ((p<=bins[i+1]) if i==n_bins-1 else (p<bins[i+1]))
        if np.any(m): out += (np.sum(m)/len(y))*abs(float(np.mean(y[m]))-float(np.mean(p[m])))
    return float(out)

def top5_metrics(y,p):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); n=len(y); k=max(1,int(math.ceil(n*0.05))); idx=np.argsort(-p)[:k]
    prev=float(np.mean(y)); rate=float(np.mean(y[idx]))
    return {"Top_5pct_n":int(k),"Top_5pct_event_rate":rate,"Top_5pct_lift":rate/prev if prev>0 else np.nan,"Top_5pct_captured_events":float(np.sum(y[idx])/np.sum(y)) if np.sum(y)>0 else np.nan}

def eval_preds(y,p,prefix=""):
    out={f"{prefix}AP":safe_average_precision(y,p),f"{prefix}ROC_AUC":safe_roc_auc(y,p),f"{prefix}Brier_score":float(brier_score_loss(y,p)),f"{prefix}ECE":ece(y,p,ECE_N_BINS),f"{prefix}N":int(len(y)),f"{prefix}Events":int(np.sum(y)),f"{prefix}Prevalence":float(np.mean(y))}
    out.update({f"{prefix}{k}":v for k,v in top5_metrics(y,p).items()})
    return out

def make_weights(y,mult):
    y=np.asarray(y).astype(int); npos=int(np.sum(y==1)); nneg=int(np.sum(y==0))
    if npos==0: raise ValueError("No positive events.")
    pw=(nneg/max(npos,1))*float(mult)
    return np.where(y==1,pw,1.0).astype(float)

def json_native(obj):
    if isinstance(obj,dict): return {str(k):json_native(v) for k,v in obj.items()}
    if isinstance(obj,(list,tuple)): return [json_native(v) for v in obj]
    if isinstance(obj,(np.integer,)): return int(obj)
    if isinstance(obj,(np.floating,)): return float(obj)
    if isinstance(obj,np.ndarray): return obj.tolist()
    try:
        if pd.isna(obj): return None
    except Exception: pass
    return obj

# ============================================================
# 3) Preprocessor with optional scaling
# ============================================================
@dataclass
class Step2Preprocessor:
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str,List[str]]
    scale_continuous: bool = False
    def __post_init__(self):
        self.continuous_imputer=None; self.continuous_scaler=None; self.discrete_imputer=None; self.nominal_imputer=None; self.onehot=None; self.output_feature_names_=[]
    def _parse_binary(self,x,f):
        sx=norm_text(x)
        if sx is None: return np.nan
        if f in BINARY_MAPS and sx in BINARY_MAPS[f]: return float(BINARY_MAPS[f][sx])
        if sx in {"1","1.0","yes","y","true","t","present","positive"}: return 1.0
        if sx in {"0","0.0","no","n","false","f","absent","negative"}: return 0.0
        try:
            v=float(sx); return float(v) if v in (0.0,1.0) else np.nan
        except Exception: return np.nan
    def _parse_ordinal(self,x,f):
        sx=norm_text(x)
        if sx is None: return np.nan
        if f in ORDINAL_MAPS and sx in ORDINAL_MAPS[f]: return float(ORDINAL_MAPS[f][sx])
        try:
            v=float(sx)
            if f=="asa": return float(min(max(int(round(v)),1),4))
            if f=="number_operated_levels": return float(min(max(int(round(v)),0),4))
            if f=="diabetes_status": return float(min(max(int(round(v)),0),2))
            if f=="institution_size": return float(min(max(int(round(v)),0),2))
            return float(v)
        except Exception: return np.nan
    def _nominal(self,f,x):
        x=clean_scalar(x)
        if pd.isna(x): return np.nan
        s=str(x).strip(); sl=s.lower()
        if f=="race": return "White" if sl=="white" else ("Black" if sl=="black" else "Other")
        if f=="institution_region": return {"south":"South","north east":"North East","northeast":"North East","north-east":"North East","west":"West","midwest":"Midwest","mid west":"Midwest"}.get(sl,s)
        if f=="payer_status":
            if sl=="medicare": return "Medicare"
            if sl in {"commercial/private","commercial","private","commercial private"}: return "Commercial/Private"
            if sl in {"medicaid/public/government","medicaid","public","government","public/government"}: return "Medicaid/Public/Government"
            return "Other"
        if f=="PatTobaccoUse": return "Never" if sl=="never" else ("Former" if sl=="former" else ("Current" if sl=="current" else "Unknown/Not reported/Multiple"))
        return s
    def _parts(self,X):
        cont=pd.DataFrame(index=X.index); disc=pd.DataFrame(index=X.index); nom=pd.DataFrame(index=X.index)
        for c in self.continuous_features: cont[c]=pd.to_numeric(X[c].map(clean_scalar),errors="coerce")
        for c in self.binary_features: disc[c]=X[c].map(lambda z:self._parse_binary(z,c)).astype(float)
        for c in self.ordinal_features: disc[c]=X[c].map(lambda z:self._parse_ordinal(z,c)).astype(float)
        for c in self.nominal_features: nom[c]=X[c].map(lambda z:self._nominal(c,z)).astype("object")
        return cont,disc,nom
    def fit(self,X):
        cont,disc,nom=self._parts(X)
        self.continuous_imputer=SimpleImputer(strategy="median")
        self.discrete_imputer=SimpleImputer(strategy="most_frequent")
        self.nominal_imputer=SimpleImputer(strategy="constant",fill_value="Missing")
        cont_imp=self.continuous_imputer.fit_transform(cont)
        if self.scale_continuous:
            self.continuous_scaler=StandardScaler(); self.continuous_scaler.fit(cont_imp)
        self.discrete_imputer.fit(disc)
        nomi=pd.DataFrame(self.nominal_imputer.fit_transform(nom),columns=self.nominal_features)
        cats=[]
        for c in self.nominal_features:
            pref=list(self.preferred_nominal_levels.get(c,[])); obs=nomi[c].astype(str).unique().tolist(); cats.append(pref+sorted([x for x in obs if x not in pref]))
        try: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse_output=False)
        except TypeError: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse=False)
        self.onehot.fit(nomi.astype(str)); self.output_feature_names_=self.continuous_features+self.binary_features+self.ordinal_features+self.onehot.get_feature_names_out(self.nominal_features).tolist(); return self
    def transform(self,X):
        cont,disc,nom=self._parts(X)
        a=self.continuous_imputer.transform(cont)
        if self.scale_continuous and self.continuous_scaler is not None: a=self.continuous_scaler.transform(a)
        b=self.discrete_imputer.transform(disc)
        nomi=pd.DataFrame(self.nominal_imputer.transform(nom),columns=self.nominal_features); c=self.onehot.transform(nomi.astype(str))
        return np.concatenate([a,b,c],axis=1).astype(float)
    def fit_transform(self,X): self.fit(X); return self.transform(X)

# ============================================================
# 4) Splits and model registry
# ============================================================
def feature_types(features):
    cont=[f for f in features if f in CONTINUOUS_FEATURES]
    binf=[f for f in features if f in BINARY_FEATURES]
    ordf=[f for f in features if f in ORDINAL_FEATURES]
    nomf=[f for f in features if f in NOMINAL_FEATURES]
    typed=set(cont+binf+ordf+nomf)
    unknown=[f for f in features if f not in typed]
    if unknown:
        raise ValueError(f"Features without preprocessing type assignment: {unknown}")
    return cont, binf, ordf, nomf

def patient_split(df,target_col,seed):
    gdf=df.groupby(GROUP_COL,dropna=False)[target_col].max().reset_index(); y=gdf[target_col].astype(int).to_numpy(); g=gdf[GROUP_COL].to_numpy()
    s1=StratifiedShuffleSplit(n_splits=1,test_size=TEST_FRACTION,random_state=seed); trcal,te=next(s1.split(g,y))
    gtc=g[trcal]; ytc=y[trcal]; s2=StratifiedShuffleSplit(n_splits=1,test_size=CALIBRATION_FRACTION_OF_REMAINING,random_state=seed+1); tr,ca=next(s2.split(gtc,ytc))
    train=set(gtc[tr]); cal=set(gtc[ca]); test=set(g[te])
    return df[GROUP_COL].isin(train).to_numpy(), df[GROUP_COL].isin(cal).to_numpy(), df[GROUP_COL].isin(test).to_numpy()

def cv_splits(y,groups,seed,n_folds=N_CV_FOLDS):
    ge=pd.DataFrame({"g":groups,"y":y}).groupby("g")["y"].max().reset_index(); nf=min(n_folds,int(np.sum(ge.y==1)),int(np.sum(ge.y==0)))
    if nf<2: raise ValueError("Not enough patient groups for CV.")
    cv=StratifiedGroupKFold(n_splits=nf,shuffle=True,random_state=seed)
    return [(tr,va) for tr,va in cv.split(np.zeros(len(y)),y,groups)]

MODEL_SEARCH_SPACES = {
    "LogisticRegression": {"C":[0.001,0.003,0.01,0.03,0.1,0.3,1,3,10,30], "positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "LinearSVC": {"C":[0.001,0.003,0.01,0.03,0.1,0.3,1,3,10,30], "positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "RandomForest": {"n_estimators":[400,700,1000],"max_depth":[None,3,5,7,10,15],"min_samples_leaf":[1,5,10,25,50],"max_features":["sqrt","log2",0.5,0.8],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "AdaBoost": {"n_estimators":[100,200,400,700,1000],"learning_rate":[0.003,0.005,0.01,0.03,0.05,0.1,0.3,0.5,1.0],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "LightGBM": {"n_estimators":[400,700,1000,1400,1800,2200,2600],"learning_rate":[0.003,0.005,0.008,0.01,0.02,0.03,0.05],"num_leaves":[7,15,31,63,127],"max_depth":[-1,2,3,5,7,9],"min_child_samples":[10,20,50,100,200,400],"subsample":[0.60,0.75,0.90,1.00],"subsample_freq":[0,1,2],"colsample_bytree":[0.60,0.75,0.90,1.00],"reg_alpha":[0,0.001,0.01,0.05,0.1,0.5,1,2],"reg_lambda":[0,0.001,0.01,0.05,0.1,0.5,1,2,5],"min_split_gain":[0,0.001,0.005,0.01,0.05,0.1],"max_bin":[63,127,255],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "XGBoost": {"n_estimators":[400,700,1000,1400],"learning_rate":[0.003,0.005,0.01,0.02,0.03,0.05],"max_depth":[2,3,5,7],"min_child_weight":[1,5,10,25],"subsample":[0.60,0.75,0.90,1.00],"colsample_bytree":[0.60,0.75,0.90,1.00],"reg_alpha":[0,0.001,0.01,0.1,1],"reg_lambda":[0.01,0.1,1,2,5],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "HistGradientBoosting": {"max_iter":[200,400,700,1000],"learning_rate":[0.003,0.005,0.01,0.03,0.05,0.1],"max_leaf_nodes":[7,15,31,63],"max_depth":[None,2,3,5,7],"min_samples_leaf":[10,20,50,100,200],"l2_regularization":[0,0.001,0.01,0.1,1,5],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "CatBoost": {"iterations":[400,700,1000,1400,1800],"learning_rate":[0.003,0.005,0.01,0.02,0.03,0.05],"depth":[2,3,5,7,9],"l2_leaf_reg":[1,3,5,10,20],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
}
SCALE_SENSITIVE_MODELS = {"LogisticRegression", "LinearSVC"}

def make_model(model_name, params, seed):
    p={k:v for k,v in params.items() if k!="positive_weight_multiplier"}
    if model_name=="LogisticRegression":
        return LogisticRegression(C=float(p["C"]), penalty="l2", solver="lbfgs", max_iter=5000, random_state=seed, n_jobs=N_JOBS)
    if model_name=="LinearSVC":
        return LinearSVC(C=float(p["C"]), max_iter=10000, random_state=seed)
    if model_name=="RandomForest":
        return RandomForestClassifier(n_estimators=int(p["n_estimators"]), max_depth=p["max_depth"], min_samples_leaf=int(p["min_samples_leaf"]), max_features=p["max_features"], random_state=seed, n_jobs=N_JOBS)
    if model_name=="AdaBoost":
        return AdaBoostClassifier(n_estimators=int(p["n_estimators"]), learning_rate=float(p["learning_rate"]), random_state=seed)
    if model_name=="LightGBM":
        return LGBMClassifier(objective="binary", boosting_type="gbdt", metric="average_precision", random_state=seed, n_jobs=N_JOBS, verbosity=-1, force_col_wise=True, **p)
    if model_name=="XGBoost":
        return XGBClassifier(objective="binary:logistic", eval_metric="aucpr", random_state=seed, n_jobs=N_JOBS, tree_method="hist", **p)
    if model_name=="HistGradientBoosting":
        return HistGradientBoostingClassifier(random_state=seed, **p)
    if model_name=="CatBoost":
        return CatBoostClassifier(loss_function="Logloss", eval_metric="PRAUC", random_seed=seed, verbose=False, allow_writing_files=False, thread_count=-1, **p)
    raise ValueError(model_name)

def fit_screening_pipeline(model_name, X, y, features, params, seed):
    cont,binf,ordf,nomf=feature_types(features)
    pre=Step2Preprocessor(cont,binf,ordf,nomf,PREFERRED_NOMINAL_LEVELS,scale_continuous=model_name in SCALE_SENSITIVE_MODELS)
    Xt=pre.fit_transform(X[features])
    model=make_model(model_name, params, seed)
    w=make_weights(y, params.get("positive_weight_multiplier", 1.0))
    try:
        model.fit(Xt, y, sample_weight=w)
    except TypeError:
        model.fit(Xt, y)
    return pre, model

def predict_raw(pre, model, X, features):
    Xt=pre.transform(X[features])
    if hasattr(model, "predict_proba"):
        return model.predict_proba(Xt)[:,1]
    if hasattr(model, "decision_function"):
        return model.decision_function(Xt)
    return model.predict(Xt).astype(float)

def calibrated_predictions(raw_train_or_test, raw_cal, y_cal):
    if CALIBRATION_METHOD == "isotonic":
        cal=IsotonicRegression(out_of_bounds="clip")
        cal.fit(raw_cal, y_cal)
        return cal.predict(raw_train_or_test)
    return raw_train_or_test

def tune_classifier(model_name, X_train, y_train, groups_train, features, folds, model_key, seed):
    space=MODEL_SEARCH_SPACES[model_name]
    n_iter=min(N_RANDOM_COMBINATIONS, math.prod([len(v) for v in space.values()]))
    candidates=list(ParameterSampler(space, n_iter=n_iter, random_state=seed))
    cand_rows=[]; fold_rows=[]
    for cid, params in enumerate(candidates, 1):
        aps=[]; aucs=[]; t0=time.time()
        for fid,(tr,va) in enumerate(folds,1):
            Xtr=X_train.iloc[tr].reset_index(drop=True); ytr=y_train[tr]
            Xva=X_train.iloc[va].reset_index(drop=True); yva=y_train[va]
            pre,model=fit_screening_pipeline(model_name,Xtr,ytr,features,params,seed+cid*1000+fid)
            s=predict_raw(pre,model,Xva,features)
            aps.append(safe_average_precision(yva,s)); aucs.append(safe_roc_auc(yva,s))
            fold_rows.append({"model_key":model_key,"classifier":model_name,"candidate_id":cid,"fold":fid,"fold_AP":aps[-1],"fold_ROC_AUC":aucs[-1],"scale_continuous":model_name in SCALE_SENSITIVE_MODELS,**params})
        row={"model_key":model_key,"classifier":model_name,"candidate_id":cid,"cv_folds":len(folds),"cv_AP_mean":float(np.nanmean(aps)),"cv_AP_SD":float(np.nanstd(aps,ddof=1)),"cv_ROC_AUC_mean":float(np.nanmean(aucs)),"cv_ROC_AUC_SD":float(np.nanstd(aucs,ddof=1)),"scale_continuous":model_name in SCALE_SENSITIVE_MODELS,"elapsed_seconds":float(time.time()-t0),**params}
        cand_rows.append(row)
        print(f"{model_key} | {model_name} | candidate {cid:03d}/{len(candidates)} | CV AP={row['cv_AP_mean']:.5f} | AUC={row['cv_ROC_AUC_mean']:.5f}")
    return pd.DataFrame(cand_rows).sort_values("cv_AP_mean",ascending=False).reset_index(drop=True), pd.DataFrame(fold_rows)

def fit_final_screening_model(model_name, Xtr, ytr, Xcal, ycal, Xte, features, params, seed):
    pre,model=fit_screening_pipeline(model_name,Xtr,ytr,features,params,seed)
    raw_cal=predict_raw(pre,model,Xcal,features)
    raw_test=predict_raw(pre,model,Xte,features)
    p_test=calibrated_predictions(raw_test, raw_cal, ycal)
    return raw_test, p_test

# ============================================================
# 5) Step 2 data preparation and cohort loop
# ============================================================
def add_dynamic_odi_features(df):
    """Derive Step 2 ODI variables using the manuscript definitions."""
    out = df.copy()
    required = ["preop_ODI", "postop_ODI", DAYS_BETWEEN_PROM_COL]
    missing = [c for c in required if c not in out.columns]
    if missing:
        raise ValueError(f"Missing columns required to derive dynamic ODI features: {missing}")
    preop = pd.to_numeric(out["preop_ODI"].map(clean_scalar), errors="coerce")
    postop = pd.to_numeric(out["postop_ODI"].map(clean_scalar), errors="coerce")
    days = pd.to_numeric(out[DAYS_BETWEEN_PROM_COL].map(clean_scalar), errors="coerce")
    out["ODI_change"] = postop - preop
    out[PROM_CHANGE_RATE_COL] = np.where(days > 0, (postop - preop) / days, np.nan)
    improvement = preop - postop
    relative = pd.Series(np.nan, index=out.index, dtype=float)
    valid = preop.notna() & postop.notna() & preop.gt(0)
    relative.loc[valid] = ((improvement.loc[valid] / preop.loc[valid]) >= RELATIVE_ODI_MCID_CUTOFF).astype(float)
    zero_base = preop.eq(0) & postop.notna()
    relative.loc[zero_base] = 0.0
    out[RELATIVE_ODI_MCID_COL] = relative
    audit = pd.DataFrame([
        {"item":"input_rows","value":int(len(out))},
        {"item":"valid_preop_postop_positive_days","value":int((preop.notna() & postop.notna() & days.gt(0)).sum())},
        {"item":"change_rate_definition","value":"(postop_ODI - preop_ODI) / days_between_PROMs"},
        {"item":"relative_MCID_definition","value":f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_CUTOFF}; preop_ODI=0 coded as 0"},
        {"item":"input_ODI_MCID_binary_present_but_not_used","value":bool(INPUT_ODI_MCID_COL in out.columns)},
    ])
    return out, audit

def split_audit(df,target_col):
    return df.groupby("split").agg(n=(target_col,"size"),events=(target_col,"sum"),patients=(GROUP_COL,"nunique")).reset_index().assign(prevalence=lambda d:d.events/d.n)

def prepare_step2_dataframe():
    if not os.path.exists(INPUT_CSV):
        raise FileNotFoundError(f"Input file not found: {INPUT_CSV}")
    df=pd.read_csv(INPUT_CSV,low_memory=False); df.columns=[str(c).strip() for c in df.columns]
    df, dynamic_audit = add_dynamic_odi_features(df)
    required = DYNAMIC_PROM_FEATURES + [TARGET_COL, GROUP_COL]
    missing=[c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Step 2 missing required columns: {missing}")
    preop = pd.to_numeric(df["preop_ODI"].map(clean_scalar), errors="coerce")
    postop = pd.to_numeric(df["postop_ODI"].map(clean_scalar), errors="coerce")
    days = pd.to_numeric(df[DAYS_BETWEEN_PROM_COL].map(clean_scalar), errors="coerce")
    eligible = preop.notna() & postop.notna() & days.gt(0)
    keep_extra=[c for c in ["InstitutionName","InstitutionNPI1"] if c in df.columns]
    work=df.loc[eligible, required+keep_extra].copy()
    before=len(work)
    work[TARGET_COL]=work[TARGET_COL].map(to_binary_target)
    work=work[work[TARGET_COL].isin([0.0,1.0]) & work[GROUP_COL].notna()].copy()
    work[TARGET_COL]=work[TARGET_COL].astype(int)
    metadata=pd.DataFrame([{
        "input_csv":INPUT_CSV,
        "rows_input":int(len(df)),
        "rows_after_PROM_eligibility":int(eligible.sum()),
        "rows_after_target_and_group_cleaning":int(len(work)),
        "dropped_after_PROM_eligibility_for_target_or_group":int(before-len(work)),
        "events":int(work[TARGET_COL].sum()),
        "prevalence":float(work[TARGET_COL].mean()) if len(work) else np.nan,
        "target_col":TARGET_COL,
        "group_col":GROUP_COL,
    }])
    return work.reset_index(drop=True), dynamic_audit, metadata

def run_step2_screening(seed):
    df,dynamic_audit,metadata=prepare_step2_dataframe()
    tr,ca,te=patient_split(df,TARGET_COL,seed)
    df["split"]=np.where(tr,"train",np.where(ca,"calibration","test"))
    if df.groupby(GROUP_COL).split.nunique().max()>1:
        raise RuntimeError("Patient leakage across splits.")
    print("\n"+"#"*100+"\nSTEP 2 DYNAMIC ODI CLASSIFIER SCREENING\n"+"#"*100)
    print(f"Input: {INPUT_CSV} | target: {TARGET_COL} | rows: {len(df)} | events: {int(df[TARGET_COL].sum())}")
    print(split_audit(df,TARGET_COL).to_string(index=False))
    train=df[df.split.eq("train")].reset_index(drop=True)
    cal=df[df.split.eq("calibration")].reset_index(drop=True)
    test=df[df.split.eq("test")].reset_index(drop=True)
    folds=cv_splits(train[TARGET_COL].astype(int).to_numpy(), train[GROUP_COL].to_numpy(), seed+10)
    rows=[]; candidates_all=[]; folds_all=[]
    feature_sets=[("baseline_only",BASELINE_ONLY_FEATURES,"not_applicable"),("dynamic_PROM_expanded",DYNAMIC_PROM_FEATURES,"dynamic_ODI_features")]
    for model_type, features, prom_label in feature_sets:
        for model_name in MODEL_SEARCH_SPACES:
            model_key=f"Step2_{model_type}_{model_name}"
            cand,fold=tune_classifier(model_name, train[features].copy(), train[TARGET_COL].astype(int).to_numpy(), train[GROUP_COL].to_numpy(), features, folds, model_key, seed+1000+len(rows))
            best=cand.iloc[0].to_dict()
            params={k:best[k] for k in MODEL_SEARCH_SPACES[model_name] if k in best}
            raw_test,p_test=fit_final_screening_model(model_name, train[features].copy(), train[TARGET_COL].astype(int).to_numpy(), cal[features].copy(), cal[TARGET_COL].astype(int).to_numpy(), test[features].copy(), features, params, seed+2000+len(rows))
            ev=eval_preds(test[TARGET_COL].astype(int).to_numpy(), p_test, prefix="Test_")
            raw_ev={"RawTest_AP":safe_average_precision(test[TARGET_COL].astype(int).to_numpy(),raw_test),"RawTest_ROC_AUC":safe_roc_auc(test[TARGET_COL].astype(int).to_numpy(),raw_test)}
            rows.append({"analysis":"Step2_dynamic_ODI","cohort":"Step2_ODI","model_type":model_type,"classifier":model_name,"model_key":model_key,"n_features":len(features),"dynamic_PROM_features":prom_label,"best_candidate_id":int(best["candidate_id"]),"cv_AP_mean":float(best["cv_AP_mean"]),"cv_AP_SD":float(best["cv_AP_SD"]),"cv_ROC_AUC_mean":float(best["cv_ROC_AUC_mean"]),"cv_ROC_AUC_SD":float(best["cv_ROC_AUC_SD"]),"scale_continuous":model_name in SCALE_SENSITIVE_MODELS,**ev,**raw_ev})
            candidates_all.append(cand.assign(analysis="Step2_dynamic_ODI", cohort="Step2_ODI", model_type=model_type))
            folds_all.append(fold.assign(analysis="Step2_dynamic_ODI", cohort="Step2_ODI", model_type=model_type))
    summary=pd.DataFrame(rows)
    candidates=pd.concat(candidates_all,ignore_index=True)
    folds=pd.concat(folds_all,ignore_index=True)
    splits=split_audit(df,TARGET_COL).assign(cohort="Step2_ODI")
    split_assignment=df[[GROUP_COL,"split",TARGET_COL]].drop_duplicates().sort_values([GROUP_COL,"split"])
    return summary,candidates,folds,splits,split_assignment,dynamic_audit,metadata

def style_excel(writer):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        for s in writer.sheets:
            ws=writer.sheets[s]; ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for cell in ws[1]: cell.font=Font(bold=True); cell.fill=PatternFill(start_color="D9EAF7",end_color="D9EAF7",fill_type="solid"); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            for col in ws.columns:
                ml=max([len(str(c.value)) for c in col if c.value is not None]+[12]); ws.column_dimensions[col[0].column_letter].width=min(max(ml+2,12),70)
    except Exception:
        pass

def methods_table():
    return pd.DataFrame([
        {"item":"Classifier-screening design","rationale":"Eight prespecified classifiers are compared in the Step 2 dynamic ODI cohort using paired baseline-only and dynamic PROM-expanded feature sets."},
        {"item":"Paired split/folds","rationale":"Baseline-only and dynamic PROM-expanded models use identical patient-level train/calibration/test splits and identical group-aware CV folds."},
        {"item":"Tuning metric","rationale":"Mean cross-validated average precision is used for classifier/hyperparameter selection because delayed reoperation is a rare-event outcome."},
        {"item":"Dynamic ODI features","rationale":"The expanded model adds preoperative ODI, postoperative ODI, ODI change, ODI change rate, relative MCID status, and postoperative PROM timing to the same 35 baseline variables."},
        {"item":"Final architecture selection","rationale":"The selected classifier is the one with the highest mean CV AP within the training split; held-out test performance is reported but not used for selection."},
    ])

def main():
    t0=time.time()
    summary,candidates,folds,splits,split_assignment,dynamic_audit,metadata=run_step2_screening(RANDOM_STATE)
    selected=(summary.sort_values(["model_type","cv_AP_mean"],ascending=[True,False]).groupby(["model_type"],as_index=False).head(1).reset_index(drop=True))
    classifier_rankings=(summary.groupby("classifier",as_index=False).agg(mean_cv_AP=("cv_AP_mean","mean"),median_cv_AP=("cv_AP_mean","median"),mean_test_AP=("Test_AP","mean"),n_models=("model_key","count")).sort_values("mean_cv_AP",ascending=False))
    xlsx=os.path.join(OUTPUT_DIR,"Step2_ClassifierScreening_All8_summary.xlsx")
    with pd.ExcelWriter(xlsx,engine="openpyxl") as w:
        methods_table().to_excel(w,"methods_rationale",index=False)
        metadata.to_excel(w,"cohort_metadata",index=False)
        dynamic_audit.to_excel(w,"dynamic_PROM_audit",index=False)
        summary.to_excel(w,"all8_model_performance",index=False)
        selected.to_excel(w,"selected_by_highest_CV_AP",index=False)
        classifier_rankings.to_excel(w,"classifier_rankings",index=False)
        candidates.to_excel(w,"cv_candidates_all8",index=False)
        folds.to_excel(w,"cv_fold_metrics_all8",index=False)
        splits.to_excel(w,"split_audit",index=False)
        split_assignment.to_excel(w,"split_assignment",index=False)
        pd.DataFrame([{"Parameter":k,"Value":v} for k,v in {"BASE_DIR":BASE_DIR,"INPUT_CSV":INPUT_CSV,"OUTPUT_DIR":OUTPUT_DIR,"RANDOM_STATE":RANDOM_STATE,"TEST_FRACTION":TEST_FRACTION,"CALIBRATION_FRACTION_OF_REMAINING":CALIBRATION_FRACTION_OF_REMAINING,"N_CV_FOLDS":N_CV_FOLDS,"N_RANDOM_COMBINATIONS":N_RANDOM_COMBINATIONS,"CALIBRATION_METHOD":CALIBRATION_METHOD,"python_version":platform.python_version(),"lightgbm_version":lgb.__version__}.items()]).to_excel(w,"run_config",index=False)
        style_excel(w)
    summary.to_csv(os.path.join(OUTPUT_DIR,"all8_model_performance.csv"),index=False)
    selected.to_csv(os.path.join(OUTPUT_DIR,"selected_by_highest_CV_AP.csv"),index=False)
    classifier_rankings.to_csv(os.path.join(OUTPUT_DIR,"classifier_rankings.csv"),index=False)
    candidates.to_csv(os.path.join(OUTPUT_DIR,"cv_candidates_all8.csv"),index=False)
    folds.to_csv(os.path.join(OUTPUT_DIR,"cv_fold_metrics_all8.csv"),index=False)
    split_assignment.to_csv(os.path.join(OUTPUT_DIR,"split_assignment_Step2_ODI.csv"),index=False)
    with open(os.path.join(OUTPUT_DIR,"run_manifest.json"),"w") as f:
        json.dump(json_native({"design":"Step 2 classifier screening: 1 dynamic ODI cohort x 2 feature sets x 8 classifiers","selection_rule":"highest mean group-aware CV average precision within training split","feature_sets":{"baseline_only":BASELINE_ONLY_FEATURES,"dynamic_PROM_expanded":DYNAMIC_PROM_FEATURES},"scale_sensitive_models":sorted(SCALE_SENSITIVE_MODELS),"runtime_minutes":float((time.time()-t0)/60),"summary_xlsx":xlsx}),f,indent=2,sort_keys=True)
    print("\n"+"="*100)
    print("STEP 2 classifier screening completed")
    print("Summary Excel:", xlsx)
    print("Selected classifier table:", os.path.join(OUTPUT_DIR,"selected_by_highest_CV_AP.csv"))
    print("="*100)

if __name__ == "__main__":
    main()

# %% [markdown] Cell 3
# #**Step2_FinalLightGBM_SHAP**

# %% Cell 4
# -*- coding: utf-8 -*-
"""
Step 2 dynamic ODI final LightGBM / SHAP analysis
========================================================

Input
-----
/content/Step 2_ODI_Cohort.csv

Target
------
final_reop_step2
    1 = reoperation from postoperative day 91 through day 365
    0 = no reoperation from postoperative day 91 through day 365

Design
------
This script runs the final paired baseline-only and dynamic ODI-expanded LightGBM
models for delayed lumbar reoperation prediction after Step 2 classifier screening has
identified LightGBM as the locked architecture. The baseline-only model
uses the 35 baseline variables used in Step 1. The dynamic ODI-expanded
model includes the same baseline variables plus preoperative ODI,
postoperative ODI, ODI change, ODI change rate, relative MCID status,
and timing of postoperative ODI assessment. Paired models use identical
patient-level train/calibration/test splits and identical group-aware
cross-validation fold construction. Hyperparameter tuning is performed
exclusively within the training split using cross-validated average
precision as the primary selection metric. Probability calibration and
threshold selection are performed only on the calibration split. The
held-out test set is reserved until the model-development pipeline is
locked.
"""

# ============================================================
# 0) Install/import
# ============================================================

import os
import sys
import json
import math
import time
import zipfile
import platform
import subprocess
import warnings
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import lightgbm as lgb
    from lightgbm import LGBMClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lightgbm"])
    import lightgbm as lgb
    from lightgbm import LGBMClassifier

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import joblib
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "joblib"])
    import joblib

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder
from sklearn.model_selection import StratifiedShuffleSplit, StratifiedGroupKFold, ParameterSampler
from sklearn.metrics import (
    average_precision_score,
    roc_auc_score,
    brier_score_loss,
    precision_recall_curve,
    roc_curve,
    confusion_matrix,
    f1_score,
)
from sklearn.isotonic import IsotonicRegression

import matplotlib.pyplot as plt

warnings.filterwarnings("ignore", category=UserWarning)


# ============================================================
# 1) User configuration
# ============================================================

BASE_DIR = "/content"
INPUT_CSV = os.path.join(BASE_DIR, "Step 2_ODI_Cohort.csv")

OUTPUT_DIR = os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(PLOT_DIR, exist_ok=True)

TARGET_COL = "final_reop_step2"
GROUP_COL = "PersonKey"

RANDOM_STATE = 20260524

# Split design: approximately 64% train, 16% calibration, 20% test
TEST_FRACTION = 0.20
CALIBRATION_FRACTION_OF_REMAINING = 0.20

# CV and tuning
N_CV_FOLDS = 5
N_RANDOM_COMBINATIONS = 300

# Calibration and threshold
CALIBRATION_METHOD = "isotonic"
THRESHOLD_STRATEGY = "max_f1"  # options: max_f1, youden, fixed, top_percent
FIXED_THRESHOLD = 0.50
THRESHOLD_TOP_PERCENT = 0.05

# Test-set uncertainty
N_BOOTSTRAPS = 2000
ECE_N_BINS = 10

# LightGBM runtime
N_JOBS = -1

# Early stopping is used only inside cross-validation to estimate a stable
# boosting length. The locked selected model is refit on the full training split
# using the median best iteration from CV folds.
USE_EARLY_STOPPING_IN_CV = True
EARLY_STOPPING_ROUNDS = 100
MIN_FINAL_N_ESTIMATORS = 50

# SHAP and threshold analysis
SHAP_MAX_EXPLAIN_ROWS = None
SHAP_BAR_MAX_DISPLAY = 30
SHAP_BEESWARM_MAX_DISPLAY = 30
SHAP_THRESHOLD_TOP_N_NUMERIC = 20
SHAP_THRESHOLD_MAX_BINS = 35

# Colors matched to the standard SHAP beeswarm palette:
# blue = lower/protective contribution, pink = higher/risk-increasing contribution.
SHAP_BEESWARM_BLUE = "#008BFB"
SHAP_BEESWARM_PINK = "#FF0051"

# Primary model selection is by cross-validated AP only.
# Test-set metrics are reported but are not used to choose the selected model.
PRIMARY_MODEL_SELECTION_RULE = "highest mean group-aware CV AP on training split; test metrics are diagnostic only and never used for model selection"

# Output archive options
AUTO_DOWNLOAD_ZIP = False
CREATE_COLAB_DOWNLOAD_LINK = True
ZIP_COMPRESSION_LEVEL = 1


# ============================================================
# 2) Step 2 feature set
# ============================================================

BASE_FEATURES = [
    "finaldx_degenerative",
    "finaldx_radicular",
    "finaldx_stenosis",
    "finaldx_deformity_instability",
    "finaldx_other_diagnosis",
    "age",
    "sex",
    "race",
    "ethnicity",
    "cancer_status",
    "chronic_pulmonary_disease",
    "congestive_heart_failure",
    "connective_tissue_rheumatic_disease",
    "diabetes_status",
    "myocardial_infarction",
    "renal_disease",
    "institution_type",
    "institution_size",
    "institution_region",
    "asa",
    "bmi",
    "payer_status",
    "alif_llif",
    "corpectomy",
    "discectomy",
    "foraminotomy",
    "instrumentation",
    "laminectomy_posterior_decompression",
    "pelvic_fixation",
    "plf",
    "tlif_plif",
    "other_lumbar_procedure",
    "number_operated_levels",
    "operative_region_extent",
    "PatTobaccoUse",
]

# Dynamic ODI features are derived explicitly from preop_ODI, postop_ODI, and days_between_PROMs below.
# The input input column ODI_MCID_binary, if present, is not used as a feature;
# it is retained only for audit and reproducibility checks.
RELATIVE_ODI_MCID_CUTOFF = 0.30
INPUT_ODI_MCID_COL = "ODI_MCID_binary"
PROM_CHANGE_RATE_COL = "ODI_change_rate"
RELATIVE_ODI_MCID_COL = "ODI_relative_MCID_binary"
DAYS_BETWEEN_PROM_COL = "days_between_PROMs"

STEP2_ODI_FEATURES = [
    "preop_ODI",
    "postop_ODI",
    "ODI_change",
    PROM_CHANGE_RATE_COL,
    RELATIVE_ODI_MCID_COL,
    "postop_ODI_day",
]

FEATURES = BASE_FEATURES + STEP2_ODI_FEATURES

# Excluded deliberately
EXCLUDED_FEATURES = {
    "reop",
    "reoptime",
    "final_reop",
    "final_reop_step2",
    "death_within_90d",
    "death_within_180d",
    "death_within_365d",
    "death_after_index_surgery",
    "death_before_or_on_index_surgery",
    "PersonDeathDate",
    "days_to_death_from_index_surgery",
    "removal_hardware",
    "any_arthroplasty",
    "final_diagnosis_complexity",
    "procedure_complexity_score",
}

bad_features = sorted(set(FEATURES) & EXCLUDED_FEATURES)
if bad_features:
    raise ValueError(f"Excluded/leakage-prone features were accidentally included: {bad_features}")

CONTINUOUS_FEATURES = [
    "age",
    "bmi",
    "preop_ODI",
    "postop_ODI",
    "ODI_change",
    PROM_CHANGE_RATE_COL,
    "postop_ODI_day",
]

BINARY_FEATURES = [
    "finaldx_degenerative",
    "finaldx_radicular",
    "finaldx_stenosis",
    "finaldx_deformity_instability",
    "finaldx_other_diagnosis",
    "sex",
    "ethnicity",
    "cancer_status",
    "chronic_pulmonary_disease",
    "congestive_heart_failure",
    "connective_tissue_rheumatic_disease",
    "myocardial_infarction",
    "renal_disease",
    "institution_type",
    "alif_llif",
    "corpectomy",
    "discectomy",
    "foraminotomy",
    "instrumentation",
    "laminectomy_posterior_decompression",
    "pelvic_fixation",
    "plf",
    "tlif_plif",
    "other_lumbar_procedure",
    "operative_region_extent",
    RELATIVE_ODI_MCID_COL,
]

ORDINAL_FEATURES = [
    "diabetes_status",
    "institution_size",
    "asa",
    "number_operated_levels",
]

NOMINAL_FEATURES = [
    "race",
    "institution_region",
    "payer_status",
    "PatTobaccoUse",
]


# ============================================================
# 3) Hyperparameter search space
# ============================================================

# positive_weight_multiplier is converted to sample weights:
# positive_weight = n_negative / n_positive * positive_weight_multiplier
LGBM_SEARCH_SPACE = {
    # Same LightGBM search space used in Step 1 final LightGBM analysis.
    "n_estimators": [400, 700, 1000, 1400, 1800, 2200, 2600],
    "learning_rate": [0.003, 0.005, 0.008, 0.01, 0.02, 0.03, 0.05],
    "num_leaves": [7, 15, 31, 63, 127],
    "max_depth": [-1, 2, 3, 5, 7, 9],
    "min_child_samples": [10, 20, 50, 100, 200, 400],
    "subsample": [0.60, 0.75, 0.90, 1.00],
    "subsample_freq": [0, 1, 2],
    "colsample_bytree": [0.60, 0.75, 0.90, 1.00],
    "reg_alpha": [0.0, 0.001, 0.01, 0.05, 0.10, 0.50, 1.00, 2.00],
    "reg_lambda": [0.0, 0.001, 0.01, 0.05, 0.10, 0.50, 1.00, 2.00, 5.00],
    "min_split_gain": [0.0, 0.001, 0.005, 0.01, 0.05, 0.10],
    "max_bin": [63, 127, 255],
    "positive_weight_multiplier": [0.25, 0.50, 0.75, 1.00, 1.50, 2.00, 3.00, 4.00, 6.00, 8.00],
}

LGBM_INT_PARAMS = {
    "n_estimators",
    "num_leaves",
    "max_depth",
    "min_child_samples",
    "subsample_freq",
    "max_bin",
}

LGBM_FLOAT_PARAMS = {
    "learning_rate",
    "subsample",
    "colsample_bytree",
    "reg_alpha",
    "reg_lambda",
    "min_split_gain",
    "positive_weight_multiplier",
}


def sanitize_lgbm_params(params: Dict[str, Any]) -> Dict[str, Any]:
    """Convert LightGBM params to native Python types, especially integer params."""
    clean: Dict[str, Any] = {}

    for k, v in params.items():
        if k in LGBM_INT_PARAMS:
            clean[k] = int(round(float(v)))
        elif k in LGBM_FLOAT_PARAMS:
            clean[k] = float(v)
        else:
            if isinstance(v, (np.integer,)):
                clean[k] = int(v)
            elif isinstance(v, (np.floating,)):
                clean[k] = float(v)
            else:
                clean[k] = v

    return clean


# ============================================================
# 4) General helpers
# ============================================================

MISSING_STRINGS = {
    "",
    " ",
    "na",
    "n/a",
    "nan",
    "none",
    "null",
    ".",
    "missing",
    "<na>",
}


def json_native(obj: Any) -> Any:
    """Recursively convert numpy/pandas objects to JSON-serializable Python objects."""
    if isinstance(obj, dict):
        return {str(k): json_native(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [json_native(v) for v in obj]
    if isinstance(obj, tuple):
        return [json_native(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return float(obj)
    if isinstance(obj, (np.ndarray,)):
        return obj.tolist()
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    return obj


def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        if s.lower() in MISSING_STRINGS:
            return np.nan
        return s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def to_binary_target(x: Any) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if sx in {"1", "1.0", "yes", "y", "true", "t"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f"}:
        return 0.0
    try:
        v = float(sx)
        if v in (0.0, 1.0):
            return float(v)
    except Exception:
        pass
    return np.nan


def count_pct(n: int, denom: int, digits: int = 2) -> str:
    if denom == 0:
        return f"{int(n):,} (NA)"
    return f"{int(n):,} ({100 * n / denom:.{digits}f}%)"


def safe_average_precision(y_true: np.ndarray, y_prob: np.ndarray) -> float:
    if len(np.unique(y_true)) < 2:
        return np.nan
    return float(average_precision_score(y_true, y_prob))


def safe_roc_auc(y_true: np.ndarray, y_prob: np.ndarray) -> float:
    if len(np.unique(y_true)) < 2:
        return np.nan
    return float(roc_auc_score(y_true, y_prob))


def expected_calibration_error(y_true: np.ndarray, y_prob: np.ndarray, n_bins: int = 10) -> float:
    y_true = np.asarray(y_true).astype(int)
    y_prob = np.asarray(y_prob).astype(float)

    bins = np.linspace(0.0, 1.0, n_bins + 1)
    ece = 0.0
    n = len(y_true)

    if n == 0:
        return np.nan

    for i in range(n_bins):
        lo, hi = bins[i], bins[i + 1]
        if i == n_bins - 1:
            mask = (y_prob >= lo) & (y_prob <= hi)
        else:
            mask = (y_prob >= lo) & (y_prob < hi)

        if np.any(mask):
            bin_conf = float(np.mean(y_prob[mask]))
            bin_acc = float(np.mean(y_true[mask]))
            ece += (np.sum(mask) / n) * abs(bin_acc - bin_conf)

    return float(ece)


def classification_metrics_at_threshold(
    y_true: np.ndarray,
    y_prob: np.ndarray,
    threshold: float,
) -> Dict[str, Any]:
    y_true = np.asarray(y_true).astype(int)
    y_pred = (np.asarray(y_prob) >= threshold).astype(int)

    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()

    sensitivity = tp / (tp + fn) if (tp + fn) > 0 else np.nan
    specificity = tn / (tn + fp) if (tn + fp) > 0 else np.nan
    ppv = tp / (tp + fp) if (tp + fp) > 0 else np.nan
    npv = tn / (tn + fn) if (tn + fn) > 0 else np.nan
    f1 = f1_score(y_true, y_pred, zero_division=0)

    return {
        "threshold": float(threshold),
        "F1": float(f1),
        "Sensitivity": float(sensitivity),
        "Specificity": float(specificity),
        "PPV": float(ppv),
        "NPV": float(npv),
        "TP": int(tp),
        "FP": int(fp),
        "TN": int(tn),
        "FN": int(fn),
        "Predicted_positive_rate": float(np.mean(y_pred)),
    }


def top_percent_metrics(
    y_true: np.ndarray,
    y_prob: np.ndarray,
    top_fraction: float = 0.05,
) -> Dict[str, Any]:
    y_true = np.asarray(y_true).astype(int)
    y_prob = np.asarray(y_prob).astype(float)

    n = len(y_true)
    k = max(1, int(math.ceil(n * top_fraction)))

    order = np.argsort(-y_prob)
    top_idx = order[:k]

    prevalence = float(np.mean(y_true)) if n > 0 else np.nan
    top_event_rate = float(np.mean(y_true[top_idx])) if k > 0 else np.nan
    lift = top_event_rate / prevalence if prevalence > 0 else np.nan
    captured = float(np.sum(y_true[top_idx]) / np.sum(y_true)) if np.sum(y_true) > 0 else np.nan

    return {
        "Top_5pct_n": int(k),
        "Top_5pct_event_rate": top_event_rate,
        "Top_5pct_lift": float(lift),
        "Top_5pct_captured_events": captured,
    }


def select_threshold(
    y_true: np.ndarray,
    y_prob: np.ndarray,
    strategy: str = "max_f1",
    fixed_threshold: float = 0.50,
    top_percent: float = 0.05,
) -> Tuple[float, Dict[str, Any]]:
    y_true = np.asarray(y_true).astype(int)
    y_prob = np.asarray(y_prob).astype(float)

    if strategy == "fixed":
        threshold = float(fixed_threshold)

    elif strategy == "top_percent":
        threshold = float(np.quantile(y_prob, 1.0 - top_percent))

    elif strategy == "youden":
        fpr, tpr, thresholds = roc_curve(y_true, y_prob)
        youden = tpr - fpr
        threshold = float(thresholds[int(np.nanargmax(youden))])

    elif strategy == "max_f1":
        precision, recall, thresholds = precision_recall_curve(y_true, y_prob)
        if len(thresholds) == 0:
            threshold = 0.50
        else:
            precision = precision[:-1]
            recall = recall[:-1]
            f1_values = 2 * precision * recall / np.maximum(precision + recall, 1e-12)
            threshold = float(thresholds[int(np.nanargmax(f1_values))])

    else:
        raise ValueError(f"Unknown threshold strategy: {strategy}")

    return threshold, classification_metrics_at_threshold(y_true, y_prob, threshold)


def make_positive_sample_weights(y: np.ndarray, multiplier: float) -> np.ndarray:
    y = np.asarray(y).astype(int)
    n_pos = int(np.sum(y == 1))
    n_neg = int(np.sum(y == 0))

    if n_pos == 0:
        raise ValueError("No positive events in training fold.")

    base_weight = n_neg / max(n_pos, 1)
    pos_weight = base_weight * float(multiplier)

    return np.where(y == 1, pos_weight, 1.0).astype(float)


def actual_positive_weight(y: np.ndarray, multiplier: float) -> float:
    y = np.asarray(y).astype(int)
    n_pos = int(np.sum(y == 1))
    n_neg = int(np.sum(y == 0))
    return float((n_neg / max(n_pos, 1)) * multiplier)


# ============================================================
# 5) Preprocessor
# ============================================================

BINARY_MAPS = {
    "sex": {
        "female": 0,
        "f": 0,
        "male": 1,
        "m": 1,
    },
    "ethnicity": {
        "non-hispanic": 0,
        "non hispanic": 0,
        "hispanic": 1,
    },
    "cancer_status": {
        "no cancer": 0,
        "no": 0,
        "none": 0,
        "any cancer": 1,
        "yes": 1,
        "cancer": 1,
    },
    "institution_type": {
        "hospital": 0,
        "non-hospital": 1,
        "non hospital": 1,
        "nonhospital": 1,
    },
    "operative_region_extent": {
        "lumbar only": 0,
        "extended_region_involvement": 1,
        "extended region involvement": 1,
        "extended": 1,
    },
}

ORDINAL_MAPS = {
    "diabetes_status": {
        "no": 0,
        "none": 0,
        "0": 0,
        "without comp": 1,
        "without complication": 1,
        "without complications": 1,
        "1": 1,
        "with comp": 2,
        "with complication": 2,
        "with complications": 2,
        "2": 2,
    },
    "institution_size": {
        "between 1-99 beds": 0,
        "1-99": 0,
        "between 100-399 beds": 1,
        "100-399": 1,
        ">= 400 beds": 2,
        ">=400 beds": 2,
        ">=400": 2,
        ">= 400": 2,
    },
    "asa": {
        "1": 1,
        "i": 1,
        "2": 2,
        "ii": 2,
        "3": 3,
        "iii": 3,
        "4": 4,
        "iv": 4,
        ">=4": 4,
        ">= 4": 4,
        "5": 4,
        "v": 4,
    },
    "number_operated_levels": {
        "0": 0,
        "1": 1,
        "2": 2,
        "3": 3,
        "4": 4,
        ">=4": 4,
        ">= 4": 4,
        "5": 4,
        "6": 4,
        "7": 4,
        "8": 4,
        "9": 4,
        "10": 4,
    },
}

PREFERRED_NOMINAL_LEVELS = {
    "race": ["White", "Black", "Other"],
    "institution_region": ["South", "North East", "West", "Midwest"],
    "payer_status": ["Medicare", "Commercial/Private", "Other", "Medicaid/Public/Government"],
    "PatTobaccoUse": ["Unknown/Not reported/Multiple", "Never", "Former", "Current"],
}

FEATURE_LABELS = {
    "finaldx_degenerative": "Degenerative diagnosis",
    "finaldx_radicular": "Radiculopathy diagnosis",
    "finaldx_stenosis": "Spinal stenosis diagnosis",
    "finaldx_deformity_instability": "Deformity or instability diagnosis",
    "finaldx_other_diagnosis": "Other lumbar diagnosis",
    "age": "Age",
    "sex": "Sex",
    "race": "Race",
    "ethnicity": "Ethnicity",
    "cancer_status": "Cancer status",
    "chronic_pulmonary_disease": "Chronic pulmonary disease",
    "congestive_heart_failure": "Congestive heart failure",
    "connective_tissue_rheumatic_disease": "Connective tissue/rheumatic disease",
    "diabetes_status": "Diabetes status",
    "myocardial_infarction": "Myocardial infarction",
    "renal_disease": "Renal disease",
    "institution_type": "Institution type",
    "institution_size": "Institution size",
    "institution_region": "Institution region",
    "asa": "ASA physical status",
    "bmi": "Body mass index",
    "payer_status": "Primary payer",
    "alif_llif": "Anterior/lateral lumbar interbody fusion",
    "corpectomy": "Corpectomy",
    "discectomy": "Discectomy",
    "foraminotomy": "Foraminotomy",
    "instrumentation": "Instrumentation",
    "laminectomy_posterior_decompression": "Posterior decompression or laminectomy",
    "pelvic_fixation": "Pelvic fixation",
    "plf": "Posterolateral fusion",
    "tlif_plif": "Posterior/transforaminal lumbar interbody fusion",
    "other_lumbar_procedure": "Other lumbar procedure",
    "number_operated_levels": "Number of operated levels",
    "operative_region_extent": "Operative region extent",
    "PatTobaccoUse": "Tobacco use",
    "preop_ODI": "Preoperative ODI",
    "postop_ODI": "Postoperative ODI",
    "ODI_change": "Change in ODI",
    PROM_CHANGE_RATE_COL: "ODI change rate",
    RELATIVE_ODI_MCID_COL: "Relative ODI MCID",
    "postop_ODI_day": "Timing of postoperative ODI assessment",
}

def pretty_feature_name(feature: str) -> str:
    return FEATURE_LABELS.get(feature, feature.replace("_", " "))


@dataclass
class Step2Preprocessor:
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str, List[str]]

    def __post_init__(self):
        self.continuous_imputer: Optional[SimpleImputer] = None
        self.discrete_imputer: Optional[SimpleImputer] = None
        self.nominal_imputer: Optional[SimpleImputer] = None
        self.onehot: Optional[OneHotEncoder] = None
        self.numeric_feature_names_: List[str] = []
        self.output_feature_names_: List[str] = []

    def _parse_binary(self, x: Any, feature: str) -> float:
        sx = norm_text(x)
        if sx is None:
            return np.nan

        if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
            return float(BINARY_MAPS[feature][sx])

        if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive"}:
            return 1.0
        if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative"}:
            return 0.0

        try:
            v = float(sx)
            if v in (0.0, 1.0):
                return float(v)
        except Exception:
            pass

        return np.nan

    def _parse_ordinal(self, x: Any, feature: str) -> float:
        sx = norm_text(x)
        if sx is None:
            return np.nan

        if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]:
            return float(ORDINAL_MAPS[feature][sx])

        try:
            v = float(sx)
            if feature == "asa":
                return float(min(max(int(round(v)), 1), 4))
            if feature == "number_operated_levels":
                return float(min(max(int(round(v)), 0), 4))
            if feature == "diabetes_status":
                return float(min(max(int(round(v)), 0), 2))
            if feature == "institution_size":
                return float(min(max(int(round(v)), 0), 2))
            return float(v)
        except Exception:
            return np.nan

    def _canonical_nominal(self, feature: str, x: Any) -> Any:
        x = clean_scalar(x)
        if pd.isna(x):
            return np.nan

        s = str(x).strip()
        sl = s.lower()

        if feature == "race":
            if sl == "white":
                return "White"
            if sl == "black":
                return "Black"
            return "Other"

        if feature == "institution_region":
            mapping = {
                "south": "South",
                "north east": "North East",
                "northeast": "North East",
                "north-east": "North East",
                "west": "West",
                "midwest": "Midwest",
                "mid west": "Midwest",
            }
            return mapping.get(sl, s)

        if feature == "payer_status":
            if sl == "medicare":
                return "Medicare"
            if sl in {"commercial/private", "commercial", "private", "commercial private"}:
                return "Commercial/Private"
            if sl in {"medicaid/public/government", "medicaid", "public", "government", "public/government"}:
                return "Medicaid/Public/Government"
            return "Other"

        if feature == "PatTobaccoUse":
            if sl == "never":
                return "Never"
            if sl == "former":
                return "Former"
            if sl == "current":
                return "Current"
            return "Unknown/Not reported/Multiple"

        return s

    def _make_parts(self, X: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        cont = pd.DataFrame(index=X.index)
        for c in self.continuous_features:
            cont[c] = pd.to_numeric(X[c].map(clean_scalar), errors="coerce")

        discrete = pd.DataFrame(index=X.index)
        for c in self.binary_features:
            discrete[c] = X[c].map(lambda z: self._parse_binary(z, c)).astype(float)
        for c in self.ordinal_features:
            discrete[c] = X[c].map(lambda z: self._parse_ordinal(z, c)).astype(float)

        nominal = pd.DataFrame(index=X.index)
        for c in self.nominal_features:
            nominal[c] = X[c].map(lambda z: self._canonical_nominal(c, z)).astype("object")

        return cont, discrete, nominal

    def fit(self, X: pd.DataFrame):
        cont, discrete, nominal = self._make_parts(X)

        self.continuous_imputer = SimpleImputer(strategy="median")
        self.discrete_imputer = SimpleImputer(strategy="most_frequent")
        self.nominal_imputer = SimpleImputer(strategy="constant", fill_value="Missing")

        self.continuous_imputer.fit(cont)
        self.discrete_imputer.fit(discrete)

        nominal_imp = self.nominal_imputer.fit_transform(nominal)
        nominal_imp = pd.DataFrame(nominal_imp, columns=self.nominal_features)

        categories = []
        for c in self.nominal_features:
            preferred = list(self.preferred_nominal_levels.get(c, []))
            observed = nominal_imp[c].astype(str).unique().tolist()
            final_cats = preferred + sorted([x for x in observed if x not in preferred])
            categories.append(final_cats)

        try:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse_output=False)
        except TypeError:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse=False)

        self.onehot.fit(nominal_imp.astype(str))

        self.numeric_feature_names_ = self.continuous_features + self.binary_features + self.ordinal_features
        self.output_feature_names_ = (
            list(self.numeric_feature_names_)
            + self.onehot.get_feature_names_out(self.nominal_features).tolist()
        )
        return self

    def transform(self, X: pd.DataFrame) -> np.ndarray:
        if self.continuous_imputer is None or self.discrete_imputer is None or self.nominal_imputer is None or self.onehot is None:
            raise RuntimeError("Preprocessor is not fitted.")

        cont, discrete, nominal = self._make_parts(X)

        cont_imp = self.continuous_imputer.transform(cont)
        discrete_imp = self.discrete_imputer.transform(discrete)

        nominal_imp = self.nominal_imputer.transform(nominal)
        nominal_imp = pd.DataFrame(nominal_imp, columns=self.nominal_features)
        nominal_oh = self.onehot.transform(nominal_imp.astype(str))

        return np.concatenate([cont_imp, discrete_imp, nominal_oh], axis=1).astype(float)

    def fit_transform(self, X: pd.DataFrame) -> np.ndarray:
        self.fit(X)
        return self.transform(X)

    @property
    def output_feature_names(self) -> List[str]:
        return self.output_feature_names_


# ============================================================
# 6) Splitting, tuning, fitting
# ============================================================

def patient_level_train_cal_test_split(
    df: pd.DataFrame,
    target_col: str,
    group_col: str,
    test_fraction: float,
    calibration_fraction_of_remaining: float,
    seed: int,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    group_df = df.groupby(group_col, dropna=False)[target_col].max().reset_index()
    y_group = group_df[target_col].astype(int).to_numpy()
    groups = group_df[group_col].to_numpy()

    if len(np.unique(y_group)) < 2:
        raise ValueError("Only one class at patient level; stratified split impossible.")

    sss1 = StratifiedShuffleSplit(n_splits=1, test_size=test_fraction, random_state=seed)
    train_cal_idx, test_idx = next(sss1.split(groups, y_group))

    groups_train_cal = groups[train_cal_idx]
    y_train_cal = y_group[train_cal_idx]

    sss2 = StratifiedShuffleSplit(
        n_splits=1,
        test_size=calibration_fraction_of_remaining,
        random_state=seed + 1,
    )
    train_idx_rel, cal_idx_rel = next(sss2.split(groups_train_cal, y_train_cal))

    train_groups = set(groups_train_cal[train_idx_rel])
    cal_groups = set(groups_train_cal[cal_idx_rel])
    test_groups = set(groups[test_idx])

    assert train_groups.isdisjoint(cal_groups)
    assert train_groups.isdisjoint(test_groups)
    assert cal_groups.isdisjoint(test_groups)

    return (
        df[group_col].isin(train_groups).to_numpy(),
        df[group_col].isin(cal_groups).to_numpy(),
        df[group_col].isin(test_groups).to_numpy(),
    )


def cv_fold_count(y: np.ndarray, groups: np.ndarray, requested_folds: int) -> int:
    group_df = pd.DataFrame({"group": groups, "y": y}).groupby("group")["y"].max().reset_index()
    n_pos_groups = int((group_df["y"] == 1).sum())
    n_neg_groups = int((group_df["y"] == 0).sum())
    n_folds = min(requested_folds, n_pos_groups, n_neg_groups)

    if n_folds < 2:
        raise ValueError("Not enough positive/negative patient groups for group-aware CV.")

    return int(n_folds)


def make_lgbm_model(
    params: Dict[str, Any],
    seed: int,
    override_n_estimators: Optional[int] = None,
) -> LGBMClassifier:
    params = sanitize_lgbm_params(params)
    model_params = {k: v for k, v in params.items() if k != "positive_weight_multiplier"}
    if override_n_estimators is not None:
        model_params["n_estimators"] = int(max(MIN_FINAL_N_ESTIMATORS, override_n_estimators))

    return LGBMClassifier(
        objective="binary",
        boosting_type="gbdt",
        metric="average_precision",
        random_state=seed,
        n_jobs=N_JOBS,
        verbosity=-1,
        force_col_wise=True,
        **model_params,
    )


def fit_model_pipeline(
    X_train: pd.DataFrame,
    y_train: np.ndarray,
    params: Dict[str, Any],
    seed: int,
    eval_set: Optional[Tuple[pd.DataFrame, np.ndarray]] = None,
    use_early_stopping: bool = False,
    override_n_estimators: Optional[int] = None,
) -> Tuple[Step2Preprocessor, LGBMClassifier, Optional[int]]:
    params = sanitize_lgbm_params(params)

    pre = Step2Preprocessor(
        continuous_features=CONTINUOUS_FEATURES,
        binary_features=BINARY_FEATURES,
        ordinal_features=ORDINAL_FEATURES,
        nominal_features=NOMINAL_FEATURES,
        preferred_nominal_levels=PREFERRED_NOMINAL_LEVELS,
    )

    Xp_train = pre.fit_transform(X_train)

    weights = make_positive_sample_weights(
        y_train,
        multiplier=float(params["positive_weight_multiplier"]),
    )

    model = make_lgbm_model(params, seed=seed, override_n_estimators=override_n_estimators)
    best_iteration: Optional[int] = None

    if eval_set is not None and use_early_stopping:
        X_val, y_val = eval_set
        Xp_val = pre.transform(X_val)
        callbacks = [
            lgb.early_stopping(stopping_rounds=EARLY_STOPPING_ROUNDS, verbose=False),
            lgb.log_evaluation(period=0),
        ]
        try:
            model.fit(
                Xp_train,
                y_train,
                sample_weight=weights,
                eval_set=[(Xp_val, y_val)],
                eval_metric="average_precision",
                callbacks=callbacks,
            )
            if hasattr(model, "best_iteration_") and model.best_iteration_ is not None:
                best_iteration = int(model.best_iteration_)
        except Exception:
            # Conservative fallback for older LightGBM builds.
            model.fit(Xp_train, y_train, sample_weight=weights)
    else:
        model.fit(Xp_train, y_train, sample_weight=weights)

    return pre, model, best_iteration


def predict_pipeline(pre: Step2Preprocessor, model: LGBMClassifier, X: pd.DataFrame) -> np.ndarray:
    Xp = pre.transform(X)
    return model.predict_proba(Xp)[:, 1]


def tune_hyperparameters(
    X_train: pd.DataFrame,
    y_train: np.ndarray,
    groups_train: np.ndarray,
    seed: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    n_folds = cv_fold_count(y_train, groups_train, N_CV_FOLDS)
    cv = StratifiedGroupKFold(n_splits=n_folds, shuffle=True, random_state=seed)

    sampler = list(ParameterSampler(
        LGBM_SEARCH_SPACE,
        n_iter=N_RANDOM_COMBINATIONS,
        random_state=seed,
    ))

    candidate_rows = []
    fold_rows = []

    for i, raw_params in enumerate(sampler, start=1):
        params = sanitize_lgbm_params(raw_params)
        fold_aps = []
        fold_aucs = []
        fold_train_aps = []
        fold_train_aucs = []
        fold_best_iterations = []
        t0 = time.time()

        for fold_id, (tr_idx, va_idx) in enumerate(cv.split(X_train, y_train, groups_train), start=1):
            X_tr = X_train.iloc[tr_idx].reset_index(drop=True)
            y_tr = y_train[tr_idx]

            X_va = X_train.iloc[va_idx].reset_index(drop=True)
            y_va = y_train[va_idx]

            pre, model, best_iter = fit_model_pipeline(
                X_tr,
                y_tr,
                params=params,
                seed=seed + i * 1000 + fold_id,
                eval_set=(X_va, y_va),
                use_early_stopping=USE_EARLY_STOPPING_IN_CV,
            )

            p_tr = predict_pipeline(pre, model, X_tr)
            p_va = predict_pipeline(pre, model, X_va)

            train_ap = safe_average_precision(y_tr, p_tr)
            train_auc = safe_roc_auc(y_tr, p_tr)
            ap = safe_average_precision(y_va, p_va)
            auc = safe_roc_auc(y_va, p_va)

            fold_train_aps.append(train_ap)
            fold_train_aucs.append(train_auc)
            fold_aps.append(ap)
            fold_aucs.append(auc)
            if best_iter is not None and best_iter > 0:
                fold_best_iterations.append(best_iter)

            fold_rows.append({
                "candidate_id": i,
                "fold": fold_id,
                "fold_train_n": int(len(y_tr)),
                "fold_train_events": int(np.sum(y_tr)),
                "fold_validation_n": int(len(y_va)),
                "fold_validation_events": int(np.sum(y_va)),
                "fold_validation_event_rate": float(np.mean(y_va)),
                "fold_train_AP": train_ap,
                "fold_train_ROC_AUC": train_auc,
                "fold_validation_AP": ap,
                "fold_validation_ROC_AUC": auc,
                "fold_train_minus_validation_AP_gap": train_ap - ap if np.isfinite(train_ap) and np.isfinite(ap) else np.nan,
                "fold_train_minus_validation_ROC_AUC_gap": train_auc - auc if np.isfinite(train_auc) and np.isfinite(auc) else np.nan,
                "fold_best_iteration": best_iter,
                "positive_weight_used": actual_positive_weight(y_tr, params["positive_weight_multiplier"]),
                **params,
            })

        elapsed = time.time() - t0
        locked_n_estimators = int(np.median(fold_best_iterations)) if fold_best_iterations else int(params["n_estimators"])

        row = {
            "candidate_id": i,
            "cv_folds": n_folds,
            "cv_train_AP_mean": float(np.nanmean(fold_train_aps)),
            "cv_train_AP_SD": float(np.nanstd(fold_train_aps, ddof=1)),
            "cv_train_ROC_AUC_mean": float(np.nanmean(fold_train_aucs)),
            "cv_train_ROC_AUC_SD": float(np.nanstd(fold_train_aucs, ddof=1)),
            "cv_AP_mean": float(np.nanmean(fold_aps)),
            "cv_AP_SD": float(np.nanstd(fold_aps, ddof=1)),
            "cv_ROC_AUC_mean": float(np.nanmean(fold_aucs)),
            "cv_ROC_AUC_SD": float(np.nanstd(fold_aucs, ddof=1)),
            "cv_train_minus_validation_AP_gap": float(np.nanmean(fold_train_aps) - np.nanmean(fold_aps)),
            "cv_train_minus_validation_ROC_AUC_gap": float(np.nanmean(fold_train_aucs) - np.nanmean(fold_aucs)),
            "mean_cv_best_iteration": float(np.nanmean(fold_best_iterations)) if fold_best_iterations else np.nan,
            "locked_n_estimators_from_cv": locked_n_estimators,
            "elapsed_seconds": float(elapsed),
            **params,
        }
        candidate_rows.append(row)

        print(
            f"Candidate {i:03d}/{len(sampler)} | "
            f"CV AP={row['cv_AP_mean']:.5f} ± {row['cv_AP_SD']:.5f} | "
            f"Train-CV AP gap={row['cv_train_minus_validation_AP_gap']:.5f} | "
            f"CV AUC={row['cv_ROC_AUC_mean']:.5f} | "
            f"locked_n={locked_n_estimators} | "
            f"pos_mult={params['positive_weight_multiplier']}"
        )

    candidates = (
        pd.DataFrame(candidate_rows)
        .sort_values("cv_AP_mean", ascending=False)
        .reset_index(drop=True)
    )
    folds = pd.DataFrame(fold_rows)

    return candidates, folds


# ============================================================
# 7) Evaluation and bootstrap
# ============================================================

def evaluate_predictions(
    y_true: np.ndarray,
    y_prob: np.ndarray,
    threshold: Optional[float] = None,
    prefix: str = "",
) -> Dict[str, Any]:
    out = {
        f"{prefix}AP": safe_average_precision(y_true, y_prob),
        f"{prefix}ROC_AUC": safe_roc_auc(y_true, y_prob),
        f"{prefix}Brier_score": float(brier_score_loss(y_true, y_prob)),
        f"{prefix}ECE": expected_calibration_error(y_true, y_prob, n_bins=ECE_N_BINS),
        f"{prefix}N": int(len(y_true)),
        f"{prefix}Events": int(np.sum(y_true)),
        f"{prefix}Prevalence": float(np.mean(y_true)),
    }

    if threshold is not None:
        cls = classification_metrics_at_threshold(y_true, y_prob, threshold)
        out.update({f"{prefix}{k}": v for k, v in cls.items()})

        top = top_percent_metrics(y_true, y_prob, top_fraction=0.05)
        out.update({f"{prefix}{k}": v for k, v in top.items()})

    return out


def patient_level_stratified_bootstrap_ci(
    y_true: np.ndarray,
    y_prob: np.ndarray,
    groups: np.ndarray,
    metric_name: str,
    threshold: Optional[float] = None,
    n_bootstraps: int = 2000,
    seed: int = 123,
) -> Tuple[float, float]:
    rng = np.random.default_rng(seed)

    d = pd.DataFrame({
        "row_id": np.arange(len(y_true)),
        "group": groups,
        "y": np.asarray(y_true).astype(int),
        "p": np.asarray(y_prob).astype(float),
    })

    group_y = d.groupby("group")["y"].max()
    pos_groups = group_y[group_y == 1].index.to_numpy()
    neg_groups = group_y[group_y == 0].index.to_numpy()

    if len(pos_groups) == 0 or len(neg_groups) == 0:
        return np.nan, np.nan

    rows_by_group = {
        g: d.loc[d["group"].eq(g), "row_id"].to_numpy()
        for g in group_y.index
    }

    values = []

    for _ in range(n_bootstraps):
        sampled_pos = rng.choice(pos_groups, size=len(pos_groups), replace=True)
        sampled_neg = rng.choice(neg_groups, size=len(neg_groups), replace=True)
        sampled_groups = np.concatenate([sampled_pos, sampled_neg])

        idx = np.concatenate([rows_by_group[g] for g in sampled_groups])
        yy = y_true[idx]
        pp = y_prob[idx]

        if len(np.unique(yy)) < 2:
            continue

        if metric_name == "AP":
            values.append(average_precision_score(yy, pp))
        elif metric_name == "ROC_AUC":
            values.append(roc_auc_score(yy, pp))
        elif metric_name == "F1":
            if threshold is None:
                raise ValueError("threshold required for F1 bootstrap")
            yhat = (pp >= threshold).astype(int)
            values.append(f1_score(yy, yhat, zero_division=0))
        else:
            raise ValueError(f"Unsupported bootstrap metric: {metric_name}")

    if len(values) == 0:
        return np.nan, np.nan

    lo, hi = np.percentile(values, [2.5, 97.5])
    return float(lo), float(hi)


def save_pr_curve(y_true: np.ndarray, y_prob: np.ndarray, path: str, title: str) -> None:
    precision, recall, _ = precision_recall_curve(y_true, y_prob)
    ap = average_precision_score(y_true, y_prob)

    plt.figure(figsize=(6, 5))
    plt.plot(recall, precision, linewidth=2)
    plt.xlabel("Recall")
    plt.ylabel("Precision")
    plt.title(f"{title}\nAP = {ap:.3f}")
    plt.tight_layout()
    plt.savefig(path, dpi=300)
    plt.close()


def save_roc_curve(y_true: np.ndarray, y_prob: np.ndarray, path: str, title: str) -> None:
    fpr, tpr, _ = roc_curve(y_true, y_prob)
    auc = roc_auc_score(y_true, y_prob)

    plt.figure(figsize=(6, 5))
    plt.plot(fpr, tpr, linewidth=2)
    plt.plot([0, 1], [0, 1], linestyle="--", linewidth=1)
    plt.xlabel("False-positive rate")
    plt.ylabel("True-positive rate")
    plt.title(f"{title}\nROC-AUC = {auc:.3f}")
    plt.tight_layout()
    plt.savefig(path, dpi=300)
    plt.close()


def save_calibration_plot(y_true: np.ndarray, y_prob: np.ndarray, path: str, title: str) -> None:
    """
    Calibration plot for rare-event prediction.

    The event prevalence in this task is low; therefore, most calibrated
    probabilities are expected to be close to zero. A conventional 0-1 axis can
    visually compress all calibration points into the lower-left corner. This
    plot uses quantile-based probability bins and a data-adaptive zoomed axis,
    while reporting Brier score and expected calibration error (ECE).
    """
    y_true = np.asarray(y_true).astype(int)
    y_prob = np.asarray(y_prob).astype(float)

    finite_mask = np.isfinite(y_prob)
    y_true = y_true[finite_mask]
    y_prob = y_prob[finite_mask]

    brier = float(brier_score_loss(y_true, y_prob))
    ece = float(expected_calibration_error(y_true, y_prob, n_bins=ECE_N_BINS))
    prevalence = float(np.mean(y_true)) if len(y_true) else np.nan

    tmp = pd.DataFrame({"y": y_true, "p": y_prob})
    n_unique = int(tmp["p"].nunique())
    n_bins = min(ECE_N_BINS, max(2, n_unique))
    tmp["bin"] = pd.qcut(tmp["p"].rank(method="first"), q=n_bins, duplicates="drop")

    cal = tmp.groupby("bin", observed=True).agg(
        mean_pred=("p", "mean"),
        observed=("y", "mean"),
        n=("y", "size"),
    ).reset_index(drop=True)

    upper_prob = float(np.nanquantile(y_prob, 0.995)) if len(y_prob) else 0.05
    upper_obs = float(max(cal["observed"].max(), cal["mean_pred"].max())) if len(cal) else 0.05
    axis_max = max(0.05, min(0.25, max(upper_prob, upper_obs, prevalence) * 1.35 + 0.005))

    fig, ax = plt.subplots(figsize=(7.2, 6.0))

    ax.plot(
        [0, axis_max],
        [0, axis_max],
        linestyle="--",
        linewidth=1.6,
        color="black",
        alpha=0.80,
        label="Perfect calibration",
    )
    ax.plot(
        cal["mean_pred"],
        cal["observed"],
        marker="o",
        linewidth=2.2,
        markersize=6.5,
        label="Model calibration",
    )
    ax.axhline(
        prevalence,
        linestyle=":",
        linewidth=1.5,
        color="gray",
        alpha=0.90,
        label=f"Observed event rate = {prevalence:.2%}",
    )

    for _, r in cal.iterrows():
        ax.annotate(
            f"n={int(r['n'])}",
            (float(r["mean_pred"]), float(r["observed"])),
            textcoords="offset points",
            xytext=(5, 4),
            fontsize=7,
            alpha=0.75,
        )

    ax.set_xlim(0, axis_max)
    ax.set_ylim(0, axis_max)
    ax.set_xlabel("Mean predicted probability")
    ax.set_ylabel("Observed event rate")
    ax.set_title(title, fontsize=14, fontweight="bold")
    ax.legend(loc="upper left", frameon=True)

    text_box = (
        f"Brier score = {brier:.4f}\n"
        f"ECE = {ece:.4f}\n"
        f"N = {len(y_true):,}\n"
        f"Events = {int(np.sum(y_true)):,}"
    )
    ax.text(
        0.98,
        0.04,
        text_box,
        transform=ax.transAxes,
        ha="right",
        va="bottom",
        fontsize=10,
        bbox=dict(boxstyle="round,pad=0.35", facecolor="white", edgecolor="gray", alpha=0.95),
    )

    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_learning_curve_grouped_ap(
    X_train: pd.DataFrame,
    y_train: np.ndarray,
    groups_train: np.ndarray,
    best_params: Dict[str, Any],
    path_png: str,
    path_csv: str,
    title: str,
    train_fractions: Tuple[float, ...] = (0.20, 0.40, 0.60, 0.80, 1.00),
    n_splits: int = 3,
    seed: int = RANDOM_STATE,
) -> pd.DataFrame:
    """
    Group-aware learning curve using Average Precision.

    This diagnostic is computed only inside the training split using the locked
    best hyperparameter configuration. It is not used for model selection and
    never uses the calibration or test sets.
    """
    X_train = X_train.reset_index(drop=True)
    y_train = np.asarray(y_train).astype(int)
    groups_train = np.asarray(groups_train)

    n_splits_eff = min(cv_fold_count(y_train, groups_train, n_splits), n_splits)
    cv = StratifiedGroupKFold(n_splits=n_splits_eff, shuffle=True, random_state=seed)

    rows = []
    rng = np.random.default_rng(seed)

    for frac in train_fractions:
        for fold_id, (tr_idx, va_idx) in enumerate(cv.split(X_train, y_train, groups_train), start=1):
            X_tr_full = X_train.iloc[tr_idx].reset_index(drop=True)
            y_tr_full = y_train[tr_idx]
            g_tr_full = groups_train[tr_idx]

            X_va = X_train.iloc[va_idx].reset_index(drop=True)
            y_va = y_train[va_idx]

            group_event = (
                pd.DataFrame({"group": g_tr_full, "y": y_tr_full})
                .groupby("group")["y"]
                .max()
                .reset_index()
            )
            pos_groups = group_event.loc[group_event["y"].eq(1), "group"].to_numpy()
            neg_groups = group_event.loc[group_event["y"].eq(0), "group"].to_numpy()

            n_pos_keep = max(1, int(np.ceil(len(pos_groups) * frac)))
            n_neg_keep = max(1, int(np.ceil(len(neg_groups) * frac)))

            sampled_pos = rng.choice(pos_groups, size=min(n_pos_keep, len(pos_groups)), replace=False)
            sampled_neg = rng.choice(neg_groups, size=min(n_neg_keep, len(neg_groups)), replace=False)
            sampled_groups = set(np.concatenate([sampled_pos, sampled_neg]))

            keep = np.array([g in sampled_groups for g in g_tr_full])
            X_tr = X_tr_full.loc[keep].reset_index(drop=True)
            y_tr = y_tr_full[keep]

            if len(np.unique(y_tr)) < 2 or len(np.unique(y_va)) < 2:
                continue

            pre_lc, model_lc, _ = fit_model_pipeline(
                X_tr,
                y_tr,
                params=best_params,
                seed=seed + fold_id + int(frac * 1000),
            )

            p_tr = predict_pipeline(pre_lc, model_lc, X_tr)
            p_va = predict_pipeline(pre_lc, model_lc, X_va)

            rows.append({
                "train_fraction": float(frac),
                "fold": int(fold_id),
                "train_n": int(len(y_tr)),
                "train_events": int(np.sum(y_tr)),
                "validation_n": int(len(y_va)),
                "validation_events": int(np.sum(y_va)),
                "train_AP": safe_average_precision(y_tr, p_tr),
                "validation_AP": safe_average_precision(y_va, p_va),
            })

    lc_raw = pd.DataFrame(rows)
    lc_raw.to_csv(path_csv, index=False)

    if lc_raw.empty:
        return lc_raw

    lc = (
        lc_raw.groupby("train_fraction", as_index=False)
        .agg(
            train_AP_mean=("train_AP", "mean"),
            train_AP_sd=("train_AP", "std"),
            validation_AP_mean=("validation_AP", "mean"),
            validation_AP_sd=("validation_AP", "std"),
        )
    ).fillna(0.0)

    fig, ax = plt.subplots(figsize=(7.2, 6.0))
    ax.plot(lc["train_fraction"], lc["train_AP_mean"], marker="o", linewidth=2.2, label="Training AP")
    ax.plot(lc["train_fraction"], lc["validation_AP_mean"], marker="o", linewidth=2.2, label="Validation AP")

    ax.fill_between(
        lc["train_fraction"].astype(float).to_numpy(),
        (lc["train_AP_mean"] - lc["train_AP_sd"]).astype(float).to_numpy(),
        (lc["train_AP_mean"] + lc["train_AP_sd"]).astype(float).to_numpy(),
        alpha=0.15,
    )
    ax.fill_between(
        lc["train_fraction"].astype(float).to_numpy(),
        (lc["validation_AP_mean"] - lc["validation_AP_sd"]).astype(float).to_numpy(),
        (lc["validation_AP_mean"] + lc["validation_AP_sd"]).astype(float).to_numpy(),
        alpha=0.15,
    )

    ax.set_xlabel("Fraction of training groups used")
    ax.set_ylabel("Average precision")
    ax.set_title(title, fontsize=14, fontweight="bold")
    ax.set_xticks(list(train_fractions))
    ax.legend(loc="best", frameon=True)
    ax.grid(alpha=0.20)

    text_box = (
        "Group-aware CV learning curve\n"
        f"Folds = {n_splits_eff}\n"
        "Metric = Average precision"
    )
    ax.text(
        0.98,
        0.04,
        text_box,
        transform=ax.transAxes,
        ha="right",
        va="bottom",
        fontsize=9.5,
        bbox=dict(boxstyle="round,pad=0.35", facecolor="white", edgecolor="gray", alpha=0.95),
    )

    fig.tight_layout()
    fig.savefig(path_png, dpi=300, bbox_inches="tight")
    plt.close(fig)

    return lc_raw


# ============================================================
# 8) Audit helpers
# ============================================================

def build_split_audit(df: pd.DataFrame, split_col: str, target_col: str) -> pd.DataFrame:
    rows = []

    for split in ["train", "calibration", "test"]:
        sub = df[df[split_col] == split].copy()
        rows.append({
            "split": split,
            "rows": len(sub),
            "events": int(sub[target_col].sum()),
            "event_rate": float(sub[target_col].mean()) if len(sub) else np.nan,
            "unique_patients": int(sub[GROUP_COL].nunique()) if GROUP_COL in sub.columns else np.nan,
        })

    return pd.DataFrame(rows)


def build_institution_audit(df: pd.DataFrame, split_col: str) -> pd.DataFrame:
    rows = []
    cols = [c for c in ["institution_type", "institution_region"] if c in df.columns]

    for col in cols:
        ct = (
            df.groupby([split_col, col], dropna=False)
            .size()
            .reset_index(name="n")
        )
        total = df.groupby(split_col).size().rename("split_n").reset_index()
        ct = ct.merge(total, on=split_col, how="left")
        ct["percent_within_split"] = 100 * ct["n"] / ct["split_n"]
        ct.insert(0, "variable", col)
        rows.append(ct)

    if rows:
        return pd.concat(rows, ignore_index=True)

    return pd.DataFrame({"note": ["No institution columns found for audit."]})


def style_excel_workbook(writer: pd.ExcelWriter) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter

                for cell in col_cells:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)
    except Exception:
        pass



# ============================================================
# 9) Grouped SHAP and threshold analysis helpers
# ============================================================

def parse_binary_global(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative"}:
        return 0.0
    try:
        v = float(sx)
        if v in (0.0, 1.0):
            return float(v)
    except Exception:
        pass
    return np.nan


def parse_ordinal_global(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]:
        return float(ORDINAL_MAPS[feature][sx])
    try:
        v = float(sx)
        if feature == "asa":
            return float(min(max(int(round(v)), 1), 4))
        if feature == "number_operated_levels":
            return float(min(max(int(round(v)), 0), 4))
        if feature == "diabetes_status":
            return float(min(max(int(round(v)), 0), 2))
        if feature == "institution_size":
            return float(min(max(int(round(v)), 0), 2))
        return float(v)
    except Exception:
        return np.nan


def raw_feature_numeric_values(X_raw: pd.DataFrame, feature: str) -> pd.Series:
    if feature in CONTINUOUS_FEATURES:
        s = pd.to_numeric(X_raw[feature].map(clean_scalar), errors="coerce")
    elif feature in BINARY_FEATURES:
        s = X_raw[feature].map(lambda z: parse_binary_global(z, feature)).astype(float)
    elif feature in ORDINAL_FEATURES:
        s = X_raw[feature].map(lambda z: parse_ordinal_global(z, feature)).astype(float)
    else:
        vals = X_raw[feature].map(lambda z: str(clean_scalar(z)) if not pd.isna(clean_scalar(z)) else "Missing")
        categories = {v: i for i, v in enumerate(sorted(vals.dropna().unique()))}
        s = vals.map(categories).astype(float)
    if s.isna().any():
        med = s.median(skipna=True)
        if pd.isna(med):
            med = 0.0
        s = s.fillna(med)
    return s.astype(float)


def encoded_to_group_mapping(pre: Step2Preprocessor) -> Dict[str, List[int]]:
    encoded_names = list(pre.output_feature_names)
    mapping: Dict[str, List[int]] = {}
    for feature in FEATURES:
        idx: List[int] = []
        if feature in encoded_names:
            idx.append(encoded_names.index(feature))
        if feature in NOMINAL_FEATURES:
            prefix = feature + "_"
            idx.extend([i for i, name in enumerate(encoded_names) if name.startswith(prefix)])
        idx = sorted(set(idx))
        if idx:
            mapping[feature] = idx
    return mapping


def compute_grouped_shap(
    pre: Step2Preprocessor,
    model: LGBMClassifier,
    X_raw: pd.DataFrame,
    y_true: np.ndarray,
    p_calibrated: np.ndarray,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, List[str]]]:
    """Compute native LightGBM TreeSHAP and aggregate one-hot variables back to raw features."""
    n = len(X_raw)
    # Explain the complete supplied dataset; no row cap or sampling is applied.
    explain_idx = np.arange(n)

    X_explain_raw = X_raw.iloc[explain_idx].reset_index(drop=True)
    y_explain = np.asarray(y_true)[explain_idx]
    p_explain = np.asarray(p_calibrated)[explain_idx]
    X_explain_enc = pre.transform(X_explain_raw)
    encoded_names = list(pre.output_feature_names)

    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X_explain_enc)
    if isinstance(shap_values, list):
        shap_values = shap_values[-1]
    shap_values = np.asarray(shap_values)
    if shap_values.ndim == 3:
        shap_values = shap_values[:, :, -1]

    mapping_idx = encoded_to_group_mapping(pre)
    grouped_values = {}
    grouped_feature_data = {}
    mapping_names: Dict[str, List[str]] = {}

    for raw_feature, idx in mapping_idx.items():
        grouped_values[pretty_feature_name(raw_feature)] = shap_values[:, idx].sum(axis=1)
        grouped_feature_data[pretty_feature_name(raw_feature)] = raw_feature_numeric_values(X_explain_raw, raw_feature).values
        mapping_names[raw_feature] = [encoded_names[i] for i in idx]

    grouped_shap_df = pd.DataFrame(grouped_values)
    grouped_data_df = pd.DataFrame(grouped_feature_data)

    importance = []
    total_abs = float(np.abs(grouped_shap_df.values).mean(axis=0).sum())
    for raw_feature in mapping_idx.keys():
        display = pretty_feature_name(raw_feature)
        mean_abs = float(np.abs(grouped_shap_df[display]).mean())
        importance.append({
            "raw_feature": raw_feature,
            "display_feature": display,
            "mean_abs_SHAP": mean_abs,
            "percent_of_grouped_SHAP": 100.0 * mean_abs / total_abs if total_abs > 0 else np.nan,
            "feature_type": (
                "Continuous" if raw_feature in CONTINUOUS_FEATURES else
                "Binary" if raw_feature in BINARY_FEATURES else
                "Ordinal" if raw_feature in ORDINAL_FEATURES else
                "Nominal" if raw_feature in NOMINAL_FEATURES else "Unknown"
            ),
            "n_encoded_columns_grouped": len(mapping_idx[raw_feature]),
        })
    importance_df = pd.DataFrame(importance).sort_values("mean_abs_SHAP", ascending=False).reset_index(drop=True)

    grouped_shap_df.insert(0, "__row_id__", explain_idx)
    grouped_shap_df.insert(1, "__y_true__", y_explain)
    grouped_shap_df.insert(2, "__p_calibrated__", p_explain)

    return grouped_shap_df, grouped_data_df, importance_df, mapping_names


def save_grouped_shap_beeswarm(grouped_shap_df: pd.DataFrame, grouped_data_df: pd.DataFrame, importance_df: pd.DataFrame) -> str:
    ordered_displays = importance_df["display_feature"].tolist()
    shap_matrix = grouped_shap_df[ordered_displays]
    data_matrix = grouped_data_df[ordered_displays]

    plt.figure(figsize=(10.5, 9.0))
    shap.summary_plot(
        shap_matrix.values,
        features=data_matrix,
        feature_names=ordered_displays,
        max_display=SHAP_BEESWARM_MAX_DISPLAY,
        show=False,
        plot_size=None,
    )
    plt.title("Step 2 ODI LightGBM: grouped SHAP beeswarm", fontsize=15, fontweight="bold")
    path = os.path.join(PLOT_DIR, "SHAP_beeswarm_GROUPED_all_features_best_model.png")
    plt.tight_layout()
    plt.savefig(path, dpi=300, bbox_inches="tight")
    plt.close()
    return path


def save_grouped_shap_bar(importance_df: pd.DataFrame) -> str:
    plot_df = importance_df.head(SHAP_BAR_MAX_DISPLAY).copy().iloc[::-1]
    fig_h = max(7.0, len(plot_df) * 0.35)
    fig, ax = plt.subplots(figsize=(10.5, fig_h))
    bars = ax.barh(plot_df["display_feature"], plot_df["mean_abs_SHAP"])
    ax.set_xlabel("Mean absolute SHAP value")
    ax.set_title("Step 2 ODI LightGBM: grouped SHAP importance", fontsize=15, fontweight="bold")
    max_x = float(plot_df["mean_abs_SHAP"].max()) if len(plot_df) else 1.0
    ax.set_xlim(0, max_x * 1.28)
    for bar, val, pct in zip(bars, plot_df["mean_abs_SHAP"], plot_df["percent_of_grouped_SHAP"]):
        ax.text(
            bar.get_width() + max_x * 0.015,
            bar.get_y() + bar.get_height() / 2,
            f"{val:.4f} ({pct:.1f}%)",
            va="center",
            fontsize=9,
        )
    fig.tight_layout()
    path = os.path.join(PLOT_DIR, "SHAP_bar_GROUPED_all_features_best_model.png")
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def build_binned_curve(x: np.ndarray, shap_y: np.ndarray, pred: np.ndarray, max_bins: int = 35) -> pd.DataFrame:
    d = pd.DataFrame({"x": x, "shap": shap_y, "pred": pred}).replace([np.inf, -np.inf], np.nan).dropna()
    if d.empty:
        return pd.DataFrame(columns=["x", "shap_mean", "pred_mean", "n"])
    unique_x = np.sort(d["x"].unique())
    if len(unique_x) <= min(max_bins, 20):
        curve = d.groupby("x", as_index=False).agg(shap_mean=("shap", "mean"), pred_mean=("pred", "mean"), n=("x", "size"))
    else:
        q = min(max_bins, max(8, int(np.sqrt(len(d)))))
        d["bin"] = pd.qcut(d["x"].rank(method="first"), q=q, duplicates="drop")
        curve = d.groupby("bin", observed=True).agg(x=("x", "mean"), shap_mean=("shap", "mean"), pred_mean=("pred", "mean"), n=("x", "size")).reset_index(drop=True)
    curve = curve.sort_values("x").reset_index(drop=True)
    return curve


def find_shap_threshold(curve: pd.DataFrame) -> Tuple[float, str]:
    if curve.empty:
        return np.nan, "unavailable"
    xs = curve["x"].astype(float).to_numpy()
    ys = curve["shap_mean"].astype(float).to_numpy()
    if len(xs) == 1:
        return float(xs[0]), "single_unique_value"
    crossings = []
    for i in range(len(xs) - 1):
        y1, y2 = ys[i], ys[i + 1]
        if not (np.isfinite(y1) and np.isfinite(y2)):
            continue
        if y1 == 0:
            crossings.append((abs(y2 - y1), xs[i]))
        elif y1 * y2 < 0:
            # Linear interpolation for zero crossing.
            t = abs(y1) / (abs(y1) + abs(y2))
            x_cross = xs[i] + t * (xs[i + 1] - xs[i])
            crossings.append((abs(y2 - y1), x_cross))
    if crossings:
        # Use the most pronounced zero-crossing.
        crossings = sorted(crossings, key=lambda z: z[0], reverse=True)
        return float(crossings[0][1]), "zero_crossing"
    idx = int(np.nanargmin(np.abs(ys)))
    return float(xs[idx]), "closest_to_zero_SHAP"


def save_threshold_plot_for_feature(
    raw_feature: str,
    display_feature: str,
    grouped_shap_df: pd.DataFrame,
    X_raw_all: pd.DataFrame,
    p_calibrated_all: np.ndarray,
) -> Dict[str, Any]:
    row_ids = grouped_shap_df["__row_id__"].astype(int).to_numpy()
    X_raw = X_raw_all.iloc[row_ids].reset_index(drop=True)
    p = np.asarray(p_calibrated_all)[row_ids]
    x = raw_feature_numeric_values(X_raw, raw_feature).to_numpy(dtype=float)
    shap_y = grouped_shap_df[display_feature].to_numpy(dtype=float)
    curve = build_binned_curve(x, shap_y, p, max_bins=SHAP_THRESHOLD_MAX_BINS)
    threshold, method = find_shap_threshold(curve)

    below = x < threshold
    above = x >= threshold
    risk_below = float(np.mean(p[below])) if np.any(below) else np.nan
    risk_above = float(np.mean(p[above])) if np.any(above) else np.nan
    abs_increase = risk_above - risk_below if np.isfinite(risk_below) and np.isfinite(risk_above) else np.nan
    rel_increase = (abs_increase / risk_below) if np.isfinite(abs_increase) and risk_below > 0 else np.nan

    fig, ax1 = plt.subplots(figsize=(8.5, 5.2))
    ax1.scatter(x, shap_y, s=9, alpha=0.12, color="gray", edgecolors="none")
    if not curve.empty:
        xs = curve["x"].to_numpy(dtype=float)
        ys = curve["shap_mean"].to_numpy(dtype=float)
        ax1.plot(xs, ys, color="black", linewidth=2.0)
        # Pink = risk-increasing/reoperation direction; blue = risk-decreasing/protective direction.
        ax1.fill_between(xs, 0, ys, where=ys >= 0, color=SHAP_BEESWARM_PINK, alpha=0.55, interpolate=True)
        ax1.fill_between(xs, 0, ys, where=ys < 0, color=SHAP_BEESWARM_BLUE, alpha=0.55, interpolate=True)
    ax1.axhline(0, color="black", linewidth=0.8, alpha=0.7)
    if np.isfinite(threshold):
        ax1.axvline(threshold, color="#1f77b4", linestyle=":", linewidth=1.6)
        ax1.text(threshold, ax1.get_ylim()[0], f"{threshold:.2f}", ha="center", va="bottom", fontsize=9, fontweight="bold")
    ax1.set_xlabel(display_feature, fontweight="bold")
    ax1.set_ylabel(f"SHAP value for {display_feature}")
    ax1.set_title(f"SHAP threshold analysis: {display_feature}", fontsize=13, fontweight="bold")

    ax2 = ax1.twinx()
    if not curve.empty:
        ax2.plot(curve["x"], curve["pred_mean"] * 100, color=SHAP_BEESWARM_PINK, linestyle="--", linewidth=1.8, alpha=0.90)
    ax2.set_ylabel("Predicted risk of reoperation (%)")

    txt = (
        f"Threshold = {threshold:.2f}\n"
        f"Predicted risk < threshold = {risk_below * 100:.2f}%\n"
        f"Predicted risk ≥ threshold = {risk_above * 100:.2f}%\n"
        f"Absolute increase = {abs_increase * 100:.2f} points\n"
        f"Relative increase = {rel_increase * 100:.1f}%\n"
        f"Method = {method}"
    )
    ax1.text(
        0.02, 0.97, txt,
        transform=ax1.transAxes,
        ha="left", va="top", fontsize=8.5,
        bbox=dict(boxstyle="round,pad=0.30", facecolor="white", alpha=0.90, edgecolor="gray"),
    )
    fig.tight_layout()
    safe_name = raw_feature.replace("/", "_").replace(" ", "_")
    path = os.path.join(PLOT_DIR, f"SHAP_threshold_{safe_name}.png")
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)

    return {
        "raw_feature": raw_feature,
        "display_feature": display_feature,
        "threshold": threshold,
        "threshold_method": method,
        "risk_below_threshold": risk_below,
        "risk_at_or_above_threshold": risk_above,
        "absolute_risk_increase": abs_increase,
        "relative_risk_increase": rel_increase,
        "n_below_threshold": int(np.sum(below)),
        "n_at_or_above_threshold": int(np.sum(above)),
        "plot_path": path,
    }



def recompute_grouped_shap_importance_from_values(
    grouped_shap_df: pd.DataFrame,
    mapping_names: Dict[str, List[str]],
) -> pd.DataFrame:
    """Recompute grouped SHAP importance from concatenated grouped SHAP values."""
    rows = []
    display_cols = [c for c in grouped_shap_df.columns if not str(c).startswith("__")]
    if not display_cols:
        return pd.DataFrame(columns=["raw_feature", "display_feature", "mean_abs_SHAP", "percent_of_grouped_SHAP", "feature_type", "n_encoded_columns_grouped"])

    total_abs = float(np.abs(grouped_shap_df[display_cols].to_numpy(dtype=float)).mean(axis=0).sum())
    for raw_feature in FEATURES:
        display = pretty_feature_name(raw_feature)
        if display not in grouped_shap_df.columns:
            continue
        mean_abs = float(np.abs(grouped_shap_df[display].astype(float)).mean())
        rows.append({
            "raw_feature": raw_feature,
            "display_feature": display,
            "mean_abs_SHAP": mean_abs,
            "percent_of_grouped_SHAP": 100.0 * mean_abs / total_abs if total_abs > 0 else np.nan,
            "feature_type": (
                "Continuous" if raw_feature in CONTINUOUS_FEATURES else
                "Binary" if raw_feature in BINARY_FEATURES else
                "Ordinal" if raw_feature in ORDINAL_FEATURES else
                "Nominal" if raw_feature in NOMINAL_FEATURES else "Unknown"
            ),
            "n_encoded_columns_grouped": len(mapping_names.get(raw_feature, [])),
        })
    return pd.DataFrame(rows).sort_values("mean_abs_SHAP", ascending=False).reset_index(drop=True)


def compute_training_oof_grouped_shap(
    X_train: pd.DataFrame,
    y_train: np.ndarray,
    groups_train: np.ndarray,
    params: Dict[str, Any],
    locked_n_estimators: int,
    seed: int,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, List[str]]]:
    """Generate out-of-fold grouped SHAP values exclusively within the training split.

    These out-of-fold SHAP values are used only for SHAP-informed threshold
    derivation. No calibration-set or held-out test-set row is used to estimate
    these thresholds.
    """
    n_folds = cv_fold_count(y_train, groups_train, N_CV_FOLDS)
    cv = StratifiedGroupKFold(n_splits=n_folds, shuffle=True, random_state=seed)

    shap_parts: List[pd.DataFrame] = []
    data_parts: List[pd.DataFrame] = []
    mapping_names_accum: Dict[str, set] = {f: set() for f in FEATURES}

    for fold_id, (tr_idx, va_idx) in enumerate(cv.split(X_train, y_train, groups_train), start=1):
        X_tr = X_train.iloc[tr_idx].reset_index(drop=True)
        y_tr = y_train[tr_idx]
        X_va = X_train.iloc[va_idx].reset_index(drop=True)
        y_va = y_train[va_idx]

        # Refit the already-selected locked configuration inside the training split only.
        pre_fold, model_fold, _ = fit_model_pipeline(
            X_train=X_tr,
            y_train=y_tr,
            params=params,
            seed=seed + 50000 + fold_id,
            eval_set=None,
            use_early_stopping=False,
            override_n_estimators=locked_n_estimators,
        )
        # For threshold derivation, the x-axis/SHAP curve is the key object. The
        # displayed risk summaries are based on out-of-fold model scores from the
        # training split, not on the held-out test set.
        p_va_oof = predict_pipeline(pre_fold, model_fold, X_va)
        g_shap, g_data, _, mapping_fold = compute_grouped_shap(
            pre=pre_fold,
            model=model_fold,
            X_raw=X_va,
            y_true=y_va,
            p_calibrated=p_va_oof,
        )
        g_shap["__row_id__"] = va_idx
        g_shap.insert(3, "__oof_fold__", fold_id)
        g_data.insert(0, "__row_id__", va_idx)
        g_data.insert(1, "__oof_fold__", fold_id)
        shap_parts.append(g_shap)
        data_parts.append(g_data)
        for raw_feature, names in mapping_fold.items():
            mapping_names_accum.setdefault(raw_feature, set()).update(names)

    grouped_shap_oof = pd.concat(shap_parts, ignore_index=True).sort_values("__row_id__").reset_index(drop=True)
    grouped_data_oof = pd.concat(data_parts, ignore_index=True).sort_values("__row_id__").reset_index(drop=True)
    mapping_names = {k: sorted(v) for k, v in mapping_names_accum.items() if v}
    importance_oof = recompute_grouped_shap_importance_from_values(grouped_shap_oof, mapping_names)
    return grouped_shap_oof, grouped_data_oof, importance_oof, mapping_names


def make_training_oof_threshold_table(
    importance_df: pd.DataFrame,
    grouped_shap_df: pd.DataFrame,
    X_train: pd.DataFrame,
    p_train_oof: np.ndarray,
) -> Tuple[pd.DataFrame, List[str]]:
    """Derive SHAP-informed numeric thresholds using training-split OOF SHAP only."""
    numeric_raw = set(CONTINUOUS_FEATURES + BINARY_FEATURES + ORDINAL_FEATURES)
    threshold_candidates = importance_df[importance_df["raw_feature"].isin(numeric_raw)].head(SHAP_THRESHOLD_TOP_N_NUMERIC)
    threshold_rows = []
    paths: List[str] = []
    for _, r in threshold_candidates.iterrows():
        row = save_threshold_plot_for_feature(
            raw_feature=r["raw_feature"],
            display_feature=r["display_feature"],
            grouped_shap_df=grouped_shap_df,
            X_raw_all=X_train,
            p_calibrated_all=p_train_oof,
        )
        row["threshold_derivation_split"] = "training"
        row["threshold_derivation_method"] = "out_of_fold_training_SHAP"
        row["held_out_test_used_for_threshold"] = False
        threshold_rows.append(row)
        if row.get("plot_path"):
            paths.append(row["plot_path"])
    return pd.DataFrame(threshold_rows), paths


def run_grouped_shap_and_threshold_analysis(
    pre: Step2Preprocessor,
    model: LGBMClassifier,
    calibrator: IsotonicRegression,
    X_test: pd.DataFrame,
    y_test: np.ndarray,
    p_test_calibrated: np.ndarray,
    X_train: pd.DataFrame,
    y_train: np.ndarray,
    groups_train: np.ndarray,
    params_for_refit: Dict[str, Any],
    locked_n_estimators: int,
    seed: int,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, List[str]], List[str]]:
    """Create descriptive held-out test SHAP outputs and training-OOF SHAP thresholds.

    Held-out test SHAP is descriptive only. SHAP-informed thresholds are derived
    exclusively from out-of-fold SHAP values generated inside the training split.
    """
    del calibrator
    print("\nRunning descriptive held-out test-set TreeSHAP for grouped features...")
    grouped_shap_test, grouped_data_test, importance_test, mapping_test = compute_grouped_shap(
        pre=pre,
        model=model,
        X_raw=X_test,
        y_true=y_test,
        p_calibrated=p_test_calibrated,
    )
    paths: List[str] = []
    grouped_shap_test.to_csv(os.path.join(OUTPUT_DIR, "grouped_shap_values_test_descriptive_best_model.csv"), index=False)
    grouped_data_test.to_csv(os.path.join(OUTPUT_DIR, "grouped_shap_feature_values_test_descriptive_best_model.csv"), index=False)
    importance_test.to_csv(os.path.join(OUTPUT_DIR, "grouped_shap_importance_test_descriptive_best_model.csv"), index=False)
    with open(os.path.join(OUTPUT_DIR, "grouped_shap_feature_mapping_test_descriptive.json"), "w") as f:
        json.dump(json_native(mapping_test), f, indent=2, sort_keys=True)

    paths.append(save_grouped_shap_beeswarm(grouped_shap_test, grouped_data_test, importance_test))
    paths.append(save_grouped_shap_bar(importance_test))

    print("Running training-split out-of-fold TreeSHAP for threshold derivation...")
    grouped_shap_oof, grouped_data_oof, importance_oof, mapping_oof = compute_training_oof_grouped_shap(
        X_train=X_train.reset_index(drop=True),
        y_train=y_train,
        groups_train=groups_train,
        params=params_for_refit,
        locked_n_estimators=locked_n_estimators,
        seed=seed,
    )
    grouped_shap_oof.to_csv(os.path.join(OUTPUT_DIR, "grouped_shap_values_training_oof_threshold_derivation.csv"), index=False)
    grouped_data_oof.to_csv(os.path.join(OUTPUT_DIR, "grouped_shap_feature_values_training_oof_threshold_derivation.csv"), index=False)
    importance_oof.to_csv(os.path.join(OUTPUT_DIR, "grouped_shap_importance_training_oof_threshold_derivation.csv"), index=False)
    with open(os.path.join(OUTPUT_DIR, "grouped_shap_feature_mapping_training_oof_threshold_derivation.json"), "w") as f:
        json.dump(json_native(mapping_oof), f, indent=2, sort_keys=True)

    p_train_oof = grouped_shap_oof.sort_values("__row_id__")["__p_calibrated__"].to_numpy(dtype=float)
    shap_threshold_df, threshold_paths = make_training_oof_threshold_table(
        importance_df=importance_oof,
        grouped_shap_df=grouped_shap_oof,
        X_train=X_train.reset_index(drop=True),
        p_train_oof=p_train_oof,
    )
    shap_threshold_df.to_csv(os.path.join(OUTPUT_DIR, "SHAP_thresholds_training_oof_top20_numeric_features.csv"), index=False)
    paths.extend(threshold_paths)

    return importance_test, shap_threshold_df, mapping_test, paths


def add_dynamic_odi_features(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Derive Step 2 dynamic ODI variables from raw preoperative/postoperative ODI values.

    Derived variables
    -----------------
    ODI_change_rate:
        (postoperative ODI - preoperative ODI) / days between ODI assessments.
        This is intentionally a per-day rate and is not multiplied by 30.
        Negative values indicate ODI improvement; positive values indicate worsening.

    Relative ODI MCID:
        (preoperative ODI - postoperative ODI) / preoperative ODI >= 0.30.
        This variable is set to missing when baseline ODI is missing; when baseline ODI is 0,
        it is coded as 0 because a 30% relative reduction is not attainable.

    The input input column ODI_MCID_binary, if present, is not used as a model
    feature. It is retained only for audit purposes.
    """
    out = df.copy()
    required = ["preop_ODI", "postop_ODI", DAYS_BETWEEN_PROM_COL]
    missing = [c for c in required if c not in out.columns]
    if missing:
        raise ValueError(
            "Cannot derive dynamic ODI features because required columns are missing: "
            + ", ".join(missing)
        )

    preop = pd.to_numeric(out["preop_ODI"].map(clean_scalar), errors="coerce")
    postop = pd.to_numeric(out["postop_ODI"].map(clean_scalar), errors="coerce")
    days_between = pd.to_numeric(out[DAYS_BETWEEN_PROM_COL].map(clean_scalar), errors="coerce")

    odi_change = postop - preop
    odi_improvement = preop - postop

    valid_rate = preop.notna() & postop.notna() & days_between.gt(0)
    prom_change_rate = pd.Series(np.nan, index=out.index, dtype="float")
    prom_change_rate.loc[valid_rate] = (odi_change.loc[valid_rate] / days_between.loc[valid_rate]).astype(float)

    valid_relative = preop.notna() & postop.notna() & preop.gt(0)
    zero_baseline_with_postop = preop.eq(0) & postop.notna()
    relative_fraction = odi_improvement / preop.replace(0, np.nan)
    relative_mcid = pd.Series(np.nan, index=out.index, dtype="float")
    relative_mcid.loc[valid_relative] = (
        relative_fraction.loc[valid_relative] >= RELATIVE_ODI_MCID_CUTOFF
    ).astype(float)
    # A 30% relative reduction is not attainable when baseline ODI is 0; code as not meeting relative MCID.
    relative_mcid.loc[zero_baseline_with_postop] = 0.0

    # Use formula-derived values to guarantee consistency across datasets.
    out["ODI_change"] = odi_change
    out[PROM_CHANGE_RATE_COL] = prom_change_rate
    out[RELATIVE_ODI_MCID_COL] = relative_mcid
    out["ODI_improvement_points_for_MCID_audit"] = odi_improvement
    out["ODI_relative_improvement_fraction_for_MCID_audit"] = relative_fraction

    audit_rows = [
        {"item": "PROM change rate definition", "value": "(postop_ODI - preop_ODI) / days_between_PROMs"},
        {"item": "Relative ODI MCID definition", "value": f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_CUTOFF}; preop_ODI=0 coded as 0 when postop_ODI is available"},
        {"item": "Rows", "value": int(len(out))},
        {"item": "Rows with valid preop/postop ODI and positive days between PROMs", "value": int(valid_rate.sum())},
        {"item": "Rows with calculable PROM change rate", "value": int(prom_change_rate.notna().sum())},
        {"item": "Rows with valid preop > 0 and postop ODI for relative MCID", "value": int(valid_relative.sum())},
        {"item": "Rows with preop ODI = 0 and postop ODI available coded as Relative ODI MCID = 0", "value": int(zero_baseline_with_postop.sum())},
        {"item": "Rows with Relative ODI MCID = 1", "value": int((relative_mcid == 1).sum())},
        {"item": "Rows with Relative ODI MCID = 0", "value": int((relative_mcid == 0).sum())},
    ]

    if INPUT_ODI_MCID_COL in out.columns:
        input = out[INPUT_ODI_MCID_COL].map(to_binary_target)
        audit_rows.extend([
            {"item": "Input ODI_MCID_binary column present", "value": True},
            {"item": "Input ODI_MCID_binary used as model feature", "value": False},
            {"item": "Input ODI_MCID_binary note", "value": "Retained only for audit; not included in FEATURES."},
            {"item": "Rows with non-missing input ODI_MCID_binary", "value": int(input.notna().sum())},
            {"item": "Rows with input ODI_MCID_binary = 1", "value": int((input == 1).sum())},
        ])
    else:
        audit_rows.append({"item": "Input ODI_MCID_binary column present", "value": False})

    return out, pd.DataFrame(audit_rows)


def methods_rationale_table() -> pd.DataFrame:
    return pd.DataFrame([
        {"item": "Primary model", "rationale": "LightGBM was tuned from scratch for the Step 2 ODI task using patient-level group-aware train/calibration/test splitting."},
        {"item": "Patient leakage control", "rationale": "PersonKey defines groups; no PersonKey overlap is allowed across train, calibration, or test."},
        {"item": "Hyperparameter tuning", "rationale": "All LightGBM and class-weight hyperparameters are selected only by StratifiedGroupKFold CV inside the training split."},
        {"item": "Early stopping", "rationale": "Early stopping is used only within CV folds to estimate boosting length; the selected model is refit on the full training split with the locked median best iteration."},
        {"item": "Primary tuning metric", "rationale": "Average Precision is used because the positive event rate is low and PR performance is more informative than accuracy."},
        {"item": "Class imbalance", "rationale": "Positive sample weight equals the natural negative/positive ratio multiplied by a tuned multiplier; the multiplier is not hand-picked."},
        {"item": "Calibration", "rationale": "Isotonic calibration is fitted only on the calibration split and never on the test split; rare-event calibration plots use quantile bins with zoomed axes and report Brier score and ECE."},
        {"item": "Threshold selection", "rationale": "The classification threshold is selected only on the calibration split using the pre-specified max-F1 rule."},
        {"item": "Final evaluation", "rationale": "The test set is isolated until the model-development pipeline is locked."},
        {"item": "Dynamic ODI features", "rationale": "PROM change rate is derived as (postop_ODI - preop_ODI) / days_between_PROMs without multiplying by 30; Relative ODI MCID is derived as at least 30% improvement from baseline ODI. The input ODI_MCID_binary column is audited but not used as a model feature."},
        {"item": "SHAP", "rationale": "Native LightGBM TreeSHAP is computed only after model selection, and one-hot categorical variables are grouped back to their raw clinical feature."},
        {"item": "SHAP thresholds", "rationale": "SHAP-informed thresholds are derived exclusively from training-split out-of-fold grouped SHAP values after model selection. Held-out test-set SHAP is descriptive only and is not used for threshold derivation."},
        {"item": "Learning curve", "rationale": "A group-aware learning curve using Average Precision is generated inside the training split for the locked best configuration and is not used for model selection."},
        {"item": "Generalization-gap audit", "rationale": "Training, cross-validation, calibration, and test performance are reported for the selected model. These diagnostics are not used for test-set model selection."},
        {"item": "Reproducibility", "rationale": "Exact best parameters, split assignments, model artifact, calibrator, selected threshold, and package versions are exported."},
    ])

# ============================================================
# 10) Main
# ============================================================


# ============================================================
# 11) Paired Step 2 model comparison helpers
# ============================================================

ROOT_OUTPUT_DIR = OUTPUT_DIR
ROOT_PLOT_DIR = PLOT_DIR

BASELINE_ONLY_FEATURES = list(BASE_FEATURES)
DYNAMIC_PROM_FEATURES = list(BASE_FEATURES) + list(STEP2_ODI_FEATURES)

BASELINE_CONTINUOUS_FEATURES = [f for f in CONTINUOUS_FEATURES if f in BASELINE_ONLY_FEATURES]
BASELINE_BINARY_FEATURES = [f for f in BINARY_FEATURES if f in BASELINE_ONLY_FEATURES]
BASELINE_ORDINAL_FEATURES = [f for f in ORDINAL_FEATURES if f in BASELINE_ONLY_FEATURES]
BASELINE_NOMINAL_FEATURES = [f for f in NOMINAL_FEATURES if f in BASELINE_ONLY_FEATURES]

DYNAMIC_CONTINUOUS_FEATURES = list(CONTINUOUS_FEATURES)
DYNAMIC_BINARY_FEATURES = list(BINARY_FEATURES)
DYNAMIC_ORDINAL_FEATURES = list(ORDINAL_FEATURES)
DYNAMIC_NOMINAL_FEATURES = list(NOMINAL_FEATURES)


def activate_feature_set(model_type: str) -> List[str]:
    """Set global feature lists used by the shared preprocessing/modeling functions."""
    global FEATURES, CONTINUOUS_FEATURES, BINARY_FEATURES, ORDINAL_FEATURES, NOMINAL_FEATURES
    if model_type == "baseline_only":
        FEATURES = list(BASELINE_ONLY_FEATURES)
        CONTINUOUS_FEATURES = list(BASELINE_CONTINUOUS_FEATURES)
        BINARY_FEATURES = list(BASELINE_BINARY_FEATURES)
        ORDINAL_FEATURES = list(BASELINE_ORDINAL_FEATURES)
        NOMINAL_FEATURES = list(BASELINE_NOMINAL_FEATURES)
    elif model_type == "dynamic_PROM_expanded":
        FEATURES = list(DYNAMIC_PROM_FEATURES)
        CONTINUOUS_FEATURES = list(DYNAMIC_CONTINUOUS_FEATURES)
        BINARY_FEATURES = list(DYNAMIC_BINARY_FEATURES)
        ORDINAL_FEATURES = list(DYNAMIC_ORDINAL_FEATURES)
        NOMINAL_FEATURES = list(DYNAMIC_NOMINAL_FEATURES)
    else:
        raise ValueError(f"Unknown model_type: {model_type}")
    return FEATURES


def paired_patient_level_delta_bootstrap_ci(
    y_true: np.ndarray,
    p_baseline: np.ndarray,
    p_expanded: np.ndarray,
    groups: np.ndarray,
    metric_name: str,
    n_bootstraps: int = N_BOOTSTRAPS,
    seed: int = RANDOM_STATE,
) -> Tuple[float, float, float, float]:
    y_true = np.asarray(y_true).astype(int)
    p_baseline = np.asarray(p_baseline).astype(float)
    p_expanded = np.asarray(p_expanded).astype(float)
    groups = np.asarray(groups)

    if metric_name == "AP":
        observed = safe_average_precision(y_true, p_expanded) - safe_average_precision(y_true, p_baseline)
    elif metric_name == "ROC_AUC":
        observed = safe_roc_auc(y_true, p_expanded) - safe_roc_auc(y_true, p_baseline)
    else:
        raise ValueError(f"Unsupported paired delta metric: {metric_name}")

    d = pd.DataFrame({"row_id": np.arange(len(y_true)), "group": groups, "y": y_true})
    group_y = d.groupby("group")["y"].max()
    pos_groups = group_y[group_y == 1].index.to_numpy()
    neg_groups = group_y[group_y == 0].index.to_numpy()
    if len(pos_groups) == 0 or len(neg_groups) == 0:
        return float(observed), np.nan, np.nan, np.nan

    rows_by_group = {g: d.loc[d["group"].eq(g), "row_id"].to_numpy() for g in group_y.index}
    rng = np.random.default_rng(seed)
    values = []
    for _ in range(n_bootstraps):
        sampled_pos = rng.choice(pos_groups, size=len(pos_groups), replace=True)
        sampled_neg = rng.choice(neg_groups, size=len(neg_groups), replace=True)
        sampled_groups = np.concatenate([sampled_pos, sampled_neg])
        idx = np.concatenate([rows_by_group[g] for g in sampled_groups])
        yy = y_true[idx]
        if len(np.unique(yy)) < 2:
            continue
        if metric_name == "AP":
            values.append(average_precision_score(yy, p_expanded[idx]) - average_precision_score(yy, p_baseline[idx]))
        else:
            values.append(roc_auc_score(yy, p_expanded[idx]) - roc_auc_score(yy, p_baseline[idx]))
    if not values:
        return float(observed), np.nan, np.nan, np.nan
    values = np.asarray(values, dtype=float)
    lo, hi = np.percentile(values, [2.5, 97.5])
    p_value = 2 * min(np.mean(values <= 0), np.mean(values >= 0))
    return float(observed), float(lo), float(hi), float(min(max(p_value, 0.0), 1.0))


def run_one_step2_model(
    model_type: str,
    work: pd.DataFrame,
    train: pd.DataFrame,
    cal: pd.DataFrame,
    test: pd.DataFrame,
    seed: int,
) -> Dict[str, Any]:
    """Tune, refit, calibrate, evaluate, and export one Step 2 model."""
    global OUTPUT_DIR, PLOT_DIR
    features = activate_feature_set(model_type)
    model_label = "Baseline-only" if model_type == "baseline_only" else "Dynamic PROM-expanded"
    model_dir = os.path.join(ROOT_OUTPUT_DIR, model_type)
    model_plot_dir = os.path.join(model_dir, "plots")
    model_artifact_dir = os.path.join(model_dir, "model_artifacts")
    os.makedirs(model_dir, exist_ok=True)
    os.makedirs(model_plot_dir, exist_ok=True)
    os.makedirs(model_artifact_dir, exist_ok=True)

    X_train = train[features].copy()
    y_train = train[TARGET_COL].to_numpy().astype(int)
    groups_train = train[GROUP_COL].to_numpy()
    X_cal = cal[features].copy()
    y_cal = cal[TARGET_COL].to_numpy().astype(int)
    X_test = test[features].copy()
    y_test = test[TARGET_COL].to_numpy().astype(int)
    groups_test = test[GROUP_COL].to_numpy()

    print("\n" + "=" * 100)
    print(f"Step 2 model: {model_label} ({model_type})")
    print(f"Features: {len(features)}")
    print("=" * 100)

    candidates, fold_metrics = tune_hyperparameters(
        X_train=X_train,
        y_train=y_train,
        groups_train=groups_train,
        seed=seed,
    )
    candidates.insert(0, "model_type", model_type)
    candidates.insert(1, "model_label", model_label)
    fold_metrics.insert(0, "model_type", model_type)
    fold_metrics.insert(1, "model_label", model_label)

    selected_cfg = candidates.iloc[0].copy()
    params = {k: selected_cfg[k] for k in LGBM_SEARCH_SPACE.keys()}
    params = sanitize_lgbm_params(params)
    locked_n_estimators = int(selected_cfg.get("locked_n_estimators_from_cv", params["n_estimators"]))
    params_for_refit = dict(params)
    params_for_refit["n_estimators"] = locked_n_estimators

    print("\nSelected configuration")
    print("Selection rule:", PRIMARY_MODEL_SELECTION_RULE)
    print(json.dumps(json_native(params_for_refit), indent=2, sort_keys=True))

    pre, model, _ = fit_model_pipeline(
        X_train,
        y_train,
        params=params,
        seed=seed + 10001,
        override_n_estimators=locked_n_estimators,
    )

    p_train_raw = predict_pipeline(pre, model, X_train)
    p_cal_raw = predict_pipeline(pre, model, X_cal)
    p_test_raw = predict_pipeline(pre, model, X_test)

    calibrator = IsotonicRegression(out_of_bounds="clip")
    calibrator.fit(p_cal_raw, y_cal)
    p_train = calibrator.transform(p_train_raw)
    p_cal = calibrator.transform(p_cal_raw)
    p_test = calibrator.transform(p_test_raw)

    selected_threshold, cal_thr_metrics = select_threshold(
        y_cal,
        p_cal,
        strategy=THRESHOLD_STRATEGY,
        fixed_threshold=FIXED_THRESHOLD,
        top_percent=THRESHOLD_TOP_PERCENT,
    )

    train_eval = evaluate_predictions(y_train, p_train, threshold=None, prefix="Train_")
    cal_eval = evaluate_predictions(y_cal, p_cal, threshold=selected_threshold, prefix="Calibration_")
    test_eval = evaluate_predictions(y_test, p_test, threshold=selected_threshold, prefix="Test_")

    ap_lo, ap_hi = patient_level_stratified_bootstrap_ci(
        y_test, p_test, groups_test, metric_name="AP", n_bootstraps=N_BOOTSTRAPS, seed=seed + 11,
    )
    auc_lo, auc_hi = patient_level_stratified_bootstrap_ci(
        y_test, p_test, groups_test, metric_name="ROC_AUC", n_bootstraps=N_BOOTSTRAPS, seed=seed + 13,
    )
    f1_lo, f1_hi = patient_level_stratified_bootstrap_ci(
        y_test, p_test, groups_test, metric_name="F1", threshold=selected_threshold, n_bootstraps=N_BOOTSTRAPS, seed=seed + 17,
    )

    pos_sample_weight_train = actual_positive_weight(y_train, multiplier=float(params["positive_weight_multiplier"]))

    summary = {
        "model_type": model_type,
        "model_label": model_label,
        "Model_selection_status": "Selected model",
        "Selection_rule": PRIMARY_MODEL_SELECTION_RULE,
        "candidate_id": int(selected_cfg["candidate_id"]),
        "Target": TARGET_COL,
        "Feature_count": len(features),
        "Train_N": int(len(y_train)),
        "Train_events": int(np.sum(y_train)),
        "Train_prevalence": float(np.mean(y_train)),
        "Calibration_N": int(len(y_cal)),
        "Calibration_events": int(np.sum(y_cal)),
        "Calibration_prevalence": float(np.mean(y_cal)),
        "Test_N": int(len(y_test)),
        "Test_events": int(np.sum(y_test)),
        "Test_prevalence": float(np.mean(y_test)),
        "Train_AP": train_eval["Train_AP"],
        "Train_ROC_AUC": train_eval["Train_ROC_AUC"],
        "Cross_validation_train_AP_mean": float(selected_cfg.get("cv_train_AP_mean", np.nan)),
        "Cross_validation_train_ROC_AUC_mean": float(selected_cfg.get("cv_train_ROC_AUC_mean", np.nan)),
        "Cross_validation_AP_mean": float(selected_cfg["cv_AP_mean"]),
        "Cross_validation_AP_SD": float(selected_cfg["cv_AP_SD"]),
        "Cross_validation_ROC_AUC_mean": float(selected_cfg["cv_ROC_AUC_mean"]),
        "Cross_validation_ROC_AUC_SD": float(selected_cfg["cv_ROC_AUC_SD"]),
        "Train_minus_CV_AP_gap": train_eval["Train_AP"] - float(selected_cfg["cv_AP_mean"]),
        "Train_minus_CV_ROC_AUC_gap": train_eval["Train_ROC_AUC"] - float(selected_cfg["cv_ROC_AUC_mean"]),
        "CV_minus_Test_AP_gap": float(selected_cfg["cv_AP_mean"]) - test_eval["Test_AP"],
        "CV_minus_Test_ROC_AUC_gap": float(selected_cfg["cv_ROC_AUC_mean"]) - test_eval["Test_ROC_AUC"],
        "CV_internal_train_minus_validation_AP_gap": float(selected_cfg.get("cv_train_minus_validation_AP_gap", np.nan)),
        "CV_internal_train_minus_validation_ROC_AUC_gap": float(selected_cfg.get("cv_train_minus_validation_ROC_AUC_gap", np.nan)),
        "locked_n_estimators_from_cv": locked_n_estimators,
        "Calibration_AP": cal_eval["Calibration_AP"],
        "Calibration_ROC_AUC": cal_eval["Calibration_ROC_AUC"],
        "Test_AP": test_eval["Test_AP"],
        "Test_AP_95CI_low_patient_bootstrap": ap_lo,
        "Test_AP_95CI_high_patient_bootstrap": ap_hi,
        "Test_ROC_AUC": test_eval["Test_ROC_AUC"],
        "Test_ROC_AUC_95CI_low_patient_bootstrap": auc_lo,
        "Test_ROC_AUC_95CI_high_patient_bootstrap": auc_hi,
        "Test_Brier_score": test_eval["Test_Brier_score"],
        "Test_ECE": test_eval["Test_ECE"],
        "Selected_threshold": selected_threshold,
        "Threshold_strategy": THRESHOLD_STRATEGY,
        "Positive_weight_multiplier": float(params["positive_weight_multiplier"]),
        "Positive_sample_weight_train": pos_sample_weight_train,
        "Test_F1": test_eval["Test_F1"],
        "Test_F1_95CI_low_patient_bootstrap": f1_lo,
        "Test_F1_95CI_high_patient_bootstrap": f1_hi,
        "Test_Sensitivity": test_eval["Test_Sensitivity"],
        "Test_Specificity": test_eval["Test_Specificity"],
        "Test_PPV": test_eval["Test_PPV"],
        "Test_NPV": test_eval["Test_NPV"],
        "Test_TP": test_eval["Test_TP"],
        "Test_FP": test_eval["Test_FP"],
        "Test_TN": test_eval["Test_TN"],
        "Test_FN": test_eval["Test_FN"],
        "Test_Top_5pct_event_rate": test_eval["Test_Top_5pct_event_rate"],
        "Test_Top_5pct_lift": test_eval["Test_Top_5pct_lift"],
        "Test_Top_5pct_captured_events": test_eval["Test_Top_5pct_captured_events"],
        "Best_params_JSON": json.dumps(json_native(params_for_refit), sort_keys=True),
    }

    calibration_row = {"model_type": model_type, "model_label": model_label, "candidate_id": int(selected_cfg["candidate_id"]), **cal_thr_metrics}

    pred_df = pd.DataFrame({
        "PersonKey": test[GROUP_COL].values,
        "row_index_in_test_split": np.arange(len(test)),
        "y_true": y_test,
        "p_test_calibrated": p_test,
        "p_test_raw": p_test_raw,
        "predicted_positive": (p_test >= selected_threshold).astype(int),
    })
    pred_df.to_csv(os.path.join(model_dir, f"test_predictions_{model_type}.csv"), index=False)
    # Root-level copies make downstream sensitivity/survival scripts easier to locate.
    pred_df.to_csv(os.path.join(ROOT_OUTPUT_DIR, f"test_predictions_Step2_{model_type}.csv"), index=False)

    save_pr_curve(y_test, p_test, os.path.join(model_plot_dir, f"PR_curve_{model_type}.png"), f"Step 2 LightGBM {model_label}: test set")
    save_roc_curve(y_test, p_test, os.path.join(model_plot_dir, f"ROC_curve_{model_type}.png"), f"Step 2 LightGBM {model_label}: test set")
    save_calibration_plot(y_test, p_test, os.path.join(model_plot_dir, f"Calibration_{model_type}.png"), f"Step 2 LightGBM {model_label}: test set")

    print(f"\nRunning group-aware AP learning curve for {model_label}...")
    save_learning_curve_grouped_ap(
        X_train=X_train,
        y_train=y_train,
        groups_train=groups_train,
        best_params=params_for_refit,
        path_png=os.path.join(model_plot_dir, f"Learning_curve_{model_type}.png"),
        path_csv=os.path.join(model_dir, f"learning_curve_{model_type}.csv"),
        title=f"Step 2 LightGBM {model_label}: group-aware learning curve",
    )

    artifact_path = os.path.join(model_artifact_dir, f"{model_type}_lightgbm_pipeline_artifact.joblib")
    joblib.dump({
        "model_type": model_type,
        "model_label": model_label,
        "preprocessor": pre,
        "model": model,
        "calibrator": calibrator,
        "selected_threshold": float(selected_threshold),
        "params": params_for_refit,
        "locked_n_estimators_from_cv": locked_n_estimators,
        "features": features,
        "target_col": TARGET_COL,
        "group_col": GROUP_COL,
        "random_state": seed,
    }, artifact_path)

    with open(os.path.join(model_artifact_dir, f"{model_type}_exact_settings.json"), "w") as f:
        json.dump(json_native({
            "model_type": model_type,
            "model_label": model_label,
            "selection_rule": PRIMARY_MODEL_SELECTION_RULE,
            "candidate_id": int(selected_cfg["candidate_id"]),
            "params": params_for_refit,
            "locked_n_estimators_from_cv": locked_n_estimators,
            "selected_threshold": float(selected_threshold),
            "features": features,
            "prom_change_rate_definition": "(postop_ODI - preop_ODI) / days_between_PROMs",
            "relative_mcid_definition": f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_CUTOFF}; preop_ODI=0 coded as 0 when postop_ODI is available",
            "search_space": LGBM_SEARCH_SPACE,
            "n_random_combinations": N_RANDOM_COMBINATIONS,
            "n_cv_folds": N_CV_FOLDS,
            "use_early_stopping_in_cv": USE_EARLY_STOPPING_IN_CV,
            "early_stopping_rounds": EARLY_STOPPING_ROUNDS,
            "threshold_strategy": THRESHOLD_STRATEGY,
            "calibration_method": CALIBRATION_METHOD,
        }), f, indent=2, sort_keys=True)

    # SHAP outputs are created for both models to mirror Step 1. The dynamic model
    # remains the primary interpretation model.
    old_out, old_plot = OUTPUT_DIR, PLOT_DIR
    try:
        OUTPUT_DIR = model_dir
        PLOT_DIR = model_plot_dir
        shap_importance_df, shap_threshold_df, shap_mapping, shap_plot_paths = run_grouped_shap_and_threshold_analysis(
            pre=pre,
            model=model,
            calibrator=calibrator,
            X_test=X_test,
            y_test=y_test,
            p_test_calibrated=p_test,
            X_train=X_train,
            y_train=y_train,
            groups_train=groups_train,
            params_for_refit=params_for_refit,
            locked_n_estimators=locked_n_estimators,
            seed=seed,
        )
        shap_importance_df.insert(0, "model_type", model_type)
        shap_importance_df.insert(1, "model_label", model_label)
        shap_threshold_df.insert(0, "model_type", model_type)
        shap_threshold_df.insert(1, "model_label", model_label)
    finally:
        OUTPUT_DIR, PLOT_DIR = old_out, old_plot

    print(
        f"{model_label} FINAL | CV AP={selected_cfg['cv_AP_mean']:.5f} ± {selected_cfg['cv_AP_SD']:.5f} | "
        f"Test AP={test_eval['Test_AP']:.5f} [{ap_lo:.5f}, {ap_hi:.5f}] | "
        f"Test AUC={test_eval['Test_ROC_AUC']:.5f} [{auc_lo:.5f}, {auc_hi:.5f}] | "
        f"threshold={selected_threshold:.6f}"
    )

    return {
        "model_type": model_type,
        "model_label": model_label,
        "features": features,
        "summary": summary,
        "calibration": calibration_row,
        "predictions": pred_df,
        "candidates": candidates,
        "fold_metrics": fold_metrics,
        "shap_importance": shap_importance_df,
        "shap_thresholds": shap_threshold_df,
        "artifact_path": artifact_path,
        "p_test": p_test,
        "p_test_raw": p_test_raw,
        "y_test": y_test,
        "groups_test": groups_test,
    }


def main() -> None:
    start = time.time()
    if not os.path.exists(INPUT_CSV):
        raise FileNotFoundError(f"Input file not found: {INPUT_CSV}")

    os.makedirs(ROOT_OUTPUT_DIR, exist_ok=True)
    os.makedirs(ROOT_PLOT_DIR, exist_ok=True)

    df = pd.read_csv(INPUT_CSV, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]

    print("Input:", INPUT_CSV)
    print("Input shape:", df.shape)

    df, dynamic_feature_audit = add_dynamic_odi_features(df)
    print("\nDynamic PROM feature derivation audit")
    print(dynamic_feature_audit.to_string(index=False))

    all_required_features = sorted(set(DYNAMIC_PROM_FEATURES))
    required_cols = all_required_features + [TARGET_COL, GROUP_COL]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        raise ValueError("Missing required columns:\n" + "\n".join(f" - {c}" for c in missing_cols))

    # The Step 2 cohort requires eligible preoperative and early postoperative PROM
    # values and a positive time interval between assessments. Relative MCID may be
    # missing when baseline PROM is zero and is handled by training-only imputation.
    preop = pd.to_numeric(df["preop_ODI"].map(clean_scalar), errors="coerce")
    postop = pd.to_numeric(df["postop_ODI"].map(clean_scalar), errors="coerce")
    days_between = pd.to_numeric(df[DAYS_BETWEEN_PROM_COL].map(clean_scalar), errors="coerce")
    eligible_prom_mask = preop.notna() & postop.notna() & days_between.gt(0)

    keep_extra_cols = [c for c in ["InstitutionName", "InstitutionNPI1"] if c in df.columns]
    work = df.loc[eligible_prom_mask, required_cols + keep_extra_cols].copy()
    work[TARGET_COL] = work[TARGET_COL].map(to_binary_target)
    before_target = len(work)
    work = work.dropna(subset=[TARGET_COL]).copy()
    work[TARGET_COL] = work[TARGET_COL].astype(int)
    dropped_target = before_target - len(work)

    if work[GROUP_COL].isna().any():
        raise ValueError(f"{int(work[GROUP_COL].isna().sum())} rows have missing {GROUP_COL}.")

    print(f"Rows after PROM eligibility filtering: {int(eligible_prom_mask.sum()):,}; dropped ineligible PROM rows: {len(df) - int(eligible_prom_mask.sum()):,}")
    print(f"Rows after target cleaning: {len(work):,}; dropped target missing: {dropped_target:,}")
    print(f"Events: {int(work[TARGET_COL].sum()):,}; prevalence: {work[TARGET_COL].mean():.5f}")

    train_mask, cal_mask, test_mask = patient_level_train_cal_test_split(
        work,
        target_col=TARGET_COL,
        group_col=GROUP_COL,
        test_fraction=TEST_FRACTION,
        calibration_fraction_of_remaining=CALIBRATION_FRACTION_OF_REMAINING,
        seed=RANDOM_STATE,
    )
    work["split"] = np.where(train_mask, "train", np.where(cal_mask, "calibration", "test"))

    train = work[work["split"] == "train"].reset_index(drop=True)
    cal = work[work["split"] == "calibration"].reset_index(drop=True)
    test = work[work["split"] == "test"].reset_index(drop=True)

    train_groups = set(train[GROUP_COL])
    cal_groups = set(cal[GROUP_COL])
    test_groups = set(test[GROUP_COL])
    assert train_groups.isdisjoint(cal_groups)
    assert train_groups.isdisjoint(test_groups)
    assert cal_groups.isdisjoint(test_groups)

    split_audit = build_split_audit(work, "split", TARGET_COL)
    institution_audit = build_institution_audit(work, "split")
    split_assignment = work[[GROUP_COL, "split", TARGET_COL]].drop_duplicates().sort_values([GROUP_COL, "split"])
    split_assignment_csv = os.path.join(ROOT_OUTPUT_DIR, "split_assignment_Step2_ODI.csv")
    split_assignment.to_csv(split_assignment_csv, index=False)

    print("\nSPLIT AUDIT")
    print(split_audit.to_string(index=False))

    baseline_result = run_one_step2_model(
        model_type="baseline_only",
        work=work,
        train=train,
        cal=cal,
        test=test,
        seed=RANDOM_STATE,
    )
    dynamic_result = run_one_step2_model(
        model_type="dynamic_PROM_expanded",
        work=work,
        train=train,
        cal=cal,
        test=test,
        seed=RANDOM_STATE,
    )

    # Paired incremental value of dynamic PROM information on the same test patients.
    y_test = baseline_result["y_test"]
    groups_test = baseline_result["groups_test"]
    p_baseline = baseline_result["p_test"]
    p_dynamic = dynamic_result["p_test"]

    delta_ap = paired_patient_level_delta_bootstrap_ci(
        y_test, p_baseline, p_dynamic, groups_test, metric_name="AP", n_bootstraps=N_BOOTSTRAPS, seed=RANDOM_STATE + 3001,
    )
    delta_auc = paired_patient_level_delta_bootstrap_ci(
        y_test, p_baseline, p_dynamic, groups_test, metric_name="ROC_AUC", n_bootstraps=N_BOOTSTRAPS, seed=RANDOM_STATE + 3003,
    )

    paired_comparison = pd.DataFrame([{
        "comparison": "dynamic_PROM_expanded_minus_baseline_only",
        "baseline_candidate_id": baseline_result["summary"]["candidate_id"],
        "dynamic_candidate_id": dynamic_result["summary"]["candidate_id"],
        "baseline_Test_AP": baseline_result["summary"]["Test_AP"],
        "dynamic_Test_AP": dynamic_result["summary"]["Test_AP"],
        "Delta_AP_dynamic_minus_baseline": delta_ap[0],
        "Delta_AP_95CI_low_patient_bootstrap": delta_ap[1],
        "Delta_AP_95CI_high_patient_bootstrap": delta_ap[2],
        "Delta_AP_bootstrap_p_value": delta_ap[3],
        "baseline_Test_ROC_AUC": baseline_result["summary"]["Test_ROC_AUC"],
        "dynamic_Test_ROC_AUC": dynamic_result["summary"]["Test_ROC_AUC"],
        "Delta_ROC_AUC_dynamic_minus_baseline": delta_auc[0],
        "Delta_ROC_AUC_95CI_low_patient_bootstrap": delta_auc[1],
        "Delta_ROC_AUC_95CI_high_patient_bootstrap": delta_auc[2],
        "Delta_ROC_AUC_bootstrap_p_value": delta_auc[3],
    }])

    summary_df = pd.DataFrame([baseline_result["summary"], dynamic_result["summary"]])
    calibration_df = pd.DataFrame([baseline_result["calibration"], dynamic_result["calibration"]])
    candidates_df = pd.concat([baseline_result["candidates"], dynamic_result["candidates"]], ignore_index=True)
    folds_df = pd.concat([baseline_result["fold_metrics"], dynamic_result["fold_metrics"]], ignore_index=True)
    shap_importance_df = pd.concat([baseline_result["shap_importance"], dynamic_result["shap_importance"]], ignore_index=True)
    shap_threshold_df = pd.concat([baseline_result["shap_thresholds"], dynamic_result["shap_thresholds"]], ignore_index=True)

    features_rows = []
    for model_type, feature_list in [("baseline_only", BASELINE_ONLY_FEATURES), ("dynamic_PROM_expanded", DYNAMIC_PROM_FEATURES)]:
        if model_type == "baseline_only":
            cont, binf, ordf, nomf = BASELINE_CONTINUOUS_FEATURES, BASELINE_BINARY_FEATURES, BASELINE_ORDINAL_FEATURES, BASELINE_NOMINAL_FEATURES
        else:
            cont, binf, ordf, nomf = DYNAMIC_CONTINUOUS_FEATURES, DYNAMIC_BINARY_FEATURES, DYNAMIC_ORDINAL_FEATURES, DYNAMIC_NOMINAL_FEATURES
        for f in feature_list:
            features_rows.append({
                "model_type": model_type,
                "Feature": f,
                "Display_name": pretty_feature_name(f),
                "Feature_group": "Dynamic_PROM" if f in STEP2_ODI_FEATURES else "Baseline",
                "Feature_type": "Continuous" if f in cont else "Binary" if f in binf else "Ordinal" if f in ordf else "Nominal" if f in nomf else "Unknown",
            })
    features_df = pd.DataFrame(features_rows)

    config_df = pd.DataFrame([
        {"Parameter": "INPUT_CSV", "Value": INPUT_CSV},
        {"Parameter": "TARGET_COL", "Value": TARGET_COL},
        {"Parameter": "GROUP_COL", "Value": GROUP_COL},
        {"Parameter": "RANDOM_STATE", "Value": RANDOM_STATE},
        {"Parameter": "TEST_FRACTION", "Value": TEST_FRACTION},
        {"Parameter": "CALIBRATION_FRACTION_OF_REMAINING", "Value": CALIBRATION_FRACTION_OF_REMAINING},
        {"Parameter": "N_CV_FOLDS", "Value": N_CV_FOLDS},
        {"Parameter": "N_RANDOM_COMBINATIONS", "Value": N_RANDOM_COMBINATIONS},
        {"Parameter": "N_BOOTSTRAPS", "Value": N_BOOTSTRAPS},
        {"Parameter": "USE_EARLY_STOPPING_IN_CV", "Value": USE_EARLY_STOPPING_IN_CV},
        {"Parameter": "EARLY_STOPPING_ROUNDS", "Value": EARLY_STOPPING_ROUNDS},
        {"Parameter": "THRESHOLD_STRATEGY", "Value": THRESHOLD_STRATEGY},
        {"Parameter": "CALIBRATION_METHOD", "Value": CALIBRATION_METHOD},
        {"Parameter": "SHAP_THRESHOLD_DERIVATION", "Value": "training_split_out_of_fold_SHAP_only"},
        {"Parameter": "HELD_OUT_TEST_USED_FOR_SHAP_THRESHOLDS", "Value": False},
        {"Parameter": "STEP2_DEATH_RETAINED_ANALYSIS_PERFORMED", "Value": False},
        {"Parameter": "PROM_CHANGE_RATE_DEFINITION", "Value": "(postop_ODI - preop_ODI) / days_between_PROMs"},
        {"Parameter": "RELATIVE_MCID_DEFINITION", "Value": f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_CUTOFF}"},
        {"Parameter": "lightgbm_version", "Value": lgb.__version__},
        {"Parameter": "sklearn_version", "Value": __import__("sklearn").__version__},
        {"Parameter": "shap_version", "Value": shap.__version__},
        {"Parameter": "python_version", "Value": platform.python_version()},
    ])

    summary_xlsx = os.path.join(ROOT_OUTPUT_DIR, "Step2_DynamicPROM_LightGBM_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        methods_rationale_table().to_excel(writer, sheet_name="methods_rationale", index=False)
        dynamic_feature_audit.to_excel(writer, sheet_name="dynamic_PROM_audit", index=False)
        summary_df.to_excel(writer, sheet_name="model_performance", index=False)
        paired_comparison.to_excel(writer, sheet_name="paired_comparison", index=False)
        candidates_df.to_excel(writer, sheet_name="cv_candidates_all_models", index=False)
        folds_df.to_excel(writer, sheet_name="cv_fold_metrics_all", index=False)
        calibration_df.to_excel(writer, sheet_name="cal_thresholds", index=False)
        features_df.to_excel(writer, sheet_name="features", index=False)
        split_audit.to_excel(writer, sheet_name="split_audit", index=False)
        institution_audit.to_excel(writer, sheet_name="institution_audit", index=False)
        split_assignment.to_excel(writer, sheet_name="split_assignment", index=False)
        config_df.to_excel(writer, sheet_name="run_config", index=False)
        shap_importance_df.to_excel(writer, sheet_name="grouped_SHAP_importance", index=False)
        shap_threshold_df.to_excel(writer, sheet_name="SHAP_thresholds", index=False)
        style_excel_workbook(writer)

    summary_df.to_csv(os.path.join(ROOT_OUTPUT_DIR, "model_performance.csv"), index=False)
    paired_comparison.to_csv(os.path.join(ROOT_OUTPUT_DIR, "paired_dynamic_PROM_comparison.csv"), index=False)
    candidates_df.to_csv(os.path.join(ROOT_OUTPUT_DIR, "cv_candidates_all_models.csv"), index=False)
    folds_df.to_csv(os.path.join(ROOT_OUTPUT_DIR, "cv_fold_metrics_all.csv"), index=False)
    shap_importance_df.to_csv(os.path.join(ROOT_OUTPUT_DIR, "SHAP_importance_all_models.csv"), index=False)
    shap_threshold_df.to_csv(os.path.join(ROOT_OUTPUT_DIR, "SHAP_thresholds_all_models.csv"), index=False)

    run_manifest = {
        "input_csv": INPUT_CSV,
        "output_dir": ROOT_OUTPUT_DIR,
        "target_col": TARGET_COL,
        "group_col": GROUP_COL,
        "design": "Step 2 paired comparison: baseline-only model vs dynamic PROM-expanded model on the same cohort and same patient-level split.",
        "baseline_features": BASELINE_ONLY_FEATURES,
        "dynamic_PROM_features": DYNAMIC_PROM_FEATURES,
        "prom_change_rate_definition": "(postop_ODI - preop_ODI) / days_between_PROMs",
        "relative_mcid_definition": f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_CUTOFF}; preop_ODI=0 coded as 0 when postop_ODI is available",
        "n_random_combinations_per_model": N_RANDOM_COMBINATIONS,
        "n_bootstraps": N_BOOTSTRAPS,
        "use_early_stopping_in_cv": USE_EARLY_STOPPING_IN_CV,
        "early_stopping_rounds": EARLY_STOPPING_ROUNDS,
        "test_fraction": TEST_FRACTION,
        "calibration_fraction_of_remaining": CALIBRATION_FRACTION_OF_REMAINING,
        "threshold_strategy": THRESHOLD_STRATEGY,
        "calibration_method": CALIBRATION_METHOD,
        "shap_threshold_derivation": "training_split_out_of_fold_SHAP_only",
        "held_out_test_used_for_shap_thresholds": False,
        "step2_death_retained_analysis_performed": False,
        "primary_model_selection_rule": PRIMARY_MODEL_SELECTION_RULE,
        "runtime_minutes": float((time.time() - start) / 60),
        "summary_xlsx": summary_xlsx,
        "paired_comparison_csv": os.path.join(ROOT_OUTPUT_DIR, "paired_dynamic_PROM_comparison.csv"),
        "split_assignment_csv": split_assignment_csv,
        "baseline_artifact": baseline_result["artifact_path"],
        "dynamic_artifact": dynamic_result["artifact_path"],
        "python_version": platform.python_version(),
        "lightgbm_version": lgb.__version__,
        "sklearn_version": __import__("sklearn").__version__,
        "shap_version": shap.__version__,
    }
    with open(os.path.join(ROOT_OUTPUT_DIR, "run_manifest.json"), "w") as f:
        json.dump(json_native(run_manifest), f, indent=2, sort_keys=True)

    zip_path = os.path.join(ROOT_OUTPUT_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip")
    tmp_zip_path = zip_path + ".tmp"
    with open(os.path.join(ROOT_OUTPUT_DIR, "DOWNLOAD_INSTRUCTIONS.txt"), "w") as f:
        f.write("All Step 2 paired LightGBM outputs were generated successfully.\n")
        f.write(f"ZIP archive: {zip_path}\n")
        f.write("If automatic Colab download is slow or stalls, download this ZIP manually from the Colab Files panel.\n")

    for pth in [zip_path, tmp_zip_path]:
        if os.path.exists(pth):
            os.remove(pth)
    try:
        zf = zipfile.ZipFile(tmp_zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL)
    except TypeError:
        zf = zipfile.ZipFile(tmp_zip_path, "w", compression=zipfile.ZIP_DEFLATED)
    n_zipped_files = 0
    with zf as z:
        for root, _, file_names in os.walk(ROOT_OUTPUT_DIR):
            for name in file_names:
                full_path = os.path.join(root, name)
                if full_path in {zip_path, tmp_zip_path}:
                    continue
                rel_path = os.path.relpath(full_path, ROOT_OUTPUT_DIR)
                z.write(full_path, arcname=rel_path)
                n_zipped_files += 1
    os.replace(tmp_zip_path, zip_path)
    zip_size_mb = os.path.getsize(zip_path) / (1024 ** 2)

    print("\n" + "=" * 100)
    print("STEP 2 PAIRED DYNAMIC PROM LIGHTGBM ANALYSIS COMPLETED")
    print("Main Excel:", summary_xlsx)
    print("Model performance:", os.path.join(ROOT_OUTPUT_DIR, "model_performance.csv"))
    print("Paired comparison:", os.path.join(ROOT_OUTPUT_DIR, "paired_dynamic_PROM_comparison.csv"))
    print("Run manifest:", os.path.join(ROOT_OUTPUT_DIR, "run_manifest.json"))
    print(f"ZIP: {zip_path}")
    print(f"ZIP size: {zip_size_mb:.2f} MB; files included: {n_zipped_files}")
    print("=" * 100)

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            href = "/files" + zip_path
            display(HTML(
                f'<p><b>Step 2 paired LightGBM output archive is ready.</b></p>'
                f'<p><a href="{href}" download>Click here to download the ZIP archive</a></p>'
                f'<p>Path: <code>{zip_path}</code></p>'
            ))
        except Exception as e:
            print("Download link display skipped:", e)

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", e)


if __name__ == "__main__":
    main()

# %% [markdown] Cell 5
# #**Step2_Sensitivity Analysis Across Different Diagnoses**

# %% Cell 6
# -*- coding: utf-8 -*-
"""
Step 2 dynamic ODI diagnosis-stratified sensitivity analysis
==================================================================

This script evaluates the locked Step 2 LightGBM models across prespecified
lumbar diagnosis subgroups.
It does not refit, retune, recalibrate, or re-optimize thresholds. Test-set
SHAP is used only for descriptive feature-attribution plots within strata.
"""

import os
import re
import sys
import json
import math
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from matplotlib import pyplot as plt
from sklearn.metrics import average_precision_score, roc_auc_score

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step2_Diagnosis_Sensitivity_DeltaAP_SHAP")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
TARGET_COL = "final_reop_step2"
N_BOOTSTRAPS = 2000
RANDOM_STATE = 20260524
DISPLAY_FEATURE_COUNT = 25
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

INPUT_CSV_CANDIDATES = [
    os.path.join(BASE_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(BASE_DIR, "Step2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step2_ODI_Cohort.csv"),
]

STEP2_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
]

STEP2_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
]

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

SUBGROUPS = [('finaldx_degenerative', 'Degenerative diagnosis'), ('finaldx_radicular', 'Radiculopathy diagnosis'), ('finaldx_stenosis', 'Spinal stenosis diagnosis'), ('finaldx_deformity_instability', 'Deformity or instability diagnosis'), ('finaldx_other_diagnosis', 'Other lumbar diagnosis')]
BINARY_MAPS = {'finaldx_degenerative': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'finaldx_radicular': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'finaldx_stenosis': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'finaldx_deformity_instability': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'finaldx_other_diagnosis': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}}

# ============================================================
# Helpers
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def parse_binary_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive", "performed"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative", "not performed"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("/", "_")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(average_precision_score(y, p))


def safe_roc_auc(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(roc_auc_score(y, p))


def top5_metrics(y: np.ndarray, p: np.ndarray) -> Dict[str, Any]:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    n = len(y)
    if n == 0:
        return {"top_5pct_n": 0, "top_5pct_event_rate": np.nan, "top_5pct_lift": np.nan}
    k = max(1, int(math.ceil(n * 0.05)))
    idx = np.argsort(-p)[:k]
    prevalence = float(np.mean(y))
    rate = float(np.mean(y[idx]))
    return {
        "top_5pct_n": int(k),
        "top_5pct_event_rate": rate,
        "top_5pct_lift": rate / prevalence if prevalence > 0 else np.nan,
        "top_5pct_captured_events": float(np.sum(y[idx]) / np.sum(y)) if np.sum(y) > 0 else np.nan,
    }

# ============================================================
# Source discovery
# ============================================================

def _archive_names(zip_path: str) -> List[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            return [n.replace(chr(92), "/") for n in zf.namelist()]
    except Exception:
        return []


def _folder_names(path: str) -> List[str]:
    names = []
    if not os.path.isdir(path):
        return names
    for root, _, files in os.walk(path):
        for fn in files:
            names.append(os.path.join(root, fn).replace(path + os.sep, "").replace(chr(92), "/"))
    return names


def prediction_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    required = [
        "test_predictions_Step2_baseline_only.csv",
        "test_predictions_Step2_dynamic_PROM.csv",
    ]
    root_ok = all(fn in basenames for fn in required)
    nested_ok = (
        any(fn == "test_predictions_baseline_only.csv" for fn in basenames)
        and any(fn == "test_predictions_dynamic_PROM.csv" for fn in basenames)
    )
    return bool(root_ok or nested_ok)


def shap_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    return (
        any("dynamic_PROM/" in n for n in names)
        and "grouped_shap_values_test_descriptive_best_model.csv" in basenames
        and "grouped_shap_feature_values_test_descriptive_best_model.csv" in basenames
    )


def ensure_step2_source_dir() -> str:
    for folder in STEP2_SOURCE_FOLDER_CANDIDATES:
        names = _folder_names(folder)
        if prediction_tables_present(names) and shap_tables_present(names):
            return folder

    candidate_archives = list(STEP2_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([
                os.path.join(root, fn) for fn in os.listdir(root)
                if fn.lower().endswith(".zip") and "step2" in fn.lower() and "lightgbm" in fn.lower()
            ])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive):
            names = _archive_names(archive)
            if prediction_tables_present(names) and shap_tables_present(names):
                archive_path = archive
                break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 2 LightGBM output folder/archive with paired test predictions and descriptive test SHAP tables was found. "
            "Run Step2_02_FinalLightGBM_SHAP_DynamicODI_TRAINING_OOF_THRESHOLDS.py first. "
            "Expected /content/Step2_DynamicPROM_LightGBM_outputs or its ZIP archive."
        )

    extract_root = os.path.join(BASE_DIR, "_step2_sensitivity_locked_lightgbm_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting Step 2 source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def find_prediction_csv(source_dir: str, model_type: str) -> str:
    preferred = f"test_predictions_Step2_{model_type}.csv"
    fallback = f"test_predictions_{model_type}.csv"
    matches = []
    for root, _, files in os.walk(source_dir):
        for fn in files:
            if fn in {preferred, fallback}:
                matches.append(os.path.join(root, fn))
    if not matches:
        raise FileNotFoundError(f"Missing Step 2 prediction table for {model_type}.")
    root_preferred = [p for p in matches if os.path.basename(p) == preferred]
    return sorted(root_preferred or matches)[0]


def find_dynamic_shap_tables(source_dir: str) -> Tuple[str, str, Optional[str]]:
    value_matches, data_matches, importance_matches = [], [], []
    for root, _, files in os.walk(source_dir):
        normalized_root = root.replace(chr(92), "/")
        if "dynamic_PROM" not in normalized_root:
            continue
        for fn in files:
            full = os.path.join(root, fn)
            if fn == "grouped_shap_values_test_descriptive_best_model.csv":
                value_matches.append(full)
            elif fn == "grouped_shap_feature_values_test_descriptive_best_model.csv":
                data_matches.append(full)
            elif fn == "grouped_shap_importance_test_descriptive_best_model.csv":
                importance_matches.append(full)
    if not value_matches or not data_matches:
        raise FileNotFoundError("Could not find Step 2 dynamic_PROM descriptive test SHAP tables.")
    return sorted(value_matches)[0], sorted(data_matches)[0], sorted(importance_matches)[0] if importance_matches else None

# ============================================================
# Data loading
# ============================================================

def find_step2_input_csv() -> str:
    for p in INPUT_CSV_CANDIDATES:
        if os.path.exists(p):
            if p.startswith(FALLBACK_DIR):
                dst = os.path.join(BASE_DIR, os.path.basename(p))
                if not os.path.exists(dst):
                    shutil.copy2(p, dst)
                return dst
            return p
    raise FileNotFoundError(f"Could not find Step 2 ODI cohort input file. Tried: {INPUT_CSV_CANDIDATES}")


def load_subgroup_map() -> pd.DataFrame:
    path = find_step2_input_csv()
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns:
        raise ValueError(f"Step 2 cohort file does not contain {GROUP_COL}.")
    missing = [c for c, _ in SUBGROUPS if c not in df.columns]
    if missing:
        raise ValueError(f"Step 2 cohort file is missing subgroup columns: {missing}")
    out = df[[GROUP_COL] + [c for c, _ in SUBGROUPS]].copy()
    for c, _ in SUBGROUPS:
        out[c] = out[c].map(lambda z: parse_binary_value(z, c)).fillna(0.0).astype(int)
    return out.groupby(GROUP_COL, dropna=False)[[c for c, _ in SUBGROUPS]].max().reset_index()


def load_paired_predictions(pred_source_dir: str) -> pd.DataFrame:
    base_path = find_prediction_csv(pred_source_dir, "baseline_only")
    dyn_path = find_prediction_csv(pred_source_dir, "dynamic_PROM")
    base = pd.read_csv(base_path)
    dyn = pd.read_csv(dyn_path)
    base.columns = [str(c).strip() for c in base.columns]
    dyn.columns = [str(c).strip() for c in dyn.columns]

    required = [GROUP_COL, "y_true", "p_test_calibrated"]
    for label, tbl in [("baseline-only", base), ("dynamic PROM", dyn)]:
        missing = [c for c in required if c not in tbl.columns]
        if missing:
            raise ValueError(f"Step 2 {label} prediction table is missing columns: {missing}")

    if "row_index_in_test_split" not in base.columns:
        base = base.reset_index().rename(columns={"index": "row_index_in_test_split"})
    if "row_index_in_test_split" not in dyn.columns:
        dyn = dyn.reset_index().rename(columns={"index": "row_index_in_test_split"})

    base_small = base[[GROUP_COL, "row_index_in_test_split", "y_true", "p_test_calibrated"]].rename(columns={"p_test_calibrated": "p_baseline"})
    dyn_small = dyn[[GROUP_COL, "row_index_in_test_split", "y_true", "p_test_calibrated"]].rename(columns={"p_test_calibrated": "p_dynamic"})

    merged = base_small.merge(
        dyn_small,
        on=[GROUP_COL, "row_index_in_test_split", "y_true"],
        how="inner",
        validate="one_to_one",
    )
    if len(merged) != len(base_small) or len(merged) != len(dyn_small):
        raise RuntimeError(
            f"Prediction-table pairing mismatch: baseline n={len(base_small)}, dynamic n={len(dyn_small)}, paired n={len(merged)}."
        )
    return merged

# ============================================================
# Statistics and plots
# ============================================================

def paired_delta_bootstrap(y: np.ndarray, p0: np.ndarray, p1: np.ndarray, groups: np.ndarray, metric: str, seed: int) -> Tuple[float, float, float, float]:
    y = np.asarray(y).astype(int)
    p0 = np.asarray(p0).astype(float)
    p1 = np.asarray(p1).astype(float)
    groups = np.asarray(groups)
    if metric == "AP":
        obs = safe_average_precision(y, p1) - safe_average_precision(y, p0)
    elif metric == "ROC_AUC":
        obs = safe_roc_auc(y, p1) - safe_roc_auc(y, p0)
    else:
        raise ValueError(metric)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return obs, np.nan, np.nan, np.nan

    d = pd.DataFrame({"idx": np.arange(len(y)), "g": groups, "y": y})
    gy = d.groupby("g", dropna=False)["y"].max()
    pos = gy[gy == 1].index.to_numpy()
    neg = gy[gy == 0].index.to_numpy()
    if len(pos) == 0 or len(neg) == 0:
        return obs, np.nan, np.nan, np.nan
    by_group = {g: d.loc[d["g"].eq(g), "idx"].to_numpy() for g in gy.index}
    rng = np.random.default_rng(seed)
    vals = []
    for _ in range(N_BOOTSTRAPS):
        sampled_groups = np.concatenate([rng.choice(pos, len(pos), replace=True), rng.choice(neg, len(neg), replace=True)])
        idx = np.concatenate([by_group[g] for g in sampled_groups])
        yy = y[idx]
        if len(np.unique(yy)) < 2:
            continue
        if metric == "AP":
            vals.append(average_precision_score(yy, p1[idx]) - average_precision_score(yy, p0[idx]))
        else:
            vals.append(roc_auc_score(yy, p1[idx]) - roc_auc_score(yy, p0[idx]))
    if not vals:
        return obs, np.nan, np.nan, np.nan
    vals = np.asarray(vals, dtype=float)
    lo, hi = np.percentile(vals, [2.5, 97.5])
    p = 2.0 * min(float(np.mean(vals <= 0)), float(np.mean(vals >= 0)))
    return float(obs), float(lo), float(hi), float(min(max(p, 0.0), 1.0))


def summarize_by_subgroup(df: pd.DataFrame, subgroup_col: str, subgroup_label: str) -> Dict[str, Any]:
    y = df["y_true"].to_numpy(dtype=int)
    p0 = df["p_baseline"].to_numpy(dtype=float)
    p1 = df["p_dynamic"].to_numpy(dtype=float)
    groups = df[GROUP_COL].to_numpy()
    ap0 = safe_average_precision(y, p0)
    ap1 = safe_average_precision(y, p1)
    auc0 = safe_roc_auc(y, p0)
    auc1 = safe_roc_auc(y, p1)
    d_ap, d_ap_lo, d_ap_hi, d_ap_p = paired_delta_bootstrap(y, p0, p1, groups, metric="AP", seed=RANDOM_STATE + 101)
    d_auc, d_auc_lo, d_auc_hi, d_auc_p = paired_delta_bootstrap(y, p0, p1, groups, metric="ROC_AUC", seed=RANDOM_STATE + 103)
    top0 = top5_metrics(y, p0)
    top1 = top5_metrics(y, p1)
    return {
        "subgroup_column": subgroup_col,
        "subgroup_label": subgroup_label,
        "n": int(len(df)),
        "events": int(np.sum(y)),
        "event_rate": float(np.mean(y)) if len(y) else np.nan,
        "baseline_AP": ap0,
        "dynamic_PROM_AP": ap1,
        "delta_AP_dynamic_minus_baseline": ap1 - ap0 if np.isfinite(ap0) and np.isfinite(ap1) else np.nan,
        "delta_AP_bootstrap": d_ap,
        "delta_AP_95CI_low": d_ap_lo,
        "delta_AP_95CI_high": d_ap_hi,
        "delta_AP_p": d_ap_p,
        "baseline_ROC_AUC": auc0,
        "dynamic_PROM_ROC_AUC": auc1,
        "delta_ROC_AUC_dynamic_minus_baseline": auc1 - auc0 if np.isfinite(auc0) and np.isfinite(auc1) else np.nan,
        "delta_ROC_AUC_bootstrap": d_auc,
        "delta_ROC_AUC_95CI_low": d_auc_lo,
        "delta_ROC_AUC_95CI_high": d_auc_hi,
        "delta_ROC_AUC_p": d_auc_p,
        "baseline_top5_event_rate": top0.get("top_5pct_event_rate", np.nan),
        "dynamic_top5_event_rate": top1.get("top_5pct_event_rate", np.nan),
        "baseline_top5_lift": top0.get("top_5pct_lift", np.nan),
        "dynamic_top5_lift": top1.get("top_5pct_lift", np.nan),
    }


def save_shap_beeswarm_for_rows(
    shap_values_df: pd.DataFrame,
    shap_data_df: pd.DataFrame,
    row_ids: np.ndarray,
    title: str,
    path: str,
    omit_display_features: Optional[List[str]] = None,
) -> Optional[str]:
    omit_display_features = set(omit_display_features or [])
    row_ids = np.asarray(row_ids, dtype=int)
    row_ids = row_ids[(row_ids >= 0) & (row_ids < len(shap_data_df))]
    row_ids = np.unique(row_ids)
    if len(row_ids) < 5:
        return None

    shap_sub = shap_values_df[shap_values_df["__row_id__"].isin(row_ids)].copy()
    shap_sub = shap_sub.sort_values("__row_id__").reset_index(drop=True)
    data_sub = shap_data_df.iloc[shap_sub["__row_id__"].astype(int).to_numpy()].reset_index(drop=True)

    feature_cols = [
        c for c in shap_sub.columns
        if not c.startswith("__") and c in data_sub.columns and c not in omit_display_features
    ]
    if len(feature_cols) == 0:
        return None
    importance = np.abs(shap_sub[feature_cols].to_numpy(dtype=float)).mean(axis=0)
    ordered_cols = [feature_cols[i] for i in np.argsort(-importance)[:DISPLAY_FEATURE_COUNT]]

    plt.figure(figsize=(10.5, 8.5))
    shap.summary_plot(
        shap_sub[ordered_cols].to_numpy(dtype=float),
        features=data_sub[ordered_cols],
        feature_names=ordered_cols,
        max_display=min(DISPLAY_FEATURE_COUNT, len(ordered_cols)),
        show=False,
        plot_size=None,
    )
    plt.title(title, fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig(path, dpi=300, bbox_inches="tight")
    plt.close()
    return path

# ============================================================
# Main
# ============================================================

def main():
    print("=" * 100)
    print("Step 2 dynamic ODI diagnosis-stratified sensitivity analysis")
    print("=" * 100)

    source_dir = ensure_step2_source_dir()
    pred = load_paired_predictions(source_dir)
    subgroup_map = load_subgroup_map()
    paired = pred.merge(subgroup_map, on=GROUP_COL, how="left")

    shap_values_path, shap_data_path, shap_importance_path = find_dynamic_shap_tables(source_dir)
    shap_values = pd.read_csv(shap_values_path)
    shap_data = pd.read_csv(shap_data_path)
    shap_importance = pd.read_csv(shap_importance_path) if shap_importance_path else pd.DataFrame()

    rows = []
    plot_rows = []
    for subgroup_col, subgroup_label in SUBGROUPS:
        if subgroup_col not in paired.columns:
            continue
        d = paired[paired[subgroup_col].fillna(0).astype(int).eq(1)].copy()
        if d.empty:
            rows.append({"subgroup_column": subgroup_col, "subgroup_label": subgroup_label, "n": 0, "events": 0})
            continue
        rows.append(summarize_by_subgroup(d, subgroup_col, subgroup_label))
        path = os.path.join(PLOT_DIR, f"SHAP_beeswarm_Step2_dynamic_PROM_{safe_filename(subgroup_col)}.png")
        saved = save_shap_beeswarm_for_rows(
            shap_values,
            shap_data,
            d["row_index_in_test_split"].to_numpy(dtype=int),
            title=f"Step 2 dynamic ODI model: {subgroup_label}",
            path=path,
            omit_display_features=[subgroup_label],
        )
        if saved:
            plot_rows.append({"subgroup_column": subgroup_col, "subgroup_label": subgroup_label, "plot_path": saved})

    summary = pd.DataFrame(rows)
    plots = pd.DataFrame(plot_rows)
    audit = pd.DataFrame([
        {"item": "step2_source_dir", "value": source_dir, "note": "Folder/archive read after primary Step 2 LightGBM run."},
        {"item": "baseline_prediction_file", "value": find_prediction_csv(source_dir, "baseline_only"), "note": ""},
        {"item": "dynamic_prediction_file", "value": find_prediction_csv(source_dir, "dynamic_PROM"), "note": ""},
        {"item": "dynamic_test_shap_values_file", "value": shap_values_path, "note": "Descriptive held-out test SHAP only; no threshold derivation."},
        {"item": "dynamic_test_shap_feature_values_file", "value": shap_data_path, "note": ""},
        {"item": "input_cohort_file", "value": find_step2_input_csv(), "note": "Used only to map subgroup labels to held-out test predictions."},
        {"item": "n_paired_test_rows", "value": int(len(pred)), "note": "Rows paired between baseline-only and dynamic PROM prediction tables."},
    ])

    summary_csv = os.path.join(TABLE_DIR, "Step2_diagnosis_stratified_deltaAP_SHAP_summary.csv")
    summary.to_csv(summary_csv, index=False)
    plots.to_csv(os.path.join(TABLE_DIR, "SHAP_beeswarm_plot_paths.csv"), index=False)
    audit.to_csv(os.path.join(AUDIT_DIR, "input_output_audit.csv"), index=False)
    if not shap_importance.empty:
        shap_importance.to_csv(os.path.join(TABLE_DIR, "dynamic_PROM_descriptive_test_SHAP_importance_source.csv"), index=False)

    xlsx_path = os.path.join(OUTPUT_DIR, "Step2_Diagnosis_Sensitivity_DeltaAP_SHAP_summary.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="subgroup_deltaAP", index=False)
        plots.to_excel(writer, sheet_name="SHAP_plots", index=False)
        audit.to_excel(writer, sheet_name="audit", index=False)
        if not shap_importance.empty:
            shap_importance.to_excel(writer, sheet_name="source_SHAP_importance", index=False)

    zip_path = os.path.join(BASE_DIR, "Step2_Diagnosis_Sensitivity_DeltaAP_SHAP.zip")
    tmp_zip = zip_path + ".tmp"
    for p in [zip_path, tmp_zip]:
        if os.path.exists(p):
            os.remove(p)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(OUTPUT_DIR):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(OUTPUT_DIR))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)

    print("DONE")
    print("Output folder:", OUTPUT_DIR)
    print("Summary CSV:", summary_csv)
    print("Summary Excel:", xlsx_path)
    print("ZIP:", zip_path)

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(
                f'<p><b>Step 2 dynamic ODI diagnosis-stratified sensitivity analysis outputs are ready.</b></p>'
                f'<p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p>'
                f'<p>Path: <code>{zip_path}</code></p>'
            ))
        except Exception as e:
            print("Download link display skipped:", repr(e))

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))


if __name__ == "__main__":
    main()

# %% [markdown] Cell 7
# #**Step2_Sensitivity Analysis Across Different Procedures**

# %% Cell 8
# -*- coding: utf-8 -*-
"""
Step 2 dynamic ODI procedure-stratified sensitivity analysis
==================================================================

This script evaluates the locked Step 2 LightGBM models across prespecified
lumbar procedure subgroups.
It does not refit, retune, recalibrate, or re-optimize thresholds. Test-set
SHAP is used only for descriptive feature-attribution plots within strata.
"""

import os
import re
import sys
import json
import math
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from matplotlib import pyplot as plt
from sklearn.metrics import average_precision_score, roc_auc_score

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step2_Procedure_Sensitivity_DeltaAP_SHAP")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
TARGET_COL = "final_reop_step2"
N_BOOTSTRAPS = 2000
RANDOM_STATE = 20260524
DISPLAY_FEATURE_COUNT = 25
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

INPUT_CSV_CANDIDATES = [
    os.path.join(BASE_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(BASE_DIR, "Step2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step2_ODI_Cohort.csv"),
]

STEP2_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
]

STEP2_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
]

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

SUBGROUPS = [('alif_llif', 'Anterior/lateral lumbar interbody fusion'), ('corpectomy', 'Corpectomy'), ('discectomy', 'Discectomy'), ('foraminotomy', 'Foraminotomy'), ('instrumentation', 'Instrumentation'), ('laminectomy_posterior_decompression', 'Posterior decompression or laminectomy'), ('pelvic_fixation', 'Pelvic fixation'), ('plf', 'Posterolateral fusion'), ('tlif_plif', 'Posterior/transforaminal lumbar interbody fusion'), ('other_lumbar_procedure', 'Other lumbar procedure')]
BINARY_MAPS = {'alif_llif': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'corpectomy': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'discectomy': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'foraminotomy': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'instrumentation': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'laminectomy_posterior_decompression': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'pelvic_fixation': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'plf': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'tlif_plif': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}, 'other_lumbar_procedure': {'yes': 1, 'no': 0, 'true': 1, 'false': 0, '1': 1, '0': 0, 'performed': 1, 'not performed': 0}}

# ============================================================
# Helpers
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def parse_binary_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive", "performed"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative", "not performed"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("/", "_")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(average_precision_score(y, p))


def safe_roc_auc(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(roc_auc_score(y, p))


def top5_metrics(y: np.ndarray, p: np.ndarray) -> Dict[str, Any]:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    n = len(y)
    if n == 0:
        return {"top_5pct_n": 0, "top_5pct_event_rate": np.nan, "top_5pct_lift": np.nan}
    k = max(1, int(math.ceil(n * 0.05)))
    idx = np.argsort(-p)[:k]
    prevalence = float(np.mean(y))
    rate = float(np.mean(y[idx]))
    return {
        "top_5pct_n": int(k),
        "top_5pct_event_rate": rate,
        "top_5pct_lift": rate / prevalence if prevalence > 0 else np.nan,
        "top_5pct_captured_events": float(np.sum(y[idx]) / np.sum(y)) if np.sum(y) > 0 else np.nan,
    }

# ============================================================
# Source discovery
# ============================================================

def _archive_names(zip_path: str) -> List[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            return [n.replace(chr(92), "/") for n in zf.namelist()]
    except Exception:
        return []


def _folder_names(path: str) -> List[str]:
    names = []
    if not os.path.isdir(path):
        return names
    for root, _, files in os.walk(path):
        for fn in files:
            names.append(os.path.join(root, fn).replace(path + os.sep, "").replace(chr(92), "/"))
    return names


def prediction_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    required = [
        "test_predictions_Step2_baseline_only.csv",
        "test_predictions_Step2_dynamic_PROM.csv",
    ]
    root_ok = all(fn in basenames for fn in required)
    nested_ok = (
        any(fn == "test_predictions_baseline_only.csv" for fn in basenames)
        and any(fn == "test_predictions_dynamic_PROM.csv" for fn in basenames)
    )
    return bool(root_ok or nested_ok)


def shap_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    return (
        any("dynamic_PROM/" in n for n in names)
        and "grouped_shap_values_test_descriptive_best_model.csv" in basenames
        and "grouped_shap_feature_values_test_descriptive_best_model.csv" in basenames
    )


def ensure_step2_source_dir() -> str:
    for folder in STEP2_SOURCE_FOLDER_CANDIDATES:
        names = _folder_names(folder)
        if prediction_tables_present(names) and shap_tables_present(names):
            return folder

    candidate_archives = list(STEP2_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([
                os.path.join(root, fn) for fn in os.listdir(root)
                if fn.lower().endswith(".zip") and "step2" in fn.lower() and "lightgbm" in fn.lower()
            ])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive):
            names = _archive_names(archive)
            if prediction_tables_present(names) and shap_tables_present(names):
                archive_path = archive
                break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 2 LightGBM output folder/archive with paired test predictions and descriptive test SHAP tables was found. "
            "Run Step2_02_FinalLightGBM_SHAP_DynamicODI_TRAINING_OOF_THRESHOLDS.py first. "
            "Expected /content/Step2_DynamicPROM_LightGBM_outputs or its ZIP archive."
        )

    extract_root = os.path.join(BASE_DIR, "_step2_sensitivity_locked_lightgbm_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting Step 2 source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def find_prediction_csv(source_dir: str, model_type: str) -> str:
    preferred = f"test_predictions_Step2_{model_type}.csv"
    fallback = f"test_predictions_{model_type}.csv"
    matches = []
    for root, _, files in os.walk(source_dir):
        for fn in files:
            if fn in {preferred, fallback}:
                matches.append(os.path.join(root, fn))
    if not matches:
        raise FileNotFoundError(f"Missing Step 2 prediction table for {model_type}.")
    root_preferred = [p for p in matches if os.path.basename(p) == preferred]
    return sorted(root_preferred or matches)[0]


def find_dynamic_shap_tables(source_dir: str) -> Tuple[str, str, Optional[str]]:
    value_matches, data_matches, importance_matches = [], [], []
    for root, _, files in os.walk(source_dir):
        normalized_root = root.replace(chr(92), "/")
        if "dynamic_PROM" not in normalized_root:
            continue
        for fn in files:
            full = os.path.join(root, fn)
            if fn == "grouped_shap_values_test_descriptive_best_model.csv":
                value_matches.append(full)
            elif fn == "grouped_shap_feature_values_test_descriptive_best_model.csv":
                data_matches.append(full)
            elif fn == "grouped_shap_importance_test_descriptive_best_model.csv":
                importance_matches.append(full)
    if not value_matches or not data_matches:
        raise FileNotFoundError("Could not find Step 2 dynamic_PROM descriptive test SHAP tables.")
    return sorted(value_matches)[0], sorted(data_matches)[0], sorted(importance_matches)[0] if importance_matches else None

# ============================================================
# Data loading
# ============================================================

def find_step2_input_csv() -> str:
    for p in INPUT_CSV_CANDIDATES:
        if os.path.exists(p):
            if p.startswith(FALLBACK_DIR):
                dst = os.path.join(BASE_DIR, os.path.basename(p))
                if not os.path.exists(dst):
                    shutil.copy2(p, dst)
                return dst
            return p
    raise FileNotFoundError(f"Could not find Step 2 ODI cohort input file. Tried: {INPUT_CSV_CANDIDATES}")


def load_subgroup_map() -> pd.DataFrame:
    path = find_step2_input_csv()
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns:
        raise ValueError(f"Step 2 cohort file does not contain {GROUP_COL}.")
    missing = [c for c, _ in SUBGROUPS if c not in df.columns]
    if missing:
        raise ValueError(f"Step 2 cohort file is missing subgroup columns: {missing}")
    out = df[[GROUP_COL] + [c for c, _ in SUBGROUPS]].copy()
    for c, _ in SUBGROUPS:
        out[c] = out[c].map(lambda z: parse_binary_value(z, c)).fillna(0.0).astype(int)
    return out.groupby(GROUP_COL, dropna=False)[[c for c, _ in SUBGROUPS]].max().reset_index()


def load_paired_predictions(pred_source_dir: str) -> pd.DataFrame:
    base_path = find_prediction_csv(pred_source_dir, "baseline_only")
    dyn_path = find_prediction_csv(pred_source_dir, "dynamic_PROM")
    base = pd.read_csv(base_path)
    dyn = pd.read_csv(dyn_path)
    base.columns = [str(c).strip() for c in base.columns]
    dyn.columns = [str(c).strip() for c in dyn.columns]

    required = [GROUP_COL, "y_true", "p_test_calibrated"]
    for label, tbl in [("baseline-only", base), ("dynamic PROM", dyn)]:
        missing = [c for c in required if c not in tbl.columns]
        if missing:
            raise ValueError(f"Step 2 {label} prediction table is missing columns: {missing}")

    if "row_index_in_test_split" not in base.columns:
        base = base.reset_index().rename(columns={"index": "row_index_in_test_split"})
    if "row_index_in_test_split" not in dyn.columns:
        dyn = dyn.reset_index().rename(columns={"index": "row_index_in_test_split"})

    base_small = base[[GROUP_COL, "row_index_in_test_split", "y_true", "p_test_calibrated"]].rename(columns={"p_test_calibrated": "p_baseline"})
    dyn_small = dyn[[GROUP_COL, "row_index_in_test_split", "y_true", "p_test_calibrated"]].rename(columns={"p_test_calibrated": "p_dynamic"})

    merged = base_small.merge(
        dyn_small,
        on=[GROUP_COL, "row_index_in_test_split", "y_true"],
        how="inner",
        validate="one_to_one",
    )
    if len(merged) != len(base_small) or len(merged) != len(dyn_small):
        raise RuntimeError(
            f"Prediction-table pairing mismatch: baseline n={len(base_small)}, dynamic n={len(dyn_small)}, paired n={len(merged)}."
        )
    return merged

# ============================================================
# Statistics and plots
# ============================================================

def paired_delta_bootstrap(y: np.ndarray, p0: np.ndarray, p1: np.ndarray, groups: np.ndarray, metric: str, seed: int) -> Tuple[float, float, float, float]:
    y = np.asarray(y).astype(int)
    p0 = np.asarray(p0).astype(float)
    p1 = np.asarray(p1).astype(float)
    groups = np.asarray(groups)
    if metric == "AP":
        obs = safe_average_precision(y, p1) - safe_average_precision(y, p0)
    elif metric == "ROC_AUC":
        obs = safe_roc_auc(y, p1) - safe_roc_auc(y, p0)
    else:
        raise ValueError(metric)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return obs, np.nan, np.nan, np.nan

    d = pd.DataFrame({"idx": np.arange(len(y)), "g": groups, "y": y})
    gy = d.groupby("g", dropna=False)["y"].max()
    pos = gy[gy == 1].index.to_numpy()
    neg = gy[gy == 0].index.to_numpy()
    if len(pos) == 0 or len(neg) == 0:
        return obs, np.nan, np.nan, np.nan
    by_group = {g: d.loc[d["g"].eq(g), "idx"].to_numpy() for g in gy.index}
    rng = np.random.default_rng(seed)
    vals = []
    for _ in range(N_BOOTSTRAPS):
        sampled_groups = np.concatenate([rng.choice(pos, len(pos), replace=True), rng.choice(neg, len(neg), replace=True)])
        idx = np.concatenate([by_group[g] for g in sampled_groups])
        yy = y[idx]
        if len(np.unique(yy)) < 2:
            continue
        if metric == "AP":
            vals.append(average_precision_score(yy, p1[idx]) - average_precision_score(yy, p0[idx]))
        else:
            vals.append(roc_auc_score(yy, p1[idx]) - roc_auc_score(yy, p0[idx]))
    if not vals:
        return obs, np.nan, np.nan, np.nan
    vals = np.asarray(vals, dtype=float)
    lo, hi = np.percentile(vals, [2.5, 97.5])
    p = 2.0 * min(float(np.mean(vals <= 0)), float(np.mean(vals >= 0)))
    return float(obs), float(lo), float(hi), float(min(max(p, 0.0), 1.0))


def summarize_by_subgroup(df: pd.DataFrame, subgroup_col: str, subgroup_label: str) -> Dict[str, Any]:
    y = df["y_true"].to_numpy(dtype=int)
    p0 = df["p_baseline"].to_numpy(dtype=float)
    p1 = df["p_dynamic"].to_numpy(dtype=float)
    groups = df[GROUP_COL].to_numpy()
    ap0 = safe_average_precision(y, p0)
    ap1 = safe_average_precision(y, p1)
    auc0 = safe_roc_auc(y, p0)
    auc1 = safe_roc_auc(y, p1)
    d_ap, d_ap_lo, d_ap_hi, d_ap_p = paired_delta_bootstrap(y, p0, p1, groups, metric="AP", seed=RANDOM_STATE + 101)
    d_auc, d_auc_lo, d_auc_hi, d_auc_p = paired_delta_bootstrap(y, p0, p1, groups, metric="ROC_AUC", seed=RANDOM_STATE + 103)
    top0 = top5_metrics(y, p0)
    top1 = top5_metrics(y, p1)
    return {
        "subgroup_column": subgroup_col,
        "subgroup_label": subgroup_label,
        "n": int(len(df)),
        "events": int(np.sum(y)),
        "event_rate": float(np.mean(y)) if len(y) else np.nan,
        "baseline_AP": ap0,
        "dynamic_PROM_AP": ap1,
        "delta_AP_dynamic_minus_baseline": ap1 - ap0 if np.isfinite(ap0) and np.isfinite(ap1) else np.nan,
        "delta_AP_bootstrap": d_ap,
        "delta_AP_95CI_low": d_ap_lo,
        "delta_AP_95CI_high": d_ap_hi,
        "delta_AP_p": d_ap_p,
        "baseline_ROC_AUC": auc0,
        "dynamic_PROM_ROC_AUC": auc1,
        "delta_ROC_AUC_dynamic_minus_baseline": auc1 - auc0 if np.isfinite(auc0) and np.isfinite(auc1) else np.nan,
        "delta_ROC_AUC_bootstrap": d_auc,
        "delta_ROC_AUC_95CI_low": d_auc_lo,
        "delta_ROC_AUC_95CI_high": d_auc_hi,
        "delta_ROC_AUC_p": d_auc_p,
        "baseline_top5_event_rate": top0.get("top_5pct_event_rate", np.nan),
        "dynamic_top5_event_rate": top1.get("top_5pct_event_rate", np.nan),
        "baseline_top5_lift": top0.get("top_5pct_lift", np.nan),
        "dynamic_top5_lift": top1.get("top_5pct_lift", np.nan),
    }


def save_shap_beeswarm_for_rows(
    shap_values_df: pd.DataFrame,
    shap_data_df: pd.DataFrame,
    row_ids: np.ndarray,
    title: str,
    path: str,
    omit_display_features: Optional[List[str]] = None,
) -> Optional[str]:
    omit_display_features = set(omit_display_features or [])
    row_ids = np.asarray(row_ids, dtype=int)
    row_ids = row_ids[(row_ids >= 0) & (row_ids < len(shap_data_df))]
    row_ids = np.unique(row_ids)
    if len(row_ids) < 5:
        return None

    shap_sub = shap_values_df[shap_values_df["__row_id__"].isin(row_ids)].copy()
    shap_sub = shap_sub.sort_values("__row_id__").reset_index(drop=True)
    data_sub = shap_data_df.iloc[shap_sub["__row_id__"].astype(int).to_numpy()].reset_index(drop=True)

    feature_cols = [
        c for c in shap_sub.columns
        if not c.startswith("__") and c in data_sub.columns and c not in omit_display_features
    ]
    if len(feature_cols) == 0:
        return None
    importance = np.abs(shap_sub[feature_cols].to_numpy(dtype=float)).mean(axis=0)
    ordered_cols = [feature_cols[i] for i in np.argsort(-importance)[:DISPLAY_FEATURE_COUNT]]

    plt.figure(figsize=(10.5, 8.5))
    shap.summary_plot(
        shap_sub[ordered_cols].to_numpy(dtype=float),
        features=data_sub[ordered_cols],
        feature_names=ordered_cols,
        max_display=min(DISPLAY_FEATURE_COUNT, len(ordered_cols)),
        show=False,
        plot_size=None,
    )
    plt.title(title, fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig(path, dpi=300, bbox_inches="tight")
    plt.close()
    return path

# ============================================================
# Main
# ============================================================

def main():
    print("=" * 100)
    print("Step 2 dynamic ODI procedure-stratified sensitivity analysis")
    print("=" * 100)

    source_dir = ensure_step2_source_dir()
    pred = load_paired_predictions(source_dir)
    subgroup_map = load_subgroup_map()
    paired = pred.merge(subgroup_map, on=GROUP_COL, how="left")

    shap_values_path, shap_data_path, shap_importance_path = find_dynamic_shap_tables(source_dir)
    shap_values = pd.read_csv(shap_values_path)
    shap_data = pd.read_csv(shap_data_path)
    shap_importance = pd.read_csv(shap_importance_path) if shap_importance_path else pd.DataFrame()

    rows = []
    plot_rows = []
    for subgroup_col, subgroup_label in SUBGROUPS:
        if subgroup_col not in paired.columns:
            continue
        d = paired[paired[subgroup_col].fillna(0).astype(int).eq(1)].copy()
        if d.empty:
            rows.append({"subgroup_column": subgroup_col, "subgroup_label": subgroup_label, "n": 0, "events": 0})
            continue
        rows.append(summarize_by_subgroup(d, subgroup_col, subgroup_label))
        path = os.path.join(PLOT_DIR, f"SHAP_beeswarm_Step2_dynamic_PROM_{safe_filename(subgroup_col)}.png")
        saved = save_shap_beeswarm_for_rows(
            shap_values,
            shap_data,
            d["row_index_in_test_split"].to_numpy(dtype=int),
            title=f"Step 2 dynamic ODI model: {subgroup_label}",
            path=path,
            omit_display_features=[subgroup_label],
        )
        if saved:
            plot_rows.append({"subgroup_column": subgroup_col, "subgroup_label": subgroup_label, "plot_path": saved})

    summary = pd.DataFrame(rows)
    plots = pd.DataFrame(plot_rows)
    audit = pd.DataFrame([
        {"item": "step2_source_dir", "value": source_dir, "note": "Folder/archive read after primary Step 2 LightGBM run."},
        {"item": "baseline_prediction_file", "value": find_prediction_csv(source_dir, "baseline_only"), "note": ""},
        {"item": "dynamic_prediction_file", "value": find_prediction_csv(source_dir, "dynamic_PROM"), "note": ""},
        {"item": "dynamic_test_shap_values_file", "value": shap_values_path, "note": "Descriptive held-out test SHAP only; no threshold derivation."},
        {"item": "dynamic_test_shap_feature_values_file", "value": shap_data_path, "note": ""},
        {"item": "input_cohort_file", "value": find_step2_input_csv(), "note": "Used only to map subgroup labels to held-out test predictions."},
        {"item": "n_paired_test_rows", "value": int(len(pred)), "note": "Rows paired between baseline-only and dynamic PROM prediction tables."},
    ])

    summary_csv = os.path.join(TABLE_DIR, "Step2_procedure_stratified_deltaAP_SHAP_summary.csv")
    summary.to_csv(summary_csv, index=False)
    plots.to_csv(os.path.join(TABLE_DIR, "SHAP_beeswarm_plot_paths.csv"), index=False)
    audit.to_csv(os.path.join(AUDIT_DIR, "input_output_audit.csv"), index=False)
    if not shap_importance.empty:
        shap_importance.to_csv(os.path.join(TABLE_DIR, "dynamic_PROM_descriptive_test_SHAP_importance_source.csv"), index=False)

    xlsx_path = os.path.join(OUTPUT_DIR, "Step2_Procedure_Sensitivity_DeltaAP_SHAP_summary.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="subgroup_deltaAP", index=False)
        plots.to_excel(writer, sheet_name="SHAP_plots", index=False)
        audit.to_excel(writer, sheet_name="audit", index=False)
        if not shap_importance.empty:
            shap_importance.to_excel(writer, sheet_name="source_SHAP_importance", index=False)

    zip_path = os.path.join(BASE_DIR, "Step2_Procedure_Sensitivity_DeltaAP_SHAP.zip")
    tmp_zip = zip_path + ".tmp"
    for p in [zip_path, tmp_zip]:
        if os.path.exists(p):
            os.remove(p)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(OUTPUT_DIR):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(OUTPUT_DIR))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)

    print("DONE")
    print("Output folder:", OUTPUT_DIR)
    print("Summary CSV:", summary_csv)
    print("Summary Excel:", xlsx_path)
    print("ZIP:", zip_path)

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(
                f'<p><b>Step 2 dynamic ODI procedure-stratified sensitivity analysis outputs are ready.</b></p>'
                f'<p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p>'
                f'<p>Path: <code>{zip_path}</code></p>'
            ))
        except Exception as e:
            print("Download link display skipped:", repr(e))

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))


if __name__ == "__main__":
    main()

# %% [markdown] Cell 9
# #**Step2_ODI_HospitalStratified_Sensitivity_DeltaAP_SHAP**

# %% Cell 10
# -*- coding: utf-8 -*-
"""
Step 2 dynamic ODI hospital-stratified sensitivity analysis
===========================================================

This script evaluates the locked Step 2 baseline-only and dynamic ODI-expanded
LightGBM models across hospital-level strata.
It uses saved held-out test-set predictions and descriptive held-out test-set
SHAP tables from the locked primary Step 2 analysis. It does not refit, retune,
recalibrate, or re-optimize thresholds.
"""

import os
import re
import sys
import json
import math
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from matplotlib import pyplot as plt
from sklearn.metrics import average_precision_score, roc_auc_score

warnings.filterwarnings("ignore")

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step2_HospitalStratified_Sensitivity_DeltaAP_SHAP")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
N_BOOTSTRAPS = 2000
RANDOM_STATE = 20260524
DISPLAY_FEATURE_COUNT = 25
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True
MIN_HOSPITAL_TEST_ROWS = 100
MIN_HOSPITAL_EVENTS = 5
LOWER_VOLUME_STRATUM_LABEL = "Lower-volume hospital stratum"

INPUT_CSV_CANDIDATES = [
    os.path.join(BASE_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(BASE_DIR, "Step2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step2_ODI_Cohort.csv"),
]
STEP2_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
]
STEP2_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
]
HOSPITAL_ID_CANDIDATES = ["InstitutionNPI1", "InstitutionName", "InstitutionKey", "HospitalID", "HospitalId", "SiteID", "FacilityID"]
HOSPITAL_LABEL_CANDIDATES = ["InstitutionName", "InstitutionNPI1", "InstitutionState"]
MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x

def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("/", "_")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"

def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int); p = np.asarray(p).astype(float)
    return np.nan if len(y) == 0 or len(np.unique(y)) < 2 else float(average_precision_score(y, p))

def safe_roc_auc(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int); p = np.asarray(p).astype(float)
    return np.nan if len(y) == 0 or len(np.unique(y)) < 2 else float(roc_auc_score(y, p))

def top5_metrics(y: np.ndarray, p: np.ndarray) -> Dict[str, Any]:
    y = np.asarray(y).astype(int); p = np.asarray(p).astype(float); n = len(y)
    if n == 0:
        return {"top_5pct_n": 0, "top_5pct_event_rate": np.nan, "top_5pct_lift": np.nan}
    k = max(1, int(math.ceil(n * 0.05))); idx = np.argsort(-p)[:k]
    prev = float(np.mean(y)); rate = float(np.mean(y[idx]))
    return {"top_5pct_n": int(k), "top_5pct_event_rate": rate, "top_5pct_lift": rate / prev if prev > 0 else np.nan, "top_5pct_captured_events": float(np.sum(y[idx]) / np.sum(y)) if np.sum(y) > 0 else np.nan}

def _archive_names(zip_path: str) -> List[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            return [n.replace(chr(92), "/") for n in zf.namelist()]
    except Exception:
        return []

def _folder_names(path: str) -> List[str]:
    names = []
    if not os.path.isdir(path): return names
    for root, _, files in os.walk(path):
        for fn in files:
            names.append(os.path.join(root, fn).replace(path + os.sep, "").replace(chr(92), "/"))
    return names

def prediction_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    root_ok = "test_predictions_Step2_baseline_only.csv" in basenames and "test_predictions_Step2_dynamic_PROM.csv" in basenames
    nested_ok = "test_predictions_baseline_only.csv" in basenames and "test_predictions_dynamic_PROM.csv" in basenames
    return root_ok or nested_ok

def shap_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    return any("dynamic_PROM/" in n for n in names) and "grouped_shap_values_test_descriptive_best_model.csv" in basenames and "grouped_shap_feature_values_test_descriptive_best_model.csv" in basenames

def ensure_step2_source_dir() -> str:
    for folder in STEP2_SOURCE_FOLDER_CANDIDATES:
        names = _folder_names(folder)
        if prediction_tables_present(names) and shap_tables_present(names):
            return folder
    candidate_archives = list(STEP2_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([os.path.join(root, fn) for fn in os.listdir(root) if fn.lower().endswith(".zip") and "step2" in fn.lower() and "lightgbm" in fn.lower()])
    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive):
            names = _archive_names(archive)
            if prediction_tables_present(names) and shap_tables_present(names):
                archive_path = archive; break
    if archive_path is None:
        raise FileNotFoundError("No Step 2 LightGBM output folder/archive was found. Run the corrected Step 2 final LightGBM script first.")
    extract_root = os.path.join(BASE_DIR, "_step2_hospital_stratified_locked_lightgbm_source")
    if os.path.isdir(extract_root): shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting Step 2 source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root

def find_prediction_csv(source_dir: str, model_type: str) -> str:
    preferred = f"test_predictions_Step2_{model_type}.csv"; fallback = f"test_predictions_{model_type}.csv"
    matches = []
    for root, _, files in os.walk(source_dir):
        for fn in files:
            if fn in {preferred, fallback}: matches.append(os.path.join(root, fn))
    if not matches: raise FileNotFoundError(f"Missing Step 2 prediction table for {model_type}.")
    root_preferred = [p for p in matches if os.path.basename(p) == preferred]
    return sorted(root_preferred or matches)[0]

def find_dynamic_shap_tables(source_dir: str) -> Tuple[str, str, Optional[str]]:
    value_matches, data_matches, imp_matches = [], [], []
    for root, _, files in os.walk(source_dir):
        if "dynamic_PROM" not in root.replace(chr(92), "/"): continue
        for fn in files:
            full = os.path.join(root, fn)
            if fn == "grouped_shap_values_test_descriptive_best_model.csv": value_matches.append(full)
            elif fn == "grouped_shap_feature_values_test_descriptive_best_model.csv": data_matches.append(full)
            elif fn == "grouped_shap_importance_test_descriptive_best_model.csv": imp_matches.append(full)
    if not value_matches or not data_matches: raise FileNotFoundError("Missing dynamic_PROM descriptive test SHAP tables.")
    return sorted(value_matches)[0], sorted(data_matches)[0], sorted(imp_matches)[0] if imp_matches else None

def find_step2_input_csv() -> str:
    for p in INPUT_CSV_CANDIDATES:
        if os.path.exists(p):
            if p.startswith(FALLBACK_DIR):
                dst = os.path.join(BASE_DIR, os.path.basename(p))
                if not os.path.exists(dst): shutil.copy2(p, dst)
                return dst
            return p
    raise FileNotFoundError(f"Could not find Step 2 ODI cohort input file. Tried: {INPUT_CSV_CANDIDATES}")

def find_existing_column(columns: List[str], candidates: List[str], required: bool = True) -> Optional[str]:
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns: return c
        if c.lower() in lookup: return lookup[c.lower()]
    if required: raise ValueError(f"Could not find required column among: {candidates}")
    return None

def load_hospital_map() -> Tuple[pd.DataFrame, pd.DataFrame]:
    path = find_step2_input_csv()
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns: raise ValueError(f"Step 2 cohort file does not contain {GROUP_COL}.")
    hospital_id_col = find_existing_column(df.columns.tolist(), HOSPITAL_ID_CANDIDATES, required=True)
    hospital_label_col = find_existing_column(df.columns.tolist(), HOSPITAL_LABEL_CANDIDATES, required=False)
    keep_cols = [GROUP_COL, hospital_id_col]
    if hospital_label_col and hospital_label_col not in keep_cols: keep_cols.append(hospital_label_col)
    raw = df[keep_cols].copy()
    raw[hospital_id_col] = raw[hospital_id_col].map(clean_scalar).astype("object")
    raw = raw[raw[hospital_id_col].notna()].copy()
    raw["hospital_id"] = raw[hospital_id_col].astype(str)
    raw["hospital_label_source"] = raw[hospital_label_col].map(clean_scalar).astype("object").astype(str) if hospital_label_col else raw["hospital_id"]
    ambiguity = raw.groupby(GROUP_COL, dropna=False)["hospital_id"].nunique(dropna=True).reset_index(name="n_hospitals_per_patient")
    n_ambiguous = int(ambiguity["n_hospitals_per_patient"].gt(1).sum())
    rows = []
    for person, g in raw.groupby(GROUP_COL, dropna=False):
        counts = g["hospital_id"].value_counts(dropna=True)
        if counts.empty: continue
        hid = str(counts.index[0])
        labels = g.loc[g["hospital_id"].astype(str).eq(hid), "hospital_label_source"].dropna().astype(str)
        rows.append({GROUP_COL: person, "hospital_id": hid, "hospital_label_source": labels.iloc[0] if len(labels) else hid})
    hospital_map = pd.DataFrame(rows)
    audit = pd.DataFrame([
        {"item": "input_file", "value": path, "note": ""},
        {"item": "hospital_id_column", "value": hospital_id_col, "note": ""},
        {"item": "hospital_label_column", "value": hospital_label_col if hospital_label_col else "none", "note": ""},
        {"item": "patients_with_hospital_assignment", "value": int(hospital_map[GROUP_COL].nunique()), "note": ""},
        {"item": "patients_with_multiple_hospital_ids", "value": n_ambiguous, "note": "modal hospital used for patient-level mapping"},
    ])
    return hospital_map, audit

def load_paired_predictions(source_dir: str) -> pd.DataFrame:
    base = pd.read_csv(find_prediction_csv(source_dir, "baseline_only"))
    dyn = pd.read_csv(find_prediction_csv(source_dir, "dynamic_PROM"))
    base.columns = [str(c).strip() for c in base.columns]; dyn.columns = [str(c).strip() for c in dyn.columns]
    required = [GROUP_COL, "y_true", "p_test_calibrated"]
    for label, tbl in [("baseline-only", base), ("dynamic PROM", dyn)]:
        missing = [c for c in required if c not in tbl.columns]
        if missing: raise ValueError(f"Step 2 {label} prediction table is missing columns: {missing}")
    if "row_index_in_test_split" not in base.columns: base = base.reset_index().rename(columns={"index": "row_index_in_test_split"})
    if "row_index_in_test_split" not in dyn.columns: dyn = dyn.reset_index().rename(columns={"index": "row_index_in_test_split"})
    b = base[[GROUP_COL, "row_index_in_test_split", "y_true", "p_test_calibrated"]].rename(columns={"p_test_calibrated": "p_baseline"})
    d = dyn[[GROUP_COL, "row_index_in_test_split", "y_true", "p_test_calibrated"]].rename(columns={"p_test_calibrated": "p_dynamic"})
    out = b.merge(d, on=[GROUP_COL, "row_index_in_test_split", "y_true"], how="inner", validate="one_to_one")
    if len(out) != len(b) or len(out) != len(d): raise RuntimeError("Prediction-table pairing mismatch between baseline-only and dynamic PROM models.")
    return out

def paired_delta_bootstrap(y: np.ndarray, p0: np.ndarray, p1: np.ndarray, groups: np.ndarray, metric: str, seed: int) -> Tuple[float, float, float, float]:
    y = np.asarray(y).astype(int); p0 = np.asarray(p0).astype(float); p1 = np.asarray(p1).astype(float); groups = np.asarray(groups)
    obs = (safe_average_precision(y, p1) - safe_average_precision(y, p0)) if metric == "AP" else (safe_roc_auc(y, p1) - safe_roc_auc(y, p0))
    if len(y) == 0 or len(np.unique(y)) < 2: return obs, np.nan, np.nan, np.nan
    d = pd.DataFrame({"idx": np.arange(len(y)), "g": groups, "y": y})
    gy = d.groupby("g", dropna=False)["y"].max(); pos = gy[gy == 1].index.to_numpy(); neg = gy[gy == 0].index.to_numpy()
    if len(pos) == 0 or len(neg) == 0: return obs, np.nan, np.nan, np.nan
    by_group = {g: d.loc[d["g"].eq(g), "idx"].to_numpy() for g in gy.index}
    rng = np.random.default_rng(seed); vals = []
    for _ in range(N_BOOTSTRAPS):
        sg = np.concatenate([rng.choice(pos, len(pos), True), rng.choice(neg, len(neg), True)])
        idx = np.concatenate([by_group[g] for g in sg]); yy = y[idx]
        if len(np.unique(yy)) < 2: continue
        vals.append((average_precision_score(yy, p1[idx]) - average_precision_score(yy, p0[idx])) if metric == "AP" else (roc_auc_score(yy, p1[idx]) - roc_auc_score(yy, p0[idx])))
    if not vals: return obs, np.nan, np.nan, np.nan
    vals = np.asarray(vals); lo, hi = np.percentile(vals, [2.5, 97.5]); p = 2 * min(float(np.mean(vals <= 0)), float(np.mean(vals >= 0)))
    return float(obs), float(lo), float(hi), float(min(max(p, 0), 1))

def summarize_group(df: pd.DataFrame, label: str, hospital_id: str) -> Dict[str, Any]:
    y = df["y_true"].to_numpy(int); p0 = df["p_baseline"].to_numpy(float); p1 = df["p_dynamic"].to_numpy(float); groups = df[GROUP_COL].to_numpy()
    ap0 = safe_average_precision(y,p0); ap1 = safe_average_precision(y,p1); auc0 = safe_roc_auc(y,p0); auc1 = safe_roc_auc(y,p1)
    d_ap, lo_ap, hi_ap, p_ap = paired_delta_bootstrap(y,p0,p1,groups,"AP",RANDOM_STATE+101)
    d_auc, lo_auc, hi_auc, p_auc = paired_delta_bootstrap(y,p0,p1,groups,"ROC_AUC",RANDOM_STATE+103)
    top0 = top5_metrics(y,p0); top1 = top5_metrics(y,p1)
    return {"hospital_stratum": label, "hospital_id": hospital_id, "n": int(len(df)), "events": int(np.sum(y)), "event_rate": float(np.mean(y)) if len(y) else np.nan, "baseline_AP": ap0, "dynamic_PROM_AP": ap1, "delta_AP_dynamic_minus_baseline": ap1-ap0 if np.isfinite(ap0) and np.isfinite(ap1) else np.nan, "delta_AP_bootstrap": d_ap, "delta_AP_95CI_low": lo_ap, "delta_AP_95CI_high": hi_ap, "delta_AP_p": p_ap, "baseline_ROC_AUC": auc0, "dynamic_PROM_ROC_AUC": auc1, "delta_ROC_AUC_dynamic_minus_baseline": auc1-auc0 if np.isfinite(auc0) and np.isfinite(auc1) else np.nan, "delta_ROC_AUC_bootstrap": d_auc, "delta_ROC_AUC_95CI_low": lo_auc, "delta_ROC_AUC_95CI_high": hi_auc, "delta_ROC_AUC_p": p_auc, "baseline_top5_event_rate": top0.get("top_5pct_event_rate", np.nan), "dynamic_top5_event_rate": top1.get("top_5pct_event_rate", np.nan), "baseline_top5_lift": top0.get("top_5pct_lift", np.nan), "dynamic_top5_lift": top1.get("top_5pct_lift", np.nan)}

def save_shap_beeswarm(shap_values: pd.DataFrame, shap_data: pd.DataFrame, row_ids: np.ndarray, title: str, path: str) -> Optional[str]:
    row_ids = np.asarray(row_ids, dtype=int); row_ids = row_ids[(row_ids >= 0) & (row_ids < len(shap_data))]; row_ids = np.unique(row_ids)
    if len(row_ids) < 5: return None
    shap_sub = shap_values[shap_values["__row_id__"].isin(row_ids)].sort_values("__row_id__").reset_index(drop=True)
    data_sub = shap_data.iloc[shap_sub["__row_id__"].astype(int).to_numpy()].reset_index(drop=True)
    cols = [c for c in shap_sub.columns if not c.startswith("__") and c in data_sub.columns]
    if not cols: return None
    imp = np.abs(shap_sub[cols].to_numpy(float)).mean(axis=0); ordered = [cols[i] for i in np.argsort(-imp)[:DISPLAY_FEATURE_COUNT]]
    plt.figure(figsize=(10.5,8.5))
    shap.summary_plot(shap_sub[ordered].to_numpy(float), features=data_sub[ordered], feature_names=ordered, max_display=min(DISPLAY_FEATURE_COUNT,len(ordered)), show=False, plot_size=None)
    plt.title(title, fontsize=13, fontweight="bold"); plt.tight_layout(); plt.savefig(path, dpi=300, bbox_inches="tight"); plt.close()
    return path

def main():
    print("="*100); print("Step 2 dynamic ODI hospital-stratified sensitivity analysis"); print("="*100)
    source_dir = ensure_step2_source_dir()
    pred = load_paired_predictions(source_dir)
    hospital_map, hospital_audit = load_hospital_map()
    paired = pred.merge(hospital_map, on=GROUP_COL, how="left")
    missing_hospital = int(paired["hospital_id"].isna().sum())
    paired = paired[paired["hospital_id"].notna()].copy()
    site_summary = paired.groupby(["hospital_id", "hospital_label_source"], dropna=False).agg(n=(GROUP_COL,"size"), events=("y_true","sum")).reset_index()
    site_summary["report_individually"] = site_summary["n"].ge(MIN_HOSPITAL_TEST_ROWS) & site_summary["events"].ge(MIN_HOSPITAL_EVENTS)
    label_lookup = dict(zip(site_summary["hospital_id"].astype(str), site_summary["hospital_label_source"].astype(str)))
    individual_ids = set(site_summary.loc[site_summary["report_individually"], "hospital_id"].astype(str))
    paired["hospital_stratum"] = paired["hospital_id"].astype(str).map(lambda z: label_lookup.get(z,z) if z in individual_ids else LOWER_VOLUME_STRATUM_LABEL)
    paired["hospital_stratum_id"] = paired["hospital_id"].astype(str).map(lambda z: z if z in individual_ids else "LOWER_VOLUME_POOL")
    shap_values_path, shap_data_path, shap_importance_path = find_dynamic_shap_tables(source_dir)
    shap_values = pd.read_csv(shap_values_path); shap_data = pd.read_csv(shap_data_path); shap_importance = pd.read_csv(shap_importance_path) if shap_importance_path else pd.DataFrame()
    rows=[]; plots=[]
    for (sid, label), d in paired.groupby(["hospital_stratum_id", "hospital_stratum"], dropna=False):
        d = d.copy()
        rows.append(summarize_group(d, label=str(label), hospital_id=str(sid)))
        path = os.path.join(PLOT_DIR, f"SHAP_beeswarm_Step2_dynamic_PROM_hospital_{safe_filename(label)}.png")
        saved = save_shap_beeswarm(shap_values, shap_data, d["row_index_in_test_split"].to_numpy(int), f"Step 2 dynamic ODI model: {label}", path)
        if saved: plots.append({"hospital_stratum": str(label), "hospital_id": str(sid), "plot_path": saved})
    summary = pd.DataFrame(rows).sort_values(["hospital_stratum_id"]).reset_index(drop=True)
    plots_df = pd.DataFrame(plots)
    audit = pd.concat([hospital_audit, pd.DataFrame([
        {"item":"step2_source_dir", "value":source_dir, "note":"Folder/archive read after primary Step 2 LightGBM run."},
        {"item":"baseline_prediction_file", "value":find_prediction_csv(source_dir,"baseline_only"), "note":""},
        {"item":"dynamic_prediction_file", "value":find_prediction_csv(source_dir,"dynamic_PROM"), "note":""},
        {"item":"dynamic_test_shap_values_file", "value":shap_values_path, "note":"Descriptive held-out test SHAP only."},
        {"item":"missing_hospital_assignment_in_test_predictions", "value":missing_hospital, "note":"Rows removed from hospital-stratified analysis."},
        {"item":"individual_hospital_min_test_rows", "value":MIN_HOSPITAL_TEST_ROWS, "note":""},
        {"item":"individual_hospital_min_events", "value":MIN_HOSPITAL_EVENTS, "note":""},
    ])], ignore_index=True)
    summary_csv = os.path.join(TABLE_DIR,"Step2_hospital_stratified_deltaAP_SHAP_summary.csv"); summary.to_csv(summary_csv,index=False)
    site_summary.to_csv(os.path.join(TABLE_DIR,"hospital_reporting_eligibility.csv"),index=False)
    plots_df.to_csv(os.path.join(TABLE_DIR,"SHAP_beeswarm_plot_paths.csv"),index=False)
    audit.to_csv(os.path.join(AUDIT_DIR,"input_output_audit.csv"),index=False)
    if not shap_importance.empty: shap_importance.to_csv(os.path.join(TABLE_DIR,"dynamic_PROM_descriptive_test_SHAP_importance_source.csv"),index=False)
    xlsx_path = os.path.join(OUTPUT_DIR,"Step2_HospitalStratified_Sensitivity_DeltaAP_SHAP_summary.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="hospital_deltaAP", index=False)
        site_summary.to_excel(writer, sheet_name="hospital_eligibility", index=False)
        plots_df.to_excel(writer, sheet_name="SHAP_plots", index=False)
        audit.to_excel(writer, sheet_name="audit", index=False)
    zip_path = os.path.join(BASE_DIR,"Step2_HospitalStratified_Sensitivity_DeltaAP_SHAP.zip"); tmp = zip_path + ".tmp"
    for p in [zip_path,tmp]:
        if os.path.exists(p): os.remove(p)
    with zipfile.ZipFile(tmp,"w",compression=zipfile.ZIP_DEFLATED,compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root,_,files in os.walk(OUTPUT_DIR):
            for fn in files:
                full = os.path.join(root,fn); rel = os.path.relpath(full, os.path.dirname(OUTPUT_DIR)); zf.write(full,rel)
    os.replace(tmp,zip_path)
    print("DONE"); print("Output folder:",OUTPUT_DIR); print("Summary Excel:",xlsx_path); print("ZIP:",zip_path)
    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(f'<p><b>Step 2 hospital-stratified outputs are ready.</b></p><p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p><p>Path: <code>{zip_path}</code></p>'))
        except Exception as e: print("Download link display skipped:",repr(e))
    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files; files.download(zip_path)
        except Exception as e: print("Automatic download skipped:",repr(e))

if __name__ == "__main__":
    main()

# %% [markdown] Cell 11
# #**Step2_ODI_HospitalHeldOut_InternalValidation**

# %% Cell 12
# -*- coding: utf-8 -*-
"""
Step 2 dynamic ODI hospital-held-out internal validation
========================================================

This script performs a true hospital-held-out internal validation for Step 2.
Hospitals are assigned exclusively to development-training, calibration, or
held-out validation splits. Held-out hospitals are never used for preprocessing,
hyperparameter tuning, model fitting, or probability calibration.

This script is intended to be run after the two primary Step 2 scripts. It reads
the primary Step 2 run manifest for audit/compatibility, but it refits models in
a hospital-level development subset because true hospital-held-out validation
requires disjoint hospitals for training/calibration/testing.
"""

import os, re, sys, json, math, time, zipfile, shutil, warnings, subprocess
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import lightgbm as lgb
    from lightgbm import LGBMClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lightgbm"])
    import lightgbm as lgb
    from lightgbm import LGBMClassifier

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder
from sklearn.model_selection import StratifiedGroupKFold, ParameterSampler
from sklearn.metrics import average_precision_score, roc_auc_score, brier_score_loss
from sklearn.isotonic import IsotonicRegression
import matplotlib.pyplot as plt
warnings.filterwarnings("ignore")

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step2_HospitalHeldOut_InternalValidation")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
MODEL_DIR = os.path.join(OUTPUT_DIR, "model_artifacts")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR, MODEL_DIR]: os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
TARGET_COL = "final_reop_step2"
RANDOM_STATE = 20260524
HOSPITAL_HELDOUT_TEST_FRACTION = 0.20
HOSPITAL_HELDOUT_CALIBRATION_FRACTION_OF_REMAINING = 0.20
HOSPITAL_HELDOUT_SPLIT_ATTEMPTS = 3000
HOSPITAL_HELDOUT_MIN_TRAIN_EVENTS = 30
HOSPITAL_HELDOUT_MIN_CALIBRATION_EVENTS = 5
HOSPITAL_HELDOUT_MIN_TEST_EVENTS = 10
N_CV_FOLDS = 5
N_RANDOM_COMBINATIONS = 300
USE_EARLY_STOPPING_IN_CV = True
EARLY_STOPPING_ROUNDS = 100
MIN_FINAL_N_ESTIMATORS = 50
N_BOOTSTRAPS = 2000
ECE_N_BINS = 10
N_JOBS = -1
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True
CALIBRATION_METHOD = "isotonic"

INPUT_CSV_CANDIDATES = [
    os.path.join(BASE_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(BASE_DIR, "Step2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step 2_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "Step2_ODI_Cohort.csv"),
]
PRIMARY_OUTPUT_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs"),
]
PRIMARY_OUTPUT_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs", "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step2_DynamicPROM_LightGBM_outputs.zip"),
]
HOSPITAL_ID_CANDIDATES = ["InstitutionNPI1", "InstitutionKey", "InstitutionName", "HospitalID", "HospitalId", "SiteID", "FacilityID"]
HOSPITAL_LABEL_CANDIDATES = ["InstitutionName", "InstitutionNPI1", "InstitutionState"]

BASELINE_FEATURES = [
    "finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis",
    "age","sex","race","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure",
    "connective_tissue_rheumatic_disease","diabetes_status","myocardial_infarction","renal_disease","institution_type",
    "institution_size","institution_region","asa","bmi","payer_status","alif_llif","corpectomy","discectomy",
    "foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif",
    "other_lumbar_procedure","number_operated_levels","operative_region_extent","PatTobaccoUse"]
assert len(BASELINE_FEATURES) == 35
RELATIVE_ODI_MCID_CUTOFF = 0.30
PROM_CHANGE_RATE_COL = "ODI_change_rate"
RELATIVE_ODI_MCID_COL = "ODI_relative_MCID_binary"
DAYS_BETWEEN_PROM_COL = "days_between_PROMs"
DYNAMIC_ODI_FEATURES = ["preop_ODI", "postop_ODI", "ODI_change", PROM_CHANGE_RATE_COL, RELATIVE_ODI_MCID_COL, "postop_ODI_day"]
CONTINUOUS_BASELINE_FEATURES = ["age", "bmi"]
CONTINUOUS_FEATURES_ALL = ["age","bmi","preop_ODI","postop_ODI","ODI_change",PROM_CHANGE_RATE_COL,"postop_ODI_day"]
BINARY_FEATURES_ALL = ["finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis","sex","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure","connective_tissue_rheumatic_disease","myocardial_infarction","renal_disease","institution_type","alif_llif","corpectomy","discectomy","foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif","other_lumbar_procedure","operative_region_extent",RELATIVE_ODI_MCID_COL]
ORDINAL_FEATURES_ALL = ["diabetes_status","institution_size","asa","number_operated_levels"]
NOMINAL_FEATURES_ALL = ["race","institution_region","payer_status","PatTobaccoUse"]
PREFERRED_NOMINAL_LEVELS = {"race":["White","Black","Other"],"institution_region":["South","North East","West","Midwest"],"payer_status":["Medicare","Commercial/Private","Other","Medicaid/Public/Government"],"PatTobaccoUse":["Unknown/Not reported/Multiple","Never","Former","Current"]}
MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}
BINARY_MAPS = {"sex":{"female":0,"f":0,"male":1,"m":1},"ethnicity":{"non-hispanic":0,"non hispanic":0,"hispanic":1},"cancer_status":{"no cancer":0,"no":0,"none":0,"any cancer":1,"yes":1,"cancer":1},"institution_type":{"hospital":0,"non-hospital":1,"non hospital":1,"nonhospital":1},"operative_region_extent":{"lumbar only":0,"extended_region_involvement":1,"extended region involvement":1,"extended":1}}
ORDINAL_MAPS = {"diabetes_status":{"no":0,"none":0,"0":0,"without comp":1,"without complication":1,"without complications":1,"1":1,"with comp":2,"with complication":2,"with complications":2,"2":2},"institution_size":{"between 1-99 beds":0,"1-99":0,"between 100-399 beds":1,"100-399":1,">= 400 beds":2,">=400 beds":2,">=400":2,">= 400":2},"asa":{"1":1,"i":1,"2":2,"ii":2,"3":3,"iii":3,"4":4,"iv":4,">=4":4,">= 4":4,"5":4,"v":4},"number_operated_levels":{"0":0,"1":1,"2":2,"3":3,"4":4,">=4":4,">= 4":4,"5":4,"6":4,"7":4,"8":4,"9":4,"10":4}}
LGBM_SEARCH_SPACE = {"n_estimators":[400,700,1000,1400,1800,2200,2600],"learning_rate":[0.003,0.005,0.008,0.01,0.02,0.03,0.05],"num_leaves":[7,15,31,63,127],"max_depth":[-1,2,3,5,7,9],"min_child_samples":[10,20,50,100,200,400],"subsample":[0.60,0.75,0.90,1.00],"subsample_freq":[0,1,2],"colsample_bytree":[0.60,0.75,0.90,1.00],"reg_alpha":[0.0,0.001,0.01,0.05,0.10,0.50,1.00,2.00],"reg_lambda":[0.0,0.001,0.01,0.05,0.10,0.50,1.00,2.00,5.00],"min_split_gain":[0.0,0.001,0.005,0.01,0.05,0.10],"max_bin":[63,127,255],"positive_weight_multiplier":[0.25,0.50,0.75,1.00,1.50,2.00,3.00,4.00,6.00,8.00]}
LGBM_INT_PARAMS = {"n_estimators","num_leaves","max_depth","min_child_samples","subsample_freq","max_bin"}
LGBM_FLOAT_PARAMS = {"learning_rate","subsample","colsample_bytree","reg_alpha","reg_lambda","min_split_gain","positive_weight_multiplier"}
PARAMETER_CANDIDATES = list(ParameterSampler(LGBM_SEARCH_SPACE, n_iter=N_RANDOM_COMBINATIONS, random_state=RANDOM_STATE))

def clean_scalar(x: Any) -> Any:
    if pd.isna(x): return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x

def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    return None if pd.isna(x) else str(x).strip().replace("≥", ">=").lower()

def to_binary_target(x: Any) -> float:
    sx = norm_text(x)
    if sx is None: return np.nan
    if sx in {"1","1.0","yes","y","true","t"}: return 1.0
    if sx in {"0","0.0","no","n","false","f"}: return 0.0
    try:
        v = float(sx); return float(v) if v in (0.0,1.0) else np.nan
    except Exception: return np.nan

def sanitize_lgbm_params(params: Dict[str, Any]) -> Dict[str, Any]:
    out={}
    for k,v in params.items():
        if k in LGBM_INT_PARAMS: out[k]=int(round(float(v)))
        elif k in LGBM_FLOAT_PARAMS: out[k]=float(v)
        elif isinstance(v,(np.integer,)): out[k]=int(v)
        elif isinstance(v,(np.floating,)): out[k]=float(v)
        else: out[k]=v
    return out

def json_native(obj: Any) -> Any:
    if isinstance(obj,dict): return {str(k):json_native(v) for k,v in obj.items()}
    if isinstance(obj,(list,tuple)): return [json_native(v) for v in obj]
    if isinstance(obj,np.ndarray): return obj.tolist()
    if isinstance(obj,(np.integer,)): return int(obj)
    if isinstance(obj,(np.floating,)): return float(obj)
    try:
        if pd.isna(obj): return None
    except Exception: pass
    return obj

def safe_average_precision(y,p):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float)
    return np.nan if len(y)==0 or len(np.unique(y))<2 else float(average_precision_score(y,p))
def safe_roc_auc(y,p):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float)
    return np.nan if len(y)==0 or len(np.unique(y))<2 else float(roc_auc_score(y,p))
def ece(y,p,n_bins=ECE_N_BINS):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); bins=np.linspace(0,1,n_bins+1); out=0.0
    if len(y)==0: return np.nan
    for i in range(n_bins):
        m=(p>=bins[i]) & ((p<=bins[i+1]) if i==n_bins-1 else (p<bins[i+1]))
        if np.any(m): out += (np.sum(m)/len(y))*abs(float(np.mean(y[m]))-float(np.mean(p[m])))
    return float(out)
def top5_metrics(y,p):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); n=len(y)
    if n==0: return {"top_5pct_n":0,"top_5pct_event_rate":np.nan,"top_5pct_lift":np.nan}
    k=max(1,int(math.ceil(n*0.05))); idx=np.argsort(-p)[:k]; prev=float(np.mean(y)); rate=float(np.mean(y[idx]))
    return {"top_5pct_n":int(k),"top_5pct_event_rate":rate,"top_5pct_lift":rate/prev if prev>0 else np.nan,"top_5pct_captured_events":float(np.sum(y[idx])/np.sum(y)) if np.sum(y)>0 else np.nan}
def make_weights(y,mult):
    y=np.asarray(y).astype(int); npos=int(np.sum(y==1)); nneg=int(np.sum(y==0))
    if npos==0: raise ValueError("No positive events in training data.")
    return np.where(y==1,(nneg/max(npos,1))*float(mult),1.0).astype(float)
def actual_positive_weight(y,mult):
    y=np.asarray(y).astype(int); return float((np.sum(y==0)/max(np.sum(y==1),1))*float(mult))
def find_existing_column(columns: List[str], candidates: List[str], what: str, required: bool=True) -> Optional[str]:
    lookup={str(c).strip().lower():str(c).strip() for c in columns}
    for c in candidates:
        if c in columns: return c
        if str(c).strip().lower() in lookup: return lookup[str(c).strip().lower()]
    if required: raise ValueError(f"Could not find {what}. Tried: {candidates}")
    return None

def ensure_primary_step2_outputs_present() -> Dict[str, Any]:
    for folder in PRIMARY_OUTPUT_FOLDER_CANDIDATES:
        manifest = os.path.join(folder, "run_manifest.json")
        if os.path.exists(manifest):
            with open(manifest, "r") as f: data = json.load(f)
            data["source_manifest_path"] = manifest; return data
    for archive in PRIMARY_OUTPUT_ARCHIVE_CANDIDATES:
        if os.path.exists(archive):
            with zipfile.ZipFile(archive, "r") as zf:
                matches = [n for n in zf.namelist() if os.path.basename(n) == "run_manifest.json"]
                if matches:
                    data = json.loads(zf.read(matches[0]).decode("utf-8")); data["source_manifest_path"] = archive + "::" + matches[0]; return data
    raise FileNotFoundError("Primary Step 2 LightGBM run_manifest.json was not found. Run Step2_02_FinalLightGBM_SHAP_DynamicODI_TRAINING_OOF_THRESHOLDS.py first.")

def find_input_csv() -> str:
    for p in INPUT_CSV_CANDIDATES:
        if os.path.exists(p):
            if p.startswith(FALLBACK_DIR):
                dst=os.path.join(BASE_DIR,os.path.basename(p))
                if not os.path.exists(dst): shutil.copy2(p,dst)
                return dst
            return p
    raise FileNotFoundError(f"Could not find Step 2 input file. Tried: {INPUT_CSV_CANDIDATES}")

def parse_binary_value(x: Any, feature: str) -> float:
    sx=norm_text(x)
    if sx is None: return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]: return float(BINARY_MAPS[feature][sx])
    if sx in {"1","1.0","yes","y","true","t","present","positive","performed"}: return 1.0
    if sx in {"0","0.0","no","n","false","f","absent","negative","not performed"}: return 0.0
    try:
        v=float(sx); return float(v) if v in (0.0,1.0) else np.nan
    except Exception: return np.nan

def parse_ordinal_value(x: Any, feature: str) -> float:
    sx=norm_text(x)
    if sx is None: return np.nan
    if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]: return float(ORDINAL_MAPS[feature][sx])
    try:
        v=float(sx)
        if feature=="asa": return float(min(max(int(round(v)),1),4))
        if feature=="number_operated_levels": return float(min(max(int(round(v)),0),4))
        if feature=="diabetes_status": return float(min(max(int(round(v)),0),2))
        if feature=="institution_size": return float(min(max(int(round(v)),0),2))
        return float(v)
    except Exception: return np.nan

def canonical_nominal(feature: str, x: Any) -> Any:
    x=clean_scalar(x)
    if pd.isna(x): return np.nan
    s=str(x).strip(); sl=s.lower()
    if feature=="race": return "White" if sl=="white" else ("Black" if sl=="black" else "Other")
    if feature=="institution_region": return {"south":"South","north east":"North East","northeast":"North East","north-east":"North East","west":"West","midwest":"Midwest","mid west":"Midwest"}.get(sl,s)
    if feature=="payer_status":
        if sl=="medicare": return "Medicare"
        if sl in {"commercial/private","commercial","private","commercial private"}: return "Commercial/Private"
        if sl in {"medicaid/public/government","medicaid","public","government","public/government"}: return "Medicaid/Public/Government"
        return "Other"
    if feature=="PatTobaccoUse": return "Never" if sl=="never" else ("Former" if sl=="former" else ("Current" if sl=="current" else "Unknown/Not reported/Multiple"))
    return s

@dataclass
class Step2Preprocessor:
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str,List[str]]
    def __post_init__(self):
        self.continuous_imputer=None; self.discrete_imputer=None; self.nominal_imputer=None; self.onehot=None; self.output_feature_names_=[]
    def _parts(self,X: pd.DataFrame):
        cont=pd.DataFrame(index=X.index); disc=pd.DataFrame(index=X.index); nom=pd.DataFrame(index=X.index)
        for c in self.continuous_features: cont[c]=pd.to_numeric(X[c].map(clean_scalar),errors="coerce")
        for c in self.binary_features: disc[c]=X[c].map(lambda z:parse_binary_value(z,c)).astype(float)
        for c in self.ordinal_features: disc[c]=X[c].map(lambda z:parse_ordinal_value(z,c)).astype(float)
        for c in self.nominal_features: nom[c]=X[c].map(lambda z:canonical_nominal(c,z)).astype("object")
        return cont,disc,nom
    def fit(self,X: pd.DataFrame):
        cont,disc,nom=self._parts(X)
        self.continuous_imputer=SimpleImputer(strategy="median"); self.discrete_imputer=SimpleImputer(strategy="most_frequent"); self.nominal_imputer=SimpleImputer(strategy="constant",fill_value="Missing")
        self.continuous_imputer.fit(cont); self.discrete_imputer.fit(disc)
        nomi=pd.DataFrame(self.nominal_imputer.fit_transform(nom),columns=self.nominal_features)
        cats=[]
        for c in self.nominal_features:
            pref=list(self.preferred_nominal_levels.get(c,[])); obs=nomi[c].astype(str).unique().tolist(); cats.append(pref+sorted([x for x in obs if x not in pref]))
        try: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse_output=False)
        except TypeError: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse=False)
        self.onehot.fit(nomi.astype(str)); self.output_feature_names_=self.continuous_features+self.binary_features+self.ordinal_features+self.onehot.get_feature_names_out(self.nominal_features).tolist(); return self
    def transform(self,X: pd.DataFrame):
        cont,disc,nom=self._parts(X); a=self.continuous_imputer.transform(cont); b=self.discrete_imputer.transform(disc); nomi=pd.DataFrame(self.nominal_imputer.transform(nom),columns=self.nominal_features); c=self.onehot.transform(nomi.astype(str)); return np.concatenate([a,b,c],axis=1).astype(float)


def add_dynamic_odi_features(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Derive Step 2 dynamic ODI variables exactly as in the primary Step 2 analysis.

    ODI_change is formula-derived as postoperative ODI minus preoperative ODI;
    any pre-existing ODI_change column in the input CSV is ignored for modeling.
    """
    out = df.copy()
    required = ["preop_ODI", "postop_ODI", DAYS_BETWEEN_PROM_COL, "postop_ODI_day"]
    missing = [c for c in required if c not in out.columns]
    if missing:
        raise ValueError(f"Missing required Step 2 dynamic ODI columns: {missing}")

    preop = pd.to_numeric(out["preop_ODI"].map(clean_scalar), errors="coerce")
    postop = pd.to_numeric(out["postop_ODI"].map(clean_scalar), errors="coerce")
    days_between = pd.to_numeric(out[DAYS_BETWEEN_PROM_COL].map(clean_scalar), errors="coerce")
    out["postop_ODI_day"] = pd.to_numeric(out["postop_ODI_day"].map(clean_scalar), errors="coerce")

    odi_change = postop - preop
    odi_improvement = preop - postop

    valid_rate = preop.notna() & postop.notna() & days_between.gt(0)
    prom_change_rate = pd.Series(np.nan, index=out.index, dtype="float")
    prom_change_rate.loc[valid_rate] = (odi_change.loc[valid_rate] / days_between.loc[valid_rate]).astype(float)

    valid_relative = preop.notna() & postop.notna() & preop.gt(0)
    zero_baseline_with_postop = preop.eq(0) & postop.notna()
    relative_fraction = odi_improvement / preop.replace(0, np.nan)
    relative_mcid = pd.Series(np.nan, index=out.index, dtype="float")
    relative_mcid.loc[valid_relative] = (
        relative_fraction.loc[valid_relative] >= RELATIVE_ODI_MCID_CUTOFF
    ).astype(float)
    relative_mcid.loc[zero_baseline_with_postop] = 0.0

    out["preop_ODI"] = preop
    out["postop_ODI"] = postop
    out["ODI_change"] = odi_change
    out[PROM_CHANGE_RATE_COL] = prom_change_rate
    out[RELATIVE_ODI_MCID_COL] = relative_mcid
    out["ODI_improvement_points_for_MCID_audit"] = odi_improvement
    out["ODI_relative_improvement_fraction_for_MCID_audit"] = relative_fraction

    audit = pd.DataFrame([
        {"audit_item": "dynamic_feature_derivation", "value": "completed", "note": "ODI_change formula-derived as postop_ODI - preop_ODI; input ODI_change ignored."},
        {"audit_item": "PROM_change_rate_definition", "value": "(postop_ODI - preop_ODI) / days_between_PROMs", "note": "Matched primary Step 2 code."},
        {"audit_item": "relative_ODI_MCID_definition", "value": f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_CUTOFF}", "note": "preop_ODI=0 coded as 0 when postop_ODI is available."},
    ])
    return out, audit

def load_and_prepare_data() -> Tuple[pd.DataFrame,pd.DataFrame]:
    path=find_input_csv(); df=pd.read_csv(path,low_memory=False); df.columns=[str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns: raise ValueError(f"Missing {GROUP_COL} in Step 2 cohort.")
    if TARGET_COL not in df.columns: raise ValueError(f"Missing {TARGET_COL} in Step 2 cohort.")
    df[TARGET_COL]=df[TARGET_COL].map(to_binary_target)
    df, dyn_audit = add_dynamic_odi_features(df)
    required=[GROUP_COL,TARGET_COL]+BASELINE_FEATURES+DYNAMIC_ODI_FEATURES
    missing=[c for c in required if c not in df.columns]
    if missing: raise ValueError(f"Missing required modeling columns: {missing}")
    before=len(df); keep=df[GROUP_COL].notna() & df[TARGET_COL].isin([0.0,1.0])
    for c in DYNAMIC_ODI_FEATURES:
        keep &= df[c].notna()
    df=df.loc[keep].copy().reset_index(drop=True); df[TARGET_COL]=df[TARGET_COL].astype(int)
    audit=pd.concat([pd.DataFrame([{"audit_item":"input_file","value":path,"note":""},{"audit_item":"input_rows","value":before,"note":""},{"audit_item":"eligible_rows_after_target_dynamic_filters","value":len(df),"note":""},{"audit_item":"events_after_filter","value":int(df[TARGET_COL].sum()),"note":""}]),dyn_audit],ignore_index=True)
    return df,audit

def hospital_map_for_rows(df: pd.DataFrame) -> Tuple[pd.DataFrame,pd.DataFrame]:
    hospital_id_col=find_existing_column(df.columns.tolist(),HOSPITAL_ID_CANDIDATES,"hospital ID",True); hospital_label_col=find_existing_column(df.columns.tolist(),HOSPITAL_LABEL_CANDIDATES,"hospital label",False)
    raw=df[[GROUP_COL,hospital_id_col]+([hospital_label_col] if hospital_label_col and hospital_label_col!=hospital_id_col else [])].copy()
    raw[hospital_id_col]=raw[hospital_id_col].map(clean_scalar).astype("object"); raw=raw[raw[hospital_id_col].notna()].copy(); raw["hospital_id"]=raw[hospital_id_col].astype(str); raw["hospital_label_source"]=raw[hospital_label_col].map(clean_scalar).astype("object").astype(str) if hospital_label_col else raw["hospital_id"]
    rows=[]
    for person,g in raw.groupby(GROUP_COL,dropna=False):
        counts=g["hospital_id"].value_counts(dropna=True)
        if counts.empty: continue
        hid=str(counts.index[0]); labels=g.loc[g["hospital_id"].astype(str).eq(hid),"hospital_label_source"].dropna().astype(str); rows.append({GROUP_COL:person,"hospital_id":hid,"hospital_label":labels.iloc[0] if len(labels) else hid})
    hm=pd.DataFrame(rows); audit=pd.DataFrame([{"audit_item":"hospital_id_column","value":hospital_id_col,"note":""},{"audit_item":"hospital_label_column","value":hospital_label_col if hospital_label_col else "none","note":""},{"audit_item":"patients_with_hospital_assignment","value":int(hm[GROUP_COL].nunique()),"note":""}])
    return hm,audit

def choose_hospital_splits(df: pd.DataFrame) -> Tuple[pd.DataFrame,pd.DataFrame]:
    hm,hospital_audit=hospital_map_for_rows(df)
    work=df.merge(hm[[GROUP_COL,"hospital_id","hospital_label"]],on=GROUP_COL,how="left")
    work=work[work["hospital_id"].notna()].copy().reset_index(drop=True)
    hosp=work.groupby("hospital_id",dropna=False).agg(events=(TARGET_COL,"sum"), n=(GROUP_COL,"size"), label=("hospital_label","first")).reset_index()
    ids=hosp["hospital_id"].astype(str).to_numpy(); y=(hosp["events"].to_numpy()>0).astype(int)
    rng=np.random.default_rng(RANDOM_STATE)
    best=None
    for attempt in range(HOSPITAL_HELDOUT_SPLIT_ATTEMPTS):
        order=rng.permutation(len(ids)); n_test=max(1,int(round(len(ids)*HOSPITAL_HELDOUT_TEST_FRACTION))); test_ids=set(ids[order[:n_test]])
        remaining=np.array([i for i in ids if i not in test_ids]); rem_order=rng.permutation(len(remaining)); n_cal=max(1,int(round(len(remaining)*HOSPITAL_HELDOUT_CALIBRATION_FRACTION_OF_REMAINING))); cal_ids=set(remaining[rem_order[:n_cal]]); train_ids=set([i for i in remaining if i not in cal_ids])
        def ev(s): return int(work.loc[work["hospital_id"].astype(str).isin(s),TARGET_COL].sum())
        tr_e, ca_e, te_e = ev(train_ids), ev(cal_ids), ev(test_ids)
        ok = tr_e>=HOSPITAL_HELDOUT_MIN_TRAIN_EVENTS and ca_e>=HOSPITAL_HELDOUT_MIN_CALIBRATION_EVENTS and te_e>=HOSPITAL_HELDOUT_MIN_TEST_EVENTS
        score = min(tr_e/HOSPITAL_HELDOUT_MIN_TRAIN_EVENTS, ca_e/HOSPITAL_HELDOUT_MIN_CALIBRATION_EVENTS, te_e/HOSPITAL_HELDOUT_MIN_TEST_EVENTS)
        candidate=(score,train_ids,cal_ids,test_ids,tr_e,ca_e,te_e)
        if best is None or candidate[0]>best[0]: best=candidate
        if ok: break
    if best is None: raise RuntimeError("Could not create hospital-level split.")
    _,train_ids,cal_ids,test_ids,tr_e,ca_e,te_e=best
    work["split"]=np.where(work["hospital_id"].astype(str).isin(train_ids),"train",np.where(work["hospital_id"].astype(str).isin(cal_ids),"calibration",np.where(work["hospital_id"].astype(str).isin(test_ids),"test",np.nan)))
    split_audit=pd.DataFrame([{"audit_item":"n_hospitals_total","value":int(len(ids)),"note":""},{"audit_item":"n_train_hospitals","value":len(train_ids),"note":""},{"audit_item":"n_calibration_hospitals","value":len(cal_ids),"note":""},{"audit_item":"n_test_hospitals","value":len(test_ids),"note":""},{"audit_item":"train_events","value":tr_e,"note":""},{"audit_item":"calibration_events","value":ca_e,"note":""},{"audit_item":"test_events","value":te_e,"note":""}])
    return work,pd.concat([hospital_audit,split_audit],ignore_index=True)

def feature_types(features: List[str]) -> Tuple[List[str],List[str],List[str],List[str]]:
    cont=[f for f in features if f in CONTINUOUS_FEATURES_ALL]
    binf=[f for f in features if f in BINARY_FEATURES_ALL]
    ordf=[f for f in features if f in ORDINAL_FEATURES_ALL]
    nomf=[f for f in features if f in NOMINAL_FEATURES_ALL]
    return cont,binf,ordf,nomf

def make_model(params, seed, override_n_estimators=None):
    p=sanitize_lgbm_params(params); mp={k:v for k,v in p.items() if k!="positive_weight_multiplier"}
    if override_n_estimators is not None: mp["n_estimators"]=int(max(MIN_FINAL_N_ESTIMATORS,override_n_estimators))
    return LGBMClassifier(objective="binary",boosting_type="gbdt",metric="average_precision",random_state=seed,n_jobs=N_JOBS,verbosity=-1,force_col_wise=True,**mp)

def fit_pipeline(X,y,features,params,seed,eval_set=None,early=False,override_n_estimators=None):
    cont,binf,ordf,nomf=feature_types(features); pre=Step2Preprocessor(cont,binf,ordf,nomf,PREFERRED_NOMINAL_LEVELS); Xt=pre.fit(X[features]).transform(X[features]); w=make_weights(y,params["positive_weight_multiplier"]); model=make_model(params,seed,override_n_estimators); best=None
    if eval_set is not None and early:
        Xv,yv=eval_set; Xpv=pre.transform(Xv[features]); callbacks=[lgb.early_stopping(EARLY_STOPPING_ROUNDS,verbose=False),lgb.log_evaluation(period=0)]
        try:
            model.fit(Xt,y,sample_weight=w,eval_set=[(Xpv,yv)],eval_metric="average_precision",callbacks=callbacks); best=int(model.best_iteration_) if getattr(model,"best_iteration_",None) else None
        except Exception: model.fit(Xt,y,sample_weight=w)
    else: model.fit(Xt,y,sample_weight=w)
    return pre,model,best

def predict(pre,model,X,features): return model.predict_proba(pre.transform(X[features]))[:,1]
def cv_splits(y,groups,seed,n_folds=N_CV_FOLDS):
    ge=pd.DataFrame({"g":groups,"y":y}).groupby("g")["y"].max().reset_index(); nf=min(n_folds,int(np.sum(ge.y==1)),int(np.sum(ge.y==0)))
    if nf<2: raise ValueError("Not enough patient groups for group-aware CV.")
    cv=StratifiedGroupKFold(n_splits=nf,shuffle=True,random_state=seed); return [(tr,va) for tr,va in cv.split(np.zeros(len(y)),y,groups)]
def tune(X_train,y_train,groups_train,features,model_key,seed):
    folds=cv_splits(y_train,groups_train,seed); cand=[]; fold_rows=[]
    for cid,raw in enumerate(PARAMETER_CANDIDATES,1):
        params=sanitize_lgbm_params(raw); aps=[]; aucs=[]; bests=[]; t0=time.time()
        for fid,(tr,va) in enumerate(folds,1):
            Xtr=X_train.iloc[tr].reset_index(drop=True); ytr=y_train[tr]; Xva=X_train.iloc[va].reset_index(drop=True); yva=y_train[va]
            pre,model,best=fit_pipeline(Xtr,ytr,features,params,seed+cid*1000+fid,eval_set=(Xva,yva),early=USE_EARLY_STOPPING_IN_CV); p=predict(pre,model,Xva,features); ap=safe_average_precision(yva,p); auc=safe_roc_auc(yva,p); aps.append(ap); aucs.append(auc)
            if best and best>0: bests.append(best)
            fold_rows.append({"model_key":model_key,"candidate_id":cid,"fold":fid,"fold_AP":ap,"fold_ROC_AUC":auc,"fold_best_iteration":best,"positive_weight_used":actual_positive_weight(ytr,params["positive_weight_multiplier"]),**params})
        locked=int(np.median(bests)) if bests else int(params["n_estimators"])
        cand.append({"model_key":model_key,"candidate_id":cid,"cv_folds":len(folds),"cv_AP_mean":float(np.nanmean(aps)),"cv_AP_SD":float(np.nanstd(aps,ddof=1)),"cv_ROC_AUC_mean":float(np.nanmean(aucs)),"cv_ROC_AUC_SD":float(np.nanstd(aucs,ddof=1)),"locked_n_estimators_from_cv":locked,"elapsed_seconds":float(time.time()-t0),**params})
        print(f"{model_key} | candidate {cid:03d}/{len(PARAMETER_CANDIDATES)} | CV AP={np.nanmean(aps):.5f} | locked_n={locked}")
    return pd.DataFrame(cand).sort_values("cv_AP_mean",ascending=False).reset_index(drop=True), pd.DataFrame(fold_rows)

def fit_final(Xtr,ytr,Xcal,ycal,Xte,features,params,locked,seed):
    pre,model,_=fit_pipeline(Xtr,ytr,features,params,seed,override_n_estimators=locked); pcal_raw=predict(pre,model,Xcal,features); pte_raw=predict(pre,model,Xte,features); ptr_raw=predict(pre,model,Xtr,features); cal=IsotonicRegression(out_of_bounds="clip"); cal.fit(pcal_raw,ycal); return {"pre":pre,"model":model,"calibrator":cal,"p_train":np.clip(cal.predict(ptr_raw),0,1),"p_cal":np.clip(cal.predict(pcal_raw),0,1),"p_test":np.clip(cal.predict(pte_raw),0,1),"p_test_raw":pte_raw}

def eval_model(y,p,prefix=""):
    out={f"{prefix}AP":safe_average_precision(y,p),f"{prefix}ROC_AUC":safe_roc_auc(y,p),f"{prefix}Brier_score":float(brier_score_loss(y,p)),f"{prefix}ECE":ece(y,p),f"{prefix}N":int(len(y)),f"{prefix}Events":int(np.sum(y)),f"{prefix}Prevalence":float(np.mean(y))}
    out.update({f"{prefix}{k}":v for k,v in top5_metrics(y,p).items()}); return out

def paired_delta_boot(y,p0,p1,groups,metric,seed):
    obs=(safe_average_precision(y,p1)-safe_average_precision(y,p0)) if metric=="AP" else (safe_roc_auc(y,p1)-safe_roc_auc(y,p0)); rng=np.random.default_rng(seed); d=pd.DataFrame({"id":np.arange(len(y)),"g":groups,"y":np.asarray(y).astype(int)}); gy=d.groupby("g").y.max(); pos=gy[gy==1].index.to_numpy(); neg=gy[gy==0].index.to_numpy()
    if len(pos)==0 or len(neg)==0: return obs,np.nan,np.nan,np.nan
    by={g:d.loc[d.g.eq(g),"id"].to_numpy() for g in gy.index}; vals=[]
    for _ in range(N_BOOTSTRAPS):
        sg=np.concatenate([rng.choice(pos,len(pos),True),rng.choice(neg,len(neg),True)]); idx=np.concatenate([by[g] for g in sg]); yy=y[idx]
        if len(np.unique(yy))<2: continue
        vals.append((average_precision_score(yy,p1[idx])-average_precision_score(yy,p0[idx])) if metric=="AP" else (roc_auc_score(yy,p1[idx])-roc_auc_score(yy,p0[idx])))
    if not vals: return obs,np.nan,np.nan,np.nan
    vals=np.asarray(vals); lo,hi=np.percentile(vals,[2.5,97.5]); p=2*min(np.mean(vals<=0),np.mean(vals>=0)); return float(obs),float(lo),float(hi),float(min(max(p,0),1))

def run_one_model(work,features,model_key,model_label,seed):
    tr=work[work.split.eq("train")].copy(); ca=work[work.split.eq("calibration")].copy(); te=work[work.split.eq("test")].copy();
    Xtr=tr.reset_index(drop=True); Xca=ca.reset_index(drop=True); Xte=te.reset_index(drop=True); ytr=Xtr[TARGET_COL].astype(int).to_numpy(); yca=Xca[TARGET_COL].astype(int).to_numpy(); yte=Xte[TARGET_COL].astype(int).to_numpy(); groups_tr=Xtr[GROUP_COL].to_numpy(); groups_te=Xte[GROUP_COL].to_numpy()
    candidates,folds=tune(Xtr,ytr,groups_tr,features,model_key,seed); sel=candidates.iloc[0].copy(); params=sanitize_lgbm_params(sel.to_dict()); params={k:params[k] for k in LGBM_SEARCH_SPACE.keys()}; locked=int(sel["locked_n_estimators_from_cv"]); final=fit_final(Xtr,ytr,Xca,yca,Xte,features,params,locked,seed); pte=final["p_test"]
    summary={"model_key":model_key,"model_label":model_label,"candidate_id":int(sel["candidate_id"]),"feature_count":len(features),"locked_n_estimators_from_cv":locked,**eval_model(yte,pte,"Test_"),"CV_AP_mean":float(sel["cv_AP_mean"]),"CV_ROC_AUC_mean":float(sel["cv_ROC_AUC_mean"]),"Best_params_JSON":json.dumps(json_native(params),sort_keys=True)}
    pred=pd.DataFrame({GROUP_COL:Xte[GROUP_COL].values,"hospital_id":Xte["hospital_id"].values,"hospital_label":Xte["hospital_label"].values,"y_true":yte,"p_test_calibrated":pte,"p_test_raw":final["p_test_raw"]})
    import joblib; joblib.dump({"model_key":model_key,"model_label":model_label,"preprocessor":final["pre"],"model":final["model"],"calibrator":final["calibrator"],"features":features,"params":params,"locked_n_estimators_from_cv":locked}, os.path.join(MODEL_DIR,f"{model_key}_hospital_heldout_artifact.joblib"))
    return {"summary":summary,"predictions":pred,"candidates":candidates,"folds":folds,"p_test":pte,"y_test":yte,"groups_test":groups_te}

def save_pr_roc(y,p,prefix):
    try:
        from sklearn.metrics import precision_recall_curve, roc_curve
        pr,rc,_=precision_recall_curve(y,p); plt.figure(figsize=(6.2,5.2)); plt.plot(rc,pr,lw=2); plt.xlabel("Recall"); plt.ylabel("Precision"); plt.title(prefix+f" PR curve: AP={average_precision_score(y,p):.3f}"); plt.tight_layout(); plt.savefig(os.path.join(PLOT_DIR,safe_name(prefix)+"_PR.png"),dpi=300,bbox_inches="tight"); plt.close()
        fpr,tpr,_=roc_curve(y,p); plt.figure(figsize=(6.2,5.2)); plt.plot(fpr,tpr,lw=2); plt.plot([0,1],[0,1],'--',lw=1); plt.xlabel("False-positive rate"); plt.ylabel("True-positive rate"); plt.title(prefix+f" ROC: AUC={roc_auc_score(y,p):.3f}"); plt.tight_layout(); plt.savefig(os.path.join(PLOT_DIR,safe_name(prefix)+"_ROC.png"),dpi=300,bbox_inches="tight"); plt.close()
    except Exception as e: print("Plot skipped:",repr(e))
def safe_name(x): return re.sub(r"[^A-Za-z0-9_.-]+","_",str(x)).strip("_")

def main():
    print("="*100); print("Step 2 dynamic ODI hospital-held-out internal validation"); print("="*100)
    manifest=ensure_primary_step2_outputs_present()
    df,input_audit=load_and_prepare_data(); work,split_audit=choose_hospital_splits(df)
    baseline_features=BASELINE_FEATURES; dynamic_features=BASELINE_FEATURES+DYNAMIC_ODI_FEATURES
    base=run_one_model(work,baseline_features,"baseline_only","Baseline-only model",RANDOM_STATE+10)
    dyn=run_one_model(work,dynamic_features,"dynamic_PROM","Dynamic ODI-expanded model",RANDOM_STATE+20)
    y=dyn["y_test"]; groups=dyn["groups_test"]; p0=base["p_test"]; p1=dyn["p_test"]
    d_ap,lo_ap,hi_ap,p_ap=paired_delta_boot(y,p0,p1,groups,"AP",RANDOM_STATE+101); d_auc,lo_auc,hi_auc,p_auc=paired_delta_boot(y,p0,p1,groups,"ROC_AUC",RANDOM_STATE+103)
    paired=pd.DataFrame([{"comparison":"dynamic_PROM vs baseline_only","delta_AP":d_ap,"delta_AP_95CI_low":lo_ap,"delta_AP_95CI_high":hi_ap,"delta_AP_p":p_ap,"delta_ROC_AUC":d_auc,"delta_ROC_AUC_95CI_low":lo_auc,"delta_ROC_AUC_95CI_high":hi_auc,"delta_ROC_AUC_p":p_auc,"test_n":int(len(y)),"test_events":int(np.sum(y))}])
    summary=pd.DataFrame([base["summary"],dyn["summary"]])
    base["predictions"].to_csv(os.path.join(TABLE_DIR,"test_predictions_hospital_heldout_baseline_only.csv"),index=False); dyn["predictions"].to_csv(os.path.join(TABLE_DIR,"test_predictions_hospital_heldout_dynamic_PROM.csv"),index=False)
    summary.to_csv(os.path.join(TABLE_DIR,"model_performance_hospital_heldout.csv"),index=False); paired.to_csv(os.path.join(TABLE_DIR,"paired_delta_hospital_heldout.csv"),index=False)
    cand=pd.concat([base["candidates"],dyn["candidates"]],ignore_index=True); folds=pd.concat([base["folds"],dyn["folds"]],ignore_index=True); cand.to_csv(os.path.join(TABLE_DIR,"cv_candidates_hospital_heldout.csv"),index=False); folds.to_csv(os.path.join(TABLE_DIR,"cv_fold_metrics_hospital_heldout.csv"),index=False)
    audit=pd.concat([input_audit,split_audit,pd.DataFrame([{"audit_item":"primary_step2_manifest_read","value":manifest.get("source_manifest_path",""),"note":"Primary Step 2 outputs were read for audit/sequence validation; models are refit for true hospital-held-out validation."}])],ignore_index=True); audit.to_csv(os.path.join(AUDIT_DIR,"hospital_heldout_audit.csv"),index=False)
    save_pr_roc(y,p0,"Step2_hospital_heldout_baseline_only"); save_pr_roc(y,p1,"Step2_hospital_heldout_dynamic_PROM")
    xlsx=os.path.join(OUTPUT_DIR,"Step2_HospitalHeldOut_InternalValidation_summary.xlsx")
    with pd.ExcelWriter(xlsx,engine="openpyxl") as writer:
        summary.to_excel(writer,sheet_name="model_performance",index=False); paired.to_excel(writer,sheet_name="paired_delta",index=False); cand.to_excel(writer,sheet_name="cv_candidates",index=False); folds.to_excel(writer,sheet_name="cv_folds",index=False); audit.to_excel(writer,sheet_name="audit",index=False)
    with open(os.path.join(AUDIT_DIR,"run_manifest.json"),"w") as f: json.dump(json_native({"analysis":"Step2 hospital-held-out internal validation","input_file":find_input_csv(),"primary_step2_manifest":manifest.get("source_manifest_path",""),"baseline_features":BASELINE_FEATURES,"dynamic_features":DYNAMIC_ODI_FEATURES,"paired_delta":paired.iloc[0].to_dict()}),f,indent=2,sort_keys=True)
    zip_path=os.path.join(BASE_DIR,"Step2_HospitalHeldOut_InternalValidation.zip"); tmp=zip_path+".tmp"
    for p in [zip_path,tmp]:
        if os.path.exists(p): os.remove(p)
    with zipfile.ZipFile(tmp,"w",compression=zipfile.ZIP_DEFLATED,compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root,_,files in os.walk(OUTPUT_DIR):
            for fn in files:
                full=os.path.join(root,fn); rel=os.path.relpath(full,os.path.dirname(OUTPUT_DIR)); zf.write(full,rel)
    os.replace(tmp,zip_path)
    print("DONE"); print("Output folder:",OUTPUT_DIR); print("Summary Excel:",xlsx); print("ZIP:",zip_path)
    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(f'<p><b>Step 2 hospital-held-out outputs are ready.</b></p><p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p><p>Path: <code>{zip_path}</code></p>'))
        except Exception as e: print("Download link display skipped:",repr(e))
    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files; files.download(zip_path)
        except Exception as e: print("Automatic download skipped:",repr(e))

if __name__=="__main__": main()

# %% [markdown] Cell 13
# #**Step 2 ODI-threshold survival analysis**

# %% Cell 14
# -*- coding: utf-8 -*-
"""
Step 2 dynamic ODI survival analysis: corrected parsimonious KM/Cox validation
==============================================================================

Purpose
-------
This script performs supportive time-to-event analyses for delayed lumbar
reoperation after the postoperative day-90 landmark.

1. ODI-derived groups use locked SHAP-informed thresholds that were derived
   before survival analysis.
2. No Kaplan-Meier, log-rank, or Cox result is used to select, optimize, or
   modify any threshold.
3. Step 2 Cox models do NOT use the full 35 baseline covariates. They use a
   parsimonious clinically prespecified adjustment set representing
   demographics, baseline health status, diagnosis, and surgical complexity.
4. Dynamic ODI variables are reconstructed exactly from the Step 2 definitions

Expected input
--------------
/content/Step 2_ODI_Cohort.csv

"""

# ============================================================
# Imports
# ============================================================

import os
import re
import sys
import json
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    from lifelines import CoxPHFitter, KaplanMeierFitter
    from lifelines.statistics import logrank_test, proportional_hazard_test
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lifelines"])
    from lifelines import CoxPHFitter, KaplanMeierFitter
    from lifelines.statistics import logrank_test, proportional_hazard_test

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

import matplotlib.pyplot as plt

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"

INPUT_CANDIDATES = [
    "Step 2_ODI_Cohort.csv",
    "Step2_ODI_Cohort.csv",
    "Step 2_ODI.csv",
    "Step2_ODI.csv",
    "Step_2_ODI_Cohort.csv",
    "Step_2_ODI.csv",
]

OUTPUT_DIR = os.path.join(BASE_DIR, "Step2_ODI_Survival_KM_Cox_Parsimonious_CORRECTED")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, PLOT_DIR, TABLE_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

LANDMARK_DAY = 90.0
MAX_FOLLOWUP_DAY = 365.0
MAX_TIME_AFTER_LANDMARK = MAX_FOLLOWUP_DAY - LANDMARK_DAY

TARGET_COL = "final_reop_step2"
GROUP_COL = "PersonKey"

COX_PENALIZER = 0.0
COX_L1_RATIO = 0.0
MIN_EVENTS_PER_PARAMETER_WARNING = 10.0

# ============================================================
# Locked Step 2 SHAP-informed thresholds
# ============================================================

# These thresholds are applied as locked thresholds. They are not derived,
# reselected, optimized, or modified in this survival script.
THRESHOLDS = {
    "preop_ODI": 40.60,
    "postop_ODI": 30.11,
    "ODI_change": -9.47,
    "ODI_change_rate": -0.27,
    "postop_ODI_day": 76.99,
}
RELATIVE_ODI_MCID_THRESHOLD_FRACTION = 0.30

FOCAL_FEATURES = [
    "preop_ODI",
    "postop_ODI",
    "ODI_change",
    "ODI_change_rate",
    "relative_ODI_MCID",
    "postop_ODI_day",
]

FOCAL_LABELS = {
    "preop_ODI": "Preoperative ODI",
    "postop_ODI": "Postoperative ODI",
    "ODI_change": "ODI improvement < 9.47 points",
    "ODI_change_rate": "ODI improvement rate < 0.27 points/day",
    "relative_ODI_MCID": "Relative ODI MCID not achieved",
    "postop_ODI_day": "Postoperative ODI assessment before day 77",
}

# Directional binary variables coded as 1 = manuscript risk stratum.
FOCAL_DUMMY_MAP = {
    "preop_ODI": "preop_ODI_ge_40_60",
    "postop_ODI": "postop_ODI_ge_30_11",
    "ODI_change": "ODI_change_ge_minus_9_47",
    "ODI_change_rate": "ODI_change_rate_ge_minus_0_27",
    "relative_ODI_MCID": "relative_ODI_MCID_not_achieved",
    "postop_ODI_day": "postop_ODI_day_lt_76_99",
}

FOCAL_DUMMY_LABELS = {
    "preop_ODI_ge_40_60": "Preoperative ODI ≥ 40.60 vs < 40.60",
    "postop_ODI_ge_30_11": "Postoperative ODI ≥ 30.11 vs < 30.11",
    "ODI_change_ge_minus_9_47": "ODI improvement < 9.47 points vs ≥ 9.47 points",
    "ODI_change_rate_ge_minus_0_27": "ODI improvement rate < 0.27 points/day vs ≥ 0.27 points/day",
    "relative_ODI_MCID_not_achieved": "Relative ODI MCID not achieved vs achieved",
    "postop_ODI_day_lt_76_99": "Postoperative ODI assessment before day 77 vs day 77 or later",
}

FOCAL_STRATUM_LABELS = {
    "preop_ODI_ge_40_60": {
        0: "Preoperative ODI < 40.60",
        1: "Preoperative ODI ≥ 40.60",
    },
    "postop_ODI_ge_30_11": {
        0: "Postoperative ODI < 30.11",
        1: "Postoperative ODI ≥ 30.11",
    },
    "ODI_change_ge_minus_9_47": {
        0: "ODI improvement ≥ 9.47 points",
        1: "ODI improvement < 9.47 points",
    },
    "ODI_change_rate_ge_minus_0_27": {
        0: "ODI improvement rate ≥ 0.27 points/day",
        1: "ODI improvement rate < 0.27 points/day",
    },
    "relative_ODI_MCID_not_achieved": {
        0: "Relative ODI MCID achieved",
        1: "Relative ODI MCID not achieved",
    },
    "postop_ODI_day_lt_76_99": {
        0: "Postoperative ODI assessment ≥ day 77",
        1: "Postoperative ODI assessment < day 77",
    },
}

# ============================================================
# Parsimonious Step 2 Cox adjustment set
# ============================================================

# These are engineered adjustment terms, not the full 35 baseline variables.
# They represent demographics, baseline health status, diagnosis, and surgical complexity.
BASELINE_COVARIATES = [
    "age_per_10y",
    "male_sex",
    "bmi_per_5",
    "asa_ordinal",
    "diabetes_ordinal",
    "current_smoking",
    "deformity_instability_dx",
    "stenosis_or_radicular_dx",
    "number_levels_ordinal",
    "fusion_or_instrumentation_complexity",
]

BASELINE_COVARIATE_LABELS = {
    "age_per_10y": "Age, per 10 years",
    "male_sex": "Male sex",
    "bmi_per_5": "BMI, per 5 kg/m²",
    "asa_ordinal": "ASA physical status, ordinal",
    "diabetes_ordinal": "Diabetes status, ordinal",
    "current_smoking": "Current smoking",
    "deformity_instability_dx": "Deformity or instability diagnosis",
    "stenosis_or_radicular_dx": "Stenosis or radiculopathy diagnosis",
    "number_levels_ordinal": "Number of operated levels, ordinal",
    "fusion_or_instrumentation_complexity": "Fusion/instrumentation complexity",
}

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

BINARY_MAPS = {
    "sex": {"female": 0, "f": 0, "male": 1, "m": 1},
    "finaldx_deformity_instability": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "finaldx_stenosis": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "finaldx_radicular": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "instrumentation": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "alif_llif": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "plf": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "tlif_plif": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "pelvic_fixation": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
    "corpectomy": {"no": 0, "0": 0, "0.0": 0, "yes": 1, "1": 1, "1.0": 1},
}

ORDINAL_MAPS = {
    "diabetes_status": {
        "no": 0, "none": 0, "0": 0, "0.0": 0,
        "without comp": 1, "without complication": 1, "without complications": 1, "1": 1, "1.0": 1,
        "with comp": 2, "with complication": 2, "with complications": 2, "2": 2, "2.0": 2,
    },
    "asa": {
        "1": 1, "1.0": 1, "i": 1,
        "2": 2, "2.0": 2, "ii": 2,
        "3": 3, "3.0": 3, "iii": 3,
        "4": 4, "4.0": 4, "iv": 4,
        ">=4": 4, ">= 4": 4, "5": 4, "5.0": 4, "v": 4,
    },
    "number_operated_levels": {
        "0": 0, "0.0": 0, "1": 1, "1.0": 1,
        "2": 2, "2.0": 2, "3": 3, "3.0": 3,
        "4": 4, "4.0": 4, ">=4": 4, ">= 4": 4,
        "5": 4, "5.0": 4, "6": 4, "6.0": 4,
        "7": 4, "8": 4, "9": 4, "10": 4,
    },
}

REOP_TIME_CANDIDATES = [
    "reoptime", "reop_time", "reoperation_time", "time_to_reoperation",
    "time_to_reoperation_days", "days_to_reoperation", "reoperation_days",
    "days_from_index_to_reoperation", "reop_days", "reoperation_time_days",
]
FOLLOWUP_TIME_CANDIDATES = [
    "followup_days", "follow_up_days", "last_followup_days", "last_follow_up_days",
    "time_to_last_followup_days", "days_to_last_followup", "days_to_last_contact",
    "days_from_index_to_last_followup", "followup_time_days", "censoring_time_days",
]
DAYS_BETWEEN_PROM_CANDIDATES = [
    "days_between_PROMs", "days_between_proms", "days_between_odi_proms",
    "days_between_preop_postop_PROMs", "days_between_preop_postop_ODI",
]
POSTOP_ODI_DAY_CANDIDATES = [
    "postop_ODI_day", "postoperative_ODI_day", "postop_odi_day",
    "postop_PROM_day", "postop_prom_day", "postoperative_PROM_day",
]

# ============================================================
# Utility functions
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def to_binary_target(x: Any) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if sx in {"1", "1.0", "yes", "y", "true", "t"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("<", "lt").replace(">", "gt")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def json_native(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {str(k): json_native(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_native(v) for v in obj]
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return float(obj)
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    return obj


def find_input_csv() -> str:
    for root in [BASE_DIR, FALLBACK_DIR]:
        for name in INPUT_CANDIDATES:
            p = os.path.join(root, name)
            if os.path.exists(p):
                if root == FALLBACK_DIR:
                    dst = os.path.join(BASE_DIR, name)
                    if not os.path.exists(dst):
                        shutil.copy2(p, dst)
                    return dst
                return p
    raise FileNotFoundError(
        "Could not find Step 2 ODI cohort CSV. Expected one of: "
        + ", ".join(INPUT_CANDIDATES)
    )


def find_existing_column(columns: List[str], candidates: List[str], label: str = "") -> Optional[str]:
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if c.lower() in lookup:
            return lookup[c.lower()]
    return None


def require_column(df: pd.DataFrame, column: str):
    if column not in df.columns:
        raise ValueError(f"Required column not found: {column}")


def parse_binary_value(x: Any, feature: str = "") -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive", "performed"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative", "not performed"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def parse_ordinal_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]:
        return float(ORDINAL_MAPS[feature][sx])
    try:
        v = float(sx)
        if feature == "asa":
            return float(min(max(int(round(v)), 1), 4))
        if feature == "number_operated_levels":
            return float(min(max(int(round(v)), 0), 4))
        if feature == "diabetes_status":
            return float(min(max(int(round(v)), 0), 2))
        return float(v)
    except Exception:
        return np.nan


def parse_current_smoking(x: Any) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if sx in {"current", "current smoker", "yes", "1", "1.0", "active"}:
        return 1.0
    if sx in {"never", "former", "no", "0", "0.0", "unknown/not reported/multiple", "unknown", "not reported", "multiple"}:
        return 0.0
    try:
        v = float(sx)
        if v in (0.0, 1.0):
            return float(v)
    except Exception:
        pass
    return np.nan


def mode_or_median_or_default(s: pd.Series, default: float = 0.0) -> float:
    s = pd.to_numeric(s, errors="coerce").replace([np.inf, -np.inf], np.nan)
    mode = s.mode(dropna=True)
    if not mode.empty:
        return float(mode.iloc[0])
    if s.notna().any():
        return float(s.median())
    return float(default)


# ============================================================
# Dynamic ODI feature construction and survival data
# ============================================================

def add_dynamic_odi_features(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Rebuild all dynamic ODI features from the primary Step 2 definitions."""
    out = df.copy()
    audit = []

    for c in ["preop_ODI", "postop_ODI"]:
        require_column(out, c)
        out[c] = pd.to_numeric(out[c], errors="coerce")

    days_col = find_existing_column(out.columns.tolist(), DAYS_BETWEEN_PROM_CANDIDATES, "days between PROMs")
    if days_col is None:
        raise ValueError(
            "days_between_PROMs is required to reconstruct ODI_change_rate exactly as in the primary Step 2 model. "
            f"Tried: {DAYS_BETWEEN_PROM_CANDIDATES}"
        )
    out["days_between_PROMs"] = pd.to_numeric(out[days_col], errors="coerce")

    postop_day_col = find_existing_column(out.columns.tolist(), POSTOP_ODI_DAY_CANDIDATES, "postoperative ODI day")
    if postop_day_col is None:
        raise ValueError(
            "postop_ODI_day is required for Step 2 survival analysis. "
            f"Tried: {POSTOP_ODI_DAY_CANDIDATES}"
        )
    out["postop_ODI_day"] = pd.to_numeric(out[postop_day_col], errors="coerce")

    original_change_available = "ODI_change" in out.columns
    original_rate_available = "ODI_change_rate" in out.columns

    # Always recompute, even if the exported dataset already contains columns.
    out["_ODI_change_original"] = pd.to_numeric(out["ODI_change"], errors="coerce") if original_change_available else np.nan
    out["_ODI_change_rate_original"] = pd.to_numeric(out["ODI_change_rate"], errors="coerce") if original_rate_available else np.nan

    out["ODI_change"] = out["postop_ODI"] - out["preop_ODI"]
    valid_days = out["days_between_PROMs"].replace(0, np.nan)
    out["ODI_change_rate"] = out["ODI_change"] / valid_days

    improvement_fraction = (out["preop_ODI"] - out["postop_ODI"]) / out["preop_ODI"].replace(0, np.nan)
    out["relative_ODI_MCID_achieved"] = np.where(
        improvement_fraction >= RELATIVE_ODI_MCID_THRESHOLD_FRACTION,
        1.0,
        0.0,
    )
    out.loc[improvement_fraction.isna(), "relative_ODI_MCID_achieved"] = np.nan
    # Match the primary Step 2 convention: if preop ODI is 0 but postop ODI is available,
    # relative MCID is coded as not achieved because a 30% reduction is not defined.
    zero_preop_with_postop = out["preop_ODI"].eq(0) & out["postop_ODI"].notna()
    out.loc[zero_preop_with_postop, "relative_ODI_MCID_achieved"] = 0.0
    out["relative_ODI_MCID"] = out["relative_ODI_MCID_achieved"]

    audit.append({
        "item": "ODI_change_definition",
        "value": "postop_ODI - preop_ODI",
        "note": "Recomputed in this survival script; existing ODI_change column is not used.",
    })
    audit.append({
        "item": "ODI_change_rate_definition",
        "value": "ODI_change / days_between_PROMs",
        "note": f"days_between_PROMs source column: {days_col}; existing ODI_change_rate column is not used.",
    })
    audit.append({
        "item": "relative_ODI_MCID_definition",
        "value": f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_THRESHOLD_FRACTION}",
        "note": "Recomputed in this survival script; existing MCID columns are not used.",
    })

    if original_change_available:
        diff = (out["_ODI_change_original"] - out["ODI_change"]).abs()
        audit.append({
            "item": "original_ODI_change_max_abs_difference_vs_recomputed",
            "value": float(diff.max(skipna=True)) if diff.notna().any() else np.nan,
            "note": "Audit only.",
        })
    if original_rate_available:
        diff = (out["_ODI_change_rate_original"] - out["ODI_change_rate"]).abs()
        audit.append({
            "item": "original_ODI_change_rate_max_abs_difference_vs_recomputed",
            "value": float(diff.max(skipna=True)) if diff.notna().any() else np.nan,
            "note": "Audit only.",
        })

    return out, pd.DataFrame(audit)


def prepare_survival_data(raw: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    work = raw.copy()
    work.columns = [str(c).strip() for c in work.columns]

    require_column(work, GROUP_COL)
    require_column(work, TARGET_COL)

    work, dynamic_audit = add_dynamic_odi_features(work)
    work[TARGET_COL] = work[TARGET_COL].map(to_binary_target)

    reop_time_col = find_existing_column(work.columns.tolist(), REOP_TIME_CANDIDATES, "reoperation time")
    followup_col = find_existing_column(work.columns.tolist(), FOLLOWUP_TIME_CANDIDATES, "follow-up time")

    if reop_time_col is None and work[TARGET_COL].fillna(0).sum() > 0:
        raise ValueError(
            "A reoperation time column is required for Step 2 survival analysis. "
            f"Tried: {REOP_TIME_CANDIDATES}"
        )

    if reop_time_col is not None:
        work["_reop_time_from_index"] = pd.to_numeric(work[reop_time_col], errors="coerce")
    else:
        work["_reop_time_from_index"] = np.nan

    if followup_col is not None:
        work["_followup_time_from_index"] = pd.to_numeric(work[followup_col], errors="coerce")
    else:
        work["_followup_time_from_index"] = MAX_FOLLOWUP_DAY

    focal_required = ["preop_ODI", "postop_ODI", "ODI_change", "ODI_change_rate", "relative_ODI_MCID", "postop_ODI_day"]
    complete_focal = work[focal_required].notna().all(axis=1)
    target_valid = work[TARGET_COL].isin([0.0, 1.0])
    group_valid = work[GROUP_COL].notna()

    # Exclude any observed reoperation at or before the day-90 landmark,
    # regardless of how final_reop_step2 was coded.
    reop_time_positive = work["_reop_time_from_index"].notna() & work["_reop_time_from_index"].gt(0)
    before_or_at_landmark = reop_time_positive & work["_reop_time_from_index"].le(LANDMARK_DAY)
    missing_time_for_event = work[TARGET_COL].eq(1.0) & work["_reop_time_from_index"].isna()

    eligible = target_valid & group_valid & complete_focal & (~before_or_at_landmark) & (~missing_time_for_event)
    surv = work.loc[eligible].copy()

    event_in_window = (
        surv[TARGET_COL].eq(1.0)
        & surv["_reop_time_from_index"].notna()
        & surv["_reop_time_from_index"].gt(LANDMARK_DAY)
        & surv["_reop_time_from_index"].le(MAX_FOLLOWUP_DAY)
    )
    surv["event"] = event_in_window.astype(int)

    censor_from_index = np.minimum(
        pd.to_numeric(surv["_followup_time_from_index"], errors="coerce").fillna(MAX_FOLLOWUP_DAY),
        MAX_FOLLOWUP_DAY,
    )
    duration = np.where(
        surv["event"].eq(1),
        surv["_reop_time_from_index"] - LANDMARK_DAY,
        censor_from_index - LANDMARK_DAY,
    )
    surv["duration_after_landmark"] = pd.to_numeric(duration, errors="coerce")

    duration_valid = surv["duration_after_landmark"].replace([np.inf, -np.inf], np.nan).notna() & surv["duration_after_landmark"].gt(0)
    excluded_nonpositive_duration = int((~duration_valid).sum())
    surv = surv.loc[duration_valid].copy()
    surv["duration_after_landmark"] = np.clip(surv["duration_after_landmark"], 1e-3, MAX_TIME_AFTER_LANDMARK)

    audit_rows = [
        {"item": "input_rows", "value": int(len(work)), "note": ""},
        {"item": "target_column", "value": TARGET_COL, "note": ""},
        {"item": "group_column", "value": GROUP_COL, "note": ""},
        {"item": "reoperation_time_column", "value": reop_time_col if reop_time_col else "NOT_FOUND", "note": ""},
        {"item": "followup_time_column", "value": followup_col if followup_col else "NOT_FOUND", "note": "If not found, non-events censored at day 365."},
        {"item": "invalid_or_missing_target_rows", "value": int((~target_valid).sum()), "note": ""},
        {"item": "missing_group_rows", "value": int((~group_valid).sum()), "note": ""},
        {"item": "rows_missing_any_required_ODI_feature", "value": int((~complete_focal).sum()), "note": ""},
        {"item": "observed_reoperations_at_or_before_landmark_excluded", "value": int(before_or_at_landmark.sum()), "note": f"landmark day = {LANDMARK_DAY}"},
        {"item": "event_positive_rows_missing_reoperation_time_excluded", "value": int(missing_time_for_event.sum()), "note": ""},
        {"item": "excluded_nonpositive_or_missing_duration_after_landmark", "value": excluded_nonpositive_duration, "note": ""},
        {"item": "analysis_rows_after_landmark_filter", "value": int(len(surv)), "note": ""},
        {"item": "delayed_reoperation_events_day91_to_day365", "value": int(surv["event"].sum()), "note": ""},
        {"item": "event_rate_after_landmark", "value": float(surv["event"].mean()) if len(surv) else np.nan, "note": ""},
        {"item": "median_duration_after_landmark_days", "value": float(surv["duration_after_landmark"].median()) if len(surv) else np.nan, "note": ""},
        {"item": "step2_death_retained_analysis_performed", "value": False, "note": "No DeathRetained analysis is performed in this Step 2 survival script."},
    ]
    audit = pd.concat([dynamic_audit, pd.DataFrame(audit_rows)], ignore_index=True, sort=False)
    return surv.reset_index(drop=True), audit


# ============================================================
# Threshold variables and baseline design
# ============================================================

def add_threshold_variables(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["preop_ODI_ge_40_60"] = (pd.to_numeric(out["preop_ODI"], errors="coerce") >= THRESHOLDS["preop_ODI"]).astype(float)
    out["postop_ODI_ge_30_11"] = (pd.to_numeric(out["postop_ODI"], errors="coerce") >= THRESHOLDS["postop_ODI"]).astype(float)
    out["ODI_change_ge_minus_9_47"] = (pd.to_numeric(out["ODI_change"], errors="coerce") >= THRESHOLDS["ODI_change"]).astype(float)
    out["ODI_change_rate_ge_minus_0_27"] = (pd.to_numeric(out["ODI_change_rate"], errors="coerce") >= THRESHOLDS["ODI_change_rate"]).astype(float)
    out["relative_ODI_MCID_not_achieved"] = 1.0 - pd.to_numeric(out["relative_ODI_MCID_achieved"], errors="coerce")
    out["postop_ODI_day_lt_76_99"] = (pd.to_numeric(out["postop_ODI_day"], errors="coerce") < THRESHOLDS["postop_ODI_day"]).astype(float)
    return out


def build_baseline_design(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str], pd.DataFrame]:
    d = df.copy()
    X = pd.DataFrame(index=d.index)
    audit_rows = []

    def add_col(name: str, values: pd.Series, default: float = 0.0):
        s = pd.to_numeric(values, errors="coerce").replace([np.inf, -np.inf], np.nan)
        fill = mode_or_median_or_default(s, default=default)
        X[name] = s.fillna(fill).astype(float)
        audit_rows.append({
            "covariate": name,
            "missing_before_fill": int(s.isna().sum()),
            "fill_value": fill,
            "label": BASELINE_COVARIATE_LABELS.get(name, name),
        })

    add_col("age_per_10y", pd.to_numeric(d["age"], errors="coerce") / 10.0 if "age" in d.columns else pd.Series(np.nan, index=d.index))
    add_col("male_sex", d["sex"].map(lambda z: parse_binary_value(z, "sex")) if "sex" in d.columns else pd.Series(np.nan, index=d.index))
    add_col("bmi_per_5", pd.to_numeric(d["bmi"], errors="coerce") / 5.0 if "bmi" in d.columns else pd.Series(np.nan, index=d.index))

    add_col("asa_ordinal", d["asa"].map(lambda z: parse_ordinal_value(z, "asa")) if "asa" in d.columns else pd.Series(np.nan, index=d.index), default=2.0)
    add_col("diabetes_ordinal", d["diabetes_status"].map(lambda z: parse_ordinal_value(z, "diabetes_status")) if "diabetes_status" in d.columns else pd.Series(np.nan, index=d.index))
    add_col("current_smoking", d["PatTobaccoUse"].map(parse_current_smoking) if "PatTobaccoUse" in d.columns else pd.Series(np.nan, index=d.index))

    add_col(
        "deformity_instability_dx",
        d["finaldx_deformity_instability"].map(lambda z: parse_binary_value(z, "finaldx_deformity_instability")) if "finaldx_deformity_instability" in d.columns else pd.Series(np.nan, index=d.index),
    )

    dx_terms = []
    for col in ["finaldx_stenosis", "finaldx_radicular"]:
        if col in d.columns:
            dx_terms.append(d[col].map(lambda z, c=col: parse_binary_value(z, c)))
    if dx_terms:
        dx_mat = pd.concat(dx_terms, axis=1)
        collapsed = dx_mat.max(axis=1, skipna=True)
        collapsed[dx_mat.isna().all(axis=1)] = np.nan
        add_col("stenosis_or_radicular_dx", collapsed)
    else:
        add_col("stenosis_or_radicular_dx", pd.Series(np.nan, index=d.index))

    add_col(
        "number_levels_ordinal",
        d["number_operated_levels"].map(lambda z: parse_ordinal_value(z, "number_operated_levels")) if "number_operated_levels" in d.columns else pd.Series(np.nan, index=d.index),
        default=1.0,
    )

    complexity_terms = []
    for col in ["instrumentation", "alif_llif", "plf", "tlif_plif", "pelvic_fixation", "corpectomy"]:
        if col in d.columns:
            complexity_terms.append(d[col].map(lambda z, c=col: parse_binary_value(z, c)))
    if complexity_terms:
        comp_mat = pd.concat(complexity_terms, axis=1)
        collapsed = comp_mat.max(axis=1, skipna=True)
        collapsed[comp_mat.isna().all(axis=1)] = np.nan
        add_col("fusion_or_instrumentation_complexity", collapsed)
    else:
        add_col("fusion_or_instrumentation_complexity", pd.Series(np.nan, index=d.index))

    X = X[BASELINE_COVARIATES].copy()
    label_map = {c: BASELINE_COVARIATE_LABELS.get(c, c) for c in BASELINE_COVARIATES}
    return X, label_map, pd.DataFrame(audit_rows)


def add_focal_design(baseline_design: pd.DataFrame, df: pd.DataFrame, focal_features: List[str]) -> Tuple[pd.DataFrame, Dict[str, str]]:
    d = add_threshold_variables(df)
    X = baseline_design.copy()
    labels = {}
    for f in focal_features:
        col = FOCAL_DUMMY_MAP[f]
        X[col] = pd.to_numeric(d[col], errors="coerce").astype(float)
        labels[col] = FOCAL_DUMMY_LABELS[col]
    return X, labels


def threshold_audit_table(df: pd.DataFrame) -> pd.DataFrame:
    d = add_threshold_variables(df)
    rows = []
    for f in FOCAL_FEATURES:
        dummy = FOCAL_DUMMY_MAP[f]
        for value in [0, 1]:
            m = d[dummy].eq(value)
            rows.append({
                "feature": f,
                "feature_label": FOCAL_LABELS[f],
                "dummy_variable": dummy,
                "stratum_value": int(value),
                "stratum_label": FOCAL_STRATUM_LABELS[dummy][value],
                "n": int(m.sum()),
                "events": int(d.loc[m, "event"].sum()),
                "event_rate": float(d.loc[m, "event"].mean()) if m.sum() else np.nan,
                "median_duration_after_landmark_days": float(d.loc[m, "duration_after_landmark"].median()) if m.sum() else np.nan,
            })
    return pd.DataFrame(rows)


# ============================================================
# Cox models
# ============================================================

def drop_unstable_nonfocal_columns(X: pd.DataFrame, y: pd.Series, focal_cols: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    focal = set(focal_cols)
    keep = []
    rows = []
    for col in X.columns:
        s = pd.to_numeric(X[col], errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)
        nunique = int(s.nunique(dropna=True))
        event_sum = float(s[y.eq(1)].sum(skipna=True))
        nonevent_sum = float(s[y.eq(0)].sum(skipna=True))
        action = "kept"
        reason = ""
        if nunique <= 1 and col not in focal:
            action = "dropped"
            reason = "zero_variance_non_focal"
        elif col not in focal and (event_sum == 0.0 or nonevent_sum == 0.0):
            action = "dropped"
            reason = "complete_or_near_separation_non_focal"

        if action == "kept":
            keep.append(col)

        rows.append({
            "covariate": col,
            "nunique": nunique,
            "event_sum": event_sum,
            "nonevent_sum": nonevent_sum,
            "protected_focal": col in focal,
            "action": action,
            "reason": reason,
        })
    return X[keep].copy(), pd.DataFrame(rows)


def fit_cox_model(df: pd.DataFrame, X: pd.DataFrame, focal_cols: List[str], model_name: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    y = df["event"].astype(int)
    X_work = X.copy().apply(pd.to_numeric, errors="coerce").replace([np.inf, -np.inf], np.nan)
    for col in X_work.columns:
        if X_work[col].isna().any():
            X_work[col] = X_work[col].fillna(mode_or_median_or_default(X_work[col]))

    X_work, sparse_audit = drop_unstable_nonfocal_columns(X_work, y, focal_cols=focal_cols)

    # If a focal term has no variation, Cox cannot estimate it; return a clean failure row.
    zero_focal = [c for c in focal_cols if c in X_work.columns and X_work[c].nunique(dropna=True) <= 1]
    if zero_focal:
        summary = pd.DataFrame([{
            "model_name": model_name,
            "term": c,
            "coef": np.nan,
            "se(coef)": np.nan,
            "z": np.nan,
            "p": np.nan,
            "HR": np.nan,
            "HR_lower_95": np.nan,
            "HR_upper_95": np.nan,
            "n": int(len(df)),
            "events": int(df["event"].sum()),
            "n_parameters": int(len(X_work.columns)),
            "events_per_parameter": float(df["event"].sum() / max(len(X_work.columns), 1)),
            "cox_penalizer": COX_PENALIZER,
            "error": "focal variable has zero variance",
        } for c in zero_focal])
        ph = pd.DataFrame([{"model_name": model_name, "term": "PH_NOT_RUN", "p": np.nan, "error": "focal variable has zero variance"}])
        audit = sparse_audit.assign(model_name=model_name)
        return summary, ph, audit

    cox_df = pd.concat(
        [df[["duration_after_landmark", "event"]].reset_index(drop=True), X_work.reset_index(drop=True)],
        axis=1,
    )

    try:
        cph = CoxPHFitter(penalizer=COX_PENALIZER, l1_ratio=COX_L1_RATIO)
        cph.fit(
            cox_df,
            duration_col="duration_after_landmark",
            event_col="event",
            robust=True,
            show_progress=False,
        )
        summary = cph.summary.reset_index().rename(columns={"covariate": "term", "index": "term"})
        summary["model_name"] = model_name
        summary["HR"] = np.exp(summary["coef"])
        summary["HR_lower_95"] = np.exp(summary["coef lower 95%"])
        summary["HR_upper_95"] = np.exp(summary["coef upper 95%"])
        summary["n"] = int(len(df))
        summary["events"] = int(df["event"].sum())
        summary["n_parameters"] = int(len(X_work.columns))
        summary["events_per_parameter"] = float(df["event"].sum() / max(len(X_work.columns), 1))
        summary["cox_penalizer"] = COX_PENALIZER
        summary["error"] = ""

        try:
            ph_test = proportional_hazard_test(cph, cox_df, time_transform="rank")
            ph = ph_test.summary.reset_index().rename(columns={"index": "term"})
            ph["model_name"] = model_name
        except Exception as e:
            ph = pd.DataFrame([{"model_name": model_name, "term": "PH_TEST_FAILED", "p": np.nan, "error": repr(e)}])

    except Exception as e:
        summary = pd.DataFrame([{
            "model_name": model_name,
            "term": c,
            "coef": np.nan,
            "se(coef)": np.nan,
            "z": np.nan,
            "p": np.nan,
            "HR": np.nan,
            "HR_lower_95": np.nan,
            "HR_upper_95": np.nan,
            "n": int(len(df)),
            "events": int(df["event"].sum()),
            "n_parameters": int(len(X_work.columns)),
            "events_per_parameter": float(df["event"].sum() / max(len(X_work.columns), 1)),
            "cox_penalizer": COX_PENALIZER,
            "error": repr(e),
        } for c in X_work.columns])
        ph = pd.DataFrame([{"model_name": model_name, "term": "PH_NOT_RUN", "p": np.nan, "error": repr(e)}])

    epv = pd.DataFrame([{
        "model_name": model_name,
        "n": int(len(df)),
        "events": int(df["event"].sum()),
        "n_parameters_after_dropping_unstable_columns": int(len(X_work.columns)),
        "events_per_parameter": float(df["event"].sum() / max(len(X_work.columns), 1)),
        "warning_events_per_parameter_lt_10": bool((df["event"].sum() / max(len(X_work.columns), 1)) < MIN_EVENTS_PER_PARAMETER_WARNING),
        "cox_penalizer": COX_PENALIZER,
        "cox_l1_ratio": COX_L1_RATIO,
    }])
    audit = pd.concat([sparse_audit.assign(model_name=model_name), epv], ignore_index=True, sort=False)
    return summary, ph, audit


def run_single_feature_cox(df: pd.DataFrame, baseline_design: pd.DataFrame, baseline_labels: Dict[str, str]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    all_summary = []
    all_ph = []
    all_audit = []
    focal_rows = []

    for f in FOCAL_FEATURES:
        dummy = FOCAL_DUMMY_MAP[f]
        model_name = f"Parsimonious baseline + {FOCAL_LABELS[f]}"
        X, focal_labels = add_focal_design(baseline_design, df, [f])
        summary, ph, audit = fit_cox_model(df, X, focal_cols=[dummy], model_name=model_name)

        display = {}
        display.update(baseline_labels)
        display.update(focal_labels)
        summary["display_label"] = summary["term"].map(lambda z: display.get(z, z))
        summary["focal_feature"] = f
        summary["modeling_form"] = "binary_threshold_or_status"

        all_summary.append(summary)
        all_ph.append(ph)
        all_audit.append(audit)

        focal = summary[summary["term"].eq(dummy)].copy()
        if not focal.empty:
            focal["feature"] = f
            focal["feature_label"] = FOCAL_LABELS[f]
            focal["comparison"] = FOCAL_DUMMY_LABELS[dummy]
            focal_rows.append(focal)

    all_summary_df = pd.concat(all_summary, ignore_index=True) if all_summary else pd.DataFrame()
    all_ph_df = pd.concat(all_ph, ignore_index=True) if all_ph else pd.DataFrame()
    all_audit_df = pd.concat(all_audit, ignore_index=True) if all_audit else pd.DataFrame()
    focal_df = pd.concat(focal_rows, ignore_index=True) if focal_rows else pd.DataFrame()

    if not focal_df.empty:
        focal_df = focal_df[[
            "feature", "feature_label", "comparison", "model_name", "term",
            "HR", "HR_lower_95", "HR_upper_95", "p",
            "n", "events", "n_parameters", "events_per_parameter", "cox_penalizer", "error",
        ]].copy()
    return all_summary_df, all_ph_df, all_audit_df, focal_df


def run_combined_cox(df: pd.DataFrame, baseline_design: pd.DataFrame, baseline_labels: Dict[str, str]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    focal_dummies = [FOCAL_DUMMY_MAP[f] for f in FOCAL_FEATURES]
    X, focal_labels = add_focal_design(baseline_design, df, FOCAL_FEATURES)
    summary, ph, audit = fit_cox_model(
        df,
        X,
        focal_cols=focal_dummies,
        model_name="Parsimonious baseline + all six dynamic ODI binary predictors",
    )

    display = {}
    display.update(baseline_labels)
    display.update(focal_labels)
    summary["display_label"] = summary["term"].map(lambda z: display.get(z, z))
    summary["modeling_form"] = "combined_binary_threshold_or_status"

    focal = summary[summary["term"].isin(focal_dummies)].copy()
    if not focal.empty:
        focal["feature"] = focal["term"].map({v: k for k, v in FOCAL_DUMMY_MAP.items()})
        focal["feature_label"] = focal["feature"].map(FOCAL_LABELS)
        focal["comparison"] = focal["term"].map(FOCAL_DUMMY_LABELS)
        focal = focal[[
            "feature", "feature_label", "comparison", "model_name", "term",
            "HR", "HR_lower_95", "HR_upper_95", "p",
            "n", "events", "n_parameters", "events_per_parameter", "cox_penalizer", "error",
        ]].copy()

    return summary, ph, audit, focal


# ============================================================
# KM/log-rank and plotting
# ============================================================

def format_p(p: float) -> str:
    if not np.isfinite(p):
        return "NA"
    if p < 0.001:
        return f"{p:.2e}"
    return f"{p:.3f}"


def plot_km_and_cumulative(df: pd.DataFrame, dummy: str, label: str, prefix: str) -> Dict[str, Any]:
    d = add_threshold_variables(df)
    labels = FOCAL_STRATUM_LABELS[dummy]
    mask0 = d[dummy].eq(0)
    mask1 = d[dummy].eq(1)

    row = {
        "dummy_variable": dummy,
        "feature_label": label,
        "stratum_0_label": labels[0],
        "stratum_1_label": labels[1],
        "n_stratum_0": int(mask0.sum()),
        "events_stratum_0": int(d.loc[mask0, "event"].sum()),
        "n_stratum_1": int(mask1.sum()),
        "events_stratum_1": int(d.loc[mask1, "event"].sum()),
        "logrank_test_statistic": np.nan,
        "logrank_p": np.nan,
        "KM_plot_path": "",
        "cumulative_plot_path": "",
    }

    if mask0.sum() == 0 or mask1.sum() == 0:
        row["error"] = "one stratum is empty"
        return row

    lr = logrank_test(
        d.loc[mask0, "duration_after_landmark"],
        d.loc[mask1, "duration_after_landmark"],
        event_observed_A=d.loc[mask0, "event"],
        event_observed_B=d.loc[mask1, "event"],
    )
    row["logrank_test_statistic"] = float(lr.test_statistic)
    row["logrank_p"] = float(lr.p_value)

    # Kaplan-Meier survival plot.
    fig, ax = plt.subplots(figsize=(8.8, 6.0))
    for value, mask in [(0, mask0), (1, mask1)]:
        km = KaplanMeierFitter()
        km.fit(d.loc[mask, "duration_after_landmark"], event_observed=d.loc[mask, "event"], label=labels[value])
        km.plot_survival_function(ax=ax, ci_show=True)
    ax.set_title(f"Kaplan-Meier analysis: {label}", fontweight="bold")
    ax.set_xlabel("Days from postoperative day 90 landmark")
    ax.set_ylabel("Probability of no delayed reoperation")
    ax.set_xlim(0, MAX_TIME_AFTER_LANDMARK)
    ax.grid(alpha=0.2)
    ax.text(
        0.02, 0.04,
        f"Log-rank p = {format_p(float(lr.p_value))}",
        transform=ax.transAxes,
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.9),
        fontsize=10,
    )
    km_path = os.path.join(PLOT_DIR, f"KM_{safe_filename(prefix)}.png")
    fig.tight_layout()
    fig.savefig(km_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    row["KM_plot_path"] = km_path

    # Cumulative reoperation risk plot.
    fig, ax = plt.subplots(figsize=(8.8, 6.0))
    for value, mask in [(0, mask0), (1, mask1)]:
        km = KaplanMeierFitter()
        km.fit(d.loc[mask, "duration_after_landmark"], event_observed=d.loc[mask, "event"], label=labels[value])
        surv = km.survival_function_.iloc[:, 0]
        ax.step(surv.index.values, 1.0 - surv.values, where="post", label=labels[value])
    ax.set_title(f"Cumulative delayed reoperation risk: {label}", fontweight="bold")
    ax.set_xlabel("Days from postoperative day 90 landmark")
    ax.set_ylabel("Cumulative probability of delayed reoperation")
    ax.set_xlim(0, MAX_TIME_AFTER_LANDMARK)
    ax.grid(alpha=0.2)
    ax.legend()
    ax.text(
        0.02, 0.92,
        f"Log-rank p = {format_p(float(lr.p_value))}",
        transform=ax.transAxes,
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.9),
        fontsize=10,
    )
    cum_path = os.path.join(PLOT_DIR, f"Cumulative_reoperation_{safe_filename(prefix)}.png")
    fig.tight_layout()
    fig.savefig(cum_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    row["cumulative_plot_path"] = cum_path

    return row


def make_km_logrank_outputs(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for f in FOCAL_FEATURES:
        dummy = FOCAL_DUMMY_MAP[f]
        rows.append(plot_km_and_cumulative(df, dummy, FOCAL_LABELS[f], prefix=f))
    return pd.DataFrame(rows)


def plot_forest(df: pd.DataFrame, label_col: str, path: str, title: str):
    d = df.copy().replace([np.inf, -np.inf], np.nan)
    d = d.dropna(subset=["HR", "HR_lower_95", "HR_upper_95"])
    if d.empty:
        return

    d = d.iloc[::-1].reset_index(drop=True)
    fig_h = max(4.6, 0.62 * len(d) + 1.8)
    fig, ax = plt.subplots(figsize=(13.8, fig_h))

    y = np.arange(len(d))
    hr = d["HR"].astype(float).to_numpy()
    lo = d["HR_lower_95"].astype(float).to_numpy()
    hi = d["HR_upper_95"].astype(float).to_numpy()

    ax.errorbar(
        hr,
        y,
        xerr=np.vstack([hr - lo, hi - hr]),
        fmt="o",
        capsize=4,
        linewidth=1.8,
    )
    ax.axvline(1.0, linestyle="--", linewidth=1.2)
    ax.set_xscale("log")
    ax.set_yticks(y)
    ax.set_yticklabels(d[label_col].tolist(), fontsize=11)
    ax.set_xlabel("Hazard ratio (log scale)")
    ax.set_title(title, fontweight="bold")
    ax.grid(axis="x", alpha=0.25)

    x_max = float(np.nanmax(hi))
    x_text = x_max * 1.12
    x_right = x_max * 3.2
    for i, row in d.iterrows():
        text = f"{row['HR']:.2f} ({row['HR_lower_95']:.2f}–{row['HR_upper_95']:.2f}), p={format_p(float(row['p']))}"
        ax.text(x_text, i, text, va="center", ha="left", fontsize=10, clip_on=False)

    ax.set_xlim(left=max(min(float(np.nanmin(lo)) * 0.8, 0.8), 0.05), right=x_right)
    fig.subplots_adjust(left=0.36, right=0.88, top=0.88, bottom=0.15)
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 100)
    print("Step 2 dynamic ODI survival analysis: corrected parsimonious KM/Cox validation")
    print("=" * 100)

    input_csv = find_input_csv()
    print("Input:", input_csv)
    raw = pd.read_csv(input_csv, low_memory=False)
    raw.columns = [str(c).strip() for c in raw.columns]

    surv, eligibility_audit = prepare_survival_data(raw)
    surv = add_threshold_variables(surv)

    print(
        f"Analysis cohort after day-90 landmark: n={len(surv):,}, "
        f"events={int(surv['event'].sum()):,}, event rate={surv['event'].mean():.3%}"
    )

    baseline_design, baseline_labels, baseline_audit = build_baseline_design(surv)

    threshold_audit = threshold_audit_table(surv)
    km_logrank = make_km_logrank_outputs(surv)

    single_all, single_ph, single_audit, single_focal = run_single_feature_cox(
        surv, baseline_design, baseline_labels
    )
    combined_all, combined_ph, combined_audit, combined_focal = run_combined_cox(
        surv, baseline_design, baseline_labels
    )

    focal_dummies = [FOCAL_DUMMY_MAP[f] for f in FOCAL_FEATURES]
    focal_corr = surv[focal_dummies].corr()

    # Forest plots corresponding to manuscript Fig. 5e and 5f.
    single_forest_path = os.path.join(PLOT_DIR, "Forest_single_feature_binary_ODI_thresholds_adjusted_Cox.png")
    if not single_focal.empty:
        plot_forest(
            single_focal.rename(columns={"comparison": "label"}),
            label_col="label",
            path=single_forest_path,
            title="Adjusted Cox models: each ODI-derived threshold/status separately",
        )
        single_focal["forest_plot_path"] = single_forest_path

    combined_forest_path = os.path.join(PLOT_DIR, "Forest_combined_six_ODI_thresholds_adjusted_Cox.png")
    if not combined_focal.empty:
        plot_forest(
            combined_focal.rename(columns={"comparison": "label"}),
            label_col="label",
            path=combined_forest_path,
            title="Adjusted Cox model: all six ODI-derived threshold/status variables",
        )
        combined_focal["forest_plot_path"] = combined_forest_path

    # Save tables.
    eligibility_audit.to_csv(os.path.join(AUDIT_DIR, "landmark_eligibility_and_dynamic_feature_audit.csv"), index=False)
    baseline_audit.to_csv(os.path.join(AUDIT_DIR, "parsimonious_baseline_covariate_audit.csv"), index=False)
    threshold_audit.to_csv(os.path.join(AUDIT_DIR, "locked_threshold_strata_event_audit.csv"), index=False)
    focal_corr.to_csv(os.path.join(AUDIT_DIR, "focal_threshold_variable_correlation.csv"))

    km_logrank.to_csv(os.path.join(TABLE_DIR, "KM_logrank_summary_all_six_ODI_thresholds.csv"), index=False)
    single_all.to_csv(os.path.join(TABLE_DIR, "Cox_single_feature_all_terms.csv"), index=False)
    single_focal.to_csv(os.path.join(TABLE_DIR, "Cox_single_feature_focal_HR_Fig5e.csv"), index=False)
    single_ph.to_csv(os.path.join(TABLE_DIR, "Schoenfeld_PH_tests_single_feature_models.csv"), index=False)
    single_audit.to_csv(os.path.join(AUDIT_DIR, "Cox_single_feature_sparse_EPV_audit.csv"), index=False)

    combined_all.to_csv(os.path.join(TABLE_DIR, "Cox_combined_six_ODI_thresholds_all_terms.csv"), index=False)
    combined_focal.to_csv(os.path.join(TABLE_DIR, "Cox_combined_six_ODI_thresholds_focal_HR_Fig5f.csv"), index=False)
    combined_ph.to_csv(os.path.join(TABLE_DIR, "Schoenfeld_PH_test_combined_six_ODI_thresholds.csv"), index=False)
    combined_audit.to_csv(os.path.join(AUDIT_DIR, "Cox_combined_sparse_EPV_audit.csv"), index=False)

    settings = {
        "input_csv": input_csv,
        "output_dir": OUTPUT_DIR,
        "landmark_day": LANDMARK_DAY,
        "max_followup_day": MAX_FOLLOWUP_DAY,
        "time_scale": "days after postoperative day 90",
        "target_col": TARGET_COL,
        "group_col": GROUP_COL,
        "thresholds": THRESHOLDS,
        "relative_ODI_MCID_definition": f"(preop_ODI - postop_ODI) / preop_ODI >= {RELATIVE_ODI_MCID_THRESHOLD_FRACTION}",
        "ODI_change_definition": "postop_ODI - preop_ODI",
        "ODI_change_rate_definition": "ODI_change / days_between_PROMs",
        "focal_features": FOCAL_FEATURES,
        "focal_dummy_variables": FOCAL_DUMMY_MAP,
        "cox_baseline_adjustment": BASELINE_COVARIATES,
        "n_parsimonious_covariates": len(BASELINE_COVARIATES),
        "full_35_baseline_covariates_used_in_step2_cox": False,
        "continuous_ODI_predictors_used_in_cox": False,
        "death_retained_analysis_performed": False,
        "cox_penalizer": COX_PENALIZER,
        "cox_l1_ratio": COX_L1_RATIO,
        "note": (
            "This script performs KM/log-rank analysis for the six locked Step 2 ODI-derived "
            "threshold/status variables, six separate parsimoniously adjusted Cox models, and one "
            "combined parsimoniously adjusted Cox model containing all six ODI-derived binary variables. "
            "No threshold is selected or optimized in this script."
        ),
    }
    with open(os.path.join(AUDIT_DIR, "analysis_settings.json"), "w") as f:
        json.dump(json_native(settings), f, indent=2, sort_keys=True)

    summary_xlsx = os.path.join(OUTPUT_DIR, "Step2_ODI_Survival_KM_Cox_Parsimonious_CORRECTED_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        pd.DataFrame([settings]).to_excel(writer, sheet_name="analysis_settings", index=False)
        eligibility_audit.to_excel(writer, sheet_name="eligibility_audit", index=False)
        baseline_audit.to_excel(writer, sheet_name="baseline_covariate_audit", index=False)
        threshold_audit.to_excel(writer, sheet_name="threshold_strata_audit", index=False)
        km_logrank.to_excel(writer, sheet_name="KM_logrank", index=False)
        single_focal.to_excel(writer, sheet_name="single_focal_HR_Fig5e", index=False)
        combined_focal.to_excel(writer, sheet_name="combined_focal_HR_Fig5f", index=False)
        single_all.to_excel(writer, sheet_name="single_all_terms", index=False)
        combined_all.to_excel(writer, sheet_name="combined_all_terms", index=False)
        single_ph.to_excel(writer, sheet_name="PH_single", index=False)
        combined_ph.to_excel(writer, sheet_name="PH_combined", index=False)
        focal_corr.reset_index().to_excel(writer, sheet_name="focal_correlation", index=False)

    zip_path = os.path.join(BASE_DIR, "Step2_ODI_Survival_KM_Cox_Parsimonious_CORRECTED.zip")
    tmp_zip = zip_path + ".tmp"
    for p in [zip_path, tmp_zip]:
        if os.path.exists(p):
            os.remove(p)

    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as z:
        for root, _, files in os.walk(OUTPUT_DIR):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(OUTPUT_DIR))
                z.write(full, rel)
    os.replace(tmp_zip, zip_path)

    print("\nDONE")
    print("Output folder:", OUTPUT_DIR)
    print("Summary Excel:", summary_xlsx)
    print("ZIP:", zip_path)

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(
                f'<p><b>Step 2 corrected survival outputs are ready.</b></p>'
                f'<p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p>'
                f'<p>Path: <code>{zip_path}</code></p>'
            ))
        except Exception as e:
            print("Download link display skipped:", repr(e))

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))


if __name__ == "__main__":
    main()
