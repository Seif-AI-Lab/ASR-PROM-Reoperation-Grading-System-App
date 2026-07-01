# -*- coding: utf-8 -*-
"""
Step 1 Analysis and Grading System

Python script exported from the corresponding notebook.
The original notebook is the primary version in the notebooks/ folder.
"""

# %% [markdown] Cell 1
# #**Step1_ModelScreening_All8Classifiers**

# %% Cell 2
# -*- coding: utf-8 -*-
"""
Step 1 PROM-specific classifier screening for 1-year lumbar reoperation prediction
=================================================================================

Purpose
-------
This script performs the Step 1 classifier-screening analysis described in the
manuscript. It compares eight prespecified machine-learning classifiers within
three PROM-specific cohorts and within paired baseline-only vs PROM-expanded
feature sets.

This script is intentionally limited to classifier screening and model-architecture
selection. The final locked LightGBM/SHAP/test-set analysis is performed in:
    Step1_02_FinalLightGBM_SHAP_and_DeathRetained.py

Inputs expected in BASE_DIR
---------------------------
PROM_ODI_Cohort.csv
PROM_BackPain_Cohort.csv
PROM_LegPain_Cohort.csv

Primary target
--------------
final_reop_step1 or one of TARGET_COL_CANDIDATES.

"""

import os, re, sys, json, math, time, platform, subprocess, warnings
from dataclasses import dataclass
from typing import Dict, List

import numpy as np
import pandas as pd

try:
    import lightgbm as lgb
    from lightgbm import LGBMClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lightgbm"])
    import lightgbm as lgb
    from lightgbm import LGBMClassifier

try:
    from xgboost import XGBClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "xgboost"])
    from xgboost import XGBClassifier

try:
    from catboost import CatBoostClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "catboost"])
    from catboost import CatBoostClassifier

try:
    import openpyxl  # noqa
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.model_selection import StratifiedShuffleSplit, StratifiedGroupKFold, ParameterSampler
from sklearn.metrics import average_precision_score, roc_auc_score, brier_score_loss
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.svm import LinearSVC
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier, HistGradientBoostingClassifier

warnings.filterwarnings("ignore")

# ============================================================
# 1) Configuration
# ============================================================
BASE_DIR = "/content"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_ClassifierScreening_All8_outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

GROUP_COL = "PersonKey"
RANDOM_STATE = 20260524
TEST_FRACTION = 0.20
CALIBRATION_FRACTION_OF_REMAINING = 0.20
N_CV_FOLDS = 5
N_RANDOM_COMBINATIONS = 300
N_JOBS = -1
CALIBRATION_METHOD = "isotonic"
ECE_N_BINS = 10

COHORT_CONFIGS = [
    dict(
        cohort_name="ODI",
        input_csv=os.path.join(BASE_DIR, "PROM_ODI_Cohort.csv"),
        prom_display_name="Preoperative ODI",
        prom_feature_candidates=["ODI_preop_value","preop_odi_within_90d","odi_preop_within_90d","odi_preop_anytime_value","preop_odi_anytime","preop_ODI","Preop_ODI","preoperative_ODI","Preoperative_ODI","ODI","odi","baseline_ODI"],
    ),
    dict(
        cohort_name="BackPain",
        input_csv=os.path.join(BASE_DIR, "PROM_BackPain_Cohort.csv"),
        prom_display_name="Preoperative back-pain NRS",
        prom_feature_candidates=["back_pain_preop_value","preop_back_pain_nrs_within_90d","back_pain_nrs_preop_within_90d","back_pain_nrs_preop_anytime_value","preop_back_pain_nrs_anytime","preop_BackPain","preop_backpain","preop_back_pain","preop_BackPain_NRS","preoperative_BackPain","preoperative_back_pain","BackPain","backpain","BackPain_NRS","back_pain_NRS","baseline_BackPain"],
    ),
    dict(
        cohort_name="LegPain",
        input_csv=os.path.join(BASE_DIR, "PROM_LegPain_Cohort.csv"),
        prom_display_name="Preoperative leg-pain NRS",
        prom_feature_candidates=["leg_pain_preop_value","preop_leg_pain_nrs_within_90d","leg_pain_nrs_preop_within_90d","leg_pain_nrs_preop_anytime_value","preop_leg_pain_nrs_anytime","preop_LegPain","preop_legpain","preop_leg_pain","preop_LegPain_NRS","preoperative_LegPain","preoperative_leg_pain","LegPain","legpain","LegPain_NRS","leg_pain_NRS","baseline_LegPain"],
    ),
]

TARGET_COL_CANDIDATES = ["final_reop_step1","final_reop_1yr","reop_1yr","reoperation_1yr","one_year_reoperation","reop365","reop_365","final_reop","reoperation","reop"]
BASELINE_FEATURES = [
    "finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis",
    "age","sex","race","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure",
    "connective_tissue_rheumatic_disease","diabetes_status","myocardial_infarction","renal_disease","institution_type",
    "institution_size","institution_region","asa","bmi","payer_status","alif_llif","corpectomy","discectomy",
    "foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif",
    "other_lumbar_procedure","number_operated_levels","operative_region_extent","PatTobaccoUse"]
assert len(BASELINE_FEATURES) == 35

CONTINUOUS_BASELINE_FEATURES = ["age","bmi"]
BINARY_FEATURES = ["finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis","sex","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure","connective_tissue_rheumatic_disease","myocardial_infarction","renal_disease","institution_type","alif_llif","corpectomy","discectomy","foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif","other_lumbar_procedure","operative_region_extent"]
ORDINAL_FEATURES = ["diabetes_status","institution_size","asa","number_operated_levels"]
NOMINAL_FEATURES = ["race","institution_region","payer_status","PatTobaccoUse"]

MISSING_STRINGS = {""," ","na","n/a","nan","none","null",".","missing","<na>"}
BINARY_MAPS = {
    "sex": {"female":0,"f":0,"male":1,"m":1},
    "ethnicity": {"non-hispanic":0,"non hispanic":0,"hispanic":1},
    "cancer_status": {"no cancer":0,"no":0,"none":0,"any cancer":1,"yes":1,"cancer":1},
    "institution_type": {"hospital":0,"non-hospital":1,"non hospital":1,"nonhospital":1},
    "operative_region_extent": {"lumbar only":0,"extended_region_involvement":1,"extended region involvement":1,"extended":1},
}
ORDINAL_MAPS = {
    "diabetes_status": {"no":0,"none":0,"0":0,"without comp":1,"without complication":1,"without complications":1,"1":1,"with comp":2,"with complication":2,"with complications":2,"2":2},
    "institution_size": {"between 1-99 beds":0,"1-99":0,"between 100-399 beds":1,"100-399":1,">= 400 beds":2,">=400 beds":2,">=400":2,">= 400":2},
    "asa": {"1":1,"i":1,"2":2,"ii":2,"3":3,"iii":3,"4":4,"iv":4,">=4":4,">= 4":4,"5":4,"v":4},
    "number_operated_levels": {"0":0,"1":1,"2":2,"3":3,"4":4,">=4":4,">= 4":4,"5":4,"6":4,"7":4,"8":4,"9":4,"10":4},
}
PREFERRED_NOMINAL_LEVELS = {"race":["White","Black","Other"],"institution_region":["South","North East","West","Midwest"],"payer_status":["Medicare","Commercial/Private","Other","Medicaid/Public/Government"],"PatTobaccoUse":["Unknown/Not reported/Multiple","Never","Former","Current"]}

# ============================================================
# 2) Utility functions
# ============================================================
def clean_scalar(x):
    if pd.isna(x): return np.nan
    if isinstance(x,str):
        s=x.strip().replace("≥",">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x

def norm_text(x):
    x=clean_scalar(x)
    if pd.isna(x): return None
    return str(x).strip().replace("≥",">=").lower()

def to_binary_target(x):
    sx=norm_text(x)
    if sx is None: return np.nan
    if sx in {"1","1.0","yes","y","true","t"}: return 1.0
    if sx in {"0","0.0","no","n","false","f"}: return 0.0
    try:
        v=float(sx); return float(v) if v in (0.0,1.0) else np.nan
    except Exception: return np.nan

def find_existing_column(columns, candidates, what):
    lookup={c.lower():c for c in columns}
    for c in candidates:
        if c in columns: return c
        if c.lower() in lookup: return lookup[c.lower()]
    raise ValueError(f"Could not find {what}. Tried: {candidates}")

def resolve_target_column(df): return find_existing_column(df.columns.tolist(), TARGET_COL_CANDIDATES, "Step 1 target column")
def safe_average_precision(y,s): return np.nan if len(np.unique(y))<2 else float(average_precision_score(y,s))
def safe_roc_auc(y,s): return np.nan if len(np.unique(y))<2 else float(roc_auc_score(y,s))

def ece(y,p,n_bins=10):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); bins=np.linspace(0,1,n_bins+1); out=0.0
    if len(y)==0: return np.nan
    for i in range(n_bins):
        m=(p>=bins[i]) & ((p<=bins[i+1]) if i==n_bins-1 else (p<bins[i+1]))
        if np.any(m): out += (np.sum(m)/len(y))*abs(float(np.mean(y[m]))-float(np.mean(p[m])))
    return float(out)

def top5_metrics(y,p):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); n=len(y); k=max(1,int(math.ceil(n*0.05))); idx=np.argsort(-p)[:k]
    prev=float(np.mean(y)); rate=float(np.mean(y[idx]))
    return {"Top_5pct_n":int(k),"Top_5pct_event_rate":rate,"Top_5pct_lift":rate/prev if prev>0 else np.nan,"Top_5pct_captured_events":float(np.sum(y[idx])/np.sum(y)) if np.sum(y)>0 else np.nan}

def eval_preds(y,p,prefix=""):
    out={f"{prefix}AP":safe_average_precision(y,p),f"{prefix}ROC_AUC":safe_roc_auc(y,p),f"{prefix}Brier_score":float(brier_score_loss(y,p)),f"{prefix}ECE":ece(y,p,ECE_N_BINS),f"{prefix}N":int(len(y)),f"{prefix}Events":int(np.sum(y)),f"{prefix}Prevalence":float(np.mean(y))}
    out.update({f"{prefix}{k}":v for k,v in top5_metrics(y,p).items()})
    return out

def make_weights(y,mult):
    y=np.asarray(y).astype(int); npos=int(np.sum(y==1)); nneg=int(np.sum(y==0))
    if npos==0: raise ValueError("No positive events.")
    pw=(nneg/max(npos,1))*float(mult)
    return np.where(y==1,pw,1.0).astype(float)

def json_native(obj):
    if isinstance(obj,dict): return {str(k):json_native(v) for k,v in obj.items()}
    if isinstance(obj,(list,tuple)): return [json_native(v) for v in obj]
    if isinstance(obj,(np.integer,)): return int(obj)
    if isinstance(obj,(np.floating,)): return float(obj)
    if isinstance(obj,np.ndarray): return obj.tolist()
    try:
        if pd.isna(obj): return None
    except Exception: pass
    return obj

# ============================================================
# 3) Preprocessor with optional scaling
# ============================================================
@dataclass
class Step1Preprocessor:
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str,List[str]]
    scale_continuous: bool = False
    def __post_init__(self):
        self.continuous_imputer=None; self.continuous_scaler=None; self.discrete_imputer=None; self.nominal_imputer=None; self.onehot=None; self.output_feature_names_=[]
    def _parse_binary(self,x,f):
        sx=norm_text(x)
        if sx is None: return np.nan
        if f in BINARY_MAPS and sx in BINARY_MAPS[f]: return float(BINARY_MAPS[f][sx])
        if sx in {"1","1.0","yes","y","true","t","present","positive"}: return 1.0
        if sx in {"0","0.0","no","n","false","f","absent","negative"}: return 0.0
        try:
            v=float(sx); return float(v) if v in (0.0,1.0) else np.nan
        except Exception: return np.nan
    def _parse_ordinal(self,x,f):
        sx=norm_text(x)
        if sx is None: return np.nan
        if f in ORDINAL_MAPS and sx in ORDINAL_MAPS[f]: return float(ORDINAL_MAPS[f][sx])
        try:
            v=float(sx)
            if f=="asa": return float(min(max(int(round(v)),1),4))
            if f=="number_operated_levels": return float(min(max(int(round(v)),0),4))
            if f=="diabetes_status": return float(min(max(int(round(v)),0),2))
            if f=="institution_size": return float(min(max(int(round(v)),0),2))
            return float(v)
        except Exception: return np.nan
    def _nominal(self,f,x):
        x=clean_scalar(x)
        if pd.isna(x): return np.nan
        s=str(x).strip(); sl=s.lower()
        if f=="race": return "White" if sl=="white" else ("Black" if sl=="black" else "Other")
        if f=="institution_region": return {"south":"South","north east":"North East","northeast":"North East","north-east":"North East","west":"West","midwest":"Midwest","mid west":"Midwest"}.get(sl,s)
        if f=="payer_status":
            if sl=="medicare": return "Medicare"
            if sl in {"commercial/private","commercial","private","commercial private"}: return "Commercial/Private"
            if sl in {"medicaid/public/government","medicaid","public","government","public/government"}: return "Medicaid/Public/Government"
            return "Other"
        if f=="PatTobaccoUse": return "Never" if sl=="never" else ("Former" if sl=="former" else ("Current" if sl=="current" else "Unknown/Not reported/Multiple"))
        return s
    def _parts(self,X):
        cont=pd.DataFrame(index=X.index); disc=pd.DataFrame(index=X.index); nom=pd.DataFrame(index=X.index)
        for c in self.continuous_features: cont[c]=pd.to_numeric(X[c].map(clean_scalar),errors="coerce")
        for c in self.binary_features: disc[c]=X[c].map(lambda z:self._parse_binary(z,c)).astype(float)
        for c in self.ordinal_features: disc[c]=X[c].map(lambda z:self._parse_ordinal(z,c)).astype(float)
        for c in self.nominal_features: nom[c]=X[c].map(lambda z:self._nominal(c,z)).astype("object")
        return cont,disc,nom
    def fit(self,X):
        cont,disc,nom=self._parts(X)
        self.continuous_imputer=SimpleImputer(strategy="median")
        self.discrete_imputer=SimpleImputer(strategy="most_frequent")
        self.nominal_imputer=SimpleImputer(strategy="constant",fill_value="Missing")
        cont_imp=self.continuous_imputer.fit_transform(cont)
        if self.scale_continuous:
            self.continuous_scaler=StandardScaler(); self.continuous_scaler.fit(cont_imp)
        self.discrete_imputer.fit(disc)
        nomi=pd.DataFrame(self.nominal_imputer.fit_transform(nom),columns=self.nominal_features)
        cats=[]
        for c in self.nominal_features:
            pref=list(self.preferred_nominal_levels.get(c,[])); obs=nomi[c].astype(str).unique().tolist(); cats.append(pref+sorted([x for x in obs if x not in pref]))
        try: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse_output=False)
        except TypeError: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse=False)
        self.onehot.fit(nomi.astype(str)); self.output_feature_names_=self.continuous_features+self.binary_features+self.ordinal_features+self.onehot.get_feature_names_out(self.nominal_features).tolist(); return self
    def transform(self,X):
        cont,disc,nom=self._parts(X)
        a=self.continuous_imputer.transform(cont)
        if self.scale_continuous and self.continuous_scaler is not None: a=self.continuous_scaler.transform(a)
        b=self.discrete_imputer.transform(disc)
        nomi=pd.DataFrame(self.nominal_imputer.transform(nom),columns=self.nominal_features); c=self.onehot.transform(nomi.astype(str))
        return np.concatenate([a,b,c],axis=1).astype(float)
    def fit_transform(self,X): self.fit(X); return self.transform(X)

# ============================================================
# 4) Splits and model registry
# ============================================================
def feature_types(features):
    cont=[f for f in features if f in CONTINUOUS_BASELINE_FEATURES or f not in BASELINE_FEATURES]
    return cont,[f for f in features if f in BINARY_FEATURES],[f for f in features if f in ORDINAL_FEATURES],[f for f in features if f in NOMINAL_FEATURES]

def patient_split(df,target_col,seed):
    gdf=df.groupby(GROUP_COL,dropna=False)[target_col].max().reset_index(); y=gdf[target_col].astype(int).to_numpy(); g=gdf[GROUP_COL].to_numpy()
    s1=StratifiedShuffleSplit(n_splits=1,test_size=TEST_FRACTION,random_state=seed); trcal,te=next(s1.split(g,y))
    gtc=g[trcal]; ytc=y[trcal]; s2=StratifiedShuffleSplit(n_splits=1,test_size=CALIBRATION_FRACTION_OF_REMAINING,random_state=seed+1); tr,ca=next(s2.split(gtc,ytc))
    train=set(gtc[tr]); cal=set(gtc[ca]); test=set(g[te])
    return df[GROUP_COL].isin(train).to_numpy(), df[GROUP_COL].isin(cal).to_numpy(), df[GROUP_COL].isin(test).to_numpy()

def cv_splits(y,groups,seed,n_folds=N_CV_FOLDS):
    ge=pd.DataFrame({"g":groups,"y":y}).groupby("g")["y"].max().reset_index(); nf=min(n_folds,int(np.sum(ge.y==1)),int(np.sum(ge.y==0)))
    if nf<2: raise ValueError("Not enough patient groups for CV.")
    cv=StratifiedGroupKFold(n_splits=nf,shuffle=True,random_state=seed)
    return [(tr,va) for tr,va in cv.split(np.zeros(len(y)),y,groups)]

MODEL_SEARCH_SPACES = {
    "LogisticRegression": {"C":[0.001,0.003,0.01,0.03,0.1,0.3,1,3,10,30], "positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "LinearSVC": {"C":[0.001,0.003,0.01,0.03,0.1,0.3,1,3,10,30], "positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "RandomForest": {"n_estimators":[400,700,1000],"max_depth":[None,3,5,7,10,15],"min_samples_leaf":[1,5,10,25,50],"max_features":["sqrt","log2",0.5,0.8],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "AdaBoost": {"n_estimators":[100,200,400,700,1000],"learning_rate":[0.003,0.005,0.01,0.03,0.05,0.1,0.3,0.5,1.0],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "LightGBM": {"n_estimators":[400,700,1000,1400,1800,2200,2600],"learning_rate":[0.003,0.005,0.008,0.01,0.02,0.03,0.05],"num_leaves":[7,15,31,63,127],"max_depth":[-1,2,3,5,7,9],"min_child_samples":[10,20,50,100,200,400],"subsample":[0.60,0.75,0.90,1.00],"subsample_freq":[0,1,2],"colsample_bytree":[0.60,0.75,0.90,1.00],"reg_alpha":[0,0.001,0.01,0.05,0.1,0.5,1,2],"reg_lambda":[0,0.001,0.01,0.05,0.1,0.5,1,2,5],"min_split_gain":[0,0.001,0.005,0.01,0.05,0.1],"max_bin":[63,127,255],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "XGBoost": {"n_estimators":[400,700,1000,1400],"learning_rate":[0.003,0.005,0.01,0.02,0.03,0.05],"max_depth":[2,3,5,7],"min_child_weight":[1,5,10,25],"subsample":[0.60,0.75,0.90,1.00],"colsample_bytree":[0.60,0.75,0.90,1.00],"reg_alpha":[0,0.001,0.01,0.1,1],"reg_lambda":[0.01,0.1,1,2,5],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "HistGradientBoosting": {"max_iter":[200,400,700,1000],"learning_rate":[0.003,0.005,0.01,0.03,0.05,0.1],"max_leaf_nodes":[7,15,31,63],"max_depth":[None,2,3,5,7],"min_samples_leaf":[10,20,50,100,200],"l2_regularization":[0,0.001,0.01,0.1,1,5],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
    "CatBoost": {"iterations":[400,700,1000,1400,1800],"learning_rate":[0.003,0.005,0.01,0.02,0.03,0.05],"depth":[2,3,5,7,9],"l2_leaf_reg":[1,3,5,10,20],"positive_weight_multiplier":[0.25,0.5,0.75,1,1.5,2,3,4,6,8]},
}
SCALE_SENSITIVE_MODELS = {"LogisticRegression", "LinearSVC"}

def make_model(model_name, params, seed):
    p={k:v for k,v in params.items() if k!="positive_weight_multiplier"}
    if model_name=="LogisticRegression":
        return LogisticRegression(C=float(p["C"]), penalty="l2", solver="lbfgs", max_iter=5000, random_state=seed, n_jobs=N_JOBS)
    if model_name=="LinearSVC":
        return LinearSVC(C=float(p["C"]), max_iter=10000, random_state=seed)
    if model_name=="RandomForest":
        return RandomForestClassifier(n_estimators=int(p["n_estimators"]), max_depth=p["max_depth"], min_samples_leaf=int(p["min_samples_leaf"]), max_features=p["max_features"], random_state=seed, n_jobs=N_JOBS)
    if model_name=="AdaBoost":
        return AdaBoostClassifier(n_estimators=int(p["n_estimators"]), learning_rate=float(p["learning_rate"]), random_state=seed)
    if model_name=="LightGBM":
        return LGBMClassifier(objective="binary", boosting_type="gbdt", metric="average_precision", random_state=seed, n_jobs=N_JOBS, verbosity=-1, force_col_wise=True, **p)
    if model_name=="XGBoost":
        return XGBClassifier(objective="binary:logistic", eval_metric="aucpr", random_state=seed, n_jobs=N_JOBS, tree_method="hist", **p)
    if model_name=="HistGradientBoosting":
        return HistGradientBoostingClassifier(random_state=seed, **p)
    if model_name=="CatBoost":
        return CatBoostClassifier(loss_function="Logloss", eval_metric="PRAUC", random_seed=seed, verbose=False, allow_writing_files=False, thread_count=-1, **p)
    raise ValueError(model_name)

def fit_screening_pipeline(model_name, X, y, features, params, seed):
    cont,binf,ordf,nomf=feature_types(features)
    pre=Step1Preprocessor(cont,binf,ordf,nomf,PREFERRED_NOMINAL_LEVELS,scale_continuous=model_name in SCALE_SENSITIVE_MODELS)
    Xt=pre.fit_transform(X[features])
    model=make_model(model_name, params, seed)
    w=make_weights(y, params.get("positive_weight_multiplier", 1.0))
    try:
        model.fit(Xt, y, sample_weight=w)
    except TypeError:
        model.fit(Xt, y)
    return pre, model

def predict_raw(pre, model, X, features):
    Xt=pre.transform(X[features])
    if hasattr(model, "predict_proba"):
        return model.predict_proba(Xt)[:,1]
    if hasattr(model, "decision_function"):
        return model.decision_function(Xt)
    return model.predict(Xt).astype(float)

def calibrated_predictions(raw_train_or_test, raw_cal, y_cal):
    if CALIBRATION_METHOD == "isotonic":
        cal=IsotonicRegression(out_of_bounds="clip")
        cal.fit(raw_cal, y_cal)
        return cal.predict(raw_train_or_test)
    return raw_train_or_test

def tune_classifier(model_name, X_train, y_train, groups_train, features, folds, model_key, seed):
    space=MODEL_SEARCH_SPACES[model_name]
    n_iter=min(N_RANDOM_COMBINATIONS, math.prod([len(v) for v in space.values()]))
    candidates=list(ParameterSampler(space, n_iter=n_iter, random_state=seed))
    cand_rows=[]; fold_rows=[]
    for cid, params in enumerate(candidates, 1):
        aps=[]; aucs=[]; t0=time.time()
        for fid,(tr,va) in enumerate(folds,1):
            Xtr=X_train.iloc[tr].reset_index(drop=True); ytr=y_train[tr]
            Xva=X_train.iloc[va].reset_index(drop=True); yva=y_train[va]
            pre,model=fit_screening_pipeline(model_name,Xtr,ytr,features,params,seed+cid*1000+fid)
            s=predict_raw(pre,model,Xva,features)
            aps.append(safe_average_precision(yva,s)); aucs.append(safe_roc_auc(yva,s))
            fold_rows.append({"model_key":model_key,"classifier":model_name,"candidate_id":cid,"fold":fid,"fold_AP":aps[-1],"fold_ROC_AUC":aucs[-1],"scale_continuous":model_name in SCALE_SENSITIVE_MODELS,**params})
        row={"model_key":model_key,"classifier":model_name,"candidate_id":cid,"cv_folds":len(folds),"cv_AP_mean":float(np.nanmean(aps)),"cv_AP_SD":float(np.nanstd(aps,ddof=1)),"cv_ROC_AUC_mean":float(np.nanmean(aucs)),"cv_ROC_AUC_SD":float(np.nanstd(aucs,ddof=1)),"scale_continuous":model_name in SCALE_SENSITIVE_MODELS,"elapsed_seconds":float(time.time()-t0),**params}
        cand_rows.append(row)
        print(f"{model_key} | {model_name} | candidate {cid:03d}/{len(candidates)} | CV AP={row['cv_AP_mean']:.5f} | AUC={row['cv_ROC_AUC_mean']:.5f}")
    return pd.DataFrame(cand_rows).sort_values("cv_AP_mean",ascending=False).reset_index(drop=True), pd.DataFrame(fold_rows)

def fit_final_screening_model(model_name, Xtr, ytr, Xcal, ycal, Xte, features, params, seed):
    pre,model=fit_screening_pipeline(model_name,Xtr,ytr,features,params,seed)
    raw_cal=predict_raw(pre,model,Xcal,features)
    raw_test=predict_raw(pre,model,Xte,features)
    p_test=calibrated_predictions(raw_test, raw_cal, ycal)
    return raw_test, p_test

# ============================================================
# 5) Cohort loop
# ============================================================
def split_audit(df,target_col):
    return df.groupby("split").agg(n=(target_col,"size"),events=(target_col,"sum"),patients=(GROUP_COL,"nunique")).reset_index().assign(prevalence=lambda d:d.events/d.n)

def run_cohort(cfg, seed):
    cohort=cfg["cohort_name"]
    df=pd.read_csv(cfg["input_csv"],low_memory=False); df.columns=[str(c).strip() for c in df.columns]
    target0=resolve_target_column(df)
    prom=find_existing_column(df.columns.tolist(),cfg["prom_feature_candidates"],f"{cohort} PROM")
    missing=[c for c in BASELINE_FEATURES+[prom,target0,GROUP_COL] if c not in df.columns]
    if missing: raise ValueError(f"{cohort} missing columns: {missing}")
    target="__target_step1__"; before=len(df); df[target]=df[target0].map(to_binary_target); df=df[df[target].isin([0.0,1.0])].copy(); df[target]=df[target].astype(int)
    if df[GROUP_COL].isna().any(): raise ValueError(f"{cohort}: missing PersonKey")
    tr,ca,te=patient_split(df,target,seed); df["split"]=np.where(tr,"train",np.where(ca,"calibration","test"))
    if df.groupby(GROUP_COL).split.nunique().max()>1: raise RuntimeError("Patient leakage across splits.")
    print("\n"+"#"*100+f"\nCOHORT {cohort}\n"+"#"*100)
    print(f"Input: {cfg['input_csv']} | target: {target0} | PROM: {prom} | rows after target cleaning: {len(df)} dropped: {before-len(df)}")
    print(split_audit(df,target).to_string(index=False))
    train=df[df.split.eq("train")].reset_index(drop=True)
    cal=df[df.split.eq("calibration")].reset_index(drop=True)
    test=df[df.split.eq("test")].reset_index(drop=True)
    folds=cv_splits(train[target].astype(int).to_numpy(), train[GROUP_COL].to_numpy(), seed+10)
    rows=[]; candidates_all=[]; folds_all=[]
    for model_type, features in [("baseline_only", BASELINE_FEATURES), ("expanded_PROM", BASELINE_FEATURES+[prom])]:
        for model_name in MODEL_SEARCH_SPACES:
            model_key=f"{cohort}_{model_type}_{model_name}"
            cand,fold=tune_classifier(model_name, train[features].copy(), train[target].astype(int).to_numpy(), train[GROUP_COL].to_numpy(), features, folds, model_key, seed+1000+len(rows))
            best=cand.iloc[0].to_dict()
            params={k:best[k] for k in MODEL_SEARCH_SPACES[model_name] if k in best}
            raw_test,p_test=fit_final_screening_model(model_name, train[features].copy(), train[target].astype(int).to_numpy(), cal[features].copy(), cal[target].astype(int).to_numpy(), test[features].copy(), features, params, seed+2000+len(rows))
            ev=eval_preds(test[target].astype(int).to_numpy(), p_test, prefix="Test_")
            raw_ev={"RawTest_AP":safe_average_precision(test[target].astype(int).to_numpy(),raw_test),"RawTest_ROC_AUC":safe_roc_auc(test[target].astype(int).to_numpy(),raw_test)}
            rows.append({"cohort":cohort,"model_type":model_type,"classifier":model_name,"model_key":model_key,"n_features":len(features),"prom_feature":prom if model_type=="expanded_PROM" else "not_applicable","best_candidate_id":int(best["candidate_id"]),"cv_AP_mean":float(best["cv_AP_mean"]),"cv_AP_SD":float(best["cv_AP_SD"]),"cv_ROC_AUC_mean":float(best["cv_ROC_AUC_mean"]),"cv_ROC_AUC_SD":float(best["cv_ROC_AUC_SD"]),"scale_continuous":model_name in SCALE_SENSITIVE_MODELS,**ev,**raw_ev})
            candidates_all.append(cand.assign(cohort=cohort, model_type=model_type, prom_feature=prom if model_type=="expanded_PROM" else "not_applicable"))
            folds_all.append(fold.assign(cohort=cohort, model_type=model_type, prom_feature=prom if model_type=="expanded_PROM" else "not_applicable"))
    return {"summary":pd.DataFrame(rows),"candidates":pd.concat(candidates_all,ignore_index=True),"folds":pd.concat(folds_all,ignore_index=True),"split_audit":split_audit(df,target).assign(cohort=cohort),"metadata":pd.DataFrame([{"cohort":cohort,"input_csv":cfg["input_csv"],"target_col_original":target0,"prom_col_used":prom,"n_after_target_cleaning":len(df)}])}

def style_excel(writer):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        for s in writer.sheets:
            ws=writer.sheets[s]; ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for cell in ws[1]: cell.font=Font(bold=True); cell.fill=PatternFill(start_color="D9EAF7",end_color="D9EAF7",fill_type="solid"); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            for col in ws.columns:
                ml=max([len(str(c.value)) for c in col if c.value is not None]+[12]); ws.column_dimensions[col[0].column_letter].width=min(max(ml+2,12),70)
    except Exception:
        pass

def methods_table():
    return pd.DataFrame([
        {"item":"Classifier-screening design","rationale":"Eight prespecified classifiers are compared within each PROM-specific cohort and paired baseline-only/PROM-expanded feature set."},
        {"item":"Paired split/folds","rationale":"Baseline-only and expanded models within a cohort use identical patient-level train/calibration/test splits and identical group-aware CV folds."},
        {"item":"Tuning metric","rationale":"Mean cross-validated average precision is used for classifier/hyperparameter selection because reoperation is a rare-event outcome."},
        {"item":"Standardization","rationale":"Continuous variables are standardized only for scale-sensitive models, with scaler parameters fitted exclusively within the training fold/split."},
        {"item":"Final architecture selection","rationale":"The selected classifier is the one with the highest mean CV AP within the training split; held-out test performance is reported but not used for selection."},
    ])

def main():
    t0=time.time()
    results=[]
    for i,cfg in enumerate(COHORT_CONFIGS):
        results.append(run_cohort(cfg, RANDOM_STATE+i*10000))
    summary=pd.concat([r["summary"] for r in results],ignore_index=True)
    candidates=pd.concat([r["candidates"] for r in results],ignore_index=True)
    folds=pd.concat([r["folds"] for r in results],ignore_index=True)
    splits=pd.concat([r["split_audit"] for r in results],ignore_index=True)
    metadata=pd.concat([r["metadata"] for r in results],ignore_index=True)
    selected=(summary.sort_values(["cohort","model_type","cv_AP_mean"],ascending=[True,True,False])
              .groupby(["cohort","model_type"],as_index=False).head(1).reset_index(drop=True))
    classifier_rankings=(summary.groupby("classifier",as_index=False)
                         .agg(mean_cv_AP=("cv_AP_mean","mean"),median_cv_AP=("cv_AP_mean","median"),mean_test_AP=("Test_AP","mean"),n_models=("model_key","count"))
                         .sort_values("mean_cv_AP",ascending=False))
    xlsx=os.path.join(OUTPUT_DIR,"Step1_ClassifierScreening_All8_summary.xlsx")
    with pd.ExcelWriter(xlsx,engine="openpyxl") as w:
        methods_table().to_excel(w,"methods_rationale",index=False)
        metadata.to_excel(w,"cohort_metadata",index=False)
        summary.to_excel(w,"all8_model_performance",index=False)
        selected.to_excel(w,"selected_by_highest_CV_AP",index=False)
        classifier_rankings.to_excel(w,"classifier_rankings",index=False)
        candidates.to_excel(w,"cv_candidates_all8",index=False)
        folds.to_excel(w,"cv_fold_metrics_all8",index=False)
        splits.to_excel(w,"split_audit",index=False)
        pd.DataFrame([{"Parameter":k,"Value":v} for k,v in {"BASE_DIR":BASE_DIR,"OUTPUT_DIR":OUTPUT_DIR,"RANDOM_STATE":RANDOM_STATE,"TEST_FRACTION":TEST_FRACTION,"CALIBRATION_FRACTION_OF_REMAINING":CALIBRATION_FRACTION_OF_REMAINING,"N_CV_FOLDS":N_CV_FOLDS,"N_RANDOM_COMBINATIONS":N_RANDOM_COMBINATIONS,"CALIBRATION_METHOD":CALIBRATION_METHOD,"python_version":platform.python_version(),"lightgbm_version":lgb.__version__}.items()]).to_excel(w,"run_config",index=False)
        style_excel(w)
    summary.to_csv(os.path.join(OUTPUT_DIR,"all8_model_performance.csv"),index=False)
    selected.to_csv(os.path.join(OUTPUT_DIR,"selected_by_highest_CV_AP.csv"),index=False)
    classifier_rankings.to_csv(os.path.join(OUTPUT_DIR,"classifier_rankings.csv"),index=False)
    candidates.to_csv(os.path.join(OUTPUT_DIR,"cv_candidates_all8.csv"),index=False)
    folds.to_csv(os.path.join(OUTPUT_DIR,"cv_fold_metrics_all8.csv"),index=False)
    json.dump(json_native({"design":"Step 1 classifier screening: 3 PROM cohorts x 2 feature sets x 8 classifiers","selection_rule":"highest mean group-aware CV average precision within training split","scale_sensitive_models":sorted(SCALE_SENSITIVE_MODELS),"runtime_minutes":float((time.time()-t0)/60),"summary_xlsx":xlsx}),open(os.path.join(OUTPUT_DIR,"run_manifest.json"),"w"),indent=2,sort_keys=True)
    print("\n"+"="*100)
    print("STEP 1 classifier screening completed")
    print("Summary Excel:", xlsx)
    print("Selected classifier table:", os.path.join(OUTPUT_DIR,"selected_by_highest_CV_AP.csv"))
    print("="*100)

if __name__ == "__main__":
    main()

# %% [markdown] Cell 3
# #**Step1_FinalLightGBM_SHAP_and_DeathRetained**

# %% Cell 4
# -*- coding: utf-8 -*-
"""
Step 1 PROM-specific 1-year lumbar reoperation prediction with LightGBM
======================================================================

Inputs
------
/content/PROM_ODI_Cohort.csv
/content/PROM_BackPain_Cohort.csv
/content/PROM_LegPain_Cohort.csv

Target
------
final_reop_step1
    1 = reoperation within 1 year after the index surgery
    0 = no reoperation within 1 year after the index surgery

Design
------
This script runs PROM-specific paired baseline-only and PROM-expanded
LightGBM models for Step 1 1-year lumbar reoperation prediction. Within
each PROM-specific cohort, paired models use identical patient-level
train/calibration/test splits and identical group-aware cross-validation
folds. Hyperparameter tuning is performed exclusively within the training
split using cross-validated average precision as the primary selection
metric. Probability calibration and threshold-dependent classification
threshold selection are performed only on the calibration split.

"""

import os, re, sys, json, math, time, zipfile, platform, subprocess, warnings
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import lightgbm as lgb
    from lightgbm import LGBMClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lightgbm"])
    import lightgbm as lgb
    from lightgbm import LGBMClassifier

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import joblib
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "joblib"])
    import joblib

try:
    import openpyxl  # noqa
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.model_selection import StratifiedShuffleSplit, StratifiedGroupKFold, ParameterSampler
from sklearn.metrics import (average_precision_score, roc_auc_score, brier_score_loss,
                             precision_recall_curve, roc_curve, confusion_matrix, f1_score)
from sklearn.isotonic import IsotonicRegression
import matplotlib.pyplot as plt
warnings.filterwarnings("ignore")

# ============================================================
# 1) Configuration
# ============================================================
BASE_DIR = "/content"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
SHAP_DIR = os.path.join(OUTPUT_DIR, "shap")
ARTIFACT_DIR = os.path.join(OUTPUT_DIR, "model_artifacts")
for d in [OUTPUT_DIR, PLOT_DIR, SHAP_DIR, ARTIFACT_DIR]: os.makedirs(d, exist_ok=True)

GROUP_COL = "PersonKey"
RANDOM_STATE = 20260524
TEST_FRACTION = 0.20
CALIBRATION_FRACTION_OF_REMAINING = 0.20
N_CV_FOLDS = 5
N_RANDOM_COMBINATIONS = 300
USE_EARLY_STOPPING_IN_CV = True
EARLY_STOPPING_ROUNDS = 100
MIN_FINAL_N_ESTIMATORS = 50
CALIBRATION_METHOD = "isotonic"
SCALE_CONTINUOUS_FOR_FINAL_LIGHTGBM = False  # LightGBM is not scale-sensitive; option retained for protocol parity.
THRESHOLD_STRATEGY = "max_f1"  # max_f1, youden, fixed, top_percent
FIXED_THRESHOLD = 0.50
THRESHOLD_TOP_PERCENT = 0.05
N_BOOTSTRAPS = 2000
ECE_N_BINS = 10
N_JOBS = -1
SHAP_MAX_EXPLAIN_ROWS = None  # SHAP values are computed on all held-out test rows; plots may downsample for readability only.
SHAP_BEESWARM_MAX_DISPLAY = 35
SHAP_BAR_MAX_DISPLAY = 35
SHAP_HEATMAP_MAX_DISPLAY = 25
SHAP_HEATMAP_MAX_ROWS = 500
SHAP_THRESHOLD_N_BINS = 20
SHAP_THRESHOLD_MIN_BIN_N = 30
HOSPITAL_MIN_TEST_N_FOR_INDIVIDUAL_REPORTING = 100
HOSPITAL_COLUMN_CANDIDATES = [
    "hospital", "Hospital", "hospital_id", "HospitalID", "HospitalId",
    "site", "Site", "site_id", "SiteID", "institution", "Institution",
    "registry_site", "RegistrySite", "facility", "Facility",
]
LEARNING_CURVE_TRAIN_FRACTIONS = (0.20, 0.40, 0.60, 0.80, 1.00)
LEARNING_CURVE_CV_FOLDS = 3
AUTO_DOWNLOAD_ZIP = False
CREATE_COLAB_DOWNLOAD_LINK = True
ZIP_COMPRESSION_LEVEL = 1

COHORT_CONFIGS = [
    dict(
        cohort_name="ODI",
        input_csv=os.path.join(BASE_DIR, "PROM_ODI_Cohort.csv"),
        prom_display_name="Preoperative ODI",
        # Exact Step 1 ETL column in current cohort file is ODI_preop_value.
        # Additional aliases are retained only to make the script robust to
        # future export-name changes.
        prom_feature_candidates=[
            "ODI_preop_value",
            "preop_odi_within_90d",
            "odi_preop_within_90d",
            "odi_preop_anytime_value",
            "preop_odi_anytime",
            "preop_ODI",
            "Preop_ODI",
            "preoperative_ODI",
            "Preoperative_ODI",
            "ODI",
            "odi",
            "baseline_ODI",
        ],
    ),
    dict(
        cohort_name="BackPain",
        input_csv=os.path.join(BASE_DIR, "PROM_BackPain_Cohort.csv"),
        prom_display_name="Preoperative back-pain NRS",
        # Exact Step 1 ETL column is expected to be back_pain_preop_value.
        prom_feature_candidates=[
            "back_pain_preop_value",
            "preop_back_pain_nrs_within_90d",
            "back_pain_nrs_preop_within_90d",
            "back_pain_nrs_preop_anytime_value",
            "preop_back_pain_nrs_anytime",
            "preop_BackPain",
            "preop_backpain",
            "preop_back_pain",
            "preop_BackPain_NRS",
            "preoperative_BackPain",
            "preoperative_back_pain",
            "BackPain",
            "backpain",
            "BackPain_NRS",
            "back_pain_NRS",
            "baseline_BackPain",
        ],
    ),
    dict(
        cohort_name="LegPain",
        input_csv=os.path.join(BASE_DIR, "PROM_LegPain_Cohort.csv"),
        prom_display_name="Preoperative leg-pain NRS",
        # Exact Step 1 ETL column is expected to be leg_pain_preop_value.
        prom_feature_candidates=[
            "leg_pain_preop_value",
            "preop_leg_pain_nrs_within_90d",
            "leg_pain_nrs_preop_within_90d",
            "leg_pain_nrs_preop_anytime_value",
            "preop_leg_pain_nrs_anytime",
            "preop_LegPain",
            "preop_legpain",
            "preop_leg_pain",
            "preop_LegPain_NRS",
            "preoperative_LegPain",
            "preoperative_leg_pain",
            "LegPain",
            "legpain",
            "LegPain_NRS",
            "leg_pain_NRS",
            "baseline_LegPain",
        ],
    ),
]

TARGET_COL_CANDIDATES = ["final_reop_step1","final_reop_1yr","reop_1yr","reoperation_1yr","one_year_reoperation","reop365","reop_365","final_reop","reoperation","reop"]

BASELINE_FEATURES = [
    "finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis",
    "age","sex","race","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure",
    "connective_tissue_rheumatic_disease","diabetes_status","myocardial_infarction","renal_disease","institution_type",
    "institution_size","institution_region","asa","bmi","payer_status","alif_llif","corpectomy","discectomy",
    "foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif",
    "other_lumbar_procedure","number_operated_levels","operative_region_extent","PatTobaccoUse"]
assert len(BASELINE_FEATURES) == 35

EXCLUDED_FEATURES = {"reop","reoptime","final_reop","final_reop_step1","final_reop_step2","death_within_90d","death_within_180d","death_within_365d","death_after_index_surgery","death_before_or_on_index_surgery","PersonDeathDate","days_to_death_from_index_surgery","removal_hardware","any_arthroplasty","final_diagnosis_complexity","procedure_complexity_score"}
CONTINUOUS_BASELINE_FEATURES = ["age","bmi"]
BINARY_FEATURES = ["finaldx_degenerative","finaldx_radicular","finaldx_stenosis","finaldx_deformity_instability","finaldx_other_diagnosis","sex","ethnicity","cancer_status","chronic_pulmonary_disease","congestive_heart_failure","connective_tissue_rheumatic_disease","myocardial_infarction","renal_disease","institution_type","alif_llif","corpectomy","discectomy","foraminotomy","instrumentation","laminectomy_posterior_decompression","pelvic_fixation","plf","tlif_plif","other_lumbar_procedure","operative_region_extent"]
ORDINAL_FEATURES = ["diabetes_status","institution_size","asa","number_operated_levels"]
NOMINAL_FEATURES = ["race","institution_region","payer_status","PatTobaccoUse"]

LGBM_SEARCH_SPACE = {
    "n_estimators": [400,700,1000,1400,1800,2200,2600],
    "learning_rate": [0.003,0.005,0.008,0.01,0.02,0.03,0.05],
    "num_leaves": [7,15,31,63,127], "max_depth": [-1,2,3,5,7,9],
    "min_child_samples": [10,20,50,100,200,400], "subsample": [0.60,0.75,0.90,1.00],
    "subsample_freq": [0,1,2], "colsample_bytree": [0.60,0.75,0.90,1.00],
    "reg_alpha": [0.0,0.001,0.01,0.05,0.10,0.50,1.00,2.00],
    "reg_lambda": [0.0,0.001,0.01,0.05,0.10,0.50,1.00,2.00,5.00],
    "min_split_gain": [0.0,0.001,0.005,0.01,0.05,0.10], "max_bin": [63,127,255],
    "positive_weight_multiplier": [0.25,0.50,0.75,1.00,1.50,2.00,3.00,4.00,6.00,8.00],
}
LGBM_INT_PARAMS = {"n_estimators","num_leaves","max_depth","min_child_samples","subsample_freq","max_bin"}
LGBM_FLOAT_PARAMS = {"learning_rate","subsample","colsample_bytree","reg_alpha","reg_lambda","min_split_gain","positive_weight_multiplier"}
PARAMETER_CANDIDATES = list(ParameterSampler(LGBM_SEARCH_SPACE, n_iter=N_RANDOM_COMBINATIONS, random_state=RANDOM_STATE))

# ============================================================
# 2) Parsing, labels, metrics
# ============================================================
MISSING_STRINGS = {""," ","na","n/a","nan","none","null",".","missing","<na>"}
BINARY_MAPS = {
    "sex": {"female":0,"f":0,"male":1,"m":1},
    "ethnicity": {"non-hispanic":0,"non hispanic":0,"hispanic":1},
    "cancer_status": {"no cancer":0,"no":0,"none":0,"any cancer":1,"yes":1,"cancer":1},
    "institution_type": {"hospital":0,"non-hospital":1,"non hospital":1,"nonhospital":1},
    "operative_region_extent": {"lumbar only":0,"extended_region_involvement":1,"extended region involvement":1,"extended":1},
}
ORDINAL_MAPS = {
    "diabetes_status": {"no":0,"none":0,"0":0,"without comp":1,"without complication":1,"without complications":1,"1":1,"with comp":2,"with complication":2,"with complications":2,"2":2},
    "institution_size": {"between 1-99 beds":0,"1-99":0,"between 100-399 beds":1,"100-399":1,">= 400 beds":2,">=400 beds":2,">=400":2,">= 400":2},
    "asa": {"1":1,"i":1,"2":2,"ii":2,"3":3,"iii":3,"4":4,"iv":4,">=4":4,">= 4":4,"5":4,"v":4},
    "number_operated_levels": {"0":0,"1":1,"2":2,"3":3,"4":4,">=4":4,">= 4":4,"5":4,"6":4,"7":4,"8":4,"9":4,"10":4},
}
PREFERRED_NOMINAL_LEVELS = {"race":["White","Black","Other"],"institution_region":["South","North East","West","Midwest"],"payer_status":["Medicare","Commercial/Private","Other","Medicaid/Public/Government"],"PatTobaccoUse":["Unknown/Not reported/Multiple","Never","Former","Current"]}
FEATURE_LABELS = {"age":"Age","bmi":"BMI","sex":"Male sex","race":"Race","ethnicity":"Hispanic ethnicity","cancer_status":"Any cancer","diabetes_status":"Diabetes status","institution_type":"Non-hospital institution","institution_size":"Institution size","institution_region":"Institution region","asa":"ASA physical status","payer_status":"Primary payer","PatTobaccoUse":"Tobacco use","finaldx_degenerative":"Degenerative diagnosis","finaldx_radicular":"Radiculopathy diagnosis","finaldx_stenosis":"Spinal stenosis diagnosis","finaldx_deformity_instability":"Deformity or instability diagnosis","finaldx_other_diagnosis":"Other lumbar diagnosis","chronic_pulmonary_disease":"Chronic pulmonary disease","congestive_heart_failure":"Congestive heart failure","connective_tissue_rheumatic_disease":"Connective tissue/rheumatic disease","myocardial_infarction":"Myocardial infarction","renal_disease":"Renal disease","alif_llif":"Anterior/lateral lumbar interbody fusion","corpectomy":"Corpectomy","discectomy":"Discectomy","foraminotomy":"Foraminotomy","instrumentation":"Instrumentation","laminectomy_posterior_decompression":"Posterior decompression or laminectomy","pelvic_fixation":"Pelvic fixation","plf":"Posterolateral fusion","tlif_plif":"Posterior/transforaminal lumbar interbody fusion","other_lumbar_procedure":"Other lumbar procedure","number_operated_levels":"Number of operated levels","operative_region_extent":"Operative region extent"}
COHORT_LABELS = {"ODI":"ODI cohort","BackPain":"Back-pain NRS cohort","LegPain":"Leg-pain NRS cohort"}

def clean_scalar(x):
    if pd.isna(x): return np.nan
    if isinstance(x,str):
        s=x.strip().replace("≥",">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x

def norm_text(x):
    x=clean_scalar(x)
    if pd.isna(x): return None
    return str(x).strip().replace("≥",">=").lower()

def to_binary_target(x):
    sx=norm_text(x)
    if sx is None: return np.nan
    if sx in {"1","1.0","yes","y","true","t"}: return 1.0
    if sx in {"0","0.0","no","n","false","f"}: return 0.0
    try:
        v=float(sx); return float(v) if v in (0.0,1.0) else np.nan
    except Exception: return np.nan

def pretty_feature_name(f): return FEATURE_LABELS.get(f, f.replace("_"," "))
def safe_filename(x): return re.sub(r"_+","_",re.sub(r"[^A-Za-z0-9_.-]+","_",str(x))).strip("_")[:180] or "file"
def json_native(obj):
    if isinstance(obj,dict): return {str(k):json_native(v) for k,v in obj.items()}
    if isinstance(obj,(list,tuple)): return [json_native(v) for v in obj]
    if isinstance(obj,(np.integer,)): return int(obj)
    if isinstance(obj,(np.floating,)): return float(obj)
    if isinstance(obj,np.ndarray): return obj.tolist()
    try:
        if pd.isna(obj): return None
    except Exception: pass
    return obj

def sanitize_lgbm_params(params):
    out={}
    for k,v in params.items():
        if k in LGBM_INT_PARAMS: out[k]=int(round(float(v)))
        elif k in LGBM_FLOAT_PARAMS: out[k]=float(v)
        elif isinstance(v,(np.integer,)): out[k]=int(v)
        elif isinstance(v,(np.floating,)): out[k]=float(v)
        else: out[k]=v
    return out

def find_existing_column(columns, candidates, what):
    lookup={c.lower():c for c in columns}
    for c in candidates:
        if c in columns: return c
        if c.lower() in lookup: return lookup[c.lower()]
    raise ValueError(f"Could not find {what}. Tried: {candidates}")

def resolve_target_column(df): return find_existing_column(df.columns.tolist(), TARGET_COL_CANDIDATES, "Step 1 target column")

def safe_average_precision(y,p): return np.nan if len(np.unique(y))<2 else float(average_precision_score(y,p))
def safe_roc_auc(y,p): return np.nan if len(np.unique(y))<2 else float(roc_auc_score(y,p))

def ece(y,p,n_bins=10):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); bins=np.linspace(0,1,n_bins+1); out=0.0
    if len(y)==0: return np.nan
    for i in range(n_bins):
        m=(p>=bins[i]) & ((p<=bins[i+1]) if i==n_bins-1 else (p<bins[i+1]))
        if np.any(m): out += (np.sum(m)/len(y))*abs(float(np.mean(y[m]))-float(np.mean(p[m])))
    return float(out)

def class_metrics(y,p,t):
    y=np.asarray(y).astype(int); yh=(np.asarray(p)>=t).astype(int); tn,fp,fn,tp=confusion_matrix(y,yh,labels=[0,1]).ravel()
    return {"threshold":float(t),"F1":float(f1_score(y,yh,zero_division=0)),"Sensitivity":tp/(tp+fn) if tp+fn>0 else np.nan,"Specificity":tn/(tn+fp) if tn+fp>0 else np.nan,"PPV":tp/(tp+fp) if tp+fp>0 else np.nan,"NPV":tn/(tn+fn) if tn+fn>0 else np.nan,"TP":int(tp),"FP":int(fp),"TN":int(tn),"FN":int(fn),"Predicted_positive_rate":float(np.mean(yh))}

def top5_metrics(y,p):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); n=len(y); k=max(1,int(math.ceil(n*0.05))); idx=np.argsort(-p)[:k]
    prev=float(np.mean(y)); rate=float(np.mean(y[idx])); return {"Top_5pct_n":int(k),"Top_5pct_event_rate":rate,"Top_5pct_lift":rate/prev if prev>0 else np.nan,"Top_5pct_captured_events":float(np.sum(y[idx])/np.sum(y)) if np.sum(y)>0 else np.nan}

def select_threshold(y,p):
    if THRESHOLD_STRATEGY=="fixed": t=float(FIXED_THRESHOLD)
    elif THRESHOLD_STRATEGY=="top_percent": t=float(np.quantile(p,1-THRESHOLD_TOP_PERCENT))
    elif THRESHOLD_STRATEGY=="youden":
        fpr,tpr,thr=roc_curve(y,p); t=float(thr[int(np.nanargmax(tpr-fpr))])
    elif THRESHOLD_STRATEGY=="max_f1":
        pr,rc,thr=precision_recall_curve(y,p)
        if len(thr)==0: t=0.5
        else:
            pr=pr[:-1]; rc=rc[:-1]; f1=2*pr*rc/np.maximum(pr+rc,1e-12); t=float(thr[int(np.nanargmax(f1))])
    else: raise ValueError(THRESHOLD_STRATEGY)
    return t, class_metrics(y,p,t)

def eval_preds(y,p,threshold=None,prefix=""):
    out={f"{prefix}AP":safe_average_precision(y,p),f"{prefix}ROC_AUC":safe_roc_auc(y,p),f"{prefix}Brier_score":float(brier_score_loss(y,p)),f"{prefix}ECE":ece(y,p,ECE_N_BINS),f"{prefix}N":int(len(y)),f"{prefix}Events":int(np.sum(y)),f"{prefix}Prevalence":float(np.mean(y))}
    if threshold is not None:
        out.update({f"{prefix}{k}":v for k,v in class_metrics(y,p,threshold).items()}); out.update({f"{prefix}{k}":v for k,v in top5_metrics(y,p).items()})
    return out

def make_weights(y,mult):
    y=np.asarray(y).astype(int); npos=int(np.sum(y==1)); nneg=int(np.sum(y==0))
    if npos==0: raise ValueError("No positive events.")
    pw=(nneg/max(npos,1))*float(mult); return np.where(y==1,pw,1.0).astype(float)
def actual_positive_weight(y,mult):
    y=np.asarray(y).astype(int); return float((np.sum(y==0)/max(np.sum(y==1),1))*float(mult))


def holm_adjust_pvalues(p_values):
    """Holm-adjust p values while preserving original row order."""
    p = np.asarray(p_values, dtype=float)
    out = np.full(p.shape, np.nan, dtype=float)
    finite = np.where(np.isfinite(p))[0]
    if len(finite) == 0:
        return out
    order = finite[np.argsort(p[finite])]
    m = len(order)
    adjusted_sorted = np.empty(m, dtype=float)
    running = 0.0
    for rank, idx in enumerate(order):
        value = min((m - rank) * p[idx], 1.0)
        running = max(running, value)
        adjusted_sorted[rank] = running
    for rank, idx in enumerate(order):
        out[idx] = adjusted_sorted[rank]
    return out

# ============================================================
# 3) Preprocessor
# ============================================================
@dataclass
class Step1Preprocessor:
    """
    Leakage-safe preprocessing fitted only on the training split.

    scale_continuous=False for tree/boosting models such as LightGBM.
    scale_continuous=True for scale-sensitive classifiers used in the
    eight-classifier screening script, such as logistic regression and
    linear SVC. The scaler is always fitted only on the training data and
    then applied unchanged to calibration/test/CV validation data.
    """
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str,List[str]]
    scale_continuous: bool = False

    def __post_init__(self):
        self.continuous_imputer = None
        self.continuous_scaler = None
        self.discrete_imputer = None
        self.nominal_imputer = None
        self.onehot = None
        self.output_feature_names_ = []

    def _parse_binary(self,x,f):
        sx=norm_text(x)
        if sx is None: return np.nan
        if f in BINARY_MAPS and sx in BINARY_MAPS[f]: return float(BINARY_MAPS[f][sx])
        if sx in {"1","1.0","yes","y","true","t","present","positive"}: return 1.0
        if sx in {"0","0.0","no","n","false","f","absent","negative"}: return 0.0
        try:
            v=float(sx); return float(v) if v in (0.0,1.0) else np.nan
        except Exception: return np.nan

    def _parse_ordinal(self,x,f):
        sx=norm_text(x)
        if sx is None: return np.nan
        if f in ORDINAL_MAPS and sx in ORDINAL_MAPS[f]: return float(ORDINAL_MAPS[f][sx])
        try:
            v=float(sx)
            if f=="asa": return float(min(max(int(round(v)),1),4))
            if f=="number_operated_levels": return float(min(max(int(round(v)),0),4))
            if f=="diabetes_status": return float(min(max(int(round(v)),0),2))
            if f=="institution_size": return float(min(max(int(round(v)),0),2))
            return float(v)
        except Exception: return np.nan

    def _nominal(self,f,x):
        x=clean_scalar(x)
        if pd.isna(x): return np.nan
        s=str(x).strip(); sl=s.lower()
        if f=="race": return "White" if sl=="white" else ("Black" if sl=="black" else "Other")
        if f=="institution_region": return {"south":"South","north east":"North East","northeast":"North East","north-east":"North East","west":"West","midwest":"Midwest","mid west":"Midwest"}.get(sl,s)
        if f=="payer_status":
            if sl=="medicare": return "Medicare"
            if sl in {"commercial/private","commercial","private","commercial private"}: return "Commercial/Private"
            if sl in {"medicaid/public/government","medicaid","public","government","public/government"}: return "Medicaid/Public/Government"
            return "Other"
        if f=="PatTobaccoUse": return "Never" if sl=="never" else ("Former" if sl=="former" else ("Current" if sl=="current" else "Unknown/Not reported/Multiple"))
        return s

    def _parts(self,X):
        cont=pd.DataFrame(index=X.index); disc=pd.DataFrame(index=X.index); nom=pd.DataFrame(index=X.index)
        for c in self.continuous_features: cont[c]=pd.to_numeric(X[c].map(clean_scalar),errors="coerce")
        for c in self.binary_features: disc[c]=X[c].map(lambda z:self._parse_binary(z,c)).astype(float)
        for c in self.ordinal_features: disc[c]=X[c].map(lambda z:self._parse_ordinal(z,c)).astype(float)
        for c in self.nominal_features: nom[c]=X[c].map(lambda z:self._nominal(c,z)).astype("object")
        return cont,disc,nom

    def fit(self,X):
        cont,disc,nom=self._parts(X)
        self.continuous_imputer=SimpleImputer(strategy="median")
        self.discrete_imputer=SimpleImputer(strategy="most_frequent")
        self.nominal_imputer=SimpleImputer(strategy="constant",fill_value="Missing")
        cont_imp=self.continuous_imputer.fit_transform(cont)
        if self.scale_continuous:
            self.continuous_scaler=StandardScaler()
            self.continuous_scaler.fit(cont_imp)
        self.discrete_imputer.fit(disc)
        nomi=pd.DataFrame(self.nominal_imputer.fit_transform(nom),columns=self.nominal_features)
        cats=[]
        for c in self.nominal_features:
            pref=list(self.preferred_nominal_levels.get(c,[])); obs=nomi[c].astype(str).unique().tolist(); cats.append(pref+sorted([x for x in obs if x not in pref]))
        try: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse_output=False)
        except TypeError: self.onehot=OneHotEncoder(categories=cats,handle_unknown="ignore",sparse=False)
        self.onehot.fit(nomi.astype(str))
        self.output_feature_names_=self.continuous_features+self.binary_features+self.ordinal_features+self.onehot.get_feature_names_out(self.nominal_features).tolist()
        return self

    def transform(self,X):
        cont,disc,nom=self._parts(X)
        a=self.continuous_imputer.transform(cont)
        if self.scale_continuous and self.continuous_scaler is not None:
            a=self.continuous_scaler.transform(a)
        b=self.discrete_imputer.transform(disc)
        nomi=pd.DataFrame(self.nominal_imputer.transform(nom),columns=self.nominal_features)
        c=self.onehot.transform(nomi.astype(str))
        return np.concatenate([a,b,c],axis=1).astype(float)

    def fit_transform(self,X): self.fit(X); return self.transform(X)

    @property
    def output_feature_names(self): return self.output_feature_names_

# ============================================================
# 4) Split, tune, fit
# ============================================================
def patient_split(df,target_col,seed):
    gdf=df.groupby(GROUP_COL,dropna=False)[target_col].max().reset_index(); y=gdf[target_col].astype(int).to_numpy(); g=gdf[GROUP_COL].to_numpy()
    s1=StratifiedShuffleSplit(n_splits=1,test_size=TEST_FRACTION,random_state=seed); trcal,te=next(s1.split(g,y))
    gtc=g[trcal]; ytc=y[trcal]; s2=StratifiedShuffleSplit(n_splits=1,test_size=CALIBRATION_FRACTION_OF_REMAINING,random_state=seed+1); tr,ca=next(s2.split(gtc,ytc))
    train=set(gtc[tr]); cal=set(gtc[ca]); test=set(g[te]); assert train.isdisjoint(cal) and train.isdisjoint(test) and cal.isdisjoint(test)
    return df[GROUP_COL].isin(train).to_numpy(), df[GROUP_COL].isin(cal).to_numpy(), df[GROUP_COL].isin(test).to_numpy()

def cv_splits(y,groups,seed,n_folds=N_CV_FOLDS):
    ge=pd.DataFrame({"g":groups,"y":y}).groupby("g")["y"].max().reset_index(); nf=min(n_folds,int(np.sum(ge.y==1)),int(np.sum(ge.y==0)))
    if nf<2: raise ValueError("Not enough patient groups for CV.")
    cv=StratifiedGroupKFold(n_splits=nf,shuffle=True,random_state=seed)
    return [(tr,va) for tr,va in cv.split(np.zeros(len(y)),y,groups)]

def feature_types(features):
    cont=[f for f in features if f in CONTINUOUS_BASELINE_FEATURES or f not in BASELINE_FEATURES]
    return cont,[f for f in features if f in BINARY_FEATURES],[f for f in features if f in ORDINAL_FEATURES],[f for f in features if f in NOMINAL_FEATURES]

def make_model(params,seed,n_estimators_override=None):
    p=sanitize_lgbm_params(params); mp={k:v for k,v in p.items() if k!="positive_weight_multiplier"}
    if n_estimators_override is not None: mp["n_estimators"]=int(max(MIN_FINAL_N_ESTIMATORS,n_estimators_override))
    return LGBMClassifier(objective="binary",boosting_type="gbdt",metric="average_precision",random_state=seed,n_jobs=N_JOBS,verbosity=-1,force_col_wise=True,**mp)

def fit_pipeline(X,y,features,params,seed,eval_set=None,early=False,n_estimators_override=None,scale_continuous=False):
    cont,binf,ordf,nomf=feature_types(features)
    pre=Step1Preprocessor(cont,binf,ordf,nomf,PREFERRED_NOMINAL_LEVELS,scale_continuous=scale_continuous)
    Xt=pre.fit_transform(X[features])
    w=make_weights(y,params["positive_weight_multiplier"])
    model=make_model(params,seed,n_estimators_override)
    best_iter=None
    if eval_set is not None and early:
        Xv,yv=eval_set
        Xpv=pre.transform(Xv[features])
        callbacks=[lgb.early_stopping(EARLY_STOPPING_ROUNDS,verbose=False),lgb.log_evaluation(period=0)]
        try:
            model.fit(Xt,y,sample_weight=w,eval_set=[(Xpv,yv)],eval_metric="average_precision",callbacks=callbacks)
            best_iter=int(model.best_iteration_) if getattr(model,"best_iteration_",None) else None
        except Exception:
            model.fit(Xt,y,sample_weight=w)
    else:
        model.fit(Xt,y,sample_weight=w)
    return pre,model,best_iter

def predict(pre,model,X,features): return model.predict_proba(pre.transform(X[features]))[:,1]

def tune(X_train,y_train,groups_train,features,folds,model_key,seed):
    cand_rows=[]; fold_rows=[]
    for cid,raw in enumerate(PARAMETER_CANDIDATES,1):
        params=sanitize_lgbm_params(raw); aps=[]; aucs=[]; bests=[]; t0=time.time()
        for fid,(tr,va) in enumerate(folds,1):
            Xtr=X_train.iloc[tr].reset_index(drop=True); ytr=y_train[tr]; Xva=X_train.iloc[va].reset_index(drop=True); yva=y_train[va]
            pre,model,best=fit_pipeline(Xtr,ytr,features,params,seed+cid*1000+fid,eval_set=(Xva,yva),early=USE_EARLY_STOPPING_IN_CV)
            p=predict(pre,model,Xva,features); ap=safe_average_precision(yva,p); auc=safe_roc_auc(yva,p); aps.append(ap); aucs.append(auc)
            if best and best>0: bests.append(best)
            fold_rows.append({"model_key":model_key,"candidate_id":cid,"fold":fid,"fold_AP":ap,"fold_ROC_AUC":auc,"fold_best_iteration":best,"positive_weight_used":actual_positive_weight(ytr,params["positive_weight_multiplier"]),**params})
        locked=int(np.median(bests)) if bests else int(params["n_estimators"])
        row={"model_key":model_key,"candidate_id":cid,"cv_folds":len(folds),"cv_AP_mean":float(np.nanmean(aps)),"cv_AP_SD":float(np.nanstd(aps,ddof=1)),"cv_ROC_AUC_mean":float(np.nanmean(aucs)),"cv_ROC_AUC_SD":float(np.nanstd(aucs,ddof=1)),"mean_cv_best_iteration":float(np.mean(bests)) if bests else np.nan,"locked_n_estimators_from_cv":locked,"elapsed_seconds":float(time.time()-t0),**params}
        cand_rows.append(row)
        print(f"{model_key} | candidate {cid:03d}/{len(PARAMETER_CANDIDATES)} | CV AP={row['cv_AP_mean']:.5f} | AUC={row['cv_ROC_AUC_mean']:.5f} | locked_n={locked}")
    return pd.DataFrame(cand_rows).sort_values("cv_AP_mean",ascending=False).reset_index(drop=True), pd.DataFrame(fold_rows)

def fit_final(Xtr,ytr,Xcal,ycal,Xte,features,params,locked_n,seed):
    pre,model,_=fit_pipeline(Xtr,ytr,features,params,seed,n_estimators_override=locked_n,scale_continuous=SCALE_CONTINUOUS_FOR_FINAL_LIGHTGBM)
    ptr_raw=predict(pre,model,Xtr,features); pcal_raw=predict(pre,model,Xcal,features); pte_raw=predict(pre,model,Xte,features)
    if CALIBRATION_METHOD=="isotonic":
        cal=IsotonicRegression(out_of_bounds="clip"); cal.fit(pcal_raw,ycal); ptr=cal.predict(ptr_raw); pcal=cal.predict(pcal_raw); pte=cal.predict(pte_raw)
    else: cal=None; ptr=ptr_raw; pcal=pcal_raw; pte=pte_raw
    thr,calm=select_threshold(ycal,pcal)
    return {"pre":pre,"model":model,"calibrator":cal,"p_train_raw":ptr_raw,"p_train":ptr,"p_cal_raw":pcal_raw,"p_cal":pcal,"p_test_raw":pte_raw,"p_test":pte,"threshold":thr,"cal_metrics":calm}

# ============================================================
# 5) Bootstrap inference
# ============================================================
def patient_boot_ci(y,p,groups,metric,threshold=None,seed=1):
    rng=np.random.default_rng(seed); d=pd.DataFrame({"id":np.arange(len(y)),"g":groups,"y":np.asarray(y).astype(int)})
    gy=d.groupby("g").y.max(); pos=gy[gy==1].index.to_numpy(); neg=gy[gy==0].index.to_numpy()
    if len(pos)==0 or len(neg)==0: return np.nan,np.nan
    by={g:d.loc[d.g.eq(g),"id"].to_numpy() for g in gy.index}; vals=[]
    for _ in range(N_BOOTSTRAPS):
        sg=np.concatenate([rng.choice(pos,len(pos),True),rng.choice(neg,len(neg),True)]); idx=np.concatenate([by[g] for g in sg]); yy=y[idx]; pp=p[idx]
        if len(np.unique(yy))<2: continue
        if metric=="AP": vals.append(average_precision_score(yy,pp))
        elif metric=="ROC_AUC": vals.append(roc_auc_score(yy,pp))
        elif metric=="F1": vals.append(f1_score(yy,(pp>=threshold).astype(int),zero_division=0))
    return tuple(map(float,np.percentile(vals,[2.5,97.5]))) if vals else (np.nan,np.nan)

def paired_delta_boot(y,p0,p1,groups,metric,seed=2):
    obs=(safe_average_precision(y,p1)-safe_average_precision(y,p0)) if metric=="AP" else (safe_roc_auc(y,p1)-safe_roc_auc(y,p0))
    rng=np.random.default_rng(seed); d=pd.DataFrame({"id":np.arange(len(y)),"g":groups,"y":np.asarray(y).astype(int)})
    gy=d.groupby("g").y.max(); pos=gy[gy==1].index.to_numpy(); neg=gy[gy==0].index.to_numpy()
    if len(pos)==0 or len(neg)==0: return obs,np.nan,np.nan,np.nan
    by={g:d.loc[d.g.eq(g),"id"].to_numpy() for g in gy.index}; vals=[]
    for _ in range(N_BOOTSTRAPS):
        sg=np.concatenate([rng.choice(pos,len(pos),True),rng.choice(neg,len(neg),True)]); idx=np.concatenate([by[g] for g in sg]); yy=y[idx]
        if len(np.unique(yy))<2: continue
        vals.append((average_precision_score(yy,p1[idx])-average_precision_score(yy,p0[idx])) if metric=="AP" else (roc_auc_score(yy,p1[idx])-roc_auc_score(yy,p0[idx])))
    if not vals: return obs,np.nan,np.nan,np.nan
    lo,hi=np.percentile(vals,[2.5,97.5]); vals=np.asarray(vals); pv=2*min(np.mean(vals<=0),np.mean(vals>=0))
    return float(obs),float(lo),float(hi),float(min(max(pv,0),1))

# ============================================================
# 6) Plots and SHAP
# ============================================================
def save_pr(y,p,path,title):
    pr,rc,_=precision_recall_curve(y,p); ap=average_precision_score(y,p); plt.figure(figsize=(6.2,5.2)); plt.plot(rc,pr,lw=2); plt.xlabel("Recall"); plt.ylabel("Precision"); plt.title(f"{title}\nAP = {ap:.3f}"); plt.tight_layout(); plt.savefig(path,dpi=300,bbox_inches="tight"); plt.close()

def save_roc(y,p,path,title):
    fpr,tpr,_=roc_curve(y,p); auc=roc_auc_score(y,p); plt.figure(figsize=(6.2,5.2)); plt.plot(fpr,tpr,lw=2); plt.plot([0,1],[0,1],'--',lw=1); plt.xlabel("False-positive rate"); plt.ylabel("True-positive rate"); plt.title(f"{title}\nROC-AUC = {auc:.3f}"); plt.tight_layout(); plt.savefig(path,dpi=300,bbox_inches="tight"); plt.close()

def save_calplot(y,p,path,title):
    y=np.asarray(y).astype(int); p=np.asarray(p).astype(float); brier=float(brier_score_loss(y,p)); ee=ece(y,p,ECE_N_BINS); prev=float(np.mean(y)); tmp=pd.DataFrame({"y":y,"p":p}); nb=min(ECE_N_BINS,max(2,tmp.p.nunique())); tmp["bin"]=pd.qcut(tmp.p.rank(method="first"),q=nb,duplicates="drop"); cal=tmp.groupby("bin",observed=True).agg(mean_pred=("p","mean"),observed=("y","mean"),n=("y","size")).reset_index(drop=True)
    axmax=max(0.05,min(0.35,max(float(np.nanquantile(p,0.995)),float(cal.observed.max()),prev)*1.35+0.005)); fig,ax=plt.subplots(figsize=(7.2,6.0)); ax.plot([0,axmax],[0,axmax],'--',color='black',lw=1.6,label="Perfect calibration"); ax.plot(cal.mean_pred,cal.observed,marker='o',lw=2.2,label="Model calibration"); ax.axhline(prev,ls=':',color='gray',label=f"Observed event rate = {prev:.2%}")
    for _,r in cal.iterrows(): ax.annotate(f"n={int(r.n)}",(float(r.mean_pred),float(r.observed)),textcoords="offset points",xytext=(5,4),fontsize=7)
    ax.set_xlim(0,axmax); ax.set_ylim(0,axmax); ax.set_xlabel("Mean predicted probability"); ax.set_ylabel("Observed event rate"); ax.set_title(title,fontweight="bold"); ax.legend(); ax.text(.98,.04,f"Brier score = {brier:.4f}\nECE = {ee:.4f}\nN = {len(y):,}\nEvents = {int(np.sum(y)):,}",transform=ax.transAxes,ha="right",va="bottom",bbox=dict(boxstyle="round",facecolor="white",alpha=.95)); fig.tight_layout(); fig.savefig(path,dpi=300,bbox_inches="tight"); plt.close(fig)

def learning_curve(X,y,groups,features,params,locked,path_png,path_csv,title,seed):
    nf=min(LEARNING_CURVE_CV_FOLDS, len(np.unique(groups))); cv=StratifiedGroupKFold(n_splits=cv_splits(y,groups,seed,nf).__len__(),shuffle=True,random_state=seed); rows=[]; rng=np.random.default_rng(seed)
    for frac in LEARNING_CURVE_TRAIN_FRACTIONS:
        for fid,(tr,va) in enumerate(cv.split(np.zeros(len(y)),y,groups),1):
            Xfull=X.iloc[tr].reset_index(drop=True); yfull=y[tr]; gfull=groups[tr]; Xva=X.iloc[va].reset_index(drop=True); yva=y[va]
            ge=pd.DataFrame({"g":gfull,"y":yfull}).groupby("g").y.max().reset_index(); pos=ge.loc[ge.y.eq(1),"g"].to_numpy(); neg=ge.loc[ge.y.eq(0),"g"].to_numpy(); keepg=set(np.concatenate([rng.choice(pos,max(1,int(np.ceil(len(pos)*frac))),False), rng.choice(neg,max(1,int(np.ceil(len(neg)*frac))),False)])); keep=np.array([g in keepg for g in gfull]); Xtr=Xfull.loc[keep].reset_index(drop=True); ytr=yfull[keep]
            if len(np.unique(ytr))<2 or len(np.unique(yva))<2: continue
            pre,model,_=fit_pipeline(Xtr,ytr,features,params,seed+fid+int(frac*1000),n_estimators_override=locked); rows.append({"train_fraction":frac,"fold":fid,"train_n":len(ytr),"train_events":int(np.sum(ytr)),"validation_n":len(yva),"validation_events":int(np.sum(yva)),"train_AP":safe_average_precision(ytr,predict(pre,model,Xtr,features)),"validation_AP":safe_average_precision(yva,predict(pre,model,Xva,features))})
    raw=pd.DataFrame(rows); raw.to_csv(path_csv,index=False)
    if raw.empty: return raw
    lc=raw.groupby("train_fraction",as_index=False).agg(train_AP_mean=("train_AP","mean"),train_AP_sd=("train_AP","std"),validation_AP_mean=("validation_AP","mean"),validation_AP_sd=("validation_AP","std")).fillna(0)
    fig,ax=plt.subplots(figsize=(7.2,6)); ax.plot(lc.train_fraction,lc.train_AP_mean,marker='o',lw=2.2,label="Training AP"); ax.plot(lc.train_fraction,lc.validation_AP_mean,marker='o',lw=2.2,label="Validation AP"); ax.fill_between(lc.train_fraction,lc.train_AP_mean-lc.train_AP_sd,lc.train_AP_mean+lc.train_AP_sd,alpha=.15); ax.fill_between(lc.train_fraction,lc.validation_AP_mean-lc.validation_AP_sd,lc.validation_AP_mean+lc.validation_AP_sd,alpha=.15); ax.set_xlabel("Fraction of training patient groups used"); ax.set_ylabel("Average precision"); ax.set_title(title,fontweight="bold"); ax.legend(); ax.grid(alpha=.2); fig.tight_layout(); fig.savefig(path_png,dpi=300,bbox_inches="tight"); plt.close(fig); return raw

def raw_value(X,feature,cont_features):
    if feature in cont_features: s=pd.to_numeric(X[feature].map(clean_scalar),errors="coerce")
    elif feature in BINARY_FEATURES: s=X[feature].map(lambda z: Step1Preprocessor([],[],[],[],{})._parse_binary(z,feature)).astype(float)
    elif feature in ORDINAL_FEATURES: s=X[feature].map(lambda z: Step1Preprocessor([],[],[],[],{})._parse_ordinal(z,feature)).astype(float)
    else:
        vals=X[feature].map(lambda z: str(clean_scalar(z)) if not pd.isna(clean_scalar(z)) else "Missing"); cats={v:i for i,v in enumerate(sorted(vals.dropna().unique()))}; s=vals.map(cats).astype(float)
    return s.fillna(s.median() if not pd.isna(s.median()) else 0).astype(float)

def compute_shap(pre,model,Xraw,y,p,features,model_key):
    # Explain the complete held-out test set; no row cap or sampling is applied.
    n=len(Xraw); idx=np.arange(n)
    Xex=Xraw.iloc[idx].reset_index(drop=True); Xenc=pre.transform(Xex[features]); names=list(pre.output_feature_names); expl=shap.TreeExplainer(model); sv=expl.shap_values(Xenc); sv=sv[-1] if isinstance(sv,list) else sv; sv=np.asarray(sv); sv=sv[:,:,-1] if sv.ndim==3 else sv
    cont,_,_,_=feature_types(features); mapping={}
    for f in features:
        ids=[]
        if f in names: ids.append(names.index(f))
        if f in NOMINAL_FEATURES: ids += [i for i,nm in enumerate(names) if nm.startswith(f+"_")]
        if ids: mapping[f]=sorted(set(ids))
    vals={}; data={}; mapnames={}
    for f,ids in mapping.items():
        disp=pretty_feature_name(f); vals[disp]=sv[:,ids].sum(axis=1); data[disp]=raw_value(Xex,f,cont).values; mapnames[f]=[names[i] for i in ids]
    shapdf=pd.DataFrame(vals); datadf=pd.DataFrame(data); total=float(np.abs(shapdf.values).mean(axis=0).sum()); imp=[]
    for f in mapping:
        disp=pretty_feature_name(f); ma=float(np.abs(shapdf[disp]).mean()); imp.append({"model_key":model_key,"raw_feature":f,"display_feature":disp,"mean_abs_SHAP":ma,"percent_of_grouped_SHAP":100*ma/total if total>0 else np.nan,"n_encoded_columns_grouped":len(mapping[f])})
    imp=pd.DataFrame(imp).sort_values("mean_abs_SHAP",ascending=False).reset_index(drop=True); shapdf.insert(0,"__row_id__",idx); shapdf.insert(1,"__y_true__",np.asarray(y)[idx]); shapdf.insert(2,"__p_calibrated__",np.asarray(p)[idx]); return shapdf,datadf,imp,mapnames

def save_shap(shapdf,datadf,imp,beeswarm_path,bar_path,title):
    ordered=imp.display_feature.tolist()[:min(SHAP_BEESWARM_MAX_DISPLAY,len(imp))]; plt.figure(figsize=(11,max(9,.32*len(ordered)+2))); shap.summary_plot(shapdf[ordered].values,features=datadf[ordered],feature_names=ordered,max_display=len(ordered),show=False,plot_size=None); plt.title(title+": grouped SHAP beeswarm",fontweight="bold"); plt.tight_layout(); plt.savefig(beeswarm_path,dpi=300,bbox_inches="tight"); plt.close()
    pdf=imp.head(SHAP_BAR_MAX_DISPLAY).copy().iloc[::-1]; fig,ax=plt.subplots(figsize=(10.5,max(7,.35*len(pdf)))); bars=ax.barh(pdf.display_feature,pdf.mean_abs_SHAP); mx=float(pdf.mean_abs_SHAP.max()) if len(pdf) else 1; ax.set_xlim(0,mx*1.3); ax.set_xlabel("Mean absolute SHAP value"); ax.set_title(title+": grouped SHAP importance",fontweight="bold")
    for b,v,pct in zip(bars,pdf.mean_abs_SHAP,pdf.percent_of_grouped_SHAP): ax.text(b.get_width()+mx*.015,b.get_y()+b.get_height()/2,f"{v:.4f} ({pct:.1f}%)",va="center",fontsize=9)
    fig.tight_layout(); fig.savefig(bar_path,dpi=300,bbox_inches="tight"); plt.close(fig)


# ============================================================
# 6b) Leakage-safe OOF SHAP thresholds, SHAP heatmaps, and sensitivity analyses
# ============================================================

def grouped_shap_values(pre, model, Xraw, features):
    """Return grouped SHAP values by raw feature for a fitted LightGBM pipeline."""
    Xenc = pre.transform(Xraw[features])
    names = list(pre.output_feature_names)
    expl = shap.TreeExplainer(model)
    sv = expl.shap_values(Xenc)
    sv = sv[-1] if isinstance(sv, list) else sv
    sv = np.asarray(sv)
    sv = sv[:, :, -1] if sv.ndim == 3 else sv

    cont, _, _, _ = feature_types(features)
    mapping = {}
    for f in features:
        ids = []
        if f in names:
            ids.append(names.index(f))
        if f in NOMINAL_FEATURES:
            ids += [i for i, nm in enumerate(names) if nm.startswith(f + "_")]
        if ids:
            mapping[f] = sorted(set(ids))

    vals, data, mapnames = {}, {}, {}
    for f, ids in mapping.items():
        disp = pretty_feature_name(f)
        vals[disp] = sv[:, ids].sum(axis=1)
        data[disp] = raw_value(Xraw.reset_index(drop=True), f, cont).values
        mapnames[f] = [names[i] for i in ids]
    shapdf = pd.DataFrame(vals)
    datadf = pd.DataFrame(data)
    total = float(np.abs(shapdf.values).mean(axis=0).sum()) if shapdf.shape[1] else 0.0
    imp = []
    for f in mapping:
        disp = pretty_feature_name(f)
        ma = float(np.abs(shapdf[disp]).mean())
        imp.append({
            "raw_feature": f,
            "display_feature": disp,
            "mean_abs_SHAP": ma,
            "percent_of_grouped_SHAP": 100.0 * ma / total if total > 0 else np.nan,
            "n_encoded_columns_grouped": len(mapping[f]),
        })
    imp = pd.DataFrame(imp).sort_values("mean_abs_SHAP", ascending=False).reset_index(drop=True)
    return shapdf, datadf, imp, mapnames


def save_shap_heatmap(shapdf, imp, path, title):
    """Save a compact SHAP heatmap using top grouped features and rows ordered by predicted risk."""
    if shapdf.empty or imp.empty:
        return None
    ordered = [f for f in imp.display_feature.tolist() if f in shapdf.columns]
    ordered = ordered[:min(SHAP_HEATMAP_MAX_DISPLAY, len(ordered))]
    if not ordered:
        return None
    work = shapdf.copy()
    if "__p_calibrated__" in work.columns:
        work = work.sort_values("__p_calibrated__", ascending=True)
    if len(work) > SHAP_HEATMAP_MAX_ROWS:
        # Deterministic evenly spaced row selection for visualization only; all SHAP values remain saved to CSV.
        take = np.linspace(0, len(work) - 1, SHAP_HEATMAP_MAX_ROWS).round().astype(int)
        work = work.iloc[take]
    mat = work[ordered].to_numpy(dtype=float)
    vmax = float(np.nanpercentile(np.abs(mat), 99)) if np.isfinite(mat).any() else 1.0
    if vmax <= 0 or not np.isfinite(vmax):
        vmax = 1.0
    fig, ax = plt.subplots(figsize=(max(8, 0.32 * len(work) / 10), max(6, 0.30 * len(ordered) + 2)))
    im = ax.imshow(mat.T, aspect="auto", cmap="coolwarm", vmin=-vmax, vmax=vmax, interpolation="nearest")
    ax.set_yticks(np.arange(len(ordered)))
    ax.set_yticklabels(ordered)
    ax.set_xticks([])
    ax.set_xlabel("Held-out test patients ordered by calibrated predicted risk")
    ax.set_title(title + ": grouped SHAP heatmap", fontweight="bold")
    cbar = fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02)
    cbar.set_label("Grouped SHAP value")
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def cv_fold_assignment_table(train_df, folds, target_col, model_key):
    rows = []
    for fid, (tr_idx, va_idx) in enumerate(folds, 1):
        for role, idxs in [("cv_train", tr_idx), ("cv_validation", va_idx)]:
            sub = train_df.iloc[idxs]
            rows.append(pd.DataFrame({
                "model_key": model_key,
                "cv_fold": fid,
                "cv_role": role,
                "row_index_in_training_split": idxs,
                GROUP_COL: sub[GROUP_COL].to_numpy(),
                "y": sub[target_col].astype(int).to_numpy(),
            }))
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def derive_threshold_from_oof_shap(oof_df):
    """Derive a PROM threshold from training-only OOF SHAP dependence, without outcomes."""
    d = oof_df[["prom_value", "prom_shap"]].copy()
    d = d.replace([np.inf, -np.inf], np.nan).dropna()
    if len(d) < max(10, SHAP_THRESHOLD_MIN_BIN_N):
        return np.nan, pd.DataFrame(), "insufficient_training_oof_shap_data"
    d = d.sort_values("prom_value").reset_index(drop=True)
    q = min(SHAP_THRESHOLD_N_BINS, int(d["prom_value"].nunique()), max(2, int(len(d) // SHAP_THRESHOLD_MIN_BIN_N)))
    q = max(2, q)
    d["bin"] = pd.qcut(d["prom_value"].rank(method="first"), q=q, duplicates="drop")
    b = d.groupby("bin", observed=True).agg(
        mean_prom=("prom_value", "mean"),
        median_prom=("prom_value", "median"),
        mean_shap=("prom_shap", "mean"),
        median_shap=("prom_shap", "median"),
        n=("prom_value", "size"),
    ).reset_index(drop=True)
    if len(b) < 2:
        return np.nan, b, "insufficient_unique_training_bins"
    b["smooth_mean_shap"] = b["mean_shap"].rolling(window=3, center=True, min_periods=1).mean()
    x = b["median_prom"].to_numpy(dtype=float)
    y = b["smooth_mean_shap"].to_numpy(dtype=float)

    # Prefer the first zero crossing from protective/negative to risk-increasing/positive SHAP.
    candidates = []
    for i in range(len(y) - 1):
        if not np.isfinite(y[i]) or not np.isfinite(y[i + 1]):
            continue
        if y[i] == 0:
            candidates.append((abs(y[i + 1] - y[i]), x[i], "zero_or_upward_crossing"))
        elif y[i] < 0 <= y[i + 1]:
            # Linear interpolation on the smoothed binned curve.
            denom = y[i + 1] - y[i]
            frac = -y[i] / denom if denom != 0 else 0.0
            xc = x[i] + frac * (x[i + 1] - x[i])
            candidates.append((abs(denom), xc, "negative_to_positive_zero_crossing"))
    if candidates:
        # Use the strongest local transition among zero-crossings.
        candidates.sort(key=lambda z: z[0], reverse=True)
        threshold = float(candidates[0][1])
        rule = candidates[0][2]
    else:
        # Fallback: most prominent local change in smoothed SHAP dependence.
        diffs = np.abs(np.diff(y))
        if len(diffs) == 0 or not np.isfinite(diffs).any():
            threshold = float(np.nanmedian(d["prom_value"]))
            rule = "fallback_median_prom_value_no_stable_shap_gradient"
        else:
            idx = int(np.nanargmax(diffs))
            threshold = float((x[idx] + x[idx + 1]) / 2.0)
            rule = "fallback_largest_smoothed_shap_gradient"
    return threshold, b, rule


def save_oof_threshold_plot(binned, threshold, path, title, rule):
    if binned is None or binned.empty or not np.isfinite(threshold):
        return None
    fig, ax = plt.subplots(figsize=(7.2, 5.6))
    ax.plot(binned["median_prom"], binned["mean_shap"], marker="o", linewidth=1.5, label="Binned OOF SHAP")
    ax.plot(binned["median_prom"], binned["smooth_mean_shap"], linewidth=2.4, label="Smoothed OOF SHAP")
    ax.axhline(0.0, linestyle="--", linewidth=1.1, color="black")
    ax.axvline(threshold, linestyle="--", linewidth=1.6, color="black", label=f"Locked threshold = {threshold:.2f}")
    ax.set_xlabel("PROM value in training split")
    ax.set_ylabel("Out-of-fold grouped SHAP contribution")
    ax.set_title(title + "\nTraining-only SHAP-threshold derivation", fontweight="bold")
    ax.text(0.98, 0.04, f"Rule: {rule}", transform=ax.transAxes, ha="right", va="bottom", fontsize=8.5,
            bbox=dict(boxstyle="round", facecolor="white", alpha=0.9))
    ax.grid(alpha=0.25)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def save_locked_threshold_test_risk_plot(test_summary, path, title):
    if test_summary is None or test_summary.empty:
        return None
    d = test_summary.copy()
    fig, ax = plt.subplots(figsize=(6.5, 5.2))
    x = np.arange(len(d))
    y = d["mean_calibrated_predicted_risk"].to_numpy(dtype=float) * 100.0
    ax.bar(x, y, alpha=0.85)
    ax.set_xticks(x)
    ax.set_xticklabels(d["threshold_group"].tolist(), rotation=15, ha="right")
    ax.set_ylabel("Mean calibrated predicted 1-year reoperation risk (%)")
    ax.set_title(title + "\nLocked training-derived PROM threshold applied to test set", fontweight="bold")
    ax.grid(axis="y", alpha=0.25)
    for i, row in d.reset_index(drop=True).iterrows():
        ax.text(i, y[i] + max(0.05, np.nanmax(y) * 0.03), f"n={int(row['n'])}\ne={int(row['events'])}", ha="center", va="bottom", fontsize=8.5)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def compute_oof_shap_threshold(train_df, target_col, features, selected_params, locked_n, folds, prom_feature, model_key, seed, out_dir):
    """Training-only out-of-fold SHAP threshold derivation for one PROM-expanded model."""
    if prom_feature is None:
        return None
    rows = []
    for fid, (tr_idx, va_idx) in enumerate(folds, 1):
        Xtr = train_df.iloc[tr_idx].reset_index(drop=True)
        ytr = Xtr[target_col].astype(int).to_numpy()
        Xva = train_df.iloc[va_idx].reset_index(drop=True)
        pre, model, _ = fit_pipeline(
            Xtr[features], ytr, features, selected_params, seed + 50000 + fid,
            eval_set=None, early=False, n_estimators_override=locked_n,
        )
        shap_va, data_va, _, _ = grouped_shap_values(pre, model, Xva[features], features)
        prom_display = pretty_feature_name(prom_feature)
        if prom_display not in shap_va.columns:
            continue
        rows.append(pd.DataFrame({
            "model_key": model_key,
            "cv_fold": fid,
            "row_index_in_training_split": va_idx,
            GROUP_COL: Xva[GROUP_COL].to_numpy() if GROUP_COL in Xva.columns else np.nan,
            "prom_feature": prom_feature,
            "prom_display_name": prom_display,
            "prom_value": raw_value(Xva, prom_feature, feature_types(features)[0]).to_numpy(dtype=float),
            "prom_shap": shap_va[prom_display].to_numpy(dtype=float),
        }))
    oof = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    os.makedirs(out_dir, exist_ok=True)
    oof_path = os.path.join(out_dir, f"OOF_training_SHAP_PROM_values_{model_key}.csv")
    oof.to_csv(oof_path, index=False)
    threshold, binned, rule = derive_threshold_from_oof_shap(oof) if not oof.empty else (np.nan, pd.DataFrame(), "no_oof_shap_rows")
    binned_path = os.path.join(out_dir, f"OOF_training_SHAP_PROM_threshold_bins_{model_key}.csv")
    binned.to_csv(binned_path, index=False)
    plot_path = os.path.join(out_dir, f"OOF_training_SHAP_PROM_threshold_{model_key}.png")
    save_oof_threshold_plot(binned, threshold, plot_path, f"{model_key}: {pretty_feature_name(prom_feature)}", rule)
    summary = {
        "model_key": model_key,
        "prom_feature": prom_feature,
        "prom_display_name": pretty_feature_name(prom_feature),
        "locked_shap_threshold": threshold,
        "threshold_derivation_split": "training_only_out_of_fold_cv",
        "threshold_derivation_uses_outcomes": False,
        "threshold_derivation_uses_test_set": False,
        "threshold_rule": rule,
        "n_oof_training_rows": int(len(oof)),
        "n_oof_bins": int(len(binned)),
        "oof_values_csv": oof_path,
        "oof_bins_csv": binned_path,
        "oof_threshold_plot": plot_path,
    }
    json.dump(json_native(summary), open(os.path.join(out_dir, f"OOF_training_SHAP_PROM_threshold_{model_key}.json"), "w"), indent=2, sort_keys=True)
    return summary


def summarize_locked_prom_threshold_in_test(te, target_col, p, prom_feature, threshold, model_key, out_dir):
    if prom_feature is None or not np.isfinite(threshold):
        return pd.DataFrame()
    values = pd.to_numeric(te[prom_feature].map(clean_scalar), errors="coerce")
    group = np.where(values >= threshold, f">= {threshold:.2f}", f"< {threshold:.2f}")
    tmp = pd.DataFrame({"threshold_group": group, "y": te[target_col].astype(int).to_numpy(), "p": np.asarray(p, dtype=float)})
    rows = []
    for lab in [f"< {threshold:.2f}", f">= {threshold:.2f}"]:
        d = tmp[tmp.threshold_group.eq(lab)]
        rows.append({
            "model_key": model_key,
            "prom_feature": prom_feature,
            "locked_shap_threshold": float(threshold),
            "threshold_group": lab,
            "n": int(len(d)),
            "events": int(d.y.sum()) if len(d) else 0,
            "observed_event_rate": float(d.y.mean()) if len(d) else np.nan,
            "mean_calibrated_predicted_risk": float(d.p.mean()) if len(d) else np.nan,
            "median_calibrated_predicted_risk": float(d.p.median()) if len(d) else np.nan,
        })
    out = pd.DataFrame(rows)
    if len(out) == 2 and out["mean_calibrated_predicted_risk"].iloc[0] > 0:
        rel = out["mean_calibrated_predicted_risk"].iloc[1] / out["mean_calibrated_predicted_risk"].iloc[0] - 1.0
        out["relative_increase_above_threshold_vs_below"] = [np.nan, float(rel)]
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"locked_SHAP_PROM_threshold_test_risk_summary_{model_key}.csv")
    out.to_csv(out_path, index=False)
    plot_path = os.path.join(out_dir, f"locked_SHAP_PROM_threshold_test_predicted_risk_{model_key}.png")
    save_locked_threshold_test_risk_plot(out, plot_path, f"{model_key}: {pretty_feature_name(prom_feature)}")
    return out


def find_optional_column(columns, candidates):
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if c.lower() in lookup:
            return lookup[c.lower()]
    return None


def build_sensitivity_strata(test_df):
    strata = []
    diagnostic = ["finaldx_degenerative", "finaldx_radicular", "finaldx_stenosis", "finaldx_deformity_instability", "finaldx_other_diagnosis"]
    procedural = ["alif_llif", "corpectomy", "discectomy", "foraminotomy", "instrumentation", "laminectomy_posterior_decompression", "pelvic_fixation", "plf", "tlif_plif", "other_lumbar_procedure"]
    for col in diagnostic:
        if col in test_df.columns:
            vals = test_df[col].map(lambda z: Step1Preprocessor([], [], [], [], {})._parse_binary(z, col)).astype(float)
            strata.append(("diagnostic", pretty_feature_name(col), vals.eq(1).to_numpy()))
    for col in procedural:
        if col in test_df.columns:
            vals = test_df[col].map(lambda z: Step1Preprocessor([], [], [], [], {})._parse_binary(z, col)).astype(float)
            strata.append(("procedural", pretty_feature_name(col), vals.eq(1).to_numpy()))
    hosp_col = find_optional_column(test_df.columns.tolist(), HOSPITAL_COLUMN_CANDIDATES)
    if hosp_col is not None:
        h = test_df[hosp_col].map(lambda x: "Missing" if pd.isna(clean_scalar(x)) else str(clean_scalar(x)))
        counts = h.value_counts(dropna=False)
        low_hospitals = set(counts[counts < HOSPITAL_MIN_TEST_N_FOR_INDIVIDUAL_REPORTING].index)
        for val in counts.index:
            if val in low_hospitals:
                continue
            strata.append(("hospital", f"{hosp_col}: {val}", h.eq(val).to_numpy()))
        if low_hospitals:
            strata.append(("hospital", f"{hosp_col}: lower-volume hospitals pooled", h.isin(low_hospitals).to_numpy()))
    return strata


def run_sensitivity_analyses(cohort, df, target_col, base_result, exp_result, out_dir):
    """Apply locked baseline and PROM-expanded predictions to prespecified test-set strata without refitting."""
    te = df[df.split.eq("test")].reset_index(drop=True)
    merged = base_result["preds"].merge(
        exp_result["preds"], on=["row_index_in_split", GROUP_COL, "y_true"],
        suffixes=("_baseline", "_expanded"), validate="one_to_one"
    )
    if len(merged) != len(te):
        raise RuntimeError(f"{cohort}: sensitivity prediction/test-row mismatch.")
    strata = build_sensitivity_strata(te)
    rows = []
    for stype, label, mask in strata:
        mask = np.asarray(mask, dtype=bool)
        if mask.sum() == 0:
            continue
        y = merged.loc[mask, "y_true"].astype(int).to_numpy()
        p0 = merged.loc[mask, "p_calibrated_baseline"].astype(float).to_numpy()
        p1 = merged.loc[mask, "p_calibrated_expanded"].astype(float).to_numpy()
        rows.append({
            "cohort": cohort,
            "stratum_type": stype,
            "stratum": label,
            "n": int(mask.sum()),
            "events": int(y.sum()),
            "event_rate": float(y.mean()) if len(y) else np.nan,
            "baseline_AP": safe_average_precision(y, p0),
            "expanded_AP": safe_average_precision(y, p1),
            "delta_AP": safe_average_precision(y, p1) - safe_average_precision(y, p0) if len(np.unique(y)) >= 2 else np.nan,
            "baseline_ROC_AUC": safe_roc_auc(y, p0),
            "expanded_ROC_AUC": safe_roc_auc(y, p1),
            "delta_ROC_AUC": safe_roc_auc(y, p1) - safe_roc_auc(y, p0) if len(np.unique(y)) >= 2 else np.nan,
        })
    out = pd.DataFrame(rows)
    os.makedirs(out_dir, exist_ok=True)
    out_csv = os.path.join(out_dir, f"sensitivity_delta_AP_by_stratum_{cohort}.csv")
    out.to_csv(out_csv, index=False)
    if not out.empty:
        plot_df = out.sort_values("delta_AP", ascending=True).copy()
        fig, ax = plt.subplots(figsize=(9.5, max(5.0, 0.34 * len(plot_df) + 1.8)))
        ax.barh(plot_df["stratum"], plot_df["delta_AP"])
        ax.axvline(0.0, linestyle="--", color="black", linewidth=1.0)
        ax.set_xlabel("Delta AP: PROM-expanded minus baseline")
        ax.set_title(f"{cohort}: locked-model sensitivity analysis by test-set stratum", fontweight="bold")
        ax.grid(axis="x", alpha=0.25)
        fig.tight_layout()
        fig.savefig(os.path.join(out_dir, f"sensitivity_delta_AP_by_stratum_{cohort}.png"), dpi=300, bbox_inches="tight")
        plt.close(fig)
    return out


# ============================================================
# 7) Main analysis per model/cohort
# ============================================================
def split_audit(df,target_col):
    return pd.DataFrame([{ "split":s,"rows":len(sub),"events":int(sub[target_col].sum()),"event_rate":float(sub[target_col].mean()) if len(sub) else np.nan,"unique_patients":int(sub[GROUP_COL].nunique())} for s,sub in df.groupby("split")])

def run_model(cohort,model_type,df,target_col,features,folds,seed,all_candidates,all_folds):
    key=f"{cohort}_{model_type}"; print("\n"+"="*100+f"\nRunning {key}\n"+"="*100)
    tr=df[df.split.eq("train")].reset_index(drop=True); ca=df[df.split.eq("calibration")].reset_index(drop=True); te=df[df.split.eq("test")].reset_index(drop=True)
    ytr=tr[target_col].astype(int).to_numpy(); yca=ca[target_col].astype(int).to_numpy(); yte=te[target_col].astype(int).to_numpy(); gtr=tr[GROUP_COL].to_numpy(); gte=te[GROUP_COL].to_numpy()
    cand,fold=tune(tr[features],ytr,gtr,features,folds,key,seed); all_candidates.append(cand); all_folds.append(fold)
    best=cand.iloc[0].to_dict(); params=sanitize_lgbm_params({k:best[k] for k in LGBM_SEARCH_SPACE}); locked=int(best["locked_n_estimators_from_cv"])
    final=fit_final(tr[features],ytr,ca[features],yca,te[features],features,params,locked,seed+777); p=final["p_test"]; thr=final["threshold"]
    train_ev=eval_preds(ytr,final["p_train"],thr,"Train_"); cal_ev=eval_preds(yca,final["p_cal"],thr,"Calibration_")
    ev=eval_preds(yte,p,thr,"Test_"); raw=eval_preds(yte,final["p_test_raw"],None,"RawTest_"); apci=patient_boot_ci(yte,p,gte,"AP",seed=seed+1); aucci=patient_boot_ci(yte,p,gte,"ROC_AUC",seed=seed+2); f1ci=patient_boot_ci(yte,p,gte,"F1",thr,seed=seed+3)
    mdir=os.path.join(OUTPUT_DIR,cohort,model_type); pdir=os.path.join(PLOT_DIR,cohort,model_type); sdir=os.path.join(SHAP_DIR,cohort,model_type); adir=os.path.join(ARTIFACT_DIR,cohort,model_type)
    for d in [mdir,pdir,sdir,adir]: os.makedirs(d,exist_ok=True)
    title=f"{COHORT_LABELS.get(cohort,cohort)}: {model_type.replace('_',' ')}"
    save_pr(yte,p,os.path.join(pdir,f"PR_curve_{key}.png"),title); save_roc(yte,p,os.path.join(pdir,f"ROC_curve_{key}.png"),title); save_calplot(yte,p,os.path.join(pdir,f"Calibration_{key}.png"),title)
    lc_png=os.path.join(pdir,f"Learning_curve_{key}.png"); lc_csv=os.path.join(mdir,f"learning_curve_{key}.csv"); learning_curve(tr[features],ytr,gtr,features,params,locked,lc_png,lc_csv,title+": learning curve",seed+4)
    # Derive any PROM threshold first, using training-only out-of-fold SHAP values.
    # Held-out test SHAP below is descriptive and is never used to choose thresholds.
    prom_features=[f for f in features if f not in BASELINE_FEATURES]
    shap_threshold_summary=None; test_threshold_summary=pd.DataFrame()
    if model_type=="expanded_PROM" and len(prom_features)==1:
        shap_threshold_summary=compute_oof_shap_threshold(tr, target_col, features, params, locked, folds, prom_features[0], key, seed+9000, sdir)
        if shap_threshold_summary is not None:
            test_threshold_summary=summarize_locked_prom_threshold_in_test(te, target_col, p, prom_features[0], float(shap_threshold_summary["locked_shap_threshold"]), key, sdir)

    shapdf,datadf,imp,mapnames=compute_shap(final["pre"],final["model"],te[features],yte,p,features,key); shapdf.to_csv(os.path.join(sdir,f"grouped_SHAP_values_{key}.csv"),index=False); datadf.to_csv(os.path.join(sdir,f"grouped_SHAP_feature_values_{key}.csv"),index=False); imp.to_csv(os.path.join(sdir,f"grouped_SHAP_importance_{key}.csv"),index=False); json.dump(json_native(mapnames),open(os.path.join(sdir,f"grouped_SHAP_mapping_{key}.json"),"w"),indent=2,sort_keys=True)
    bees=os.path.join(sdir,f"SHAP_beeswarm_GROUPED_{key}.png"); bar=os.path.join(sdir,f"SHAP_bar_GROUPED_{key}.png"); heat=os.path.join(sdir,f"SHAP_heatmap_GROUPED_{key}.png"); save_shap(shapdf,datadf,imp,bees,bar,title); save_shap_heatmap(shapdf,imp,heat,title)
    preds=pd.DataFrame({"cohort":cohort,"model_type":model_type,"row_index_in_split":np.arange(len(te)),GROUP_COL:gte,"y_true":yte,"p_raw":final["p_test_raw"],"p_calibrated":p,"selected_threshold":thr,"threshold_source":"calibration_split_max_F1_for_classification_metrics_only","y_pred":(p>=thr).astype(int)}); preds.to_csv(os.path.join(mdir,f"test_predictions_{key}.csv"),index=False)
    joblib.dump({"cohort":cohort,"model_type":model_type,"model_key":key,"features":features,"target_col":target_col,"group_col":GROUP_COL,"best_candidate_id":int(best["candidate_id"]),"best_params":params,"locked_n_estimators_from_cv":locked,"selected_threshold":thr,"preprocessor":final["pre"],"model":final["model"],"calibrator":final["calibrator"],"random_state":seed},os.path.join(adir,f"{key}_pipeline_artifact.joblib"))
    json.dump(json_native({"cohort":cohort,"model_type":model_type,"model_key":key,"features":features,"best_candidate_id":int(best["candidate_id"]),"best_params":params,"locked_n_estimators_from_cv":locked,"selected_threshold":thr,"threshold_source":"calibration_split_max_F1_for_classification_metrics_only","search_space":LGBM_SEARCH_SPACE,"n_random_combinations":N_RANDOM_COMBINATIONS,"n_cv_folds":N_CV_FOLDS,"calibration_method":CALIBRATION_METHOD,"threshold_strategy":THRESHOLD_STRATEGY,"test_set_used_for_model_development":False,"shap_threshold_summary":shap_threshold_summary}),open(os.path.join(adir,f"{key}_exact_settings.json"),"w"),indent=2,sort_keys=True)
    summ={"cohort":cohort,"model_type":model_type,"model_key":key,"n_features":len(features),"features":json.dumps(features),"best_candidate_id":int(best["candidate_id"]),"cv_AP_mean":float(best["cv_AP_mean"]),"cv_AP_SD":float(best["cv_AP_SD"]),"cv_ROC_AUC_mean":float(best["cv_ROC_AUC_mean"]),"cv_ROC_AUC_SD":float(best["cv_ROC_AUC_SD"]),"locked_n_estimators_from_cv":locked,"positive_weight_multiplier":float(params["positive_weight_multiplier"]),"positive_weight_used_full_training":actual_positive_weight(ytr,params["positive_weight_multiplier"]),"selected_threshold":thr,"Train_AP":train_ev["Train_AP"],"Train_ROC_AUC":train_ev["Train_ROC_AUC"],"Train_Brier_score":train_ev["Train_Brier_score"],"Train_ECE":train_ev["Train_ECE"],"Calibration_AP":cal_ev["Calibration_AP"],"Calibration_ROC_AUC":cal_ev["Calibration_ROC_AUC"],"Calibration_F1":cal_ev.get("Calibration_F1"),"Calibration_Brier_score":cal_ev["Calibration_Brier_score"],"Calibration_ECE":cal_ev["Calibration_ECE"],"Train_minus_CV_AP_gap":train_ev["Train_AP"]-float(best["cv_AP_mean"]),"Train_minus_CV_ROC_AUC_gap":train_ev["Train_ROC_AUC"]-float(best["cv_ROC_AUC_mean"]),"CV_minus_Test_AP_gap":float(best["cv_AP_mean"])-ev["Test_AP"],"CV_minus_Test_ROC_AUC_gap":float(best["cv_ROC_AUC_mean"])-ev["Test_ROC_AUC"],"Test_AP":ev["Test_AP"],"Test_AP_CI_lower":apci[0],"Test_AP_CI_upper":apci[1],"Test_ROC_AUC":ev["Test_ROC_AUC"],"Test_ROC_AUC_CI_lower":aucci[0],"Test_ROC_AUC_CI_upper":aucci[1],"Test_F1":ev.get("Test_F1"),"Test_F1_CI_lower":f1ci[0],"Test_F1_CI_upper":f1ci[1],"Test_Brier_score":ev["Test_Brier_score"],"Test_ECE":ev["Test_ECE"],"RawTest_AP":raw["RawTest_AP"],"RawTest_ROC_AUC":raw["RawTest_ROC_AUC"],"learning_curve_png":lc_png,"shap_beeswarm_png":bees,"shap_bar_png":bar,"shap_heatmap_png":heat,"locked_training_oof_shap_threshold":shap_threshold_summary.get("locked_shap_threshold") if shap_threshold_summary else np.nan,"shap_threshold_derivation_split":shap_threshold_summary.get("threshold_derivation_split") if shap_threshold_summary else "not_applicable","artifact_path":os.path.join(adir,f"{key}_pipeline_artifact.joblib")}
    for src in (train_ev,cal_ev,ev):
        for k,v in src.items():
            if k.startswith(("Train_S","Train_P","Train_N","Train_Top_5pct","Train_TP","Train_FP","Train_TN","Train_FN","Train_Predicted","Calibration_S","Calibration_P","Calibration_N","Calibration_Top_5pct","Calibration_TP","Calibration_FP","Calibration_TN","Calibration_FN","Calibration_Predicted","Test_S","Test_P","Test_N","Test_Top_5pct","Test_TP","Test_FP","Test_TN","Test_FN","Test_Predicted")): summ[k]=v
    print(f"{key} FINAL | AP={summ['Test_AP']:.5f} [{apci[0]:.5f},{apci[1]:.5f}] | AUC={summ['Test_ROC_AUC']:.5f} [{aucci[0]:.5f},{aucci[1]:.5f}]")
    return {"summary":summ,"preds":preds,"shap_importance":imp,"shap_values_test":shapdf,"shap_feature_values_test":datadf,"shap_threshold_summary":shap_threshold_summary,"test_threshold_summary":test_threshold_summary}

def run_cohort(cfg,seed):
    cohort=cfg["cohort_name"]; df=pd.read_csv(cfg["input_csv"],low_memory=False); df.columns=[str(c).strip() for c in df.columns]
    target0=resolve_target_column(df); prom=find_existing_column(df.columns.tolist(),cfg["prom_feature_candidates"],f"{cohort} PROM")
    FEATURE_LABELS[prom]=cfg["prom_display_name"]
    missing=[c for c in BASELINE_FEATURES+[prom,target0,GROUP_COL] if c not in df.columns]
    if missing: raise ValueError(f"{cohort} missing columns: {missing}")
    if set(BASELINE_FEATURES+[prom]) & (EXCLUDED_FEATURES-{target0}): raise ValueError("Leakage feature included.")
    target="__target_step1__"; before=len(df); df[target]=df[target0].map(to_binary_target); df=df[df[target].isin([0.0,1.0])].copy(); df[target]=df[target].astype(int)
    if df[GROUP_COL].isna().any(): raise ValueError(f"{cohort}: missing PersonKey")
    tr,ca,te=patient_split(df,target,seed); df["split"]=np.where(tr,"train",np.where(ca,"calibration","test"));
    if df.groupby(GROUP_COL).split.nunique().max()>1: raise RuntimeError("Patient leakage across splits.")
    print("\n"+"#"*100+f"\nCOHORT {cohort}\n"+"#"*100); print(f"Input: {cfg['input_csv']} | target: {target0} | PROM: {prom} | rows after target cleaning: {len(df)} dropped: {before-len(df)}"); print(split_audit(df,target).to_string(index=False))
    train=df[df.split.eq("train")].reset_index(drop=True); folds=cv_splits(train[target].astype(int).to_numpy(),train[GROUP_COL].to_numpy(),seed+10)
    cdir=os.path.join(OUTPUT_DIR,cohort); os.makedirs(cdir,exist_ok=True)
    cv_fold_assignment_table(train, folds, target, f"{cohort}_shared_training_CV_folds").to_csv(os.path.join(cdir,f"shared_group_aware_CV_fold_assignments_{cohort}.csv"),index=False)
    allc=[]; allf=[]; base=run_model(cohort,"baseline_only",df,target,BASELINE_FEATURES,folds,seed+100,allc,allf); exp=run_model(cohort,"expanded_PROM",df,target,BASELINE_FEATURES+[prom],folds,seed+200,allc,allf)
    m=base["preds"].merge(exp["preds"],on=["row_index_in_split",GROUP_COL,"y_true"],suffixes=("_baseline","_expanded"),validate="one_to_one")
    y=m.y_true.astype(int).to_numpy(); p0=m.p_calibrated_baseline.astype(float).to_numpy(); p1=m.p_calibrated_expanded.astype(float).to_numpy(); groups=m[GROUP_COL].to_numpy()
    dap=paired_delta_boot(y,p0,p1,groups,"AP",seed+1000); dauc=paired_delta_boot(y,p0,p1,groups,"ROC_AUC",seed+2000)
    comp=pd.DataFrame([{ "cohort":cohort,"prom_feature":prom,"baseline_Test_AP":base["summary"]["Test_AP"],"expanded_Test_AP":exp["summary"]["Test_AP"],"Delta_AP_expanded_minus_baseline":dap[0],"Delta_AP_CI_lower":dap[1],"Delta_AP_CI_upper":dap[2],"Delta_AP_bootstrap_p_value":dap[3],"baseline_Test_ROC_AUC":base["summary"]["Test_ROC_AUC"],"expanded_Test_ROC_AUC":exp["summary"]["Test_ROC_AUC"],"Delta_ROC_AUC_expanded_minus_baseline":dauc[0],"Delta_ROC_AUC_CI_lower":dauc[1],"Delta_ROC_AUC_CI_upper":dauc[2],"Delta_ROC_AUC_bootstrap_p_value":dauc[3]}])
    df[[GROUP_COL,"split"]].drop_duplicates().to_csv(os.path.join(cdir,f"split_assignment_{cohort}.csv"),index=False); split_audit(df,target).to_csv(os.path.join(cdir,f"split_audit_{cohort}.csv"),index=False); comp.to_csv(os.path.join(cdir,f"paired_PROM_comparison_{cohort}.csv"),index=False)
    sensitivity=run_sensitivity_analyses(cohort, df, target, base, exp, cdir)
    miss=pd.DataFrame([{"column":c,"n_missing_or_blank":int(df[c].isna().sum()),"percent_missing_or_blank":100*float(df[c].isna().sum())/len(df)} for c in BASELINE_FEATURES+[prom,target0,GROUP_COL] if c in df.columns])
    return {"metadata":{"cohort":cohort,"input_csv":cfg["input_csv"],"target_col_original":target0,"prom_col_used":prom,"baseline_feature_count":35,"expanded_feature_count":36},"split_audit":split_audit(df,target).assign(cohort=cohort),"missingness":miss.assign(cohort=cohort),"base":base,"exp":exp,"comparison":comp,"sensitivity":sensitivity,"candidates":pd.concat(allc,ignore_index=True),"folds":pd.concat(allf,ignore_index=True)}


# ============================================================
# 8b) Death-retained binary sensitivity analysis
# ============================================================
DEATH_RETAINED_TARGET_COL_CANDIDATES = [
    "final_reop_step1_death_retained",
    "final_reop_1yr_death_retained",
    "reop_1yr_death_retained",
    "reoperation_1yr_death_retained",
    "observed_reoperation_before_death_or_within_1yr",
    "reop_before_death_or_365d",
]

def derived_death_retained_csv_path(primary_csv):
    root, ext = os.path.splitext(primary_csv)
    return root + "_death_retained" + ext

def resolve_death_retained_source(cfg):
    explicit = cfg.get("death_retained_input_csv")
    if explicit and os.path.exists(explicit):
        return explicit, "explicit_death_retained_input_csv"
    derived = derived_death_retained_csv_path(cfg["input_csv"])
    if os.path.exists(derived):
        return derived, "derived_death_retained_input_csv"
    # Fallback: allow the primary cohort file to contain an already-defined death-retained target.
    try:
        cols = pd.read_csv(cfg["input_csv"], nrows=0).columns.astype(str).str.strip().tolist()
        _ = find_existing_column(cols, DEATH_RETAINED_TARGET_COL_CANDIDATES, "death-retained Step 1 target")
        return cfg["input_csv"], "primary_input_contains_death_retained_target"
    except Exception:
        return None, "death_retained_file_or_target_not_found"

def run_model_performance_only(cohort, analysis_label, model_type, df, target_col, features, folds, seed, out_dir):
    """Same LightGBM tuning/splitting/calibration protocol, but no SHAP/plots/grading."""
    key=f"{cohort}_{analysis_label}_{model_type}_LightGBM"
    tr=df[df.split.eq("train")].reset_index(drop=True)
    ca=df[df.split.eq("calibration")].reset_index(drop=True)
    te=df[df.split.eq("test")].reset_index(drop=True)
    ytr=tr[target_col].astype(int).to_numpy(); yca=ca[target_col].astype(int).to_numpy(); yte=te[target_col].astype(int).to_numpy(); gte=te[GROUP_COL].to_numpy()
    cand,fold=tune(tr[features].copy(),ytr,tr[GROUP_COL].to_numpy(),features,folds,key,seed)
    best=cand.iloc[0].to_dict(); params=sanitize_lgbm_params({k:best[k] for k in LGBM_SEARCH_SPACE if k in best}); locked=int(best["locked_n_estimators_from_cv"])
    final=fit_final(tr[features].copy(),ytr,ca[features].copy(),yca,te[features].copy(),features,params,locked,seed+3)
    p=final["p_test"]; thr=final["threshold"]
    ev=eval_preds(yte,p,thr,"Test_")
    raw=eval_preds(yte,final["p_test_raw"],None,"RawTest_")
    apci=patient_boot_ci(yte,p,gte,"AP",seed=seed+11); aucci=patient_boot_ci(yte,p,gte,"ROC_AUC",seed=seed+22)
    summ={
        "cohort":cohort,"analysis_label":analysis_label,"model_type":model_type,"model_key":key,
        "n_features":len(features),"best_candidate_id":int(best["candidate_id"]),
        "cv_AP_mean":float(best["cv_AP_mean"]),"cv_AP_SD":float(best["cv_AP_SD"]),
        "cv_ROC_AUC_mean":float(best["cv_ROC_AUC_mean"]),"cv_ROC_AUC_SD":float(best["cv_ROC_AUC_SD"]),
        "locked_n_estimators_from_cv":locked,"positive_weight_multiplier":float(params["positive_weight_multiplier"]),
        "selected_threshold":thr,"Test_AP":ev["Test_AP"],"Test_AP_CI_lower":apci[0],"Test_AP_CI_upper":apci[1],
        "Test_ROC_AUC":ev["Test_ROC_AUC"],"Test_ROC_AUC_CI_lower":aucci[0],"Test_ROC_AUC_CI_upper":aucci[1],
        "Test_Brier_score":ev["Test_Brier_score"],"Test_ECE":ev["Test_ECE"],"RawTest_AP":raw["RawTest_AP"],"RawTest_ROC_AUC":raw["RawTest_ROC_AUC"],
    }
    for k,v in ev.items():
        if k.startswith(("Test_Top_5pct","Test_P","Test_S","Test_N","Test_TP","Test_FP","Test_TN","Test_FN")):
            summ[k]=v
    preds=pd.DataFrame({"cohort":cohort,"analysis_label":analysis_label,"model_type":model_type,"row_index_in_split":np.arange(len(te)),GROUP_COL:gte,"y_true":yte,"p_raw":final["p_test_raw"],"p_calibrated":p,"selected_threshold":thr,"y_pred":(p>=thr).astype(int)})
    os.makedirs(out_dir, exist_ok=True)
    preds.to_csv(os.path.join(out_dir,f"test_predictions_{key}.csv"),index=False)
    cand.to_csv(os.path.join(out_dir,f"cv_candidates_{key}.csv"),index=False)
    fold.to_csv(os.path.join(out_dir,f"cv_fold_metrics_{key}.csv"),index=False)
    return {"summary":summ,"preds":preds,"candidates":cand,"folds":fold}

def run_death_retained_sensitivity_cohort(cfg,seed):
    cohort=cfg["cohort_name"]
    source, source_rule = resolve_death_retained_source(cfg)
    out_dir=os.path.join(OUTPUT_DIR,"death_retained_sensitivity",cohort)
    os.makedirs(out_dir, exist_ok=True)
    if source is None:
        skip=pd.DataFrame([{"cohort":cohort,"status":"skipped","reason":source_rule,"expected_file":derived_death_retained_csv_path(cfg["input_csv"]),"expected_target_candidates":";".join(DEATH_RETAINED_TARGET_COL_CANDIDATES)}])
        skip.to_csv(os.path.join(out_dir,f"death_retained_sensitivity_SKIPPED_{cohort}.csv"),index=False)
        print(f"{cohort} death-retained sensitivity skipped: {source_rule}")
        return {"cohort":cohort,"status":"skipped","skip":skip,"summary":pd.DataFrame(),"comparison":pd.DataFrame(),"candidates":pd.DataFrame(),"folds":pd.DataFrame()}
    df=pd.read_csv(source,low_memory=False); df.columns=[str(c).strip() for c in df.columns]
    target0=find_existing_column(df.columns.tolist(),DEATH_RETAINED_TARGET_COL_CANDIDATES,"death-retained Step 1 target")
    prom=find_existing_column(df.columns.tolist(),cfg["prom_feature_candidates"],f"{cohort} PROM")
    FEATURE_LABELS[prom]=cfg["prom_display_name"]
    missing=[c for c in BASELINE_FEATURES+[prom,target0,GROUP_COL] if c not in df.columns]
    if missing: raise ValueError(f"{cohort} death-retained sensitivity missing columns: {missing}")
    target="__target_step1_death_retained__"
    before=len(df); df[target]=df[target0].map(to_binary_target); df=df[df[target].isin([0.0,1.0])].copy(); df[target]=df[target].astype(int)
    if df[GROUP_COL].isna().any(): raise ValueError(f"{cohort} death-retained sensitivity: missing PersonKey")
    tr,ca,te=patient_split(df,target,seed); df["split"]=np.where(tr,"train",np.where(ca,"calibration","test"))
    if df.groupby(GROUP_COL).split.nunique().max()>1: raise RuntimeError(f"{cohort} death-retained sensitivity: patient leakage across splits")
    train=df[df.split.eq("train")].reset_index(drop=True); folds=cv_splits(train[target].astype(int).to_numpy(),train[GROUP_COL].to_numpy(),seed+10)
    cv_fold_assignment_table(train,folds,target,f"{cohort}_death_retained_shared_training_CV_folds").to_csv(os.path.join(out_dir,f"shared_group_aware_CV_fold_assignments_death_retained_{cohort}.csv"),index=False)
    print("\n"+"#"*100+f"\nDEATH-RETAINED SENSITIVITY: {cohort}\n"+"#"*100)
    print(f"Input: {source} | source rule: {source_rule} | target: {target0} | PROM: {prom} | rows after target cleaning: {len(df)} dropped: {before-len(df)}")
    print(split_audit(df,target).to_string(index=False))
    base=run_model_performance_only(cohort,"death_retained_binary_sensitivity","baseline_only",df,target,BASELINE_FEATURES,folds,seed+100,out_dir)
    exp=run_model_performance_only(cohort,"death_retained_binary_sensitivity","expanded_PROM",df,target,BASELINE_FEATURES+[prom],folds,seed+200,out_dir)
    m=base["preds"].merge(exp["preds"],on=["row_index_in_split",GROUP_COL,"y_true"],suffixes=("_baseline","_expanded"),validate="one_to_one")
    y=m.y_true.astype(int).to_numpy(); p0=m.p_calibrated_baseline.astype(float).to_numpy(); p1=m.p_calibrated_expanded.astype(float).to_numpy(); groups=m[GROUP_COL].to_numpy()
    dap=paired_delta_boot(y,p0,p1,groups,"AP",seed+1000); dauc=paired_delta_boot(y,p0,p1,groups,"ROC_AUC",seed+2000)
    comp=pd.DataFrame([{ "cohort":cohort,"analysis_label":"death_retained_binary_sensitivity","source_csv":source,"source_rule":source_rule,"target_col_original":target0,"prom_feature":prom,"baseline_Test_AP":base["summary"]["Test_AP"],"expanded_Test_AP":exp["summary"]["Test_AP"],"Delta_AP_expanded_minus_baseline":dap[0],"Delta_AP_CI_lower":dap[1],"Delta_AP_CI_upper":dap[2],"Delta_AP_bootstrap_p_value":dap[3],"baseline_Test_ROC_AUC":base["summary"]["Test_ROC_AUC"],"expanded_Test_ROC_AUC":exp["summary"]["Test_ROC_AUC"],"Delta_ROC_AUC_expanded_minus_baseline":dauc[0],"Delta_ROC_AUC_CI_lower":dauc[1],"Delta_ROC_AUC_CI_upper":dauc[2],"Delta_ROC_AUC_bootstrap_p_value":dauc[3]}])
    summary=pd.DataFrame([base["summary"], exp["summary"]])
    summary.to_csv(os.path.join(out_dir,f"death_retained_model_performance_{cohort}.csv"),index=False)
    comp.to_csv(os.path.join(out_dir,f"death_retained_paired_PROM_comparison_{cohort}.csv"),index=False)
    split_audit(df,target).to_csv(os.path.join(out_dir,f"split_audit_death_retained_{cohort}.csv"),index=False)
    return {"cohort":cohort,"status":"completed","summary":summary,"comparison":comp,"candidates":pd.concat([base["candidates"],exp["candidates"]],ignore_index=True),"folds":pd.concat([base["folds"],exp["folds"]],ignore_index=True)}

def methods_table():
    return pd.DataFrame([{"item":"Six-model design","rationale":"Within each PROM-specific cohort, baseline-only and PROM-expanded models are tuned separately under the same prespecified LightGBM protocol."},{"item":"Paired split/folds","rationale":"Baseline-only and expanded models within each cohort use identical patient-level train/calibration/test splits and identical group-aware CV folds."},{"item":"Tuning metric","rationale":"Average precision is the primary model-selection metric because reoperation is a rare-event outcome."},{"item":"Class imbalance","rationale":"Positive-class sample weighting is tuned through a multiplier applied to the natural negative-to-positive ratio."},{"item":"Calibration and threshold","rationale":"Isotonic calibration and threshold selection are performed only on the calibration split."},{"item":"Held-out test set","rationale":"The test set is isolated until the model-development pipeline is locked."},{"item":"Paired inference","rationale":"Delta AP and Delta ROC-AUC use paired patient-level bootstrap; p values are Holm-adjusted separately for AP and ROC-AUC across the three prespecified PROM comparisons."},{"item":"SHAP","rationale":"Grouped TreeSHAP feature attribution is calculated after CV-selected model locking. PROM thresholds are derived from training-only out-of-fold SHAP dependence; held-out test SHAP is descriptive only."}])

def style_excel(writer):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        for s in writer.sheets:
            ws=writer.sheets[s]; ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for cell in ws[1]: cell.font=Font(bold=True); cell.fill=PatternFill(start_color="D9EAF7",end_color="D9EAF7",fill_type="solid"); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            for col in ws.columns:
                ml=max([len(str(c.value)) for c in col if c.value is not None]+[12]); ws.column_dimensions[col[0].column_letter].width=min(max(ml+2,12),70)
    except Exception: pass

def main():
    t0=time.time(); results=[]
    for i,cfg in enumerate(COHORT_CONFIGS): results.append(run_cohort(cfg,RANDOM_STATE+i*10000))
    death_results=[]
    for i,cfg in enumerate(COHORT_CONFIGS): death_results.append(run_death_retained_sensitivity_cohort(cfg,RANDOM_STATE+i*10000+700000))
    summary=pd.DataFrame([r["base"]["summary"] for r in results]+[r["exp"]["summary"] for r in results]); comps=pd.concat([r["comparison"] for r in results],ignore_index=True)
    # Multiplicity control for the prespecified Step 1 PROM incremental-performance family.
    # AP and ROC-AUC are corrected separately across the three PROM comparisons.
    comps["Delta_AP_Holm_adjusted_p_value"] = holm_adjust_pvalues(comps["Delta_AP_bootstrap_p_value"].to_numpy(dtype=float))
    comps["Delta_ROC_AUC_Holm_adjusted_p_value"] = holm_adjust_pvalues(comps["Delta_ROC_AUC_bootstrap_p_value"].to_numpy(dtype=float))
    comps["Delta_AP_multiplicity_family"] = "three_prespecified_preoperative_PROM_comparisons"
    comps["Delta_ROC_AUC_multiplicity_family"] = "three_prespecified_preoperative_PROM_comparisons"
    cand=pd.concat([r["candidates"] for r in results],ignore_index=True); folds=pd.concat([r["folds"] for r in results],ignore_index=True); splits=pd.concat([r["split_audit"] for r in results],ignore_index=True); miss=pd.concat([r["missingness"] for r in results],ignore_index=True); shapimp=pd.concat([r["base"]["shap_importance"] for r in results]+[r["exp"]["shap_importance"] for r in results],ignore_index=True); meta=pd.DataFrame([r["metadata"] for r in results])
    shap_thresholds=pd.DataFrame([r["exp"]["shap_threshold_summary"] for r in results if r["exp"].get("shap_threshold_summary") is not None])
    locked_threshold_test_summaries=pd.concat([r["exp"]["test_threshold_summary"] for r in results if isinstance(r["exp"].get("test_threshold_summary"), pd.DataFrame) and not r["exp"]["test_threshold_summary"].empty], ignore_index=True) if any(isinstance(r["exp"].get("test_threshold_summary"), pd.DataFrame) and not r["exp"]["test_threshold_summary"].empty for r in results) else pd.DataFrame()
    sensitivity_all=pd.concat([r["sensitivity"] for r in results if isinstance(r.get("sensitivity"), pd.DataFrame) and not r["sensitivity"].empty],ignore_index=True) if any(isinstance(r.get("sensitivity"), pd.DataFrame) and not r["sensitivity"].empty for r in results) else pd.DataFrame()
    death_sensitivity_summary=pd.concat([r["summary"] for r in death_results if isinstance(r.get("summary"), pd.DataFrame) and not r["summary"].empty],ignore_index=True) if any(isinstance(r.get("summary"), pd.DataFrame) and not r["summary"].empty for r in death_results) else pd.DataFrame()
    death_sensitivity_comparisons=pd.concat([r["comparison"] for r in death_results if isinstance(r.get("comparison"), pd.DataFrame) and not r["comparison"].empty],ignore_index=True) if any(isinstance(r.get("comparison"), pd.DataFrame) and not r["comparison"].empty for r in death_results) else pd.DataFrame()
    if not death_sensitivity_comparisons.empty:
        death_sensitivity_comparisons["Delta_AP_Holm_adjusted_p_value"] = holm_adjust_pvalues(death_sensitivity_comparisons["Delta_AP_bootstrap_p_value"].to_numpy(dtype=float))
        death_sensitivity_comparisons["Delta_ROC_AUC_Holm_adjusted_p_value"] = holm_adjust_pvalues(death_sensitivity_comparisons["Delta_ROC_AUC_bootstrap_p_value"].to_numpy(dtype=float))
    death_sensitivity_skips=pd.concat([r["skip"] for r in death_results if isinstance(r.get("skip"), pd.DataFrame) and not r["skip"].empty],ignore_index=True) if any(isinstance(r.get("skip"), pd.DataFrame) and not r["skip"].empty for r in death_results) else pd.DataFrame()
    features=pd.DataFrame({"Feature":BASELINE_FEATURES,"Feature_group":"Baseline","Feature_type":["Continuous" if f in CONTINUOUS_BASELINE_FEATURES else "Binary" if f in BINARY_FEATURES else "Ordinal" if f in ORDINAL_FEATURES else "Nominal" if f in NOMINAL_FEATURES else "Unknown" for f in BASELINE_FEATURES]})
    cfgdf=pd.DataFrame([{"Parameter":k,"Value":v} for k,v in {"OUTPUT_DIR":OUTPUT_DIR,"RANDOM_STATE":RANDOM_STATE,"TEST_FRACTION":TEST_FRACTION,"CALIBRATION_FRACTION_OF_REMAINING":CALIBRATION_FRACTION_OF_REMAINING,"N_CV_FOLDS":N_CV_FOLDS,"N_RANDOM_COMBINATIONS":N_RANDOM_COMBINATIONS,"USE_EARLY_STOPPING_IN_CV":USE_EARLY_STOPPING_IN_CV,"EARLY_STOPPING_ROUNDS":EARLY_STOPPING_ROUNDS,"CALIBRATION_METHOD":CALIBRATION_METHOD,"SCALE_CONTINUOUS_FOR_FINAL_LIGHTGBM":SCALE_CONTINUOUS_FOR_FINAL_LIGHTGBM,"THRESHOLD_STRATEGY":THRESHOLD_STRATEGY,"N_BOOTSTRAPS":N_BOOTSTRAPS,"lightgbm_version":lgb.__version__,"shap_version":shap.__version__,"python_version":platform.python_version()}.items()])
    xlsx=os.path.join(OUTPUT_DIR,"Step1_PROM_LightGBM_summary.xlsx")
    with pd.ExcelWriter(xlsx,engine="openpyxl") as w:
        methods_table().to_excel(w,"methods_rationale",index=False); meta.to_excel(w,"cohort_metadata",index=False); summary.to_excel(w,"model_performance_all6",index=False); comps.to_excel(w,"paired_PROM_comparisons",index=False); cand.to_excel(w,"cv_candidates_all_models",index=False); folds.to_excel(w,"cv_fold_metrics_all",index=False); splits.to_excel(w,"split_audit",index=False); miss.to_excel(w,"missingness_audit",index=False); features.to_excel(w,"baseline_features",index=False); shapimp.to_excel(w,"SHAP_importance_all6",index=False); shap_thresholds.to_excel(w,"OOF_SHAP_PROM_thresholds",index=False); locked_threshold_test_summaries.to_excel(w,"locked_threshold_test_summary",index=False); sensitivity_all.to_excel(w,"sensitivity_strata",index=False); death_sensitivity_summary.to_excel(w,"death_retained_performance",index=False); death_sensitivity_comparisons.to_excel(w,"death_retained_comparison",index=False); death_sensitivity_skips.to_excel(w,"death_retained_skipped",index=False); cfgdf.to_excel(w,"run_config",index=False); style_excel(w)
    summary.to_csv(os.path.join(OUTPUT_DIR,"model_performance_all6.csv"),index=False); comps.to_csv(os.path.join(OUTPUT_DIR,"paired_PROM_comparisons.csv"),index=False); cand.to_csv(os.path.join(OUTPUT_DIR,"cv_candidates_all_models.csv"),index=False); folds.to_csv(os.path.join(OUTPUT_DIR,"cv_fold_metrics_all.csv"),index=False); shapimp.to_csv(os.path.join(OUTPUT_DIR,"SHAP_importance_all6.csv"),index=False); shap_thresholds.to_csv(os.path.join(OUTPUT_DIR,"OOF_SHAP_PROM_thresholds.csv"),index=False); locked_threshold_test_summaries.to_csv(os.path.join(OUTPUT_DIR,"locked_SHAP_threshold_test_summaries.csv"),index=False); sensitivity_all.to_csv(os.path.join(OUTPUT_DIR,"sensitivity_strata_all_cohorts.csv"),index=False); death_sensitivity_summary.to_csv(os.path.join(OUTPUT_DIR,"death_retained_model_performance_all_cohorts.csv"),index=False); death_sensitivity_comparisons.to_csv(os.path.join(OUTPUT_DIR,"death_retained_paired_PROM_comparisons_all_cohorts.csv"),index=False); death_sensitivity_skips.to_csv(os.path.join(OUTPUT_DIR,"death_retained_sensitivity_skipped.csv"),index=False); meta.to_csv(os.path.join(OUTPUT_DIR,"cohort_metadata.csv"),index=False)
    manifest={"input_files":[c["input_csv"] for c in COHORT_CONFIGS],"output_dir":OUTPUT_DIR,"group_col":GROUP_COL,"target_col_candidates":TARGET_COL_CANDIDATES,"baseline_features":BASELINE_FEATURES,"design":"Step 1 only: 3 PROM cohorts x 2 paired final LightGBM models = 6 models; optional death-retained binary sensitivity repeats the same paired LightGBM protocol when death-retained inputs are present","model_selection":"highest mean group-aware CV average precision inside training split","calibration":CALIBRATION_METHOD,"scale_continuous_for_final_lightgbm":SCALE_CONTINUOUS_FOR_FINAL_LIGHTGBM,"threshold_strategy":THRESHOLD_STRATEGY,"bootstrap":"patient-level stratified bootstrap for AP/AUC CIs and paired delta CIs; Holm-adjusted p values reported across the three prespecified PROM comparisons separately for AP and ROC-AUC","shap":"grouped TreeSHAP for all six selected LightGBM models on the full held-out test set; PROM thresholds derived from training-only out-of-fold SHAP values","runtime_minutes":float((time.time()-t0)/60),"summary_xlsx":xlsx}
    json.dump(json_native(manifest),open(os.path.join(OUTPUT_DIR,"run_manifest.json"),"w"),indent=2,sort_keys=True)
    zip_path=os.path.join(OUTPUT_DIR,"Step1_PROM_LightGBM_outputs.zip"); tmp=zip_path+".tmp"
    for p in [zip_path,tmp]:
        if os.path.exists(p): os.remove(p)
    open(os.path.join(OUTPUT_DIR,"DOWNLOAD_INSTRUCTIONS.txt"),"w").write(f"All Step 1 LightGBM outputs were generated successfully.\nZIP archive: {zip_path}\nIf Colab download stalls, download this ZIP manually from the Files panel.\n")
    print("\nCreating ZIP archive..."); zf=zipfile.ZipFile(tmp,"w",compression=zipfile.ZIP_DEFLATED,compresslevel=ZIP_COMPRESSION_LEVEL); n=0
    with zf as z:
        for root,_,files in os.walk(OUTPUT_DIR):
            for name in files:
                full=os.path.join(root,name)
                if full in {zip_path,tmp}: continue
                z.write(full,os.path.relpath(full,OUTPUT_DIR)); n+=1
    os.replace(tmp,zip_path); size=os.path.getsize(zip_path)/(1024**2)
    print("\n"+"="*100); print("STEP 1 PROM LightGBM analysis completed"); print("Summary Excel:",xlsx); print("Paired comparisons:",os.path.join(OUTPUT_DIR,"paired_PROM_comparisons.csv")); print("ZIP archive:",zip_path); print(f"ZIP size: {size:.2f} MB; files included: {n}"); print("="*100)
    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(f'<p><b>Step 1 LightGBM output archive is ready.</b></p><p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p><p>Path: <code>{zip_path}</code></p>'))
        except Exception as e: print("Download link display skipped:",e)
    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e: print("Automatic download skipped:",e)

if __name__ == "__main__": main()

# %% [markdown] Cell 5
# 
# #**Step1_Sensitivity Analysis Across Different Diagnoses**

# %% Cell 6
# -*- coding: utf-8 -*-
"""
Step 1 ODI diagnosis-stratified sensitivity analysis
====================================================

This script evaluates the locked Step 1 LightGBM ODI models across prespecified
lumbar diagnosis subgroups. It uses saved held-out test-set predictions and
saved SHAP tables from the locked primary analysis. It does not refit, retune,
recalibrate, or re-optimize thresholds.

Outputs
-------
1. Diagnosis-stratified ΔAP for the ODI PROM-expanded model relative to the
   ODI baseline-only model.
2. Diagnosis-stratified SHAP beeswarm plots for the locked ODI PROM-expanded
   model, omitting the subgroup-defining diagnosis feature from the corresponding
   visualization only. The locked prediction model is not modified.

"""

# ============================================================
# Imports
# ============================================================

import os
import re
import sys
import json
import math
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from matplotlib import pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from sklearn.metrics import average_precision_score

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_ODI_Diagnosis_Sensitivity_DeltaAP_SHAP")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
COHORT = "ODI"
N_BOOTSTRAPS = 2000
RANDOM_STATE = 20260524
DISPLAY_FEATURE_COUNT = 25
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

SHAP_BEESWARM_BLUE = "#008BFB"
SHAP_BEESWARM_PINK = "#FF0051"
SHAP_COLOR_CMAP = LinearSegmentedColormap.from_list(
    "shap_blue_pink", [SHAP_BEESWARM_BLUE, SHAP_BEESWARM_PINK]
)

ODI_INPUT_CANDIDATES = ["PROM_ODI_Cohort.csv", "PROM_ODI_90.csv"]

DIAGNOSIS_SUBGROUPS = [
    ("finaldx_degenerative", "Degenerative diagnosis"),
    ("finaldx_radicular", "Radiculopathy diagnosis"),
    ("finaldx_stenosis", "Spinal stenosis diagnosis"),
    ("finaldx_deformity_instability", "Deformity or instability diagnosis"),
    ("finaldx_other_diagnosis", "Other lumbar diagnosis"),
]

PREDICTION_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
]

PREDICTION_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
]

SHAP_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final"),
]

SHAP_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final.zip"),
]

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

BINARY_MAPS = {
    "finaldx_degenerative": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "finaldx_radicular": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "finaldx_stenosis": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "finaldx_deformity_instability": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "finaldx_other_diagnosis": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
}

# ============================================================
# Basic helpers
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def parse_binary_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("/", "_")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(average_precision_score(y, p))

# ============================================================
# Source discovery
# ============================================================

def _archive_names(zip_path: str) -> List[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            return [n.replace(chr(92), "/") for n in zf.namelist()]
    except Exception:
        return []


def _folder_names(path: str) -> List[str]:
    names = []
    if not os.path.isdir(path):
        return names
    for root, _, files in os.walk(path):
        for fn in files:
            names.append(os.path.join(root, fn).replace(path + os.sep, "").replace(chr(92), "/"))
    return names


def prediction_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    required = [
        "test_predictions_ODI_baseline_only",
        "test_predictions_ODI_expanded_PROM",
    ]
    return all(any(fn.startswith(token) and fn.endswith(".csv") for fn in basenames) for token in required)


def shap_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    return (
        any("shap/ODI/expanded_PROM/" in n for n in names)
        and any(fn.startswith("grouped_SHAP_values_") and fn.endswith(".csv") for fn in basenames)
        and any(fn.startswith("grouped_SHAP_feature_values_") and fn.endswith(".csv") for fn in basenames)
    )


def ensure_prediction_source_dir() -> str:
    for folder in PREDICTION_SOURCE_FOLDER_CANDIDATES:
        if prediction_tables_present(_folder_names(folder)):
            return folder

    candidate_archives = list(PREDICTION_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([os.path.join(root, fn) for fn in os.listdir(root) if fn.endswith(".zip")])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive) and prediction_tables_present(_archive_names(archive)):
            archive_path = archive
            break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 1 LightGBM output folder/archive with ODI held-out test prediction tables was found. "
            "Expected files beginning with test_predictions_ODI_baseline_only and test_predictions_ODI_expanded_PROM."
        )

    extract_root = os.path.join(BASE_DIR, "_step1_odi_sensitivity_prediction_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting prediction source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def ensure_shap_source_dir() -> str:
    for folder in SHAP_SOURCE_FOLDER_CANDIDATES:
        if shap_tables_present(_folder_names(folder)):
            return folder

    candidate_archives = list(SHAP_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([os.path.join(root, fn) for fn in os.listdir(root) if fn.endswith(".zip")])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive) and shap_tables_present(_archive_names(archive)):
            archive_path = archive
            break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 1 LightGBM source folder/archive with ODI grouped SHAP tables was found. "
            "Expected grouped_SHAP_values and grouped_SHAP_feature_values under shap/ODI/expanded_PROM."
        )

    extract_root = os.path.join(BASE_DIR, "_step1_odi_sensitivity_shap_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting SHAP source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def find_prediction_csv(source_dir: str, model_type: str) -> str:
    token = f"test_predictions_ODI_{model_type}"
    matches = []
    for root, _, files in os.walk(source_dir):
        for fn in files:
            if fn.startswith(token) and fn.endswith(".csv"):
                matches.append(os.path.join(root, fn))
    if not matches:
        raise FileNotFoundError(f"Missing ODI prediction table for {model_type}. Expected file beginning with {token}.")
    final_matches = [p for p in matches if p.endswith("_Final.csv")]
    return sorted(final_matches or matches)[0]


def find_odi_shap_tables(source_dir: str) -> Tuple[str, str, Optional[str]]:
    value_matches = []
    data_matches = []
    importance_matches = []
    for root, _, files in os.walk(source_dir):
        normalized_root = root.replace(chr(92), "/")
        if "shap/ODI/expanded_PROM" not in normalized_root:
            continue
        for fn in files:
            full = os.path.join(root, fn)
            if fn.startswith("grouped_SHAP_values_") and fn.endswith(".csv"):
                value_matches.append(full)
            elif fn.startswith("grouped_SHAP_feature_values_") and fn.endswith(".csv"):
                data_matches.append(full)
            elif fn.startswith("grouped_SHAP_importance_") and fn.endswith(".csv"):
                importance_matches.append(full)

    if not value_matches or not data_matches:
        raise FileNotFoundError("Could not find ODI grouped SHAP value and feature-value tables.")

    def _select(matches: List[str]) -> str:
        final = [p for p in matches if p.endswith("_Final.csv")]
        top01 = [p for p in (final or matches) if "top01_" in p]
        return sorted(top01 or final or matches)[0]

    return _select(value_matches), _select(data_matches), _select(importance_matches) if importance_matches else None

# ============================================================
# Data loading
# ============================================================

def find_odi_input_csv() -> str:
    for root in [BASE_DIR, FALLBACK_DIR]:
        for fn in ODI_INPUT_CANDIDATES:
            p = os.path.join(root, fn)
            if os.path.exists(p):
                if root == FALLBACK_DIR:
                    dst = os.path.join(BASE_DIR, fn)
                    if not os.path.exists(dst):
                        shutil.copy2(p, dst)
                    return dst
                return p
    raise FileNotFoundError(f"Could not find ODI cohort input file. Tried: {ODI_INPUT_CANDIDATES}")


def load_diagnosis_map() -> pd.DataFrame:
    path = find_odi_input_csv()
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns:
        raise ValueError(f"ODI cohort file does not contain {GROUP_COL}.")
    missing = [c for c, _ in DIAGNOSIS_SUBGROUPS if c not in df.columns]
    if missing:
        raise ValueError(f"ODI cohort file is missing diagnosis columns: {missing}")

    out = df[[GROUP_COL] + [c for c, _ in DIAGNOSIS_SUBGROUPS]].copy()
    for c, _ in DIAGNOSIS_SUBGROUPS:
        out[c] = out[c].map(lambda z: parse_binary_value(z, c)).fillna(0.0).astype(int)
    return out.groupby(GROUP_COL, dropna=False)[[c for c, _ in DIAGNOSIS_SUBGROUPS]].max().reset_index()


def load_paired_predictions(pred_source_dir: str) -> pd.DataFrame:
    base_path = find_prediction_csv(pred_source_dir, "baseline_only")
    exp_path = find_prediction_csv(pred_source_dir, "expanded_PROM")
    base = pd.read_csv(base_path)
    exp = pd.read_csv(exp_path)
    base.columns = [str(c).strip() for c in base.columns]
    exp.columns = [str(c).strip() for c in exp.columns]

    required = [GROUP_COL, "y_true", "p_calibrated"]
    for label, tbl in [("baseline-only", base), ("PROM-expanded", exp)]:
        missing = [c for c in required if c not in tbl.columns]
        if missing:
            raise ValueError(f"ODI {label} prediction table is missing columns: {missing}")

    if "row_index_in_split" not in base.columns:
        base = base.reset_index().rename(columns={"index": "row_index_in_split"})
    if "row_index_in_split" not in exp.columns:
        exp = exp.reset_index().rename(columns={"index": "row_index_in_split"})

    key_cols = ["row_index_in_split", GROUP_COL, "y_true"]
    base_small = base[key_cols + ["p_calibrated"]].rename(columns={"p_calibrated": "p_baseline"})
    exp_small = exp[key_cols + ["p_calibrated"]].rename(columns={"p_calibrated": "p_expanded"})
    paired = base_small.merge(exp_small, on=key_cols, how="inner", validate="one_to_one")

    if len(paired) != len(base_small) or len(paired) != len(exp_small):
        raise RuntimeError("ODI baseline-only and PROM-expanded prediction tables could not be paired one-to-one.")

    paired["y_true"] = paired["y_true"].astype(int)
    paired["p_baseline"] = pd.to_numeric(paired["p_baseline"], errors="coerce")
    paired["p_expanded"] = pd.to_numeric(paired["p_expanded"], errors="coerce")
    return paired.dropna(subset=["p_baseline", "p_expanded"]).copy()

# ============================================================
# ΔAP analysis
# ============================================================

def bootstrap_delta_ap(y: np.ndarray, p_base: np.ndarray, p_exp: np.ndarray, groups: np.ndarray, seed: int) -> Tuple[float, float, float, float]:
    y = np.asarray(y).astype(int)
    p_base = np.asarray(p_base).astype(float)
    p_exp = np.asarray(p_exp).astype(float)
    groups = np.asarray(groups)

    base_ap = safe_average_precision(y, p_base)
    exp_ap = safe_average_precision(y, p_exp)
    obs_delta = exp_ap - base_ap if np.isfinite(base_ap) and np.isfinite(exp_ap) else np.nan
    if not np.isfinite(obs_delta) or len(np.unique(y)) < 2:
        return obs_delta, np.nan, np.nan, np.nan

    group_df = pd.DataFrame({"group": groups, "y": y}).groupby("group", dropna=False)["y"].max().reset_index()
    pos_groups = group_df.loc[group_df["y"].eq(1), "group"].to_numpy()
    neg_groups = group_df.loc[group_df["y"].eq(0), "group"].to_numpy()
    if len(pos_groups) == 0 or len(neg_groups) == 0:
        return obs_delta, np.nan, np.nan, np.nan

    index_by_group = {g: np.where(groups == g)[0] for g in np.unique(groups)}
    rng = np.random.default_rng(seed)
    deltas = []
    for _ in range(N_BOOTSTRAPS):
        sampled_groups = np.concatenate([
            rng.choice(pos_groups, size=len(pos_groups), replace=True),
            rng.choice(neg_groups, size=len(neg_groups), replace=True),
        ])
        idx = np.concatenate([index_by_group[g] for g in sampled_groups])
        yy = y[idx]
        if len(np.unique(yy)) < 2:
            continue
        b = safe_average_precision(yy, p_base[idx])
        e = safe_average_precision(yy, p_exp[idx])
        if np.isfinite(b) and np.isfinite(e):
            deltas.append(e - b)

    if not deltas:
        return float(obs_delta), np.nan, np.nan, np.nan
    deltas = np.asarray(deltas, dtype=float)
    lo, hi = np.percentile(deltas, [2.5, 97.5])
    p_val = 2.0 * min(np.mean(deltas <= 0), np.mean(deltas >= 0))
    return float(obs_delta), float(lo), float(hi), float(min(max(p_val, 0.0), 1.0))


def run_delta_ap_analysis(pred_source_dir: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    preds = load_paired_predictions(pred_source_dir)
    dx_map = load_diagnosis_map()
    work = preds.merge(dx_map, on=GROUP_COL, how="left", validate="many_to_one")
    for c, _ in DIAGNOSIS_SUBGROUPS:
        work[c] = work[c].fillna(0).astype(int)

    rows = []
    audit_rows = [
        {"item": "test_prediction_rows", "value": int(len(preds)), "note": "ODI held-out test-set predictions"},
        {"item": "test_events", "value": int(preds["y_true"].sum()), "note": ""},
        {"item": "test_prevalence", "value": float(preds["y_true"].mean()), "note": ""},
    ]

    for i, (dx_col, dx_label) in enumerate(DIAGNOSIS_SUBGROUPS, start=1):
        sub = work[work[dx_col].eq(1)].copy()
        y = sub["y_true"].to_numpy(dtype=int)
        p0 = sub["p_baseline"].to_numpy(dtype=float)
        p1 = sub["p_expanded"].to_numpy(dtype=float)
        groups = sub[GROUP_COL].to_numpy()
        base_ap = safe_average_precision(y, p0)
        exp_ap = safe_average_precision(y, p1)
        delta, lo, hi, p_val = bootstrap_delta_ap(y, p0, p1, groups, seed=RANDOM_STATE + i * 100)
        rows.append({
            "cohort": "ODI",
            "subgroup_variable": dx_col,
            "subgroup_label": dx_label,
            "n": int(len(sub)),
            "events": int(np.sum(y)) if len(sub) else 0,
            "prevalence": float(np.mean(y)) if len(sub) else np.nan,
            "baseline_AP": base_ap,
            "expanded_AP": exp_ap,
            "delta_AP": delta,
            "delta_AP_CI_lower": lo,
            "delta_AP_CI_upper": hi,
            "delta_AP_p_value_two_sided": p_val,
        })
        audit_rows.append({"item": f"{dx_col}_n", "value": int(len(sub)), "note": dx_label})
        audit_rows.append({"item": f"{dx_col}_events", "value": int(np.sum(y)) if len(sub) else 0, "note": dx_label})

    work.to_csv(os.path.join(TABLE_DIR, "ODI_paired_test_predictions_with_diagnosis.csv"), index=False)
    return pd.DataFrame(rows), pd.DataFrame(audit_rows)

# ============================================================
# SHAP subgroup beeswarm analysis
# ============================================================

def load_odi_shap_tables(shap_source_dir: str) -> Tuple[pd.DataFrame, pd.DataFrame, Optional[pd.DataFrame]]:
    shap_path, data_path, imp_path = find_odi_shap_tables(shap_source_dir)
    shap_df = pd.read_csv(shap_path)
    data_df = pd.read_csv(data_path)
    imp_df = pd.read_csv(imp_path) if imp_path else None
    shap_df.columns = [str(c).strip() for c in shap_df.columns]
    data_df.columns = [str(c).strip() for c in data_df.columns]
    if imp_df is not None:
        imp_df.columns = [str(c).strip() for c in imp_df.columns]
    if len(shap_df) != len(data_df):
        raise RuntimeError("ODI grouped SHAP value and feature-value tables have different row counts.")
    return shap_df, data_df, imp_df


def diagnosis_mask_for_shap(data_df: pd.DataFrame, shap_df: pd.DataFrame, dx_col: str, dx_label: str) -> pd.Series:
    candidates = [dx_label, dx_col, dx_col.replace("_", " ")]
    for c in candidates:
        if c in data_df.columns:
            return pd.to_numeric(data_df[c], errors="coerce").fillna(0).astype(float).ge(0.5)
        if c in shap_df.columns:
            return pd.to_numeric(shap_df[c], errors="coerce").fillna(0).astype(float).ge(0.5)
    raise ValueError(
        f"Could not find diagnosis feature for SHAP subgroup filtering: {dx_label}. "
        "The ODI grouped SHAP feature-value table must include diagnosis columns."
    )




def subgroup_defining_feature_candidates(subgroup_col: str, subgroup_label: str) -> List[str]:
    """Return possible grouped-SHAP column names for the subgroup-defining feature."""
    candidates = [
        subgroup_label,
        subgroup_col,
        subgroup_col.replace("_", " "),
        subgroup_col.replace("_", " ").title(),
    ]
    # Preserve order while removing duplicates.
    return list(dict.fromkeys(candidates))


def identify_omitted_plot_features(shap_sub: pd.DataFrame, data_sub: pd.DataFrame, subgroup_col: str, subgroup_label: str) -> List[str]:
    """Identify the subgroup-defining feature columns to omit from SHAP display.

    This affects visualization only. Predictions, ΔAP, and all subgroup membership
    definitions still use the unchanged locked model and original subgroup columns.
    """
    candidates = subgroup_defining_feature_candidates(subgroup_col, subgroup_label)
    return [c for c in candidates if c in shap_sub.columns and c in data_sub.columns]

def subgroup_feature_order(
    shap_sub: pd.DataFrame,
    data_sub: pd.DataFrame,
    exclude_features: Optional[List[str]] = None,
) -> List[str]:
    """Rank SHAP features for plotting after excluding tautological subgroup features."""
    exclude_set = set(exclude_features or [])
    feature_cols = [
        c for c in shap_sub.columns
        if not c.startswith("__") and c in data_sub.columns and c not in exclude_set
    ]
    if not feature_cols:
        raise RuntimeError("No shared SHAP/data feature columns were available for beeswarm plotting after exclusions.")
    imp = pd.Series({
        c: float(np.nanmean(np.abs(pd.to_numeric(shap_sub[c], errors="coerce"))))
        for c in feature_cols
    })
    return imp.sort_values(ascending=False).head(DISPLAY_FEATURE_COUNT).index.tolist()


def save_beeswarm(shap_sub: pd.DataFrame, data_sub: pd.DataFrame, features: List[str], path: str, title: str) -> None:
    plt.figure(figsize=(11, max(8.8, 0.32 * len(features) + 2)))
    shap_values = shap_sub[features].apply(pd.to_numeric, errors="coerce").fillna(0.0).values
    feature_values = data_sub[features].apply(pd.to_numeric, errors="coerce")
    try:
        shap.summary_plot(
            shap_values,
            features=feature_values,
            feature_names=features,
            max_display=len(features),
            show=False,
            plot_size=None,
            color=SHAP_COLOR_CMAP,
        )
    except TypeError:
        shap.summary_plot(
            shap_values,
            features=feature_values,
            feature_names=features,
            max_display=len(features),
            show=False,
            plot_size=None,
        )
    plt.title(title, fontsize=15, fontweight="bold")
    plt.tight_layout()
    plt.savefig(path, dpi=300, bbox_inches="tight")
    plt.close()


def run_shap_beeswarm_analysis(shap_source_dir: str) -> pd.DataFrame:
    shap_df, data_df, _ = load_odi_shap_tables(shap_source_dir)
    rows = []
    for dx_col, dx_label in DIAGNOSIS_SUBGROUPS:
        mask = diagnosis_mask_for_shap(data_df, shap_df, dx_col, dx_label)
        shap_sub = shap_df.loc[mask].reset_index(drop=True)
        data_sub = data_df.loc[mask].reset_index(drop=True)
        if len(shap_sub) == 0:
            rows.append({
                "cohort": "ODI",
                "subgroup_variable": dx_col,
                "subgroup_label": dx_label,
                "n_shap_rows": 0,
                "plot_path": "",
                "status": "no_rows",
            })
            continue
        omitted_features = identify_omitted_plot_features(shap_sub, data_sub, dx_col, dx_label)
        features = subgroup_feature_order(shap_sub, data_sub, exclude_features=omitted_features)
        plot_path = os.path.join(PLOT_DIR, f"ODI_SHAP_beeswarm_{safe_filename(dx_label)}.png")
        save_beeswarm(
            shap_sub=shap_sub,
            data_sub=data_sub,
            features=features,
            path=plot_path,
            title=f"ODI PROM-expanded model: {dx_label}",
        )
        rows.append({
            "cohort": "ODI",
            "subgroup_variable": dx_col,
            "subgroup_label": dx_label,
            "n_shap_rows": int(len(shap_sub)),
            "plot_path": plot_path,
            "status": "created",
            "features_displayed": "; ".join(features),
            "subgroup_defining_features_omitted_from_plot": "; ".join(omitted_features),
            "visualization_note": "Subgroup-defining feature omitted from this SHAP beeswarm only; locked model and ΔAP analysis unchanged.",
        })
    return pd.DataFrame(rows)

# ============================================================
# Output
# ============================================================

def make_zip(src_dir: str, zip_path: str) -> None:
    if os.path.exists(zip_path):
        os.remove(zip_path)
    tmp_zip = zip_path + ".tmp"
    if os.path.exists(tmp_zip):
        os.remove(tmp_zip)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(src_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(src_dir))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)


def main() -> None:
    print("=" * 90)
    print("Step 1 ODI diagnosis-stratified sensitivity analysis")
    print("=" * 90)

    pred_source_dir = ensure_prediction_source_dir()
    shap_source_dir = ensure_shap_source_dir()
    print("Using prediction source directory:", pred_source_dir)
    print("Using SHAP source directory:", shap_source_dir)

    delta_ap, audit = run_delta_ap_analysis(pred_source_dir)
    shap_manifest = run_shap_beeswarm_analysis(shap_source_dir)

    delta_csv = os.path.join(TABLE_DIR, "ODI_diagnosis_delta_AP.csv")
    audit_csv = os.path.join(AUDIT_DIR, "ODI_diagnosis_sensitivity_audit.csv")
    shap_manifest_csv = os.path.join(TABLE_DIR, "ODI_diagnosis_SHAP_beeswarm_manifest.csv")
    delta_ap.to_csv(delta_csv, index=False)
    audit.to_csv(audit_csv, index=False)
    shap_manifest.to_csv(shap_manifest_csv, index=False)

    summary_xlsx = os.path.join(OUTPUT_DIR, "Step1_ODI_Diagnosis_Sensitivity_DeltaAP_SHAP_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        delta_ap.to_excel(writer, sheet_name="delta_AP", index=False)
        shap_manifest.to_excel(writer, sheet_name="SHAP_beeswarm", index=False)
        audit.to_excel(writer, sheet_name="audit", index=False)

    manifest = {
        "analysis": "Step 1 ODI diagnosis-stratified sensitivity analysis",
        "cohort": "ODI",
        "model": "locked LightGBM",
        "prediction_source_dir": pred_source_dir,
        "shap_source_dir": shap_source_dir,
        "refit": False,
        "retune": False,
        "recalibrate": False,
        "threshold_reoptimization": False,
        "reported_metrics": ["delta_AP"],
        "reported_plots": ["SHAP beeswarm"],
        "shap_visualization_note": "The subgroup-defining feature is omitted from the corresponding SHAP beeswarm plot only; the locked model, predictions, and ΔAP estimates are unchanged.",
        "subgroups": [{"column": c, "label": l, "definition": "diagnosis present"} for c, l in DIAGNOSIS_SUBGROUPS],
        "bootstrap_iterations_for_delta_AP_CI": N_BOOTSTRAPS,
    }
    with open(os.path.join(OUTPUT_DIR, "run_manifest.json"), "w") as f:
        json.dump(manifest, f, indent=2)

    zip_path = os.path.join(BASE_DIR, "Step1_ODI_Diagnosis_Sensitivity_DeltaAP_SHAP_omit_stratifying_feature.zip")
    make_zip(OUTPUT_DIR, zip_path)

    print("\nDONE")
    print("Delta AP table:", delta_csv)
    print("SHAP beeswarm manifest:", shap_manifest_csv)
    print("Summary Excel:", summary_xlsx)
    print("ZIP:", zip_path)

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(f'<p><b>ODI diagnosis sensitivity output archive is ready. SHAP plots omit subgroup-defining features for display only.</b></p><p><a href="/files{zip_path}" download>Download ZIP archive</a></p><p>Path: <code>{zip_path}</code></p>'))
        except Exception:
            pass


if __name__ == "__main__":
    main()

# %% [markdown] Cell 7
# #**Step1_Sensitivity Analysis Across Different Procedures**

# %% Cell 8
# -*- coding: utf-8 -*-
"""
Step 1 ODI procedure-stratified sensitivity analysis
====================================================

This script evaluates the locked Step 1 LightGBM ODI models across prespecified
lumbar procedural subgroups. It uses saved held-out test-set predictions and
saved SHAP tables from the locked primary analysis. It does not refit, retune,
recalibrate, or re-optimize thresholds.

Outputs
-------
1. Procedure-stratified ΔAP for the ODI PROM-expanded model relative to the
   ODI baseline-only model.
2. Procedure-stratified SHAP beeswarm plots for the locked ODI PROM-expanded
   model, omitting the subgroup-defining procedure feature from the corresponding
   visualization only. The locked prediction model is not modified.

"""

# ============================================================
# Imports
# ============================================================

import os
import re
import sys
import json
import math
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from matplotlib import pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from sklearn.metrics import average_precision_score

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_ODI_Procedure_Sensitivity_DeltaAP_SHAP")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
COHORT = "ODI"
N_BOOTSTRAPS = 2000
RANDOM_STATE = 20260524
DISPLAY_FEATURE_COUNT = 25
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

SHAP_BEESWARM_BLUE = "#008BFB"
SHAP_BEESWARM_PINK = "#FF0051"
SHAP_COLOR_CMAP = LinearSegmentedColormap.from_list(
    "shap_blue_pink", [SHAP_BEESWARM_BLUE, SHAP_BEESWARM_PINK]
)

ODI_INPUT_CANDIDATES = ["PROM_ODI_Cohort.csv", "PROM_ODI_90.csv"]

PROCEDURE_SUBGROUPS = [
    ("alif_llif", "Anterior/lateral lumbar interbody fusion"),
    ("corpectomy", "Corpectomy"),
    ("discectomy", "Discectomy"),
    ("foraminotomy", "Foraminotomy"),
    ("instrumentation", "Instrumentation"),
    ("laminectomy_posterior_decompression", "Posterior decompression or laminectomy"),
    ("pelvic_fixation", "Pelvic fixation"),
    ("plf", "Posterolateral fusion"),
    ("tlif_plif", "Posterior/transforaminal lumbar interbody fusion"),
    ("other_lumbar_procedure", "Other lumbar procedure"),
]
PREDICTION_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
]

PREDICTION_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
]

SHAP_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final"),
]

SHAP_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final.zip"),
]

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

BINARY_MAPS = {
    "alif_llif": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "corpectomy": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "discectomy": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "foraminotomy": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "instrumentation": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "laminectomy_posterior_decompression": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "pelvic_fixation": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "plf": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "tlif_plif": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
    "other_lumbar_procedure": {"yes": 1, "no": 0, "true": 1, "false": 0, "1": 1, "0": 0},
}

# ============================================================
# Basic helpers
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def parse_binary_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("/", "_")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(average_precision_score(y, p))

# ============================================================
# Source discovery
# ============================================================

def _archive_names(zip_path: str) -> List[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            return [n.replace(chr(92), "/") for n in zf.namelist()]
    except Exception:
        return []


def _folder_names(path: str) -> List[str]:
    names = []
    if not os.path.isdir(path):
        return names
    for root, _, files in os.walk(path):
        for fn in files:
            names.append(os.path.join(root, fn).replace(path + os.sep, "").replace(chr(92), "/"))
    return names


def prediction_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    required = [
        "test_predictions_ODI_baseline_only",
        "test_predictions_ODI_expanded_PROM",
    ]
    return all(any(fn.startswith(token) and fn.endswith(".csv") for fn in basenames) for token in required)


def shap_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    return (
        any("shap/ODI/expanded_PROM/" in n for n in names)
        and any(fn.startswith("grouped_SHAP_values_") and fn.endswith(".csv") for fn in basenames)
        and any(fn.startswith("grouped_SHAP_feature_values_") and fn.endswith(".csv") for fn in basenames)
    )


def ensure_prediction_source_dir() -> str:
    for folder in PREDICTION_SOURCE_FOLDER_CANDIDATES:
        if prediction_tables_present(_folder_names(folder)):
            return folder

    candidate_archives = list(PREDICTION_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([os.path.join(root, fn) for fn in os.listdir(root) if fn.endswith(".zip")])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive) and prediction_tables_present(_archive_names(archive)):
            archive_path = archive
            break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 1 LightGBM output folder/archive with ODI held-out test prediction tables was found. "
            "Expected files beginning with test_predictions_ODI_baseline_only and test_predictions_ODI_expanded_PROM."
        )

    extract_root = os.path.join(BASE_DIR, "_step1_odi_sensitivity_prediction_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting prediction source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def ensure_shap_source_dir() -> str:
    for folder in SHAP_SOURCE_FOLDER_CANDIDATES:
        if shap_tables_present(_folder_names(folder)):
            return folder

    candidate_archives = list(SHAP_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([os.path.join(root, fn) for fn in os.listdir(root) if fn.endswith(".zip")])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive) and shap_tables_present(_archive_names(archive)):
            archive_path = archive
            break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 1 LightGBM source folder/archive with ODI grouped SHAP tables was found. "
            "Expected grouped_SHAP_values and grouped_SHAP_feature_values under shap/ODI/expanded_PROM."
        )

    extract_root = os.path.join(BASE_DIR, "_step1_odi_sensitivity_shap_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting SHAP source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def find_prediction_csv(source_dir: str, model_type: str) -> str:
    token = f"test_predictions_ODI_{model_type}"
    matches = []
    for root, _, files in os.walk(source_dir):
        for fn in files:
            if fn.startswith(token) and fn.endswith(".csv"):
                matches.append(os.path.join(root, fn))
    if not matches:
        raise FileNotFoundError(f"Missing ODI prediction table for {model_type}. Expected file beginning with {token}.")
    final_matches = [p for p in matches if p.endswith("_Final.csv")]
    return sorted(final_matches or matches)[0]


def find_odi_shap_tables(source_dir: str) -> Tuple[str, str, Optional[str]]:
    value_matches = []
    data_matches = []
    importance_matches = []
    for root, _, files in os.walk(source_dir):
        normalized_root = root.replace(chr(92), "/")
        if "shap/ODI/expanded_PROM" not in normalized_root:
            continue
        for fn in files:
            full = os.path.join(root, fn)
            if fn.startswith("grouped_SHAP_values_") and fn.endswith(".csv"):
                value_matches.append(full)
            elif fn.startswith("grouped_SHAP_feature_values_") and fn.endswith(".csv"):
                data_matches.append(full)
            elif fn.startswith("grouped_SHAP_importance_") and fn.endswith(".csv"):
                importance_matches.append(full)

    if not value_matches or not data_matches:
        raise FileNotFoundError("Could not find ODI grouped SHAP value and feature-value tables.")

    def _select(matches: List[str]) -> str:
        final = [p for p in matches if p.endswith("_Final.csv")]
        top01 = [p for p in (final or matches) if "top01_" in p]
        return sorted(top01 or final or matches)[0]

    return _select(value_matches), _select(data_matches), _select(importance_matches) if importance_matches else None

# ============================================================
# Data loading
# ============================================================

def find_odi_input_csv() -> str:
    for root in [BASE_DIR, FALLBACK_DIR]:
        for fn in ODI_INPUT_CANDIDATES:
            p = os.path.join(root, fn)
            if os.path.exists(p):
                if root == FALLBACK_DIR:
                    dst = os.path.join(BASE_DIR, fn)
                    if not os.path.exists(dst):
                        shutil.copy2(p, dst)
                    return dst
                return p
    raise FileNotFoundError(f"Could not find ODI cohort input file. Tried: {ODI_INPUT_CANDIDATES}")


def load_procedure_map() -> pd.DataFrame:
    path = find_odi_input_csv()
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns:
        raise ValueError(f"ODI cohort file does not contain {GROUP_COL}.")
    missing = [c for c, _ in PROCEDURE_SUBGROUPS if c not in df.columns]
    if missing:
        raise ValueError(f"ODI cohort file is missing procedure columns: {missing}")

    out = df[[GROUP_COL] + [c for c, _ in PROCEDURE_SUBGROUPS]].copy()
    for c, _ in PROCEDURE_SUBGROUPS:
        out[c] = out[c].map(lambda z: parse_binary_value(z, c)).fillna(0.0).astype(int)
    return out.groupby(GROUP_COL, dropna=False)[[c for c, _ in PROCEDURE_SUBGROUPS]].max().reset_index()


def load_paired_predictions(pred_source_dir: str) -> pd.DataFrame:
    base_path = find_prediction_csv(pred_source_dir, "baseline_only")
    exp_path = find_prediction_csv(pred_source_dir, "expanded_PROM")
    base = pd.read_csv(base_path)
    exp = pd.read_csv(exp_path)
    base.columns = [str(c).strip() for c in base.columns]
    exp.columns = [str(c).strip() for c in exp.columns]

    required = [GROUP_COL, "y_true", "p_calibrated"]
    for label, tbl in [("baseline-only", base), ("PROM-expanded", exp)]:
        missing = [c for c in required if c not in tbl.columns]
        if missing:
            raise ValueError(f"ODI {label} prediction table is missing columns: {missing}")

    if "row_index_in_split" not in base.columns:
        base = base.reset_index().rename(columns={"index": "row_index_in_split"})
    if "row_index_in_split" not in exp.columns:
        exp = exp.reset_index().rename(columns={"index": "row_index_in_split"})

    key_cols = ["row_index_in_split", GROUP_COL, "y_true"]
    base_small = base[key_cols + ["p_calibrated"]].rename(columns={"p_calibrated": "p_baseline"})
    exp_small = exp[key_cols + ["p_calibrated"]].rename(columns={"p_calibrated": "p_expanded"})
    paired = base_small.merge(exp_small, on=key_cols, how="inner", validate="one_to_one")

    if len(paired) != len(base_small) or len(paired) != len(exp_small):
        raise RuntimeError("ODI baseline-only and PROM-expanded prediction tables could not be paired one-to-one.")

    paired["y_true"] = paired["y_true"].astype(int)
    paired["p_baseline"] = pd.to_numeric(paired["p_baseline"], errors="coerce")
    paired["p_expanded"] = pd.to_numeric(paired["p_expanded"], errors="coerce")
    return paired.dropna(subset=["p_baseline", "p_expanded"]).copy()

# ============================================================
# ΔAP analysis
# ============================================================

def bootstrap_delta_ap(y: np.ndarray, p_base: np.ndarray, p_exp: np.ndarray, groups: np.ndarray, seed: int) -> Tuple[float, float, float, float]:
    y = np.asarray(y).astype(int)
    p_base = np.asarray(p_base).astype(float)
    p_exp = np.asarray(p_exp).astype(float)
    groups = np.asarray(groups)

    base_ap = safe_average_precision(y, p_base)
    exp_ap = safe_average_precision(y, p_exp)
    obs_delta = exp_ap - base_ap if np.isfinite(base_ap) and np.isfinite(exp_ap) else np.nan
    if not np.isfinite(obs_delta) or len(np.unique(y)) < 2:
        return obs_delta, np.nan, np.nan, np.nan

    group_df = pd.DataFrame({"group": groups, "y": y}).groupby("group", dropna=False)["y"].max().reset_index()
    pos_groups = group_df.loc[group_df["y"].eq(1), "group"].to_numpy()
    neg_groups = group_df.loc[group_df["y"].eq(0), "group"].to_numpy()
    if len(pos_groups) == 0 or len(neg_groups) == 0:
        return obs_delta, np.nan, np.nan, np.nan

    index_by_group = {g: np.where(groups == g)[0] for g in np.unique(groups)}
    rng = np.random.default_rng(seed)
    deltas = []
    for _ in range(N_BOOTSTRAPS):
        sampled_groups = np.concatenate([
            rng.choice(pos_groups, size=len(pos_groups), replace=True),
            rng.choice(neg_groups, size=len(neg_groups), replace=True),
        ])
        idx = np.concatenate([index_by_group[g] for g in sampled_groups])
        yy = y[idx]
        if len(np.unique(yy)) < 2:
            continue
        b = safe_average_precision(yy, p_base[idx])
        e = safe_average_precision(yy, p_exp[idx])
        if np.isfinite(b) and np.isfinite(e):
            deltas.append(e - b)

    if not deltas:
        return float(obs_delta), np.nan, np.nan, np.nan
    deltas = np.asarray(deltas, dtype=float)
    lo, hi = np.percentile(deltas, [2.5, 97.5])
    p_val = 2.0 * min(np.mean(deltas <= 0), np.mean(deltas >= 0))
    return float(obs_delta), float(lo), float(hi), float(min(max(p_val, 0.0), 1.0))


def run_delta_ap_analysis(pred_source_dir: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    preds = load_paired_predictions(pred_source_dir)
    dx_map = load_procedure_map()
    work = preds.merge(dx_map, on=GROUP_COL, how="left", validate="many_to_one")
    for c, _ in PROCEDURE_SUBGROUPS:
        work[c] = work[c].fillna(0).astype(int)

    rows = []
    audit_rows = [
        {"item": "test_prediction_rows", "value": int(len(preds)), "note": "ODI held-out test-set predictions"},
        {"item": "test_events", "value": int(preds["y_true"].sum()), "note": ""},
        {"item": "test_prevalence", "value": float(preds["y_true"].mean()), "note": ""},
    ]

    for i, (dx_col, dx_label) in enumerate(PROCEDURE_SUBGROUPS, start=1):
        sub = work[work[dx_col].eq(1)].copy()
        y = sub["y_true"].to_numpy(dtype=int)
        p0 = sub["p_baseline"].to_numpy(dtype=float)
        p1 = sub["p_expanded"].to_numpy(dtype=float)
        groups = sub[GROUP_COL].to_numpy()
        base_ap = safe_average_precision(y, p0)
        exp_ap = safe_average_precision(y, p1)
        delta, lo, hi, p_val = bootstrap_delta_ap(y, p0, p1, groups, seed=RANDOM_STATE + i * 100)
        rows.append({
            "cohort": "ODI",
            "subgroup_variable": dx_col,
            "subgroup_label": dx_label,
            "n": int(len(sub)),
            "events": int(np.sum(y)) if len(sub) else 0,
            "prevalence": float(np.mean(y)) if len(sub) else np.nan,
            "baseline_AP": base_ap,
            "expanded_AP": exp_ap,
            "delta_AP": delta,
            "delta_AP_CI_lower": lo,
            "delta_AP_CI_upper": hi,
            "delta_AP_p_value_two_sided": p_val,
        })
        audit_rows.append({"item": f"{dx_col}_n", "value": int(len(sub)), "note": dx_label})
        audit_rows.append({"item": f"{dx_col}_events", "value": int(np.sum(y)) if len(sub) else 0, "note": dx_label})

    work.to_csv(os.path.join(TABLE_DIR, "ODI_paired_test_predictions_with_procedure.csv"), index=False)
    return pd.DataFrame(rows), pd.DataFrame(audit_rows)

# ============================================================
# SHAP subgroup beeswarm analysis
# ============================================================

def load_odi_shap_tables(shap_source_dir: str) -> Tuple[pd.DataFrame, pd.DataFrame, Optional[pd.DataFrame]]:
    shap_path, data_path, imp_path = find_odi_shap_tables(shap_source_dir)
    shap_df = pd.read_csv(shap_path)
    data_df = pd.read_csv(data_path)
    imp_df = pd.read_csv(imp_path) if imp_path else None
    shap_df.columns = [str(c).strip() for c in shap_df.columns]
    data_df.columns = [str(c).strip() for c in data_df.columns]
    if imp_df is not None:
        imp_df.columns = [str(c).strip() for c in imp_df.columns]
    if len(shap_df) != len(data_df):
        raise RuntimeError("ODI grouped SHAP value and feature-value tables have different row counts.")
    return shap_df, data_df, imp_df


def procedure_mask_for_shap(data_df: pd.DataFrame, shap_df: pd.DataFrame, dx_col: str, dx_label: str) -> pd.Series:
    candidates = [dx_label, dx_col, dx_col.replace("_", " ")]
    for c in candidates:
        if c in data_df.columns:
            return pd.to_numeric(data_df[c], errors="coerce").fillna(0).astype(float).ge(0.5)
        if c in shap_df.columns:
            return pd.to_numeric(shap_df[c], errors="coerce").fillna(0).astype(float).ge(0.5)
    raise ValueError(
        f"Could not find procedure feature for SHAP subgroup filtering: {dx_label}. "
        "The ODI grouped SHAP feature-value table must include procedure columns."
    )




def subgroup_defining_feature_candidates(subgroup_col: str, subgroup_label: str) -> List[str]:
    """Return possible grouped-SHAP column names for the subgroup-defining feature."""
    candidates = [
        subgroup_label,
        subgroup_col,
        subgroup_col.replace("_", " "),
        subgroup_col.replace("_", " ").title(),
    ]
    # Preserve order while removing duplicates.
    return list(dict.fromkeys(candidates))


def identify_omitted_plot_features(shap_sub: pd.DataFrame, data_sub: pd.DataFrame, subgroup_col: str, subgroup_label: str) -> List[str]:
    """Identify the subgroup-defining feature columns to omit from SHAP display.

    This affects visualization only. Predictions, ΔAP, and all subgroup membership
    definitions still use the unchanged locked model and original subgroup columns.
    """
    candidates = subgroup_defining_feature_candidates(subgroup_col, subgroup_label)
    return [c for c in candidates if c in shap_sub.columns and c in data_sub.columns]

def subgroup_feature_order(
    shap_sub: pd.DataFrame,
    data_sub: pd.DataFrame,
    exclude_features: Optional[List[str]] = None,
) -> List[str]:
    """Rank SHAP features for plotting after excluding tautological subgroup features."""
    exclude_set = set(exclude_features or [])
    feature_cols = [
        c for c in shap_sub.columns
        if not c.startswith("__") and c in data_sub.columns and c not in exclude_set
    ]
    if not feature_cols:
        raise RuntimeError("No shared SHAP/data feature columns were available for beeswarm plotting after exclusions.")
    imp = pd.Series({
        c: float(np.nanmean(np.abs(pd.to_numeric(shap_sub[c], errors="coerce"))))
        for c in feature_cols
    })
    return imp.sort_values(ascending=False).head(DISPLAY_FEATURE_COUNT).index.tolist()


def save_beeswarm(shap_sub: pd.DataFrame, data_sub: pd.DataFrame, features: List[str], path: str, title: str) -> None:
    plt.figure(figsize=(11, max(8.8, 0.32 * len(features) + 2)))
    shap_values = shap_sub[features].apply(pd.to_numeric, errors="coerce").fillna(0.0).values
    feature_values = data_sub[features].apply(pd.to_numeric, errors="coerce")
    try:
        shap.summary_plot(
            shap_values,
            features=feature_values,
            feature_names=features,
            max_display=len(features),
            show=False,
            plot_size=None,
            color=SHAP_COLOR_CMAP,
        )
    except TypeError:
        shap.summary_plot(
            shap_values,
            features=feature_values,
            feature_names=features,
            max_display=len(features),
            show=False,
            plot_size=None,
        )
    plt.title(title, fontsize=15, fontweight="bold")
    plt.tight_layout()
    plt.savefig(path, dpi=300, bbox_inches="tight")
    plt.close()


def run_shap_beeswarm_analysis(shap_source_dir: str) -> pd.DataFrame:
    shap_df, data_df, _ = load_odi_shap_tables(shap_source_dir)
    rows = []
    for dx_col, dx_label in PROCEDURE_SUBGROUPS:
        mask = procedure_mask_for_shap(data_df, shap_df, dx_col, dx_label)
        shap_sub = shap_df.loc[mask].reset_index(drop=True)
        data_sub = data_df.loc[mask].reset_index(drop=True)
        if len(shap_sub) == 0:
            rows.append({
                "cohort": "ODI",
                "subgroup_variable": dx_col,
                "subgroup_label": dx_label,
                "n_shap_rows": 0,
                "plot_path": "",
                "status": "no_rows",
            })
            continue
        omitted_features = identify_omitted_plot_features(shap_sub, data_sub, dx_col, dx_label)
        features = subgroup_feature_order(shap_sub, data_sub, exclude_features=omitted_features)
        plot_path = os.path.join(PLOT_DIR, f"ODI_SHAP_beeswarm_{safe_filename(dx_label)}.png")
        save_beeswarm(
            shap_sub=shap_sub,
            data_sub=data_sub,
            features=features,
            path=plot_path,
            title=f"ODI PROM-expanded model: {dx_label}",
        )
        rows.append({
            "cohort": "ODI",
            "subgroup_variable": dx_col,
            "subgroup_label": dx_label,
            "n_shap_rows": int(len(shap_sub)),
            "plot_path": plot_path,
            "status": "created",
            "features_displayed": "; ".join(features),
            "subgroup_defining_features_omitted_from_plot": "; ".join(omitted_features),
            "visualization_note": "Subgroup-defining feature omitted from this SHAP beeswarm only; locked model and ΔAP analysis unchanged.",
        })
    return pd.DataFrame(rows)

# ============================================================
# Output
# ============================================================

def make_zip(src_dir: str, zip_path: str) -> None:
    if os.path.exists(zip_path):
        os.remove(zip_path)
    tmp_zip = zip_path + ".tmp"
    if os.path.exists(tmp_zip):
        os.remove(tmp_zip)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(src_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(src_dir))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)


def main() -> None:
    print("=" * 90)
    print("Step 1 ODI procedure-stratified sensitivity analysis")
    print("=" * 90)

    pred_source_dir = ensure_prediction_source_dir()
    shap_source_dir = ensure_shap_source_dir()
    print("Using prediction source directory:", pred_source_dir)
    print("Using SHAP source directory:", shap_source_dir)

    delta_ap, audit = run_delta_ap_analysis(pred_source_dir)
    shap_manifest = run_shap_beeswarm_analysis(shap_source_dir)

    delta_csv = os.path.join(TABLE_DIR, "ODI_procedure_delta_AP.csv")
    audit_csv = os.path.join(AUDIT_DIR, "ODI_procedure_sensitivity_audit.csv")
    shap_manifest_csv = os.path.join(TABLE_DIR, "ODI_procedure_SHAP_beeswarm_manifest.csv")
    delta_ap.to_csv(delta_csv, index=False)
    audit.to_csv(audit_csv, index=False)
    shap_manifest.to_csv(shap_manifest_csv, index=False)

    summary_xlsx = os.path.join(OUTPUT_DIR, "Step1_ODI_Procedure_Sensitivity_DeltaAP_SHAP_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        delta_ap.to_excel(writer, sheet_name="delta_AP", index=False)
        shap_manifest.to_excel(writer, sheet_name="SHAP_beeswarm", index=False)
        audit.to_excel(writer, sheet_name="audit", index=False)

    manifest = {
        "analysis": "Step 1 ODI procedure-stratified sensitivity analysis",
        "cohort": "ODI",
        "model": "locked LightGBM",
        "prediction_source_dir": pred_source_dir,
        "shap_source_dir": shap_source_dir,
        "refit": False,
        "retune": False,
        "recalibrate": False,
        "threshold_reoptimization": False,
        "reported_metrics": ["delta_AP"],
        "reported_plots": ["SHAP beeswarm"],
        "shap_visualization_note": "The subgroup-defining feature is omitted from the corresponding SHAP beeswarm plot only; the locked model, predictions, and ΔAP estimates are unchanged.",
        "subgroups": [{"column": c, "label": l, "definition": "procedure performed"} for c, l in PROCEDURE_SUBGROUPS],
        "bootstrap_iterations_for_delta_AP_CI": N_BOOTSTRAPS,
    }
    with open(os.path.join(OUTPUT_DIR, "run_manifest.json"), "w") as f:
        json.dump(manifest, f, indent=2)

    zip_path = os.path.join(BASE_DIR, "Step1_ODI_Procedure_Sensitivity_DeltaAP_SHAP_omit_stratifying_feature.zip")
    make_zip(OUTPUT_DIR, zip_path)

    print("\nDONE")
    print("Delta AP table:", delta_csv)
    print("SHAP beeswarm manifest:", shap_manifest_csv)
    print("Summary Excel:", summary_xlsx)
    print("ZIP:", zip_path)

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(f'<p><b>ODI procedure sensitivity output archive is ready. SHAP plots omit subgroup-defining features for display only.</b></p><p><a href="/files{zip_path}" download>Download ZIP archive</a></p><p>Path: <code>{zip_path}</code></p>'))
        except Exception:
            pass


if __name__ == "__main__":
    main()

# %% [markdown] Cell 9
# #**Step1_ODI_HospitalStratified_Sensitivity_DeltaAP_SHAP**

# %% Cell 10
# -*- coding: utf-8 -*-
"""
Step 1 ODI hospital-stratified sensitivity analysis
=============================================

This script evaluates the locked Step 1 LightGBM ODI models across hospital-level
strata. It uses saved held-out test-set predictions and saved SHAP tables from
the locked primary analysis. It does not refit, retune, recalibrate, or
re-optimize thresholds.

Outputs
-------
1. Hospital-stratified ΔAP for the ODI PROM-expanded model relative to the ODI
   baseline-only model.
2. Hospital-stratified SHAP beeswarm plots for the locked ODI PROM-expanded
   model.

"""

# ============================================================
# Imports
# ============================================================

import os
import re
import sys
import json
import math
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import shap
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "shap"])
    import shap

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from matplotlib import pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from sklearn.metrics import average_precision_score

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_ODI_HospitalStratified_Sensitivity_DeltaAP_SHAP")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
COHORT = "ODI"
N_BOOTSTRAPS = 2000
RANDOM_STATE = 20260524
DISPLAY_FEATURE_COUNT = 25
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

# Prespecified hospital-level reporting criteria for individual sites.
# Hospitals not meeting both criteria are pooled into a lower-volume stratum.
MIN_HOSPITAL_TEST_ROWS = 100
MIN_HOSPITAL_EVENTS = 5
LOWER_VOLUME_STRATUM_LABEL = "Lower-volume hospital stratum"

SHAP_BEESWARM_BLUE = "#008BFB"
SHAP_BEESWARM_PINK = "#FF0051"
SHAP_COLOR_CMAP = LinearSegmentedColormap.from_list(
    "shap_blue_pink", [SHAP_BEESWARM_BLUE, SHAP_BEESWARM_PINK]
)

ODI_INPUT_CANDIDATES = ["PROM_ODI_Cohort.csv", "PROM_ODI_90.csv", "PROM_ODI_Cohort(1).csv"]

HOSPITAL_ID_CANDIDATES = [
    "InstitutionNPI1",
    "InstitutionName",
    "InstitutionKey",
    "HospitalID",
    "SiteID",
    "FacilityID",
]
HOSPITAL_LABEL_CANDIDATES = ["InstitutionName", "InstitutionNPI1", "InstitutionState"]

PREDICTION_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
]

PREDICTION_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
]

SHAP_SOURCE_FOLDER_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final"),
]

SHAP_SOURCE_ARCHIVE_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs", "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_outputs.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_TopConfig_Final.zip"),
    os.path.join(BASE_DIR, "Step1_PROM_LightGBM_Final.zip"),
    os.path.join(FALLBACK_DIR, "Step1_PROM_LightGBM_Final.zip"),
]

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

# ============================================================
# Basic helpers
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("/", "_")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(average_precision_score(y, p))

# ============================================================
# Source discovery
# ============================================================

def _archive_names(zip_path: str) -> List[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            return [n.replace(chr(92), "/") for n in zf.namelist()]
    except Exception:
        return []


def _folder_names(path: str) -> List[str]:
    names = []
    if not os.path.isdir(path):
        return names
    for root, _, files in os.walk(path):
        for fn in files:
            names.append(os.path.join(root, fn).replace(path + os.sep, "").replace(chr(92), "/"))
    return names


def prediction_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    required = [
        "test_predictions_ODI_baseline_only",
        "test_predictions_ODI_expanded_PROM",
    ]
    return all(any(fn.startswith(token) and fn.endswith(".csv") for fn in basenames) for token in required)


def shap_tables_present(names: List[str]) -> bool:
    basenames = [os.path.basename(n) for n in names]
    return (
        any("shap/ODI/expanded_PROM/" in n for n in names)
        and any(fn.startswith("grouped_SHAP_values_") and fn.endswith(".csv") for fn in basenames)
        and any(fn.startswith("grouped_SHAP_feature_values_") and fn.endswith(".csv") for fn in basenames)
    )


def ensure_prediction_source_dir() -> str:
    for folder in PREDICTION_SOURCE_FOLDER_CANDIDATES:
        if prediction_tables_present(_folder_names(folder)):
            return folder

    candidate_archives = list(PREDICTION_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([os.path.join(root, fn) for fn in os.listdir(root) if fn.endswith(".zip")])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive) and prediction_tables_present(_archive_names(archive)):
            archive_path = archive
            break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 1 LightGBM output folder/archive with ODI held-out test prediction tables was found. "
            "Expected files beginning with test_predictions_ODI_baseline_only and test_predictions_ODI_expanded_PROM."
        )

    extract_root = os.path.join(BASE_DIR, "_step1_odi_hospital_prediction_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting prediction source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def ensure_shap_source_dir() -> str:
    for folder in SHAP_SOURCE_FOLDER_CANDIDATES:
        if shap_tables_present(_folder_names(folder)):
            return folder

    candidate_archives = list(SHAP_SOURCE_ARCHIVE_CANDIDATES)
    for root in [BASE_DIR, FALLBACK_DIR]:
        if os.path.isdir(root):
            candidate_archives.extend([os.path.join(root, fn) for fn in os.listdir(root) if fn.endswith(".zip")])

    archive_path = None
    for archive in candidate_archives:
        if os.path.exists(archive) and shap_tables_present(_archive_names(archive)):
            archive_path = archive
            break

    if archive_path is None:
        raise FileNotFoundError(
            "No Step 1 LightGBM source folder/archive with ODI grouped SHAP tables was found. "
            "Expected grouped_SHAP_values and grouped_SHAP_feature_values under shap/ODI/expanded_PROM."
        )

    extract_root = os.path.join(BASE_DIR, "_step1_odi_hospital_shap_source")
    if os.path.isdir(extract_root):
        shutil.rmtree(extract_root)
    os.makedirs(extract_root, exist_ok=True)
    print(f"Extracting SHAP source archive: {archive_path}")
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(extract_root)
    return extract_root


def find_prediction_csv(source_dir: str, model_type: str) -> str:
    token = f"test_predictions_ODI_{model_type}"
    matches = []
    for root, _, files in os.walk(source_dir):
        for fn in files:
            if fn.startswith(token) and fn.endswith(".csv"):
                matches.append(os.path.join(root, fn))
    if not matches:
        raise FileNotFoundError(f"Missing ODI prediction table for {model_type}. Expected file beginning with {token}.")
    final_matches = [p for p in matches if p.endswith("_Final.csv")]
    return sorted(final_matches or matches)[0]


def find_odi_shap_tables(source_dir: str) -> Tuple[str, str, Optional[str]]:
    value_matches = []
    data_matches = []
    importance_matches = []
    for root, _, files in os.walk(source_dir):
        normalized_root = root.replace(chr(92), "/")
        if "shap/ODI/expanded_PROM" not in normalized_root:
            continue
        for fn in files:
            full = os.path.join(root, fn)
            if fn.startswith("grouped_SHAP_values_") and fn.endswith(".csv"):
                value_matches.append(full)
            elif fn.startswith("grouped_SHAP_feature_values_") and fn.endswith(".csv"):
                data_matches.append(full)
            elif fn.startswith("grouped_SHAP_importance_") and fn.endswith(".csv"):
                importance_matches.append(full)

    if not value_matches or not data_matches:
        raise FileNotFoundError("Could not find ODI grouped SHAP value and feature-value tables.")

    def _select(matches: List[str]) -> str:
        final = [p for p in matches if p.endswith("_Final.csv")]
        top01 = [p for p in (final or matches) if "top01_" in p]
        return sorted(top01 or final or matches)[0]

    return _select(value_matches), _select(data_matches), _select(importance_matches) if importance_matches else None

# ============================================================
# Data loading and hospital strata
# ============================================================

def find_odi_input_csv() -> str:
    for root in [BASE_DIR, FALLBACK_DIR]:
        for fn in ODI_INPUT_CANDIDATES:
            p = os.path.join(root, fn)
            if os.path.exists(p):
                if root == FALLBACK_DIR:
                    dst = os.path.join(BASE_DIR, fn)
                    if not os.path.exists(dst):
                        shutil.copy2(p, dst)
                    return dst
                return p
    raise FileNotFoundError(f"Could not find ODI cohort input file. Tried: {ODI_INPUT_CANDIDATES}")


def find_existing_column(columns: List[str], candidates: List[str], required: bool = True) -> Optional[str]:
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if c.lower() in lookup:
            return lookup[c.lower()]
    if required:
        raise ValueError(f"Could not find any required column among: {candidates}")
    return None


def load_hospital_map() -> Tuple[pd.DataFrame, pd.DataFrame, str, Optional[str]]:
    path = find_odi_input_csv()
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns:
        raise ValueError(f"ODI cohort file does not contain {GROUP_COL}.")

    hospital_id_col = find_existing_column(df.columns.tolist(), HOSPITAL_ID_CANDIDATES, required=True)
    hospital_label_col = find_existing_column(df.columns.tolist(), HOSPITAL_LABEL_CANDIDATES, required=False)

    keep_cols = [GROUP_COL, hospital_id_col]
    if hospital_label_col and hospital_label_col not in keep_cols:
        keep_cols.append(hospital_label_col)
    raw = df[keep_cols].copy()
    raw[hospital_id_col] = raw[hospital_id_col].map(clean_scalar).astype("object")
    raw = raw[raw[hospital_id_col].notna()].copy()
    raw["hospital_id"] = raw[hospital_id_col].astype(str)
    if hospital_label_col:
        raw["hospital_label_source"] = raw[hospital_label_col].map(clean_scalar).astype("object").astype(str)
    else:
        raw["hospital_label_source"] = raw["hospital_id"]

    ambiguity = (
        raw.groupby(GROUP_COL, dropna=False)["hospital_id"]
        .nunique(dropna=True)
        .reset_index(name="n_hospitals_per_patient")
    )
    n_ambiguous = int(ambiguity["n_hospitals_per_patient"].gt(1).sum())

    # Use the modal hospital within each patient group if multiple surgical rows exist.
    rows = []
    for person, g in raw.groupby(GROUP_COL, dropna=False):
        counts = g["hospital_id"].value_counts(dropna=True)
        if counts.empty:
            continue
        hid = str(counts.index[0])
        label_candidates = g.loc[g["hospital_id"].astype(str).eq(hid), "hospital_label_source"].dropna().astype(str)
        hlabel = label_candidates.iloc[0] if len(label_candidates) else hid
        rows.append({GROUP_COL: person, "hospital_id": hid, "hospital_label_source": hlabel})
    hospital_map = pd.DataFrame(rows)

    audit = pd.DataFrame([
        {"item": "input_file", "value": path, "note": ""},
        {"item": "hospital_id_column", "value": hospital_id_col, "note": ""},
        {"item": "hospital_label_column", "value": hospital_label_col if hospital_label_col else "none", "note": ""},
        {"item": "patients_with_hospital_assignment", "value": int(hospital_map[GROUP_COL].nunique()), "note": ""},
        {"item": "patients_with_multiple_hospital_ids", "value": n_ambiguous, "note": "modal hospital used for patient-level mapping"},
    ])
    return hospital_map, audit, hospital_id_col, hospital_label_col


def load_paired_predictions(pred_source_dir: str) -> pd.DataFrame:
    base_path = find_prediction_csv(pred_source_dir, "baseline_only")
    exp_path = find_prediction_csv(pred_source_dir, "expanded_PROM")
    base = pd.read_csv(base_path)
    exp = pd.read_csv(exp_path)
    base.columns = [str(c).strip() for c in base.columns]
    exp.columns = [str(c).strip() for c in exp.columns]

    required = [GROUP_COL, "y_true", "p_calibrated"]
    for label, tbl in [("baseline-only", base), ("PROM-expanded", exp)]:
        missing = [c for c in required if c not in tbl.columns]
        if missing:
            raise ValueError(f"ODI {label} prediction table is missing columns: {missing}")

    if "row_index_in_split" not in base.columns:
        base = base.reset_index().rename(columns={"index": "row_index_in_split"})
    if "row_index_in_split" not in exp.columns:
        exp = exp.reset_index().rename(columns={"index": "row_index_in_split"})

    key_cols = ["row_index_in_split", GROUP_COL, "y_true"]
    base_small = base[key_cols + ["p_calibrated"]].rename(columns={"p_calibrated": "p_baseline"})
    exp_small = exp[key_cols + ["p_calibrated"]].rename(columns={"p_calibrated": "p_expanded"})
    paired = base_small.merge(exp_small, on=key_cols, how="inner", validate="one_to_one")

    if len(paired) != len(base_small) or len(paired) != len(exp_small):
        raise RuntimeError("ODI baseline-only and PROM-expanded prediction tables could not be paired one-to-one.")

    paired["y_true"] = paired["y_true"].astype(int)
    paired["p_baseline"] = pd.to_numeric(paired["p_baseline"], errors="coerce")
    paired["p_expanded"] = pd.to_numeric(paired["p_expanded"], errors="coerce")
    return paired.dropna(subset=["p_baseline", "p_expanded"]).copy()


def assign_hospital_strata(paired: pd.DataFrame, hospital_map: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    work = paired.merge(hospital_map, on=GROUP_COL, how="left", validate="many_to_one")
    missing_hospital = int(work["hospital_id"].isna().sum())
    work = work.dropna(subset=["hospital_id"]).copy()
    work["hospital_id"] = work["hospital_id"].astype(str)

    hospital_counts = (
        work.groupby("hospital_id", dropna=False)
        .agg(
            n=("y_true", "size"),
            events=("y_true", "sum"),
            hospital_label_source=("hospital_label_source", "first"),
        )
        .reset_index()
    )
    hospital_counts["meets_individual_reporting_criteria"] = (
        hospital_counts["n"].ge(MIN_HOSPITAL_TEST_ROWS)
        & hospital_counts["events"].ge(MIN_HOSPITAL_EVENTS)
    )
    eligible = hospital_counts[hospital_counts["meets_individual_reporting_criteria"]].copy()
    eligible = eligible.sort_values(["events", "n", "hospital_id"], ascending=[False, False, True]).reset_index(drop=True)
    eligible["hospital_stratum"] = [f"Hospital {i:02d}" for i in range(1, len(eligible) + 1)]

    stratum_map = dict(zip(eligible["hospital_id"], eligible["hospital_stratum"]))
    work["hospital_stratum"] = work["hospital_id"].map(stratum_map).fillna(LOWER_VOLUME_STRATUM_LABEL)
    work["hospital_stratum_type"] = np.where(
        work["hospital_stratum"].eq(LOWER_VOLUME_STRATUM_LABEL),
        "pooled_lower_volume",
        "individual_hospital",
    )

    crosswalk = hospital_counts.merge(
        eligible[["hospital_id", "hospital_stratum"]],
        on="hospital_id",
        how="left",
    )
    crosswalk["hospital_stratum"] = crosswalk["hospital_stratum"].fillna(LOWER_VOLUME_STRATUM_LABEL)
    crosswalk["minimum_test_rows_required"] = MIN_HOSPITAL_TEST_ROWS
    crosswalk["minimum_events_required"] = MIN_HOSPITAL_EVENTS
    crosswalk["missing_hospital_rows_in_test_predictions"] = missing_hospital
    crosswalk = crosswalk.sort_values(["hospital_stratum", "events", "n"], ascending=[True, False, False])
    return work, crosswalk

# ============================================================
# ΔAP analysis
# ============================================================

def bootstrap_delta_ap(y: np.ndarray, p_base: np.ndarray, p_exp: np.ndarray, groups: np.ndarray, seed: int) -> Tuple[float, float, float, float]:
    y = np.asarray(y).astype(int)
    p_base = np.asarray(p_base).astype(float)
    p_exp = np.asarray(p_exp).astype(float)
    groups = np.asarray(groups)

    base_ap = safe_average_precision(y, p_base)
    exp_ap = safe_average_precision(y, p_exp)
    obs_delta = exp_ap - base_ap if np.isfinite(base_ap) and np.isfinite(exp_ap) else np.nan
    if not np.isfinite(obs_delta) or len(np.unique(y)) < 2:
        return obs_delta, np.nan, np.nan, np.nan

    group_df = pd.DataFrame({"group": groups, "y": y}).groupby("group", dropna=False)["y"].max().reset_index()
    pos_groups = group_df.loc[group_df["y"].eq(1), "group"].to_numpy()
    neg_groups = group_df.loc[group_df["y"].eq(0), "group"].to_numpy()
    if len(pos_groups) == 0 or len(neg_groups) == 0:
        return obs_delta, np.nan, np.nan, np.nan

    index_by_group = {g: np.where(groups == g)[0] for g in np.unique(groups)}
    rng = np.random.default_rng(seed)
    deltas = []
    for _ in range(N_BOOTSTRAPS):
        sampled_groups = np.concatenate([
            rng.choice(pos_groups, size=len(pos_groups), replace=True),
            rng.choice(neg_groups, size=len(neg_groups), replace=True),
        ])
        idx = np.concatenate([index_by_group[g] for g in sampled_groups])
        yy = y[idx]
        if len(np.unique(yy)) < 2:
            continue
        b = safe_average_precision(yy, p_base[idx])
        e = safe_average_precision(yy, p_exp[idx])
        if np.isfinite(b) and np.isfinite(e):
            deltas.append(e - b)

    if not deltas:
        return float(obs_delta), np.nan, np.nan, np.nan
    deltas = np.asarray(deltas, dtype=float)
    lo, hi = np.percentile(deltas, [2.5, 97.5])
    p_val = 2.0 * min(np.mean(deltas <= 0), np.mean(deltas >= 0))
    return float(obs_delta), float(lo), float(hi), float(min(max(p_val, 0.0), 1.0))


def run_delta_ap_analysis(pred_source_dir: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    preds = load_paired_predictions(pred_source_dir)
    hospital_map, hospital_audit, hospital_id_col, hospital_label_col = load_hospital_map()
    work, crosswalk = assign_hospital_strata(preds, hospital_map)

    rows = []
    audit_rows = [
        {"item": "test_prediction_rows", "value": int(len(preds)), "note": "ODI held-out test-set predictions"},
        {"item": "test_events", "value": int(preds["y_true"].sum()), "note": ""},
        {"item": "test_prevalence", "value": float(preds["y_true"].mean()), "note": ""},
        {"item": "hospital_id_column", "value": hospital_id_col, "note": ""},
        {"item": "hospital_label_column", "value": hospital_label_col if hospital_label_col else "none", "note": ""},
        {"item": "minimum_test_rows_for_individual_hospital", "value": int(MIN_HOSPITAL_TEST_ROWS), "note": ""},
        {"item": "minimum_events_for_individual_hospital", "value": int(MIN_HOSPITAL_EVENTS), "note": ""},
        {"item": "number_of_individual_hospital_strata", "value": int(crosswalk["meets_individual_reporting_criteria"].sum()), "note": ""},
    ]
    audit_rows.extend(hospital_audit.to_dict("records"))

    strata_order = (
        work.groupby("hospital_stratum")
        .agg(n=("y_true", "size"), events=("y_true", "sum"))
        .reset_index()
        .sort_values(["hospital_stratum"], ascending=True)
    )
    if LOWER_VOLUME_STRATUM_LABEL in strata_order["hospital_stratum"].values:
        ordered = [s for s in strata_order["hospital_stratum"].tolist() if s != LOWER_VOLUME_STRATUM_LABEL] + [LOWER_VOLUME_STRATUM_LABEL]
    else:
        ordered = strata_order["hospital_stratum"].tolist()

    for i, stratum in enumerate(ordered, start=1):
        sub = work[work["hospital_stratum"].eq(stratum)].copy()
        y = sub["y_true"].to_numpy(dtype=int)
        p0 = sub["p_baseline"].to_numpy(dtype=float)
        p1 = sub["p_expanded"].to_numpy(dtype=float)
        groups = sub[GROUP_COL].to_numpy()
        base_ap = safe_average_precision(y, p0)
        exp_ap = safe_average_precision(y, p1)
        delta, lo, hi, p_val = bootstrap_delta_ap(y, p0, p1, groups, seed=RANDOM_STATE + i * 100)
        rows.append({
            "cohort": "ODI",
            "hospital_stratum": stratum,
            "hospital_stratum_type": "pooled_lower_volume" if stratum == LOWER_VOLUME_STRATUM_LABEL else "individual_hospital",
            "n": int(len(sub)),
            "events": int(np.sum(y)) if len(sub) else 0,
            "prevalence": float(np.mean(y)) if len(sub) else np.nan,
            "baseline_AP": base_ap,
            "expanded_AP": exp_ap,
            "delta_AP": delta,
            "delta_AP_CI_lower": lo,
            "delta_AP_CI_upper": hi,
            "delta_AP_p_value_two_sided": p_val,
        })
        audit_rows.append({"item": f"{safe_filename(stratum)}_n", "value": int(len(sub)), "note": stratum})
        audit_rows.append({"item": f"{safe_filename(stratum)}_events", "value": int(np.sum(y)) if len(sub) else 0, "note": stratum})

    work.to_csv(os.path.join(TABLE_DIR, "ODI_paired_test_predictions_with_hospital_strata.csv"), index=False)
    crosswalk.to_csv(os.path.join(AUDIT_DIR, "ODI_hospital_strata_crosswalk.csv"), index=False)
    return pd.DataFrame(rows), pd.DataFrame(audit_rows), work

# ============================================================
# SHAP subgroup beeswarm analysis
# ============================================================

def load_odi_shap_tables(shap_source_dir: str) -> Tuple[pd.DataFrame, pd.DataFrame, Optional[pd.DataFrame]]:
    shap_path, data_path, imp_path = find_odi_shap_tables(shap_source_dir)
    shap_df = pd.read_csv(shap_path)
    data_df = pd.read_csv(data_path)
    imp_df = pd.read_csv(imp_path) if imp_path else None
    shap_df.columns = [str(c).strip() for c in shap_df.columns]
    data_df.columns = [str(c).strip() for c in data_df.columns]
    if imp_df is not None:
        imp_df.columns = [str(c).strip() for c in imp_df.columns]
    if len(shap_df) != len(data_df):
        raise RuntimeError("ODI grouped SHAP value and feature-value tables have different row counts.")
    return shap_df, data_df, imp_df


def attach_hospital_strata_to_shap(shap_df: pd.DataFrame, data_df: pd.DataFrame, paired_with_hospital: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, str]:
    shap_out = shap_df.copy()
    data_out = data_df.copy()
    paired = paired_with_hospital.copy()

    if "__row_id__" in shap_out.columns:
        left = pd.DataFrame({"__shap_order__": np.arange(len(shap_out)), "row_index_in_split": pd.to_numeric(shap_out["__row_id__"], errors="coerce")})
        right = paired[["row_index_in_split", "hospital_stratum", "hospital_stratum_type"]].copy()
        merged = left.merge(right, on="row_index_in_split", how="left", validate="one_to_one")
        if merged["hospital_stratum"].notna().mean() >= 0.95:
            merged = merged.sort_values("__shap_order__")
            shap_out["__hospital_stratum__"] = merged["hospital_stratum"].values
            data_out["__hospital_stratum__"] = merged["hospital_stratum"].values
            return shap_out, data_out, "matched_by_row_id"

    if len(shap_out) == len(paired):
        ordered = paired.sort_values("row_index_in_split").reset_index(drop=True)
        shap_out["__hospital_stratum__"] = ordered["hospital_stratum"].values
        data_out["__hospital_stratum__"] = ordered["hospital_stratum"].values
        return shap_out, data_out, "matched_by_row_order"

    raise RuntimeError(
        "Could not attach hospital strata to SHAP rows. The SHAP table must contain __row_id__ matching "
        "row_index_in_split, or the SHAP table length must equal the held-out prediction table length."
    )


def subgroup_feature_order(shap_sub: pd.DataFrame, data_sub: pd.DataFrame) -> List[str]:
    feature_cols = [c for c in shap_sub.columns if not c.startswith("__") and c in data_sub.columns]
    if not feature_cols:
        raise RuntimeError("No shared SHAP/data feature columns were available for beeswarm plotting.")
    imp = pd.Series({c: float(np.nanmean(np.abs(pd.to_numeric(shap_sub[c], errors="coerce")))) for c in feature_cols})
    return imp.sort_values(ascending=False).head(DISPLAY_FEATURE_COUNT).index.tolist()


def save_beeswarm(shap_sub: pd.DataFrame, data_sub: pd.DataFrame, features: List[str], path: str, title: str) -> None:
    plt.figure(figsize=(11, max(8.8, 0.32 * len(features) + 2)))
    shap_values = shap_sub[features].apply(pd.to_numeric, errors="coerce").fillna(0.0).values
    feature_values = data_sub[features].apply(pd.to_numeric, errors="coerce")
    try:
        shap.summary_plot(
            shap_values,
            features=feature_values,
            feature_names=features,
            max_display=len(features),
            show=False,
            plot_size=None,
            color=SHAP_COLOR_CMAP,
        )
    except TypeError:
        shap.summary_plot(
            shap_values,
            features=feature_values,
            feature_names=features,
            max_display=len(features),
            show=False,
            plot_size=None,
        )
    plt.title(title, fontsize=15, fontweight="bold")
    plt.tight_layout()
    plt.savefig(path, dpi=300, bbox_inches="tight")
    plt.close()


def run_shap_beeswarm_analysis(shap_source_dir: str, paired_with_hospital: pd.DataFrame, delta_ap_table: pd.DataFrame) -> pd.DataFrame:
    shap_df, data_df, _ = load_odi_shap_tables(shap_source_dir)
    shap_df, data_df, match_method = attach_hospital_strata_to_shap(shap_df, data_df, paired_with_hospital)

    rows = []
    for _, r in delta_ap_table.iterrows():
        stratum = str(r["hospital_stratum"])
        mask = shap_df["__hospital_stratum__"].astype(str).eq(stratum)
        shap_sub = shap_df.loc[mask].reset_index(drop=True)
        data_sub = data_df.loc[mask].reset_index(drop=True)
        if len(shap_sub) == 0:
            rows.append({
                "cohort": "ODI",
                "hospital_stratum": stratum,
                "hospital_stratum_type": r["hospital_stratum_type"],
                "n_shap_rows": 0,
                "plot_path": "",
                "status": "no_rows",
                "match_method": match_method,
            })
            continue
        features = subgroup_feature_order(shap_sub, data_sub)
        plot_path = os.path.join(PLOT_DIR, f"ODI_SHAP_beeswarm_{safe_filename(stratum)}.png")
        save_beeswarm(
            shap_sub=shap_sub,
            data_sub=data_sub,
            features=features,
            path=plot_path,
            title=f"ODI PROM-expanded model: {stratum}",
        )
        rows.append({
            "cohort": "ODI",
            "hospital_stratum": stratum,
            "hospital_stratum_type": r["hospital_stratum_type"],
            "n_shap_rows": int(len(shap_sub)),
            "plot_path": plot_path,
            "status": "created",
            "match_method": match_method,
            "features_displayed": "; ".join(features),
        })
    return pd.DataFrame(rows)

# ============================================================
# Output
# ============================================================

def make_zip(src_dir: str, zip_path: str) -> None:
    if os.path.exists(zip_path):
        os.remove(zip_path)
    tmp_zip = zip_path + ".tmp"
    if os.path.exists(tmp_zip):
        os.remove(tmp_zip)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(src_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(src_dir))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)


def main() -> None:
    print("=" * 90)
    print("Step 1 ODI hospital-stratified sensitivity analysis")
    print("=" * 90)

    pred_source_dir = ensure_prediction_source_dir()
    shap_source_dir = ensure_shap_source_dir()
    print("Using prediction source directory:", pred_source_dir)
    print("Using SHAP source directory:", shap_source_dir)

    delta_ap, audit, paired_with_hospital = run_delta_ap_analysis(pred_source_dir)
    shap_manifest = run_shap_beeswarm_analysis(shap_source_dir, paired_with_hospital, delta_ap)

    delta_csv = os.path.join(TABLE_DIR, "ODI_hospital_stratified_delta_AP.csv")
    audit_csv = os.path.join(AUDIT_DIR, "ODI_hospital_stratified_sensitivity_audit.csv")
    shap_manifest_csv = os.path.join(TABLE_DIR, "ODI_hospital_stratified_SHAP_beeswarm_manifest.csv")
    delta_ap.to_csv(delta_csv, index=False)
    audit.to_csv(audit_csv, index=False)
    shap_manifest.to_csv(shap_manifest_csv, index=False)

    summary_xlsx = os.path.join(OUTPUT_DIR, "Step1_ODI_HospitalStratified_Sensitivity_DeltaAP_SHAP_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        delta_ap.to_excel(writer, sheet_name="delta_AP", index=False)
        shap_manifest.to_excel(writer, sheet_name="SHAP_beeswarm", index=False)
        audit.to_excel(writer, sheet_name="audit", index=False)

    manifest = {
        "analysis": "Step 1 ODI hospital-stratified sensitivity analysis",
        "cohort": "ODI",
        "model": "locked LightGBM",
        "prediction_source_dir": pred_source_dir,
        "shap_source_dir": shap_source_dir,
        "refit": False,
        "retune": False,
        "recalibrate": False,
        "threshold_reoptimization": False,
        "reported_metrics": ["delta_AP"],
        "reported_plots": ["SHAP beeswarm"],
        "hospital_id_candidates": HOSPITAL_ID_CANDIDATES,
        "individual_hospital_minimum_test_rows": MIN_HOSPITAL_TEST_ROWS,
        "individual_hospital_minimum_events": MIN_HOSPITAL_EVENTS,
        "pooled_stratum_label": LOWER_VOLUME_STRATUM_LABEL,
        "bootstrap_iterations_for_delta_AP_CI": N_BOOTSTRAPS,
        "note": "This is a hospital-stratified sensitivity analysis using the locked held-out test-set predictions from the primary Step 1 LightGBM analysis. It is not a true hospital-held-out validation."
    }
    with open(os.path.join(OUTPUT_DIR, "run_manifest.json"), "w") as f:
        json.dump(manifest, f, indent=2)

    zip_path = os.path.join(BASE_DIR, "Step1_ODI_HospitalStratified_Sensitivity_DeltaAP_SHAP.zip")
    make_zip(OUTPUT_DIR, zip_path)

    print("\nDONE")
    print("Delta AP table:", delta_csv)
    print("SHAP beeswarm manifest:", shap_manifest_csv)
    print("Summary Excel:", summary_xlsx)
    print("ZIP:", zip_path)

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(f'<p><b>ODI hospital-stratified sensitivity output archive is ready.</b></p><p><a href="/files{zip_path}" download>Download ZIP archive</a></p><p>Path: <code>{zip_path}</code></p>'))
        except Exception:
            pass


if __name__ == "__main__":
    main()

# %% [markdown] Cell 11
# #**Step1_ODI_HospitalHeldOut_InternalValidation**

# %% Cell 12
# -*- coding: utf-8 -*-
"""
Step 1 ODI hospital-held-out internal validation
================================================

This script performs a true hospital-held-out internal validation for the Step 1
ODI comparison. Hospitals are assigned exclusively to development-training,
calibration, or held-out validation splits. The held-out hospitals are never used
for preprocessing, hyperparameter tuning, model fitting, or probability
calibration.
"""

import os
import re
import sys
import json
import math
import time
import zipfile
import shutil
import warnings
import subprocess
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import lightgbm as lgb
    from lightgbm import LGBMClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lightgbm"])
    import lightgbm as lgb
    from lightgbm import LGBMClassifier

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder
from sklearn.model_selection import StratifiedGroupKFold, ParameterSampler
from sklearn.metrics import (
    average_precision_score,
    roc_auc_score,
    brier_score_loss,
)
from sklearn.isotonic import IsotonicRegression
import matplotlib.pyplot as plt

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"
OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_ODI_HospitalHeldOut_InternalValidation")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
MODEL_DIR = os.path.join(OUTPUT_DIR, "model_artifacts")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR, MODEL_DIR]:
    os.makedirs(_d, exist_ok=True)

GROUP_COL = "PersonKey"
RANDOM_STATE = 20260524

# Hospital-level split. Fractions are by hospital count, not row count.
HOSPITAL_HELDOUT_TEST_FRACTION = 0.20
HOSPITAL_HELDOUT_CALIBRATION_FRACTION_OF_REMAINING = 0.20
HOSPITAL_HELDOUT_SPLIT_ATTEMPTS = 3000
HOSPITAL_HELDOUT_MIN_TRAIN_EVENTS = 30
HOSPITAL_HELDOUT_MIN_CALIBRATION_EVENTS = 5
HOSPITAL_HELDOUT_MIN_TEST_EVENTS = 10

# Tuning protocol. This mirrors the Step 1 LightGBM protocol but is run only in
# the hospital-training split.
N_CV_FOLDS = 5
N_RANDOM_COMBINATIONS = 300
USE_EARLY_STOPPING_IN_CV = True
EARLY_STOPPING_ROUNDS = 100
MIN_FINAL_N_ESTIMATORS = 50
N_BOOTSTRAPS = 2000
ECE_N_BINS = 10
N_JOBS = -1
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

ODI_INPUT_CANDIDATES = [
    "PROM_ODI_Cohort.csv",
    "PROM_ODI_90.csv",
    "PROM_ODI_Cohort(1).csv",
]
TARGET_COL_CANDIDATES = [
    "final_reop_step1",
    "final_reop_1yr",
    "reop_1yr",
    "reoperation_1yr",
    "one_year_reoperation",
    "reop365",
    "reop_365",
    "final_reop",
    "reoperation",
    "reop",
]
PROM_COL_CANDIDATES = [
    "ODI_preop_value",
    "preop_odi_within_90d",
    "odi_preop_within_90d",
    "odi_preop_anytime_value",
    "preop_odi_anytime",
    "preop_ODI",
    "Preop_ODI",
    "preoperative_ODI",
    "Preoperative_ODI",
    "ODI",
    "odi",
    "baseline_ODI",
]
HOSPITAL_ID_CANDIDATES = [
    "InstitutionNPI1",
    "InstitutionKey",
    "InstitutionName",
    "HospitalID",
    "HospitalId",
    "SiteID",
    "FacilityID",
]
HOSPITAL_LABEL_CANDIDATES = ["InstitutionName", "InstitutionNPI1", "InstitutionState"]

BASELINE_FEATURES = [
    "finaldx_degenerative",
    "finaldx_radicular",
    "finaldx_stenosis",
    "finaldx_deformity_instability",
    "finaldx_other_diagnosis",
    "age",
    "sex",
    "race",
    "ethnicity",
    "cancer_status",
    "chronic_pulmonary_disease",
    "congestive_heart_failure",
    "connective_tissue_rheumatic_disease",
    "diabetes_status",
    "myocardial_infarction",
    "renal_disease",
    "institution_type",
    "institution_size",
    "institution_region",
    "asa",
    "bmi",
    "payer_status",
    "alif_llif",
    "corpectomy",
    "discectomy",
    "foraminotomy",
    "instrumentation",
    "laminectomy_posterior_decompression",
    "pelvic_fixation",
    "plf",
    "tlif_plif",
    "other_lumbar_procedure",
    "number_operated_levels",
    "operative_region_extent",
    "PatTobaccoUse",
]
assert len(BASELINE_FEATURES) == 35

CONTINUOUS_BASELINE_FEATURES = ["age", "bmi"]
BINARY_FEATURES = [
    "finaldx_degenerative",
    "finaldx_radicular",
    "finaldx_stenosis",
    "finaldx_deformity_instability",
    "finaldx_other_diagnosis",
    "sex",
    "ethnicity",
    "cancer_status",
    "chronic_pulmonary_disease",
    "congestive_heart_failure",
    "connective_tissue_rheumatic_disease",
    "myocardial_infarction",
    "renal_disease",
    "institution_type",
    "alif_llif",
    "corpectomy",
    "discectomy",
    "foraminotomy",
    "instrumentation",
    "laminectomy_posterior_decompression",
    "pelvic_fixation",
    "plf",
    "tlif_plif",
    "other_lumbar_procedure",
    "operative_region_extent",
]
ORDINAL_FEATURES = ["diabetes_status", "institution_size", "asa", "number_operated_levels"]
NOMINAL_FEATURES = ["race", "institution_region", "payer_status", "PatTobaccoUse"]

PREFERRED_NOMINAL_LEVELS = {
    "race": ["White", "Black", "Other"],
    "institution_region": ["South", "North East", "West", "Midwest"],
    "payer_status": ["Medicare", "Commercial/Private", "Other", "Medicaid/Public/Government"],
    "PatTobaccoUse": ["Unknown/Not reported/Multiple", "Never", "Former", "Current"],
}

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}
BINARY_MAPS = {
    "sex": {"female": 0, "f": 0, "male": 1, "m": 1},
    "ethnicity": {"non-hispanic": 0, "non hispanic": 0, "hispanic": 1},
    "cancer_status": {"no cancer": 0, "no": 0, "none": 0, "any cancer": 1, "yes": 1, "cancer": 1},
    "institution_type": {"hospital": 0, "non-hospital": 1, "non hospital": 1, "nonhospital": 1},
    "operative_region_extent": {"lumbar only": 0, "extended_region_involvement": 1, "extended region involvement": 1, "extended": 1},
}
ORDINAL_MAPS = {
    "diabetes_status": {
        "no": 0,
        "none": 0,
        "0": 0,
        "without comp": 1,
        "without complication": 1,
        "without complications": 1,
        "1": 1,
        "with comp": 2,
        "with complication": 2,
        "with complications": 2,
        "2": 2,
    },
    "institution_size": {
        "between 1-99 beds": 0,
        "1-99": 0,
        "between 100-399 beds": 1,
        "100-399": 1,
        ">= 400 beds": 2,
        ">=400 beds": 2,
        ">=400": 2,
        ">= 400": 2,
    },
    "asa": {"1": 1, "i": 1, "2": 2, "ii": 2, "3": 3, "iii": 3, "4": 4, "iv": 4, ">=4": 4, ">= 4": 4, "5": 4, "v": 4},
    "number_operated_levels": {"0": 0, "1": 1, "2": 2, "3": 3, "4": 4, ">=4": 4, ">= 4": 4, "5": 4, "6": 4, "7": 4, "8": 4, "9": 4, "10": 4},
}

LGBM_SEARCH_SPACE = {
    "n_estimators": [400, 700, 1000, 1400, 1800, 2200, 2600],
    "learning_rate": [0.003, 0.005, 0.008, 0.01, 0.02, 0.03, 0.05],
    "num_leaves": [7, 15, 31, 63, 127],
    "max_depth": [-1, 2, 3, 5, 7, 9],
    "min_child_samples": [10, 20, 50, 100, 200, 400],
    "subsample": [0.60, 0.75, 0.90, 1.00],
    "subsample_freq": [0, 1, 2],
    "colsample_bytree": [0.60, 0.75, 0.90, 1.00],
    "reg_alpha": [0.0, 0.001, 0.01, 0.05, 0.10, 0.50, 1.00, 2.00],
    "reg_lambda": [0.0, 0.001, 0.01, 0.05, 0.10, 0.50, 1.00, 2.00, 5.00],
    "min_split_gain": [0.0, 0.001, 0.005, 0.01, 0.05, 0.10],
    "max_bin": [63, 127, 255],
    "positive_weight_multiplier": [0.25, 0.50, 0.75, 1.00, 1.50, 2.00, 3.00, 4.00, 6.00, 8.00],
}
LGBM_INT_PARAMS = {"n_estimators", "num_leaves", "max_depth", "min_child_samples", "subsample_freq", "max_bin"}
LGBM_FLOAT_PARAMS = {
    "learning_rate",
    "subsample",
    "colsample_bytree",
    "reg_alpha",
    "reg_lambda",
    "min_split_gain",
    "positive_weight_multiplier",
}
PARAMETER_CANDIDATES = list(
    ParameterSampler(LGBM_SEARCH_SPACE, n_iter=N_RANDOM_COMBINATIONS, random_state=RANDOM_STATE)
)

# ============================================================
# Helpers
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("/", "_")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def find_existing_column(columns: List[str], candidates: List[str], what: str, required: bool = True) -> Optional[str]:
    lookup = {str(c).strip().lower(): str(c).strip() for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if str(c).strip().lower() in lookup:
            return lookup[str(c).strip().lower()]
    if required:
        raise ValueError(f"Could not find {what}. Tried: {candidates}")
    return None


def to_binary_target(x: Any) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if sx in {"1", "1.0", "yes", "y", "true", "t"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(average_precision_score(y, p))


def safe_roc_auc(y: np.ndarray, p: np.ndarray) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0 or len(np.unique(y)) < 2:
        return np.nan
    return float(roc_auc_score(y, p))


def ece(y: np.ndarray, p: np.ndarray, n_bins: int = ECE_N_BINS) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    if len(y) == 0:
        return np.nan
    bins = np.linspace(0, 1, n_bins + 1)
    out = 0.0
    for i in range(n_bins):
        m = (p >= bins[i]) & ((p <= bins[i + 1]) if i == n_bins - 1 else (p < bins[i + 1]))
        if np.any(m):
            out += (np.sum(m) / len(y)) * abs(float(np.mean(y[m])) - float(np.mean(p[m])))
    return float(out)


def top5_metrics(y: np.ndarray, p: np.ndarray) -> Dict[str, Any]:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    n = len(y)
    if n == 0:
        return {"top_5pct_n": 0, "top_5pct_event_rate": np.nan, "top_5pct_lift": np.nan}
    k = max(1, int(math.ceil(n * 0.05)))
    idx = np.argsort(-p)[:k]
    prevalence = float(np.mean(y))
    rate = float(np.mean(y[idx]))
    return {
        "top_5pct_n": int(k),
        "top_5pct_event_rate": rate,
        "top_5pct_lift": rate / prevalence if prevalence > 0 else np.nan,
        "top_5pct_captured_events": float(np.sum(y[idx]) / np.sum(y)) if np.sum(y) > 0 else np.nan,
    }


def sanitize_lgbm_params(params: Dict[str, Any]) -> Dict[str, Any]:
    out = {}
    for k, v in params.items():
        if k in LGBM_INT_PARAMS:
            out[k] = int(round(float(v)))
        elif k in LGBM_FLOAT_PARAMS:
            out[k] = float(v)
        elif isinstance(v, (np.integer,)):
            out[k] = int(v)
        elif isinstance(v, (np.floating,)):
            out[k] = float(v)
        else:
            out[k] = v
    return out


def make_weights(y: np.ndarray, multiplier: float) -> np.ndarray:
    y = np.asarray(y).astype(int)
    n_pos = int(np.sum(y == 1))
    n_neg = int(np.sum(y == 0))
    if n_pos == 0:
        raise ValueError("No positive events in training data.")
    positive_weight = (n_neg / max(n_pos, 1)) * float(multiplier)
    return np.where(y == 1, positive_weight, 1.0).astype(float)


def actual_positive_weight(y: np.ndarray, multiplier: float) -> float:
    y = np.asarray(y).astype(int)
    return float((np.sum(y == 0) / max(np.sum(y == 1), 1)) * float(multiplier))

# ============================================================
# Preprocessor
# ============================================================

@dataclass
class Step1Preprocessor:
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str, List[str]]

    def __post_init__(self):
        self.continuous_imputer = None
        self.discrete_imputer = None
        self.nominal_imputer = None
        self.onehot = None
        self.output_feature_names_ = []

    def _parse_binary(self, x: Any, feature: str) -> float:
        sx = norm_text(x)
        if sx is None:
            return np.nan
        if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
            return float(BINARY_MAPS[feature][sx])
        if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive", "performed"}:
            return 1.0
        if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative", "not performed"}:
            return 0.0
        try:
            v = float(sx)
            return float(v) if v in (0.0, 1.0) else np.nan
        except Exception:
            return np.nan

    def _parse_ordinal(self, x: Any, feature: str) -> float:
        sx = norm_text(x)
        if sx is None:
            return np.nan
        if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]:
            return float(ORDINAL_MAPS[feature][sx])
        try:
            v = float(sx)
            if feature == "asa":
                return float(min(max(int(round(v)), 1), 4))
            if feature == "number_operated_levels":
                return float(min(max(int(round(v)), 0), 4))
            if feature == "diabetes_status":
                return float(min(max(int(round(v)), 0), 2))
            if feature == "institution_size":
                return float(min(max(int(round(v)), 0), 2))
            return float(v)
        except Exception:
            return np.nan

    def _canonical_nominal(self, feature: str, x: Any) -> Any:
        x = clean_scalar(x)
        if pd.isna(x):
            return np.nan
        s = str(x).strip()
        sl = s.lower()
        if feature == "race":
            if sl == "white":
                return "White"
            if sl == "black":
                return "Black"
            return "Other"
        if feature == "institution_region":
            mapping = {
                "south": "South",
                "north east": "North East",
                "northeast": "North East",
                "north-east": "North East",
                "west": "West",
                "midwest": "Midwest",
                "mid west": "Midwest",
            }
            return mapping.get(sl, s)
        if feature == "payer_status":
            if sl == "medicare":
                return "Medicare"
            if sl in {"commercial/private", "commercial", "private", "commercial private"}:
                return "Commercial/Private"
            if sl in {"medicaid/public/government", "medicaid", "public", "government", "public/government"}:
                return "Medicaid/Public/Government"
            return "Other"
        if feature == "PatTobaccoUse":
            if sl == "never":
                return "Never"
            if sl == "former":
                return "Former"
            if sl == "current":
                return "Current"
            return "Unknown/Not reported/Multiple"
        return s

    def _parts(self, X: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        cont = pd.DataFrame(index=X.index)
        discrete = pd.DataFrame(index=X.index)
        nominal = pd.DataFrame(index=X.index)
        for c in self.continuous_features:
            cont[c] = pd.to_numeric(X[c].map(clean_scalar), errors="coerce")
        for c in self.binary_features:
            discrete[c] = X[c].map(lambda z: self._parse_binary(z, c)).astype(float)
        for c in self.ordinal_features:
            discrete[c] = X[c].map(lambda z: self._parse_ordinal(z, c)).astype(float)
        for c in self.nominal_features:
            nominal[c] = X[c].map(lambda z: self._canonical_nominal(c, z)).astype("object")
        return cont, discrete, nominal

    def fit(self, X: pd.DataFrame):
        cont, discrete, nominal = self._parts(X)
        self.continuous_imputer = SimpleImputer(strategy="median")
        self.discrete_imputer = SimpleImputer(strategy="most_frequent")
        self.nominal_imputer = SimpleImputer(strategy="constant", fill_value="Missing")
        self.continuous_imputer.fit(cont)
        self.discrete_imputer.fit(discrete)
        nominal_imp = pd.DataFrame(self.nominal_imputer.fit_transform(nominal), columns=self.nominal_features)
        categories = []
        for c in self.nominal_features:
            preferred = list(self.preferred_nominal_levels.get(c, []))
            observed = nominal_imp[c].astype(str).unique().tolist()
            categories.append(preferred + sorted([x for x in observed if x not in preferred]))
        try:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse_output=False)
        except TypeError:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse=False)
        self.onehot.fit(nominal_imp.astype(str))
        self.output_feature_names_ = (
            self.continuous_features
            + self.binary_features
            + self.ordinal_features
            + self.onehot.get_feature_names_out(self.nominal_features).tolist()
        )
        return self

    def transform(self, X: pd.DataFrame) -> np.ndarray:
        cont, discrete, nominal = self._parts(X)
        cont_imp = self.continuous_imputer.transform(cont)
        discrete_imp = self.discrete_imputer.transform(discrete)
        nominal_imp = pd.DataFrame(self.nominal_imputer.transform(nominal), columns=self.nominal_features)
        nominal_oh = self.onehot.transform(nominal_imp.astype(str))
        return np.concatenate([cont_imp, discrete_imp, nominal_oh], axis=1).astype(float)

    def fit_transform(self, X: pd.DataFrame) -> np.ndarray:
        self.fit(X)
        return self.transform(X)

    @property
    def output_feature_names(self) -> List[str]:
        return self.output_feature_names_


def feature_types(features: List[str]) -> Tuple[List[str], List[str], List[str], List[str]]:
    continuous = [f for f in features if f in CONTINUOUS_BASELINE_FEATURES or f not in BASELINE_FEATURES]
    binary = [f for f in features if f in BINARY_FEATURES]
    ordinal = [f for f in features if f in ORDINAL_FEATURES]
    nominal = [f for f in features if f in NOMINAL_FEATURES]
    return continuous, binary, ordinal, nominal


def make_preprocessor(features: List[str]) -> Step1Preprocessor:
    continuous, binary, ordinal, nominal = feature_types(features)
    return Step1Preprocessor(continuous, binary, ordinal, nominal, PREFERRED_NOMINAL_LEVELS)

# ============================================================
# Data and hospital split
# ============================================================

def find_odi_input_csv() -> str:
    for root in [BASE_DIR, FALLBACK_DIR]:
        for fn in ODI_INPUT_CANDIDATES:
            path = os.path.join(root, fn)
            if os.path.exists(path):
                if root == FALLBACK_DIR:
                    dst = os.path.join(BASE_DIR, fn)
                    if not os.path.exists(dst):
                        shutil.copy2(path, dst)
                    return dst
                return path
    raise FileNotFoundError(f"Could not find ODI cohort input file. Tried: {ODI_INPUT_CANDIDATES}")


def load_step1_odi_cohort() -> Tuple[pd.DataFrame, Dict[str, Any]]:
    path = find_odi_input_csv()
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if GROUP_COL not in df.columns:
        raise ValueError(f"ODI cohort file does not contain {GROUP_COL}.")

    target_col = find_existing_column(df.columns.tolist(), TARGET_COL_CANDIDATES, "Step 1 target column")
    prom_col = find_existing_column(df.columns.tolist(), PROM_COL_CANDIDATES, "preoperative ODI column")
    hospital_id_col = find_existing_column(df.columns.tolist(), HOSPITAL_ID_CANDIDATES, "hospital ID column")
    hospital_label_col = find_existing_column(df.columns.tolist(), HOSPITAL_LABEL_CANDIDATES, "hospital label column", required=False)

    missing_features = [f for f in BASELINE_FEATURES if f not in df.columns]
    if missing_features:
        raise ValueError(f"ODI cohort file is missing baseline feature columns: {missing_features}")

    work = df.copy()
    work[target_col] = work[target_col].map(to_binary_target)
    work["preop_ODI"] = pd.to_numeric(work[prom_col].map(clean_scalar), errors="coerce")
    work["__hospital_id__"] = work[hospital_id_col].map(clean_scalar)
    if hospital_label_col and hospital_label_col in work.columns:
        work["__hospital_label__"] = work[hospital_label_col].map(clean_scalar).astype("object")
    else:
        work["__hospital_label__"] = work["__hospital_id__"]

    keep = (
        work[GROUP_COL].notna()
        & work[target_col].isin([0.0, 1.0])
        & work["preop_ODI"].notna()
        & work["__hospital_id__"].notna()
    )
    work = work.loc[keep].copy().reset_index(drop=True)
    work[target_col] = work[target_col].astype(int)
    work["__hospital_id__"] = work["__hospital_id__"].astype(str)
    work["__hospital_label__"] = work["__hospital_label__"].astype(str)

    audit = {
        "input_csv": path,
        "target_column": target_col,
        "preoperative_ODI_column": prom_col,
        "hospital_id_column": hospital_id_col,
        "hospital_label_column": hospital_label_col if hospital_label_col else "none",
        "n_rows_after_required_fields": int(len(work)),
        "n_patients_after_required_fields": int(work[GROUP_COL].nunique()),
        "n_hospitals_after_required_fields": int(work["__hospital_id__"].nunique()),
        "events_after_required_fields": int(work[target_col].sum()),
        "event_rate_after_required_fields": float(work[target_col].mean()),
    }
    return work, audit


def choose_hospital_split(df: pd.DataFrame, target_col: str, seed: int) -> Tuple[pd.Series, pd.DataFrame, pd.DataFrame]:
    hospital_stats = (
        df.groupby("__hospital_id__", dropna=False)
        .agg(n=(target_col, "size"), events=(target_col, "sum"), label=("__hospital_label__", "first"))
        .reset_index()
    )
    hospital_ids = hospital_stats["__hospital_id__"].astype(str).to_numpy()
    if len(hospital_ids) < 3:
        raise RuntimeError("Hospital-held-out validation requires at least 3 hospitals with usable data.")

    rng = np.random.default_rng(seed)
    n_hospitals = len(hospital_ids)
    n_test = max(1, int(round(HOSPITAL_HELDOUT_TEST_FRACTION * n_hospitals)))
    best = None
    best_score = -np.inf

    def counts_for_map(split_map: Dict[str, str]) -> Dict[str, Tuple[int, int, int]]:
        tmp = hospital_stats.copy()
        tmp["split"] = tmp["__hospital_id__"].astype(str).map(split_map)
        out = {}
        for split_name in ["train", "calibration", "hospital_heldout_test"]:
            d = tmp[tmp["split"].eq(split_name)]
            out[split_name] = (int(len(d)), int(d["n"].sum()), int(d["events"].sum()))
        return out

    for attempt in range(HOSPITAL_HELDOUT_SPLIT_ATTEMPTS):
        perm = rng.permutation(hospital_ids)
        test_ids = set(perm[:n_test])
        remaining = np.array([h for h in perm if h not in test_ids], dtype=object)
        n_cal = max(1, int(round(HOSPITAL_HELDOUT_CALIBRATION_FRACTION_OF_REMAINING * len(remaining))))
        cal_ids = set(remaining[:n_cal])
        train_ids = set([h for h in remaining if h not in cal_ids])
        split_map = {h: "train" for h in train_ids}
        split_map.update({h: "calibration" for h in cal_ids})
        split_map.update({h: "hospital_heldout_test" for h in test_ids})
        counts = counts_for_map(split_map)
        ok = (
            counts["train"][2] >= HOSPITAL_HELDOUT_MIN_TRAIN_EVENTS
            and counts["calibration"][2] >= HOSPITAL_HELDOUT_MIN_CALIBRATION_EVENTS
            and counts["hospital_heldout_test"][2] >= HOSPITAL_HELDOUT_MIN_TEST_EVENTS
        )
        score = min(
            counts["train"][2] / max(HOSPITAL_HELDOUT_MIN_TRAIN_EVENTS, 1),
            counts["calibration"][2] / max(HOSPITAL_HELDOUT_MIN_CALIBRATION_EVENTS, 1),
            counts["hospital_heldout_test"][2] / max(HOSPITAL_HELDOUT_MIN_TEST_EVENTS, 1),
        )
        if score > best_score:
            best = (split_map, counts, attempt, ok)
            best_score = score
        if ok:
            break

    split_map, counts, attempt, ok = best
    split_series = df["__hospital_id__"].astype(str).map(split_map)

    split_audit_rows = []
    for split_name, (n_hosp, n_rows, n_events) in counts.items():
        split_audit_rows.append({
            "split": split_name,
            "n_hospitals": n_hosp,
            "n_rows": n_rows,
            "events": n_events,
            "selection_attempt": int(attempt),
            "met_minimum_event_criteria": bool(ok),
            "minimum_train_events": HOSPITAL_HELDOUT_MIN_TRAIN_EVENTS,
            "minimum_calibration_events": HOSPITAL_HELDOUT_MIN_CALIBRATION_EVENTS,
            "minimum_test_events": HOSPITAL_HELDOUT_MIN_TEST_EVENTS,
        })
    split_audit = pd.DataFrame(split_audit_rows)
    hospital_stats["hospital_heldout_split"] = hospital_stats["__hospital_id__"].astype(str).map(split_map)
    hospital_stats = hospital_stats.sort_values(["hospital_heldout_split", "events", "n"], ascending=[True, False, False])
    return split_series, split_audit, hospital_stats

# ============================================================
# Model development
# ============================================================

def cv_splits(y: np.ndarray, groups: np.ndarray, seed: int, n_folds: int = N_CV_FOLDS):
    group_df = pd.DataFrame({"g": groups, "y": y}).groupby("g", dropna=False)["y"].max().reset_index()
    n_folds_eff = min(n_folds, int(np.sum(group_df["y"].eq(1))), int(np.sum(group_df["y"].eq(0))))
    if n_folds_eff < 2:
        raise ValueError("Not enough patient groups/events for group-aware CV.")
    cv = StratifiedGroupKFold(n_splits=n_folds_eff, shuffle=True, random_state=seed)
    return [(tr, va) for tr, va in cv.split(np.zeros(len(y)), y, groups)]


def make_model(params: Dict[str, Any], seed: int, n_estimators_override: Optional[int] = None) -> LGBMClassifier:
    p = sanitize_lgbm_params(params)
    model_params = {k: v for k, v in p.items() if k != "positive_weight_multiplier"}
    if n_estimators_override is not None:
        model_params["n_estimators"] = int(max(MIN_FINAL_N_ESTIMATORS, n_estimators_override))
    return LGBMClassifier(
        objective="binary",
        boosting_type="gbdt",
        metric="average_precision",
        random_state=seed,
        n_jobs=N_JOBS,
        verbosity=-1,
        force_col_wise=True,
        **model_params,
    )


def fit_pipeline(
    X: pd.DataFrame,
    y: np.ndarray,
    features: List[str],
    params: Dict[str, Any],
    seed: int,
    eval_set: Optional[Tuple[pd.DataFrame, np.ndarray]] = None,
    early: bool = False,
    n_estimators_override: Optional[int] = None,
):
    pre = make_preprocessor(features)
    X_enc = pre.fit_transform(X[features])
    weights = make_weights(y, params["positive_weight_multiplier"])
    model = make_model(params, seed, n_estimators_override=n_estimators_override)
    best_iteration = None
    if eval_set is not None and early:
        X_val, y_val = eval_set
        X_val_enc = pre.transform(X_val[features])
        callbacks = [lgb.early_stopping(EARLY_STOPPING_ROUNDS, verbose=False), lgb.log_evaluation(period=0)]
        try:
            model.fit(X_enc, y, sample_weight=weights, eval_set=[(X_val_enc, y_val)], eval_metric="average_precision", callbacks=callbacks)
            best_iteration = int(model.best_iteration_) if getattr(model, "best_iteration_", None) else None
        except Exception:
            model.fit(X_enc, y, sample_weight=weights)
    else:
        model.fit(X_enc, y, sample_weight=weights)
    return pre, model, best_iteration


def predict(pre: Step1Preprocessor, model: LGBMClassifier, X: pd.DataFrame, features: List[str]) -> np.ndarray:
    return model.predict_proba(pre.transform(X[features]))[:, 1]


def tune_lightgbm(
    X_train: pd.DataFrame,
    y_train: np.ndarray,
    groups_train: np.ndarray,
    features: List[str],
    model_key: str,
    seed: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    folds = cv_splits(y_train, groups_train, seed=seed)
    candidate_rows = []
    fold_rows = []
    for candidate_id, raw_params in enumerate(PARAMETER_CANDIDATES, start=1):
        params = sanitize_lgbm_params(raw_params)
        aps = []
        aucs = []
        best_iterations = []
        t0 = time.time()
        for fold_id, (tr_idx, val_idx) in enumerate(folds, start=1):
            X_tr = X_train.iloc[tr_idx].reset_index(drop=True)
            y_tr = y_train[tr_idx]
            X_val = X_train.iloc[val_idx].reset_index(drop=True)
            y_val = y_train[val_idx]
            pre, model, best_iter = fit_pipeline(
                X_tr,
                y_tr,
                features,
                params,
                seed + candidate_id * 1000 + fold_id,
                eval_set=(X_val, y_val),
                early=USE_EARLY_STOPPING_IN_CV,
            )
            p_val = predict(pre, model, X_val, features)
            ap = safe_average_precision(y_val, p_val)
            auc = safe_roc_auc(y_val, p_val)
            aps.append(ap)
            aucs.append(auc)
            if best_iter and best_iter > 0:
                best_iterations.append(best_iter)
            fold_rows.append({
                "model_key": model_key,
                "candidate_id": candidate_id,
                "fold": fold_id,
                "fold_AP": ap,
                "fold_ROC_AUC": auc,
                "fold_best_iteration": best_iter,
                "positive_weight_used": actual_positive_weight(y_tr, params["positive_weight_multiplier"]),
                **params,
            })
        locked_n = int(np.median(best_iterations)) if best_iterations else int(params["n_estimators"])
        row = {
            "model_key": model_key,
            "candidate_id": candidate_id,
            "cv_folds": len(folds),
            "cv_AP_mean": float(np.nanmean(aps)),
            "cv_AP_SD": float(np.nanstd(aps, ddof=1)) if len(aps) > 1 else np.nan,
            "cv_ROC_AUC_mean": float(np.nanmean(aucs)),
            "cv_ROC_AUC_SD": float(np.nanstd(aucs, ddof=1)) if len(aucs) > 1 else np.nan,
            "locked_n_estimators_from_cv": int(locked_n),
            "elapsed_seconds": float(time.time() - t0),
            **params,
        }
        candidate_rows.append(row)
        print(f"{model_key} | candidate {candidate_id:03d}/{len(PARAMETER_CANDIDATES)} | CV AP={row['cv_AP_mean']:.5f} | locked_n={locked_n}")
    candidate_df = pd.DataFrame(candidate_rows).sort_values("cv_AP_mean", ascending=False).reset_index(drop=True)
    fold_df = pd.DataFrame(fold_rows)
    return candidate_df, fold_df


def fit_final_and_predict(
    train_df: pd.DataFrame,
    cal_df: pd.DataFrame,
    test_df: pd.DataFrame,
    features: List[str],
    target_col: str,
    params: Dict[str, Any],
    locked_n_estimators: int,
    seed: int,
) -> Dict[str, Any]:
    y_train = train_df[target_col].astype(int).to_numpy()
    y_cal = cal_df[target_col].astype(int).to_numpy()
    pre, model, _ = fit_pipeline(
        train_df,
        y_train,
        features,
        params,
        seed,
        n_estimators_override=locked_n_estimators,
    )
    p_train_raw = predict(pre, model, train_df, features)
    p_cal_raw = predict(pre, model, cal_df, features)
    p_test_raw = predict(pre, model, test_df, features)

    if len(np.unique(y_cal)) < 2:
        train_rate = float(np.mean(y_train))
        p_train = np.full_like(p_train_raw, train_rate, dtype=float)
        p_cal = np.full_like(p_cal_raw, train_rate, dtype=float)
        p_test = np.full_like(p_test_raw, train_rate, dtype=float)
        calibrator = None
        calibration_method = "constant_train_event_rate_due_to_single_class_calibration_split"
    else:
        calibrator = IsotonicRegression(out_of_bounds="clip")
        calibrator.fit(p_cal_raw, y_cal)
        p_train = np.clip(calibrator.predict(p_train_raw), 0.0, 1.0)
        p_cal = np.clip(calibrator.predict(p_cal_raw), 0.0, 1.0)
        p_test = np.clip(calibrator.predict(p_test_raw), 0.0, 1.0)
        calibration_method = "isotonic_regression_fit_in_hospital_calibration_split"

    return {
        "preprocessor": pre,
        "model": model,
        "calibrator": calibrator,
        "calibration_method": calibration_method,
        "p_train_raw": p_train_raw,
        "p_train": p_train,
        "p_cal_raw": p_cal_raw,
        "p_cal": p_cal,
        "p_test_raw": p_test_raw,
        "p_test": p_test,
        "n_encoded_features": int(len(pre.output_feature_names)),
        "encoded_feature_names": pre.output_feature_names,
    }

# ============================================================
# Evaluation
# ============================================================

def eval_predictions(y: np.ndarray, p: np.ndarray, prefix: str) -> Dict[str, Any]:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    out = {
        f"{prefix}AP": safe_average_precision(y, p),
        f"{prefix}ROC_AUC": safe_roc_auc(y, p),
        f"{prefix}Brier_score": float(brier_score_loss(y, p)) if len(y) else np.nan,
        f"{prefix}ECE_10bins": ece(y, p, n_bins=10),
    }
    out.update({f"{prefix}{k}": v for k, v in top5_metrics(y, p).items()})
    return out


def paired_delta_bootstrap(
    y: np.ndarray,
    p_base: np.ndarray,
    p_expanded: np.ndarray,
    groups: np.ndarray,
    metric: str,
    seed: int,
) -> Tuple[float, float, float, float]:
    y = np.asarray(y).astype(int)
    p_base = np.asarray(p_base).astype(float)
    p_expanded = np.asarray(p_expanded).astype(float)
    obs = (
        safe_average_precision(y, p_expanded) - safe_average_precision(y, p_base)
        if metric == "AP"
        else safe_roc_auc(y, p_expanded) - safe_roc_auc(y, p_base)
    )
    if not np.isfinite(obs) or len(np.unique(y)) < 2:
        return float(obs), np.nan, np.nan, np.nan
    group_df = pd.DataFrame({"idx": np.arange(len(y)), "group": groups, "y": y})
    group_y = group_df.groupby("group", dropna=False)["y"].max()
    pos_groups = group_y[group_y.eq(1)].index.to_numpy()
    neg_groups = group_y[group_y.eq(0)].index.to_numpy()
    if len(pos_groups) == 0 or len(neg_groups) == 0:
        return float(obs), np.nan, np.nan, np.nan
    idx_by_group = {g: group_df.loc[group_df["group"].eq(g), "idx"].to_numpy() for g in group_y.index}
    rng = np.random.default_rng(seed)
    deltas = []
    for _ in range(N_BOOTSTRAPS):
        sampled_groups = np.concatenate([
            rng.choice(pos_groups, size=len(pos_groups), replace=True),
            rng.choice(neg_groups, size=len(neg_groups), replace=True),
        ])
        idx = np.concatenate([idx_by_group[g] for g in sampled_groups])
        yy = y[idx]
        if len(np.unique(yy)) < 2:
            continue
        if metric == "AP":
            delta = safe_average_precision(yy, p_expanded[idx]) - safe_average_precision(yy, p_base[idx])
        else:
            delta = safe_roc_auc(yy, p_expanded[idx]) - safe_roc_auc(yy, p_base[idx])
        if np.isfinite(delta):
            deltas.append(delta)
    if not deltas:
        return float(obs), np.nan, np.nan, np.nan
    deltas = np.asarray(deltas, dtype=float)
    lo, hi = np.percentile(deltas, [2.5, 97.5])
    p_value = 2.0 * min(np.mean(deltas <= 0), np.mean(deltas >= 0))
    return float(obs), float(lo), float(hi), float(min(max(p_value, 0.0), 1.0))


def save_ap_plot(metrics_df: pd.DataFrame, path: str) -> None:
    vals = [float(metrics_df["baseline_AP"].iloc[0]), float(metrics_df["ODI_expanded_AP"].iloc[0])]
    fig, ax = plt.subplots(figsize=(6.8, 4.6))
    ax.bar(["Baseline", "ODI-expanded"], vals)
    ax.set_ylabel("Average precision")
    ax.set_title("Step 1 ODI hospital-held-out validation", fontweight="bold")
    delta = float(metrics_df["delta_AP"].iloc[0])
    y_text = max(vals) * 1.04 if max(vals) > 0 else 0.01
    ax.text(0.5, y_text, f"ΔAP = {delta:+.4f}", ha="center", va="bottom", fontsize=10)
    ax.set_ylim(0, max(vals) * 1.25 if max(vals) > 0 else 0.1)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)

# ============================================================
# Output helpers
# ============================================================

def make_zip(src_dir: str, zip_path: str) -> None:
    if os.path.exists(zip_path):
        os.remove(zip_path)
    tmp_zip = zip_path + ".tmp"
    if os.path.exists(tmp_zip):
        os.remove(tmp_zip)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(src_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(src_dir))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)


def maybe_save_artifact(obj: Any, path: str) -> None:
    try:
        import joblib
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "joblib"])
        import joblib
    joblib.dump(obj, path)

# ============================================================
# Main
# ============================================================

def main() -> None:
    print("=" * 90)
    print("Step 1 ODI hospital-held-out internal validation")
    print("=" * 90)

    data, input_audit = load_step1_odi_cohort()
    target_col = input_audit["target_column"]
    split_series, split_audit, hospital_counts = choose_hospital_split(data, target_col, seed=RANDOM_STATE + 9001)
    data["hospital_heldout_split"] = split_series

    train_df = data[data["hospital_heldout_split"].eq("train")].reset_index(drop=True)
    cal_df = data[data["hospital_heldout_split"].eq("calibration")].reset_index(drop=True)
    test_df = data[data["hospital_heldout_split"].eq("hospital_heldout_test")].reset_index(drop=True)

    if train_df.empty or cal_df.empty or test_df.empty:
        raise RuntimeError("Hospital split produced an empty train, calibration, or held-out test set.")

    baseline_features = [f for f in BASELINE_FEATURES if f in data.columns]
    expanded_features = baseline_features + ["preop_ODI"]

    print("\nTuning baseline-only LightGBM model inside hospital-training split...")
    y_train = train_df[target_col].astype(int).to_numpy()
    groups_train = train_df[GROUP_COL].to_numpy()
    baseline_candidates, baseline_folds = tune_lightgbm(
        train_df,
        y_train,
        groups_train,
        baseline_features,
        model_key="baseline_only",
        seed=RANDOM_STATE + 1000,
    )
    best_baseline = baseline_candidates.iloc[0].to_dict()
    baseline_params = sanitize_lgbm_params({k: best_baseline[k] for k in LGBM_SEARCH_SPACE.keys()})
    baseline_locked_n = int(best_baseline["locked_n_estimators_from_cv"])

    print("\nTuning ODI-expanded LightGBM model inside hospital-training split...")
    expanded_candidates, expanded_folds = tune_lightgbm(
        train_df,
        y_train,
        groups_train,
        expanded_features,
        model_key="ODI_expanded",
        seed=RANDOM_STATE + 2000,
    )
    best_expanded = expanded_candidates.iloc[0].to_dict()
    expanded_params = sanitize_lgbm_params({k: best_expanded[k] for k in LGBM_SEARCH_SPACE.keys()})
    expanded_locked_n = int(best_expanded["locked_n_estimators_from_cv"])

    print("\nFitting final paired models on hospital-training split and calibrating on hospital-calibration split...")
    baseline_final = fit_final_and_predict(
        train_df,
        cal_df,
        test_df,
        baseline_features,
        target_col,
        baseline_params,
        baseline_locked_n,
        seed=RANDOM_STATE + 3000,
    )
    expanded_final = fit_final_and_predict(
        train_df,
        cal_df,
        test_df,
        expanded_features,
        target_col,
        expanded_params,
        expanded_locked_n,
        seed=RANDOM_STATE + 4000,
    )

    y_test = test_df[target_col].astype(int).to_numpy()
    p_base = baseline_final["p_test"]
    p_exp = expanded_final["p_test"]
    test_groups = test_df[GROUP_COL].to_numpy()

    base_metrics = eval_predictions(y_test, p_base, prefix="baseline_")
    exp_metrics = eval_predictions(y_test, p_exp, prefix="ODI_expanded_")
    delta_ap, delta_ap_lo, delta_ap_hi, delta_ap_p = paired_delta_bootstrap(
        y_test, p_base, p_exp, test_groups, metric="AP", seed=RANDOM_STATE + 5000
    )
    delta_auc, delta_auc_lo, delta_auc_hi, delta_auc_p = paired_delta_bootstrap(
        y_test, p_base, p_exp, test_groups, metric="ROC_AUC", seed=RANDOM_STATE + 6000
    )

    metrics = {
        "analysis": "Step 1 ODI hospital-held-out internal validation",
        "n_train": int(len(train_df)),
        "events_train": int(train_df[target_col].sum()),
        "n_train_hospitals": int(train_df["__hospital_id__"].nunique()),
        "n_calibration": int(len(cal_df)),
        "events_calibration": int(cal_df[target_col].sum()),
        "n_calibration_hospitals": int(cal_df["__hospital_id__"].nunique()),
        "n_heldout_test": int(len(test_df)),
        "events_heldout_test": int(test_df[target_col].sum()),
        "heldout_test_event_rate": float(np.mean(y_test)) if len(y_test) else np.nan,
        "n_heldout_test_hospitals": int(test_df["__hospital_id__"].nunique()),
        "baseline_AP": base_metrics["baseline_AP"],
        "ODI_expanded_AP": exp_metrics["ODI_expanded_AP"],
        "delta_AP": delta_ap,
        "delta_AP_CI_lower": delta_ap_lo,
        "delta_AP_CI_upper": delta_ap_hi,
        "delta_AP_p_value_two_sided": delta_ap_p,
        "baseline_ROC_AUC": base_metrics["baseline_ROC_AUC"],
        "ODI_expanded_ROC_AUC": exp_metrics["ODI_expanded_ROC_AUC"],
        "delta_ROC_AUC": delta_auc,
        "delta_ROC_AUC_CI_lower": delta_auc_lo,
        "delta_ROC_AUC_CI_upper": delta_auc_hi,
        "delta_ROC_AUC_p_value_two_sided": delta_auc_p,
        "baseline_Brier_score": base_metrics["baseline_Brier_score"],
        "ODI_expanded_Brier_score": exp_metrics["ODI_expanded_Brier_score"],
        "baseline_ECE_10bins": base_metrics["baseline_ECE_10bins"],
        "ODI_expanded_ECE_10bins": exp_metrics["ODI_expanded_ECE_10bins"],
        "baseline_top_5pct_event_rate": base_metrics["baseline_top_5pct_event_rate"],
        "ODI_expanded_top_5pct_event_rate": exp_metrics["ODI_expanded_top_5pct_event_rate"],
        "baseline_top_5pct_lift": base_metrics["baseline_top_5pct_lift"],
        "ODI_expanded_top_5pct_lift": exp_metrics["ODI_expanded_top_5pct_lift"],
        "baseline_calibration_method": baseline_final["calibration_method"],
        "ODI_expanded_calibration_method": expanded_final["calibration_method"],
        "baseline_best_cv_AP": float(best_baseline["cv_AP_mean"]),
        "ODI_expanded_best_cv_AP": float(best_expanded["cv_AP_mean"]),
        "baseline_locked_n_estimators": baseline_locked_n,
        "ODI_expanded_locked_n_estimators": expanded_locked_n,
        "baseline_n_encoded_features": baseline_final["n_encoded_features"],
        "ODI_expanded_n_encoded_features": expanded_final["n_encoded_features"],
    }
    metrics_df = pd.DataFrame([metrics])

    predictions = test_df[[GROUP_COL, target_col, "preop_ODI", "__hospital_id__", "__hospital_label__", "hospital_heldout_split"]].copy()
    predictions["row_index_in_heldout_test"] = np.arange(len(predictions))
    predictions["p_baseline_raw"] = baseline_final["p_test_raw"]
    predictions["p_baseline_calibrated"] = p_base
    predictions["p_ODI_expanded_raw"] = expanded_final["p_test_raw"]
    predictions["p_ODI_expanded_calibrated"] = p_exp

    print("\nSaving outputs...")
    metrics_path = os.path.join(TABLE_DIR, "Step1_ODI_hospital-held-out_metrics.csv")
    predictions_path = os.path.join(TABLE_DIR, "Step1_ODI_hospital-held-out_predictions.csv")
    baseline_candidates_path = os.path.join(TABLE_DIR, "baseline_tuning_candidates.csv")
    expanded_candidates_path = os.path.join(TABLE_DIR, "ODI_expanded_tuning_candidates.csv")
    baseline_folds_path = os.path.join(TABLE_DIR, "baseline_tuning_folds.csv")
    expanded_folds_path = os.path.join(TABLE_DIR, "ODI_expanded_tuning_folds.csv")
    split_audit_path = os.path.join(AUDIT_DIR, "hospital_split_audit.csv")
    hospital_counts_path = os.path.join(AUDIT_DIR, "hospital_counts_by_split.csv")
    input_audit_path = os.path.join(AUDIT_DIR, "input_audit.json")

    metrics_df.to_csv(metrics_path, index=False)
    predictions.to_csv(predictions_path, index=False)
    baseline_candidates.to_csv(baseline_candidates_path, index=False)
    expanded_candidates.to_csv(expanded_candidates_path, index=False)
    baseline_folds.to_csv(baseline_folds_path, index=False)
    expanded_folds.to_csv(expanded_folds_path, index=False)
    split_audit.to_csv(split_audit_path, index=False)
    hospital_counts.to_csv(hospital_counts_path, index=False)
    with open(input_audit_path, "w") as f:
        json.dump(input_audit, f, indent=2)

    # Save final model artifacts for reproducibility.
    maybe_save_artifact(
        {
            "preprocessor": baseline_final["preprocessor"],
            "model": baseline_final["model"],
            "calibrator": baseline_final["calibrator"],
            "features": baseline_features,
            "params": baseline_params,
            "locked_n_estimators": baseline_locked_n,
        },
        os.path.join(MODEL_DIR, "baseline_hospital_heldout_model.joblib"),
    )
    maybe_save_artifact(
        {
            "preprocessor": expanded_final["preprocessor"],
            "model": expanded_final["model"],
            "calibrator": expanded_final["calibrator"],
            "features": expanded_features,
            "params": expanded_params,
            "locked_n_estimators": expanded_locked_n,
        },
        os.path.join(MODEL_DIR, "ODI_expanded_hospital_heldout_model.joblib"),
    )

    plot_path = os.path.join(PLOT_DIR, "Step1_ODI_hospital-held-out_AP_comparison.png")
    save_ap_plot(metrics_df, plot_path)

    summary_xlsx = os.path.join(OUTPUT_DIR, "Step1_ODI_HospitalHeldOut_InternalValidation_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        metrics_df.to_excel(writer, sheet_name="metrics", index=False)
        predictions.to_excel(writer, sheet_name="heldout_predictions", index=False)
        split_audit.to_excel(writer, sheet_name="split_audit", index=False)
        hospital_counts.to_excel(writer, sheet_name="hospital_counts", index=False)
        baseline_candidates.head(50).to_excel(writer, sheet_name="baseline_top50_tuning", index=False)
        expanded_candidates.head(50).to_excel(writer, sheet_name="ODI_top50_tuning", index=False)

    manifest = {
        "analysis": "Step 1 ODI hospital-held-out internal validation",
        "input_csv": input_audit["input_csv"],
        "target_column": input_audit["target_column"],
        "preoperative_ODI_column": input_audit["preoperative_ODI_column"],
        "hospital_id_column": input_audit["hospital_id_column"],
        "hospital_label_column": input_audit["hospital_label_column"],
        "heldout_unit": "hospital",
        "test_hospitals_seen_during_training_or_calibration": False,
        "train_calibration_test_design": "Hospitals assigned exclusively to training, calibration, or held-out validation.",
        "hyperparameter_tuning": "Performed separately for baseline-only and ODI-expanded LightGBM models within hospital-training split using group-aware CV and AP.",
        "probability_calibration": "Fitted only in hospital-calibration split.",
        "test_evaluation": "Performed only in held-out hospitals.",
        "threshold_reoptimization": False,
        "n_random_hyperparameter_combinations": N_RANDOM_COMBINATIONS,
        "cv_folds": N_CV_FOLDS,
        "bootstrap_iterations": N_BOOTSTRAPS,
        "outputs": {
            "metrics_csv": metrics_path,
            "predictions_csv": predictions_path,
            "summary_xlsx": summary_xlsx,
            "ap_plot": plot_path,
        },
    }
    with open(os.path.join(OUTPUT_DIR, "run_manifest.json"), "w") as f:
        json.dump(manifest, f, indent=2)

    zip_path = os.path.join(BASE_DIR, "Step1_ODI_HospitalHeldOut_InternalValidation.zip")
    make_zip(OUTPUT_DIR, zip_path)

    print("\nDONE")
    print("Metrics:", metrics_path)
    print("Predictions:", predictions_path)
    print("Summary Excel:", summary_xlsx)
    print("ZIP:", zip_path)

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(f'<p><b>ODI hospital-held-out validation output archive is ready.</b></p><p><a href="/files{zip_path}" download>Download ZIP archive</a></p><p>Path: <code>{zip_path}</code></p>'))
        except Exception:
            pass


if __name__ == "__main__":
    main()

# %% [markdown] Cell 13
# #**Step 1 PROM-threshold survival analysis**

# %% Cell 14
# -*- coding: utf-8 -*-
"""
Step 1 PROM-stratified survival analysis
========================================

This script performs supportive time-to-event analyses for 1-year lumbar
reoperation in the preoperative PROM cohorts. PROM-derived risk strata are
defined using prespecified SHAP-informed thresholds. Kaplan-Meier curves,
log-rank tests, and adjusted Cox proportional hazards models are generated
using the prespecified baseline covariates.
"""

# ============================================================
# Imports
# ============================================================

import os
import re
import sys
import json
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    from lifelines import CoxPHFitter, KaplanMeierFitter
    from lifelines.statistics import logrank_test, proportional_hazard_test
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lifelines"])
    from lifelines import CoxPHFitter, KaplanMeierFitter
    from lifelines.statistics import logrank_test, proportional_hazard_test

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from matplotlib import pyplot as plt
from matplotlib.offsetbox import AnchoredOffsetbox, VPacker, HPacker, TextArea, DrawingArea
from matplotlib.lines import Line2D

warnings.filterwarnings("ignore")


# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"

OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_PROM_Survival_KM_Cox_BinaryContinuous_Final")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, PLOT_DIR, TABLE_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

COX_PENALIZER = 0.0
COX_L1_RATIO = 0.0
MIN_EVENTS_PER_PARAMETER_WARNING = 10.0

TARGET_WINDOW_DAYS = 365.0
LAST_KNOWN_FOLLOWUP_DATE = "2026-03-31"

# Match the primary Step 1 analysis: patients who died before 1-year ascertainment
# are excluded from the primary reported survival analysis. If the cohort CSV has
# already excluded these patients, this setting has no effect.
PRIMARY_EXCLUDE_DEATHS_WITHIN_365 = True

AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True
ZIP_COMPRESSION_LEVEL = 1

GROUP_COL_CANDIDATES = ["PersonKey", "PatientKey"]

REOP_COL_CANDIDATES = [
    "final_reop_step1", "final_reop", "reop_1y", "reop_within_1y",
    "reoperation_1y", "reop365", "reop"
]

REOP_TIME_CANDIDATES = [
    "reoptime", "reop_time", "reoperation_time", "time_to_reoperation",
    "time_to_reoperation_days", "days_to_reoperation", "reoperation_days",
    "days_from_index_to_reoperation", "reop_days", "reoperation_time_days",
]

FOLLOWUP_TIME_CANDIDATES = [
    "followup_days", "follow_up_days", "last_followup_days", "last_follow_up_days",
    "time_to_last_followup_days", "days_to_last_followup", "days_to_last_contact",
    "days_from_index_to_last_followup", "followup_time_days", "censoring_time_days",
]

SURGERY_DATE_CANDIDATES = ["index_surgery_date", "ProcedureDate", "surgery_date", "procedure_date"]

DEATH_DAYS_CANDIDATES = [
    "days_to_death_from_index_surgery", "days_to_death", "death_time_days"
]

SURVIVAL_TIME_COL = "time_to_reoperation_or_censor_days"
SURVIVAL_EVENT_COL = "event_reoperation_within_365d"

COHORTS = {
    "ODI": {
        "file_candidates": ["PROM_ODI_Cohort.csv", "PROM_ODI_90.csv"],
        "prom_candidates": [
            "ODI_preop_value", "preop_odi_within_90d", "preop_ODI",
            "preoperative_ODI", "preop_odi", "ODI", "odi", "baseline_ODI"
        ],
        "threshold": 40.0,
        "display_name": "Preoperative ODI",
        "short_name": "ODI",
        "binary_col": "preop_odi_ge_40",
        "term_label": "Preoperative ODI ≥ 40 vs < 40",
        "continuous_col": "preop_odi_continuous_per_10",
        "continuous_scale": 10.0,
        "continuous_term_label": "Preoperative ODI, per 10-point increase",
    },
    "BackPain": {
        "file_candidates": ["PROM_BackPain_Cohort.csv", "PROM_BackPain_90.csv"],
        "prom_candidates": [
            "back_pain_preop_value", "preop_back_pain_nrs_within_90d",
            "back_pain_nrs_preop_within_90d", "preop_BackPain",
            "preop_backpain", "preop_back_pain_nrs", "preoperative_back_pain_nrs",
            "BackPain", "back_pain_preop", "BackPain_NRS", "back_pain_NRS"
        ],
        "threshold": 3.93,
        "display_name": "Preoperative Back Pain NRS",
        "short_name": "Back Pain NRS",
        "binary_col": "preop_back_pain_nrs_ge_3_93",
        "term_label": "Preoperative Back Pain NRS ≥ 3.93 vs < 3.93",
        "continuous_col": "preop_back_pain_nrs_continuous",
        "continuous_scale": 1.0,
        "continuous_term_label": "Preoperative Back Pain NRS, per 1-point increase",
    },
    "LegPain": {
        "file_candidates": ["PROM_LegPain_Cohort.csv", "PROM_LegPain_90.csv"],
        "prom_candidates": [
            "leg_pain_preop_value", "preop_leg_pain_nrs_within_90d",
            "leg_pain_nrs_preop_within_90d", "preop_LegPain",
            "preop_legpain", "preop_leg_pain_nrs", "preoperative_leg_pain_nrs",
            "LegPain", "leg_pain_preop", "LegPain_NRS", "leg_pain_NRS"
        ],
        "threshold": 3.46,
        "display_name": "Preoperative Leg Pain NRS",
        "short_name": "Leg Pain NRS",
        "binary_col": "preop_leg_pain_nrs_ge_3_46",
        "term_label": "Preoperative Leg Pain NRS ≥ 3.46 vs < 3.46",
        "continuous_col": "preop_leg_pain_nrs_continuous",
        "continuous_scale": 1.0,
        "continuous_term_label": "Preoperative Leg Pain NRS, per 1-point increase",
    },
}


# ============================================================
# Baseline covariates
# ============================================================

BASELINE_FEATURES = [
    "finaldx_degenerative", "finaldx_radicular", "finaldx_stenosis",
    "finaldx_deformity_instability", "finaldx_other_diagnosis",
    "age", "sex", "race", "ethnicity", "cancer_status",
    "chronic_pulmonary_disease", "congestive_heart_failure",
    "connective_tissue_rheumatic_disease",
    "diabetes_status", "myocardial_infarction", "renal_disease",
    "institution_type", "institution_size", "institution_region",
    "asa", "bmi", "payer_status",
    "alif_llif", "corpectomy", "discectomy", "foraminotomy",
    "instrumentation", "laminectomy_posterior_decompression",
    "pelvic_fixation", "plf", "tlif_plif", "other_lumbar_procedure",
    "number_operated_levels", "operative_region_extent", "PatTobaccoUse",
]

BINARY_FEATURES = [
    "finaldx_degenerative", "finaldx_radicular", "finaldx_stenosis",
    "finaldx_deformity_instability", "finaldx_other_diagnosis",
    "sex", "ethnicity", "cancer_status", "chronic_pulmonary_disease",
    "congestive_heart_failure", "connective_tissue_rheumatic_disease",
    "myocardial_infarction", "renal_disease", "institution_type",
    "alif_llif", "corpectomy", "discectomy", "foraminotomy",
    "instrumentation", "laminectomy_posterior_decompression",
    "pelvic_fixation", "plf", "tlif_plif", "other_lumbar_procedure",
    "operative_region_extent",
]

CONTINUOUS_FEATURES = ["age", "bmi"]
ORDINAL_FEATURES = ["diabetes_status", "institution_size", "asa", "number_operated_levels"]
NOMINAL_FEATURES = ["race", "institution_region", "payer_status", "PatTobaccoUse"]

FEATURE_LABELS = {
    "finaldx_degenerative": "Degenerative diagnosis",
    "finaldx_radicular": "Radiculopathy diagnosis",
    "finaldx_stenosis": "Spinal stenosis diagnosis",
    "finaldx_deformity_instability": "Deformity or instability diagnosis",
    "finaldx_other_diagnosis": "Other lumbar diagnosis",
    "age": "Age",
    "bmi": "BMI",
    "sex": "Male sex",
    "race": "Race",
    "ethnicity": "Hispanic ethnicity",
    "cancer_status": "Any cancer",
    "chronic_pulmonary_disease": "Chronic pulmonary disease",
    "congestive_heart_failure": "Congestive heart failure",
    "connective_tissue_rheumatic_disease": "Connective tissue/rheumatic disease",
    "diabetes_status": "Diabetes status",
    "myocardial_infarction": "Myocardial infarction",
    "renal_disease": "Renal disease",
    "institution_type": "Non-hospital institution",
    "institution_size": "Institution size",
    "institution_region": "Institution region",
    "asa": "ASA physical status",
    "payer_status": "Primary payer",
    "PatTobaccoUse": "Tobacco use",
    "alif_llif": "Anterior/lateral lumbar interbody fusion",
    "corpectomy": "Corpectomy",
    "discectomy": "Discectomy",
    "foraminotomy": "Foraminotomy",
    "instrumentation": "Instrumentation",
    "laminectomy_posterior_decompression": "Posterior decompression or laminectomy",
    "pelvic_fixation": "Pelvic fixation",
    "plf": "Posterolateral fusion",
    "tlif_plif": "Posterior/transforaminal lumbar interbody fusion",
    "other_lumbar_procedure": "Other lumbar procedure",
    "number_operated_levels": "Number of operated levels",
    "operative_region_extent": "Extended operative region",
}


# ============================================================
# Parsing helpers
# ============================================================

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

BINARY_MAPS = {
    "sex": {"female": 0, "f": 0, "male": 1, "m": 1},
    "ethnicity": {"non-hispanic": 0, "non hispanic": 0, "hispanic": 1},
    "cancer_status": {"no cancer": 0, "no": 0, "none": 0, "any cancer": 1, "yes": 1, "cancer": 1},
    "institution_type": {"hospital": 0, "non-hospital": 1, "non hospital": 1, "nonhospital": 1},
    "operative_region_extent": {
        "lumbar only": 0,
        "extended_region_involvement": 1,
        "extended region involvement": 1,
        "extended": 1,
        "thoracolumbar extension only": 1,
        "sacrum pelvis extension only": 1,
        "thoracolumbar and sacrum pelvis extension": 1,
    },
}

ORDINAL_MAPS = {
    "diabetes_status": {
        "no": 0, "none": 0, "0": 0,
        "without comp": 1, "without complication": 1, "without complications": 1, "1": 1,
        "with comp": 2, "with complication": 2, "with complications": 2, "2": 2,
    },
    "institution_size": {
        "between 1-99 beds": 0, "1-99": 0,
        "between 100-399 beds": 1, "100-399": 1,
        ">= 400 beds": 2, ">=400 beds": 2, ">=400": 2, ">= 400": 2,
        "unspecified": np.nan,
    },
    "asa": {
        "1": 1, "i": 1, "2": 2, "ii": 2, "3": 3, "iii": 3,
        "4": 4, "iv": 4, ">=4": 4, ">= 4": 4, "5": 4, "v": 4,
    },
    "number_operated_levels": {
        "0": 0, "1": 1, "2": 2, "3": 3, "4": 4,
        ">=4": 4, ">= 4": 4, "5": 4, "6": 4, "7": 4, "8": 4, "9": 4, "10": 4,
    },
}

ORDINAL_BOUNDS = {
    "diabetes_status": (0, 2),
    "institution_size": (0, 2),
    "asa": (1, 4),
    "number_operated_levels": (0, 4),
}


def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def to_binary_target(x: Any) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if sx in {"1", "1.0", "yes", "y", "true", "t"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def parse_binary_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive", "performed"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative", "not performed"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def parse_ordinal_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]:
        return float(ORDINAL_MAPS[feature][sx])
    try:
        v = float(sx)
        lo, hi = ORDINAL_BOUNDS.get(feature, (-np.inf, np.inf))
        return float(min(max(int(round(v)), lo), hi))
    except Exception:
        return np.nan


def canonical_nominal(feature: str, x: Any) -> str:
    x = clean_scalar(x)
    if pd.isna(x):
        return "Missing"
    s = str(x).strip()
    sl = s.lower()

    if feature == "race":
        if sl == "white":
            return "White"
        if sl == "black":
            return "Black"
        return "Other"

    if feature == "institution_region":
        return {
            "south": "South",
            "north east": "North East",
            "northeast": "North East",
            "north-east": "North East",
            "west": "West",
            "midwest": "Midwest",
            "mid west": "Midwest",
        }.get(sl, "Missing" if sl in MISSING_STRINGS else s)

    if feature == "payer_status":
        if sl == "medicare":
            return "Medicare"
        if sl in {"commercial/private", "commercial", "private", "commercial private"}:
            return "Commercial/Private"
        if sl in {"medicaid/public/government", "medicaid", "public", "government", "public/government"}:
            return "Medicaid/Public/Government"
        return "Missing" if sl in MISSING_STRINGS else "Other"

    if feature == "PatTobaccoUse":
        if sl == "never":
            return "Never"
        if sl == "former":
            return "Former"
        if sl == "current":
            return "Current"
        return "Unknown/Not reported/Multiple"

    return s


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le").replace("<", "lt").replace(">", "gt")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def find_existing_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if c.lower() in lookup:
            return lookup[c.lower()]
    return None


def find_input_file(cohort: str) -> str:
    for root in [BASE_DIR, FALLBACK_DIR]:
        for name in COHORTS[cohort]["file_candidates"]:
            p = os.path.join(root, name)
            if os.path.exists(p):
                if root == FALLBACK_DIR:
                    dst = os.path.join(BASE_DIR, name)
                    if not os.path.exists(dst):
                        shutil.copy2(p, dst)
                    return dst
                return p
    raise FileNotFoundError(f"{cohort}: could not find input file. Tried: {COHORTS[cohort]['file_candidates']}")


def choose_group_col(df: pd.DataFrame, cohort: str) -> str:
    for c in GROUP_COL_CANDIDATES:
        if c in df.columns and df[c].notna().any():
            return c
    raise ValueError(f"{cohort}: no patient-level group column found. Tried {GROUP_COL_CANDIDATES}.")


# ============================================================
# Survival outcome construction
# ============================================================

def add_survival_outcome(df: pd.DataFrame, cohort: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    out = df.copy()
    columns = out.columns.tolist()

    target_col = find_existing_column(columns, REOP_COL_CANDIDATES)
    reop_time_col = find_existing_column(columns, REOP_TIME_CANDIDATES)
    followup_col = find_existing_column(columns, FOLLOWUP_TIME_CANDIDATES)
    surgery_date_col = find_existing_column(columns, SURGERY_DATE_CANDIDATES)
    death_days_col = find_existing_column(columns, DEATH_DAYS_CANDIDATES)

    if target_col is None:
        raise ValueError(f"{cohort}: no reoperation target column found. Tried {REOP_COL_CANDIDATES}.")
    if reop_time_col is None:
        raise ValueError(f"{cohort}: no reoperation time column found. Tried {REOP_TIME_CANDIDATES}.")

    target = out[target_col].map(to_binary_target)
    reop_time = pd.to_numeric(out[reop_time_col], errors="coerce")

    # Follow-up/censoring time.
    if followup_col is not None:
        followup = pd.to_numeric(out[followup_col], errors="coerce")
    elif surgery_date_col is not None:
        surgery_date = pd.to_datetime(out[surgery_date_col], errors="coerce")
        followup = (pd.to_datetime(LAST_KNOWN_FOLLOWUP_DATE) - surgery_date).dt.days.astype(float)
    else:
        followup = pd.Series(TARGET_WINDOW_DAYS, index=out.index, dtype=float)

    # Primary death exclusion flag. The primary manuscript analysis excludes
    # patients who died before 1-year reoperation ascertainment. Death-retained
    # sensitivity analysis should be run separately.
    death_days = pd.Series(np.nan, index=out.index, dtype=float)
    if death_days_col is not None:
        death_days = pd.to_numeric(out[death_days_col], errors="coerce")
        death_days = death_days.where(death_days.gt(0), np.nan)

    death_exclusion = pd.Series(False, index=out.index)
    if PRIMARY_EXCLUDE_DEATHS_WITHIN_365 and death_days_col is not None:
        death_exclusion = death_days.notna() & death_days.gt(0) & death_days.le(TARGET_WINDOW_DAYS)

    censor_base = np.minimum(followup, TARGET_WINDOW_DAYS)
    censor_base = pd.Series(censor_base, index=out.index, dtype=float)

    if death_days_col is not None:
        censor_base = pd.concat([censor_base, death_days], axis=1).min(axis=1, skipna=True)

    event = (
        target.eq(1.0)
        & reop_time.notna()
        & reop_time.gt(0)
        & reop_time.le(TARGET_WINDOW_DAYS)
        & reop_time.le(censor_base)
    )

    missing_event_time = target.eq(1.0) & reop_time.isna()
    invalid_target = ~target.isin([0.0, 1.0])

    duration = censor_base.copy()
    duration[event] = reop_time[event]

    out[SURVIVAL_TIME_COL] = duration.astype(float)
    out[SURVIVAL_EVENT_COL] = event.astype(int)
    out["_target_col_used"] = target_col
    out["_reop_time_col_used"] = reop_time_col
    out["_followup_col_used"] = followup_col if followup_col else ("computed_from_surgery_date" if surgery_date_col else "administrative_365_days")
    out["_surgery_date_col_used"] = surgery_date_col if surgery_date_col else ""
    out["_death_days_col_used"] = death_days_col if death_days_col else ""

    before = len(out)
    keep = (
        (~invalid_target)
        & (~missing_event_time)
        & (~death_exclusion)
        & out[SURVIVAL_TIME_COL].notna()
        & np.isfinite(out[SURVIVAL_TIME_COL])
        & out[SURVIVAL_TIME_COL].gt(0)
    )
    out = out.loc[keep].copy()

    audit = pd.DataFrame([
        {"cohort": cohort, "audit_item": "input_rows", "value": int(before), "note": ""},
        {"cohort": cohort, "audit_item": "target_col_used", "value": target_col, "note": ""},
        {"cohort": cohort, "audit_item": "reop_time_col_used", "value": reop_time_col, "note": ""},
        {"cohort": cohort, "audit_item": "followup_col_used", "value": followup_col if followup_col else "none", "note": ""},
        {"cohort": cohort, "audit_item": "surgery_date_col_used", "value": surgery_date_col if surgery_date_col else "none", "note": ""},
        {"cohort": cohort, "audit_item": "death_days_col_used", "value": death_days_col if death_days_col else "none", "note": ""},
        {"cohort": cohort, "audit_item": "rows_after_survival_filter", "value": int(len(out)), "note": ""},
        {"cohort": cohort, "audit_item": "invalid_or_missing_target_excluded", "value": int(invalid_target.sum()), "note": ""},
        {"cohort": cohort, "audit_item": "event_positive_missing_reop_time_excluded", "value": int(missing_event_time.sum()), "note": ""},
        {"cohort": cohort, "audit_item": "deaths_within_365d_excluded_primary", "value": int(death_exclusion.sum()), "note": "PRIMARY_EXCLUDE_DEATHS_WITHIN_365=True"},
        {"cohort": cohort, "audit_item": "events_within_365d", "value": int(out[SURVIVAL_EVENT_COL].sum()), "note": ""},
        {"cohort": cohort, "audit_item": "censored", "value": int((1 - out[SURVIVAL_EVENT_COL]).sum()), "note": ""},
        {"cohort": cohort, "audit_item": "event_rate", "value": float(out[SURVIVAL_EVENT_COL].mean()), "note": ""},
    ])
    return out.reset_index(drop=True), audit


def load_and_prepare_cohort(cohort: str) -> Tuple[pd.DataFrame, str, str, pd.DataFrame]:
    path = find_input_file(cohort)
    df = pd.read_csv(path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]

    group_col = choose_group_col(df, cohort)
    prom_col = find_existing_column(df.columns.tolist(), COHORTS[cohort]["prom_candidates"])
    if prom_col is None:
        raise ValueError(f"{cohort}: no PROM column found. Tried {COHORTS[cohort]['prom_candidates']}.")

    df, survival_audit = add_survival_outcome(df, cohort)

    prom = pd.to_numeric(df[prom_col].map(clean_scalar), errors="coerce")
    before = len(df)
    df = df.loc[prom.notna()].copy()
    prom = pd.to_numeric(df[prom_col].map(clean_scalar), errors="coerce")

    threshold = COHORTS[cohort]["threshold"]
    binary_col = COHORTS[cohort]["binary_col"]

    df["PROM_value_used"] = prom.astype(float)
    df["PROM_threshold_used"] = threshold
    df[binary_col] = (prom >= threshold).astype(int)

    prom_audit = pd.DataFrame([
        {"cohort": cohort, "audit_item": "input_file", "value": path, "note": ""},
        {"cohort": cohort, "audit_item": "group_col_used", "value": group_col, "note": ""},
        {"cohort": cohort, "audit_item": "prom_col_used", "value": prom_col, "note": ""},
        {"cohort": cohort, "audit_item": "prom_threshold", "value": float(threshold), "note": ""},
        {"cohort": cohort, "audit_item": "rows_before_PROM_nonmissing_filter", "value": int(before), "note": ""},
        {"cohort": cohort, "audit_item": "rows_after_PROM_nonmissing_filter", "value": int(len(df)), "note": ""},
        {"cohort": cohort, "audit_item": "PROM_missing_excluded", "value": int(before - len(df)), "note": ""},
        {"cohort": cohort, "audit_item": "unique_patients", "value": int(df[group_col].nunique()), "note": ""},
        {"cohort": cohort, "audit_item": "events_after_PROM_filter", "value": int(df[SURVIVAL_EVENT_COL].sum()), "note": ""},
    ])

    return df.reset_index(drop=True), group_col, prom_col, pd.concat([survival_audit, prom_audit], ignore_index=True)


# ============================================================
# Design matrix
# ============================================================

def build_baseline_design(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    available_features = [f for f in BASELINE_FEATURES if f in df.columns]
    design = pd.DataFrame(index=df.index)
    labels = {}

    for f in available_features:
        if f in CONTINUOUS_FEATURES:
            s = pd.to_numeric(df[f].map(clean_scalar), errors="coerce")
            fill = float(s.median()) if s.notna().any() else 0.0
            design[f] = s.fillna(fill).astype(float)
            labels[f] = FEATURE_LABELS.get(f, f)

        elif f in BINARY_FEATURES:
            s = df[f].map(lambda z: parse_binary_value(z, f)).astype(float)
            mode = s.mode(dropna=True)
            fill = float(mode.iloc[0]) if not mode.empty else 0.0
            design[f] = s.fillna(fill).astype(float)
            labels[f] = FEATURE_LABELS.get(f, f)

        elif f in ORDINAL_FEATURES:
            s = df[f].map(lambda z: parse_ordinal_value(z, f)).astype(float)
            mode = s.mode(dropna=True)
            fill = float(mode.iloc[0]) if not mode.empty else (float(s.median()) if s.notna().any() else 0.0)
            s = s.fillna(fill)
            cat = s.astype("Int64").astype(str).replace("<NA>", "Missing")
            tmp = pd.get_dummies(cat, prefix=f, drop_first=True, dtype=float)
            for c in tmp.columns:
                design[c] = tmp[c]
                labels[c] = f"{FEATURE_LABELS.get(f, f)}: {c.replace(f + '_', '')}"

        elif f in NOMINAL_FEATURES:
            cat = df[f].map(lambda z: canonical_nominal(f, z)).astype(str)
            tmp = pd.get_dummies(cat, prefix=f, drop_first=True, dtype=float)
            for c in tmp.columns:
                design[c] = tmp[c]
                labels[c] = f"{FEATURE_LABELS.get(f, f)}: {c.replace(f + '_', '')}"

    return design, labels


def drop_unstable_columns(X: pd.DataFrame, y: pd.Series, protected_cols: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    audit = []
    keep = []
    protected = set(protected_cols)

    for col in X.columns:
        s = pd.to_numeric(X[col], errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)
        nunique = int(s.nunique(dropna=True))
        event_sum = float(s[y.eq(1)].sum())
        nonevent_sum = float(s[y.eq(0)].sum())

        action = "kept"
        reason = ""

        if nunique <= 1:
            action = "dropped"
            reason = "zero_variance"
        elif col not in protected:
            if event_sum == 0 or nonevent_sum == 0:
                action = "dropped"
                reason = "complete_or_near_separation_non_focal"

        if action == "kept" or col in protected:
            keep.append(col)

        audit.append({
            "covariate": col,
            "nunique": nunique,
            "event_sum": event_sum,
            "nonevent_sum": nonevent_sum,
            "action": "kept" if col in keep else "dropped",
            "reason": reason,
            "protected_focal": col in protected,
        })

    return X[keep].copy(), pd.DataFrame(audit)


# ============================================================
# Plot helpers
# ============================================================

def add_logrank_caption_with_line_keys(ax, p_value: float, label0: str, label1: str, n0: int, e0: int, n1: int, e1: int, loc: str):
    def line_row(color: str, text_value: str):
        da = DrawingArea(28, 10, 0, 0)
        da.add_artist(Line2D([2, 26], [5, 5], color=color, linewidth=2.4))
        ta = TextArea(text_value, textprops={"fontsize": 9})
        return HPacker(children=[da, ta], align="center", pad=0, sep=5)

    p_text = "NA" if not np.isfinite(p_value) else f"{p_value:.3g}"
    title = TextArea(f"Log-rank p = {p_text}", textprops={"fontsize": 9})
    row0 = line_row("tab:blue", f"{label0}")
    row1 = line_row("tab:orange", f"{label1}")
    box = VPacker(children=[title, row0, row1], align="left", pad=0, sep=2)

    anchored = AnchoredOffsetbox(loc=loc, child=box, pad=0.35, borderpad=0.5, frameon=True)
    anchored.patch.set_boxstyle("round,pad=0.35")
    anchored.patch.set_facecolor("white")
    anchored.patch.set_edgecolor("gray")
    anchored.patch.set_alpha(0.92)
    ax.add_artist(anchored)


def plot_km_and_cumulative(df: pd.DataFrame, cohort: str) -> Tuple[pd.DataFrame, List[str]]:
    binary_col = COHORTS[cohort]["binary_col"]
    short = COHORTS[cohort]["short_name"]
    display = COHORTS[cohort]["display_name"]
    threshold = COHORTS[cohort]["threshold"]

    mask0 = df[binary_col].eq(0)
    mask1 = df[binary_col].eq(1)

    label0 = f"{display} < {threshold:g}"
    label1 = f"{display} ≥ {threshold:g}"

    lr = logrank_test(
        df.loc[mask0, SURVIVAL_TIME_COL],
        df.loc[mask1, SURVIVAL_TIME_COL],
        event_observed_A=df.loc[mask0, SURVIVAL_EVENT_COL],
        event_observed_B=df.loc[mask1, SURVIVAL_EVENT_COL],
    )

    km0 = KaplanMeierFitter()
    km1 = KaplanMeierFitter()
    km0.fit(df.loc[mask0, SURVIVAL_TIME_COL], event_observed=df.loc[mask0, SURVIVAL_EVENT_COL], label=label0)
    km1.fit(df.loc[mask1, SURVIVAL_TIME_COL], event_observed=df.loc[mask1, SURVIVAL_EVENT_COL], label=label1)

    plot_paths = []

    # KM reoperation-free survival.
    fig, ax = plt.subplots(figsize=(8.8, 6.0))
    km0.plot_survival_function(ax=ax, ci_show=True, legend=False, color="tab:blue")
    km1.plot_survival_function(ax=ax, ci_show=True, legend=False, color="tab:orange")
    ax.set_title(f"Kaplan–Meier analysis: {display}", fontweight="bold")
    ax.set_xlabel("Days from index surgery to 1-year follow-up")
    ax.set_ylabel("Probability of no reoperation")
    ax.set_xlim(0, TARGET_WINDOW_DAYS)
    ax.grid(alpha=0.2)
    add_logrank_caption_with_line_keys(
        ax, float(lr.p_value), label0, label1,
        int(mask0.sum()), int(df.loc[mask0, SURVIVAL_EVENT_COL].sum()),
        int(mask1.sum()), int(df.loc[mask1, SURVIVAL_EVENT_COL].sum()),
        loc="lower left",
    )
    path = os.path.join(PLOT_DIR, f"KM_{cohort}_preoperative_PROM_threshold_Final.png")
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    plot_paths.append(path)

    # Cumulative reoperation risk.
    fig, ax = plt.subplots(figsize=(8.8, 6.0))
    for km, color in [(km0, "tab:blue"), (km1, "tab:orange")]:
        sf = km.survival_function_
        ci = km.confidence_interval_
        t = sf.index.values
        cum = 1.0 - sf.iloc[:, 0].values
        ax.plot(t, cum, color=color, label=km._label)
        try:
            lower_surv = ci.iloc[:, 0].values
            upper_surv = ci.iloc[:, 1].values
            ax.fill_between(t, 1.0 - upper_surv, 1.0 - lower_surv, alpha=0.15, color=color)
        except Exception:
            pass
    ax.set_title(f"Cumulative reoperation risk: {display}", fontweight="bold")
    ax.set_xlabel("Days from index surgery to 1-year follow-up")
    ax.set_ylabel("Cumulative probability of reoperation")
    ax.set_xlim(0, TARGET_WINDOW_DAYS)
    ax.grid(alpha=0.2)
    add_logrank_caption_with_line_keys(
        ax, float(lr.p_value), label0, label1,
        int(mask0.sum()), int(df.loc[mask0, SURVIVAL_EVENT_COL].sum()),
        int(mask1.sum()), int(df.loc[mask1, SURVIVAL_EVENT_COL].sum()),
        loc="upper left",
    )
    path = os.path.join(PLOT_DIR, f"Cumulative_reoperation_{cohort}_preoperative_PROM_threshold_Final.png")
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    plot_paths.append(path)

    rows = []
    for group_label, mask, km in [(label0, mask0, km0), (label1, mask1, km1)]:
        surv365 = float(km.survival_function_at_times(TARGET_WINDOW_DAYS).iloc[0])
        rows.append({
            "cohort": cohort,
            "PROM": display,
            "threshold": threshold,
            "group": group_label,
            "n": int(mask.sum()),
            "events": int(df.loc[mask, SURVIVAL_EVENT_COL].sum()),
            "KM_survival_365d": surv365,
            "KM_cumulative_reoperation_365d": 1.0 - surv365,
            "logrank_test_statistic": float(lr.test_statistic),
            "logrank_p": float(lr.p_value),
            "KM_plot_path": plot_paths[0],
            "cumulative_plot_path": plot_paths[1],
        })

    return pd.DataFrame(rows), plot_paths


def plot_forest(df: pd.DataFrame, label_col: str, path: str, title: str):
    d = df.copy()
    d = d.replace([np.inf, -np.inf], np.nan)
    d = d.dropna(subset=["HR", "HR_lower_95", "HR_upper_95"])
    if d.empty:
        return

    d = d.iloc[::-1].reset_index(drop=True)
    fig_h = max(4.0, 0.46 * len(d) + 1.6)
    fig, ax = plt.subplots(figsize=(10.0, fig_h))

    y = np.arange(len(d))
    hr = d["HR"].astype(float).to_numpy()
    lo = d["HR_lower_95"].astype(float).to_numpy()
    hi = d["HR_upper_95"].astype(float).to_numpy()

    significant = (lo > 1.0) | (hi < 1.0)

    for yi, h, l, u, is_sig in zip(y, hr, lo, hi, significant):
        ax.errorbar(
            [h],
            [yi],
            xerr=np.array([[h - l], [u - h]]),
            fmt="o",
            color="#2563EB",
            ecolor="#2563EB",
            capsize=4 if is_sig else 3,
            markersize=7.5 if is_sig else 5.5,
            elinewidth=2.6 if is_sig else 1.2,
            markeredgewidth=1.6 if is_sig else 1.0,
        )

    ax.axvline(1.0, color="gray", linestyle="--", linewidth=1.0)
    ax.set_xscale("log")
    ax.set_yticks(y)
    ax.set_yticklabels(d[label_col].tolist())
    ax.set_xlabel("Hazard ratio (log scale)")
    ax.set_title(title, fontweight="bold")

    x_max = max(float(np.nanmax(hi)), 1.1)
    for i, row in d.iterrows():
        is_sig = (float(row["HR_lower_95"]) > 1.0) or (float(row["HR_upper_95"]) < 1.0)
        txt = f"{row['HR']:.2f} ({row['HR_lower_95']:.2f}–{row['HR_upper_95']:.2f}), p={row['p']:.3g}"
        ax.text(
            x_max * 1.05,
            i,
            txt,
            va="center",
            fontsize=8.8,
            fontweight="bold" if is_sig else "normal",
        )

    ax.set_xlim(left=max(min(float(np.nanmin(lo)) * 0.8, 0.8), 0.05), right=x_max * 2.45)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


# ============================================================
# Cox model
# ============================================================

def fit_adjusted_cox(
    df: pd.DataFrame,
    cohort: str,
    prom_form: str = "binary",
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Fit one adjusted Cox model for a Step 1 cohort.

    prom_form="binary" uses the prespecified SHAP-informed binary PROM threshold.
    prom_form="continuous" uses the continuous preoperative PROM value as the
    predictor of interest. ODI is scaled per 10-point increase; pain NRS values
    are scaled per 1-point increase.
    """
    if prom_form not in {"binary", "continuous"}:
        raise ValueError("prom_form must be either 'binary' or 'continuous'.")

    baseline_design, label_map = build_baseline_design(df)
    cfg = COHORTS[cohort]
    design = baseline_design.copy()

    if prom_form == "binary":
        focal_col = cfg["binary_col"]
        design[focal_col] = df[focal_col].astype(float).to_numpy()
        label_map[focal_col] = cfg["term_label"]
        comparison = cfg["term_label"]
        model_suffix = "binary_threshold"
    else:
        focal_col = cfg["continuous_col"]
        scale = float(cfg.get("continuous_scale", 1.0))
        design[focal_col] = pd.to_numeric(df["PROM_value_used"], errors="coerce").astype(float) / scale
        label_map[focal_col] = cfg["continuous_term_label"]
        comparison = cfg["continuous_term_label"]
        model_suffix = "continuous"

    y = df[SURVIVAL_EVENT_COL].astype(int)
    design = design.apply(pd.to_numeric, errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)

    design, unstable_audit = drop_unstable_columns(design, y, protected_cols=[focal_col])
    unstable_audit["cohort"] = cohort
    unstable_audit["PROM_modeling_form"] = prom_form
    unstable_audit["model_name"] = f"{cohort}_{model_suffix}"

    cox_df = pd.concat(
        [df[[SURVIVAL_TIME_COL, SURVIVAL_EVENT_COL]].reset_index(drop=True), design.reset_index(drop=True)],
        axis=1,
    )

    cph = CoxPHFitter(penalizer=COX_PENALIZER, l1_ratio=COX_L1_RATIO)
    cph.fit(cox_df, duration_col=SURVIVAL_TIME_COL, event_col=SURVIVAL_EVENT_COL, robust=True)

    summary = cph.summary.reset_index().rename(columns={"covariate": "term"})
    if "index" in summary.columns and "term" not in summary.columns:
        summary = summary.rename(columns={"index": "term"})

    summary["cohort"] = cohort
    summary["PROM_modeling_form"] = prom_form
    summary["model_name"] = f"{cohort}_{model_suffix}"
    summary["display_label"] = summary["term"].map(lambda x: label_map.get(x, x))
    summary["HR"] = np.exp(summary["coef"])
    summary["HR_lower_95"] = np.exp(summary["coef lower 95%"])
    summary["HR_upper_95"] = np.exp(summary["coef upper 95%"])
    summary["n"] = int(len(df))
    summary["events"] = int(df[SURVIVAL_EVENT_COL].sum())
    summary["n_parameters"] = int(len(design.columns))
    summary["events_per_parameter"] = float(df[SURVIVAL_EVENT_COL].sum() / max(len(design.columns), 1))
    summary["cox_penalizer"] = COX_PENALIZER

    try:
        ph = proportional_hazard_test(cph, cox_df, time_transform="rank")
        ph_df = ph.summary.reset_index().rename(columns={"index": "term"})
        ph_df["cohort"] = cohort
        ph_df["PROM_modeling_form"] = prom_form
        ph_df["model_name"] = f"{cohort}_{model_suffix}"
    except Exception as err:
        ph_df = pd.DataFrame([{
            "cohort": cohort,
            "PROM_modeling_form": prom_form,
            "model_name": f"{cohort}_{model_suffix}",
            "term": "PH_TEST_FAILED",
            "test_statistic": np.nan,
            "p": np.nan,
            "error": repr(err),
        }])

    focal = summary[summary["term"].eq(focal_col)].copy()
    if not focal.empty:
        focal["PROM"] = cfg["display_name"]
        focal["comparison"] = comparison
        focal["PROM_modeling_form"] = prom_form

    epv_audit = pd.DataFrame([{
        "cohort": cohort,
        "PROM_modeling_form": prom_form,
        "model_name": f"{cohort}_{model_suffix}",
        "n": int(len(df)),
        "events": int(df[SURVIVAL_EVENT_COL].sum()),
        "n_parameters_after_dropping_unstable_columns": int(len(design.columns)),
        "events_per_parameter": float(df[SURVIVAL_EVENT_COL].sum() / max(len(design.columns), 1)),
        "warning_events_per_parameter_lt_10": bool((df[SURVIVAL_EVENT_COL].sum() / max(len(design.columns), 1)) < MIN_EVENTS_PER_PARAMETER_WARNING),
        "cox_penalizer": COX_PENALIZER,
        "cox_l1_ratio": COX_L1_RATIO,
    }])

    audit = pd.concat([unstable_audit, epv_audit], ignore_index=True, sort=False)
    return summary, focal, ph_df, audit


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 100)
    print("Step 1 PROM survival analysis: Kaplan–Meier + adjusted Cox")
    print("=" * 100)

    all_audits = []
    all_km = []
    all_cox = []
    all_focal = []
    all_ph = []
    all_cox_audits = []

    for cohort in ["ODI", "BackPain", "LegPain"]:
        print(f"\nProcessing {cohort}...")
        df, group_col, prom_col, audit = load_and_prepare_cohort(cohort)
        all_audits.append(audit)

        print(
            f"{cohort}: n={len(df):,}, events={int(df[SURVIVAL_EVENT_COL].sum()):,}, "
            f"event rate={df[SURVIVAL_EVENT_COL].mean():.3%}, PROM={prom_col}, threshold={COHORTS[cohort]['threshold']}"
        )

        # KM and cumulative curves use the prespecified binary SHAP-informed threshold.
        km_df, _ = plot_km_and_cumulative(df, cohort)
        all_km.append(km_df)

        # Six total PROM-specific Cox analyses are produced across the three cohorts:
        #   3 binary threshold models + 3 continuous PROM models.
        for prom_form in ["binary", "continuous"]:
            cox_summary, focal, ph_df, cox_audit = fit_adjusted_cox(df, cohort, prom_form=prom_form)
            all_cox.append(cox_summary)
            all_focal.append(focal)
            all_ph.append(ph_df)
            all_cox_audits.append(cox_audit)

            sig = cox_summary[
                (cox_summary["p"] < 0.05)
                & cox_summary["HR"].notna()
                & cox_summary["HR_lower_95"].notna()
                & cox_summary["HR_upper_95"].notna()
            ].copy()
            sig = sig.sort_values("p").reset_index(drop=True)
            sig_path = os.path.join(PLOT_DIR, f"Forest_{cohort}_{prom_form}_significant_adjusted_Cox_terms_Final.png")
            if not sig.empty:
                plot_forest(
                    sig.rename(columns={"display_label": "label"}),
                    label_col="label",
                    path=sig_path,
                    title=f"{cohort}: significant adjusted Cox terms ({prom_form} PROM model)",
                )
                sig["forest_plot_path"] = sig_path
            sig.to_csv(os.path.join(TABLE_DIR, f"Cox_{cohort}_{prom_form}_significant_terms_Final.csv"), index=False)

    audit_df = pd.concat(all_audits, ignore_index=True)
    km_summary = pd.concat(all_km, ignore_index=True)
    cox_all = pd.concat(all_cox, ignore_index=True)
    focal_all = pd.concat(all_focal, ignore_index=True)
    ph_all = pd.concat(all_ph, ignore_index=True)
    cox_audit_all = pd.concat(all_cox_audits, ignore_index=True, sort=False)

    # Combined forest plot for all six PROM-specific Cox analyses.
    focal_forest = focal_all.copy()
    focal_forest["label"] = focal_forest["comparison"]
    cohort_order = {"ODI": 1, "BackPain": 2, "LegPain": 3}
    form_order = {"binary": 1, "continuous": 2}
    focal_forest["cohort_order"] = focal_forest["cohort"].map(cohort_order)
    focal_forest["form_order"] = focal_forest["PROM_modeling_form"].map(form_order)
    focal_forest = focal_forest.sort_values(["cohort_order", "form_order"]).drop(columns=["cohort_order", "form_order"])
    focal_forest = focal_forest.reset_index(drop=True)

    combined_prom_forest_path = os.path.join(PLOT_DIR, "Forest_all_six_preoperative_PROM_adjusted_Cox_Final.png")
    plot_forest(
        focal_forest,
        label_col="label",
        path=combined_prom_forest_path,
        title="Step 1 adjusted Cox models: binary and continuous preoperative PROMs",
    )
    focal_forest["forest_plot_path"] = combined_prom_forest_path

    # Save tables.
    audit_df.to_csv(os.path.join(AUDIT_DIR, "cohort_survival_PROM_audit_Final.csv"), index=False)
    km_summary.to_csv(os.path.join(TABLE_DIR, "KM_logrank_summary_all_PROMs_Final.csv"), index=False)
    cox_all.to_csv(os.path.join(TABLE_DIR, "Cox_all_terms_all_PROM_cohorts_Final.csv"), index=False)
    focal_forest.to_csv(os.path.join(TABLE_DIR, "Cox_focal_PROM_HR_all_six_analyses_Final.csv"), index=False)
    ph_all.to_csv(os.path.join(TABLE_DIR, "Schoenfeld_PH_tests_all_PROM_cohorts_Final.csv"), index=False)
    cox_audit_all.to_csv(os.path.join(AUDIT_DIR, "Cox_sparse_EPV_audit_all_PROM_cohorts_Final.csv"), index=False)

    settings = {
        "analysis": "Step 1 PROM survival analysis",
        "time_scale": "days after index surgery",
        "target_window_days": TARGET_WINDOW_DAYS,
        "PROM_thresholds": {k: COHORTS[k]["threshold"] for k in COHORTS},
        "PROM_binary_terms": {k: COHORTS[k]["term_label"] for k in COHORTS},
        "PROM_continuous_terms": {k: COHORTS[k]["continuous_term_label"] for k in COHORTS},
        "continuous_scaling": {k: COHORTS[k]["continuous_scale"] for k in COHORTS},
        "cox_penalizer": COX_PENALIZER,
        "cox_l1_ratio": COX_L1_RATIO,
        "primary_exclude_deaths_within_365": PRIMARY_EXCLUDE_DEATHS_WITHIN_365,
        "baseline_features": BASELINE_FEATURES,
        "note": (
            "Each row in the combined PROM forest plot comes from a separate adjusted Cox model. "
            "For each cohort, the PROM was evaluated twice: once as a binary SHAP-informed threshold "
            "and once as a continuous predictor. The combined forest plot pools focal PROM HRs from "
            "six separate PROM-specific Cox analyses; it is not one joint model across all three datasets."
        ),
    }
    with open(os.path.join(AUDIT_DIR, "analysis_settings_Final.json"), "w") as f:
        json.dump(settings, f, indent=2, sort_keys=True)

    summary_xlsx = os.path.join(OUTPUT_DIR, "Step1_PROM_Survival_KM_Cox_BinaryContinuous_summary_Final.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        km_summary.to_excel(writer, sheet_name="KM_logrank_PROMs", index=False)
        focal_forest.to_excel(writer, sheet_name="focal_PROM_HR_6_models", index=False)
        cox_all.to_excel(writer, sheet_name="Cox_all_terms", index=False)
        ph_all.to_excel(writer, sheet_name="PH_tests", index=False)
        cox_audit_all.to_excel(writer, sheet_name="Cox_EPV_sparse_audit", index=False)
        audit_df.to_excel(writer, sheet_name="cohort_survival_audit", index=False)
        pd.DataFrame([settings]).to_excel(writer, sheet_name="settings", index=False)

    zip_path = os.path.join(BASE_DIR, "Step1_PROM_Survival_KM_Cox_BinaryContinuous_Final.zip")
    tmp_zip = zip_path + ".tmp"
    for p in [zip_path, tmp_zip]:
        if os.path.exists(p):
            os.remove(p)

    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as z:
        for root, _, files in os.walk(OUTPUT_DIR):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(OUTPUT_DIR))
                z.write(full, rel)
    os.replace(tmp_zip, zip_path)

    print("\nDONE")
    print("Output folder:", OUTPUT_DIR)
    print("Summary Excel:", summary_xlsx)
    print("ZIP:", zip_path)
    print("\nFocal PROM HRs from all six Cox analyses:")
    if not focal_forest.empty:
        print(focal_forest[["cohort", "PROM_modeling_form", "comparison", "HR", "HR_lower_95", "HR_upper_95", "p", "n", "events"]].to_string(index=False))

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(
                f'<p><b>Step 1 PROM survival outputs are ready.</b></p>'
                f'<p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p>'
                f'<p>Path: <code>{zip_path}</code></p>'
            ))
        except Exception as e:
            print("Download link display skipped:", repr(e))

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))


if __name__ == "__main__":
    main()

# %% [markdown] Cell 15
# #**ODI_ASR_Grading_DCA_Streamlit_LOCKED_SPLIT_ONLY**

# %% Cell 16
# -*- coding: utf-8 -*-
"""
Step 1 ODI-based ASR reoperation risk grading and Streamlit application: risk-distribution-derived 3-grade version
================================================================================================

This script develops an ODI-specific Step 1 ASR reoperation risk-grade system
using the locked ODI>=40 LightGBM model. Hyperparameters are not retuned.
Three risk strata are derived in the locked calibration split only using tertiles of calibrated predicted risk. Decision-curve
analysis (DCA) is used only to evaluate clinical utility after the grade boundaries are locked.

Note
---------------------------
This version requires the upstream locked split_assignment_ODI.csv generated by
the primary Step 1 LightGBM analysis. It will stop with an error if the locked
split cannot be found. It never creates a new train/calibration/test split.

Outputs include locked model artifacts, calibrated predictions, risk-distribution-derived
risk-grade cutoffs, calibration- and test-set decision-curve tables and plots, held-out test-set
grade summaries, validation plots, and a minimal Streamlit application for
reviewer-facing patient-level risk stratification.
"""

# ============================================================
# Imports
# ============================================================

import os
import re
import sys
import json
import math
import time
import zipfile
import shutil
import warnings
import subprocess
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import lightgbm as lgb  # noqa: F401
    from lightgbm import LGBMClassifier
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lightgbm"])
    import lightgbm as lgb  # noqa: F401
    from lightgbm import LGBMClassifier

try:
    import joblib
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "joblib"])
    import joblib

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder
from sklearn.metrics import average_precision_score, roc_auc_score, brier_score_loss
from sklearn.isotonic import IsotonicRegression
from sklearn.calibration import calibration_curve
from matplotlib import pyplot as plt

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"

OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_ODI_RiskGrade_ODI40_StreamlitApp_RiskTertile3Grades_DCAValidation")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
ARTIFACT_DIR = os.path.join(OUTPUT_DIR, "model_artifacts")
APP_DIR = os.path.join(OUTPUT_DIR, "streamlit_app")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, ARTIFACT_DIR, APP_DIR]:
    os.makedirs(_d, exist_ok=True)

INPUT_CSV_CANDIDATES = [
    os.path.join(BASE_DIR, "PROM_ODI_Cohort.csv"),
    os.path.join(BASE_DIR, "PROM_ODI_Cohort(1).csv"),
    os.path.join(FALLBACK_DIR, "PROM_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "PROM_ODI_Cohort(1).csv"),
]

# Candidate upstream Step 1 LightGBM output folders/archives that may contain
# the locked split assignment. The code requires split_assignment_ODI.csv.
SOURCE_FOLDER_CANDIDATES = [
    "Step1_PROM_LightGBM_outputs",
    "Step1_PROM_LightGBM_Final",
    "Step1_PROM_LightGBM_TopConfig_Final",
]
SOURCE_ARCHIVE_CANDIDATES = [
    "Step1_PROM_LightGBM_outputs.zip",
    "Step1_PROM_LightGBM_Final.zip",
    "Step1_PROM_LightGBM_TopConfig_Final.zip",
]
REQUIRE_LOCKED_SPLIT_ASSIGNMENT = True

GROUP_COL_CANDIDATES = ["PersonKey", "PatientKey"]
TARGET_COL_CANDIDATES = [
    "final_reop_step1", "final_reop_1yr", "reop_1yr", "reoperation_1yr",
    "one_year_reoperation", "reop365", "reop_365", "final_reop",
    "reoperation", "reop",
]
PROM_COL_CANDIDATES = [
    "ODI_preop_value", "preop_odi_within_90d", "odi_preop_within_90d",
    "odi_preop_anytime_value", "preop_odi_anytime", "preop_ODI",
    "Preop_ODI", "preoperative_ODI", "Preoperative_ODI", "ODI", "odi",
    "baseline_ODI",
]

GROUP_COL = "PersonKey"
RANDOM_STATE = 20260524
TEST_FRACTION = 0.20
CALIBRATION_FRACTION_OF_REMAINING = 0.20
ODI_THRESHOLD = 40.0
ODI_BINARY_COL = "preop_ODI_ge_40"
N_BOOTSTRAPS = 2000
ECE_N_BINS = 10
ZIP_COMPRESSION_LEVEL = 1
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True

# Risk-grade cut points are derived only in the calibration split using
# prespecified tertiles of calibrated predicted risk. DCA is used downstream
# only as a clinical-utility validation analysis after the grade boundaries are locked.
N_RISK_GRADES = 3
RISK_GRADE_QUANTILES = (1.0 / 3.0, 2.0 / 3.0)
MIN_PATIENTS_PER_GRADE = 25
DCA_THRESHOLD_GRID = np.arange(0.005, 0.151, 0.001)  # 0.5% to 15%
GRADE_CUTOFF_MODE = "calibration_set_predicted_risk_tertile_boundaries"

GRADE_SCORE_SOURCE = "calibrated_predicted_risk"
GRADE_LABELS = {
    0: "Grade 0 / Low risk",
    1: "Grade 1 / Intermediate risk",
    2: "Grade 2 / High risk",
}
GRADE_SHORT_LABELS = {
    0: "Low risk",
    1: "Intermediate risk",
    2: "High risk",
}

# Locked final Step 1 ODI LightGBM settings from the primary analysis.
LOCKED_ODI_LIGHTGBM_SETTINGS = {
    "candidate_id": 13,
    "cv_AP_mean": 0.22169515127877565,
    "cv_AP_SD": 0.04074959826689169,
    "cv_ROC_AUC_mean": 0.7306789191033995,
    "cv_ROC_AUC_SD": 0.023494265157821013,
    "calibration_method": "isotonic",
    "locked_n_estimators_from_cv": 45,
    "lightgbm_params": {
        "colsample_bytree": 0.6,
        "learning_rate": 0.05,
        "max_bin": 127,
        "max_depth": -1,
        "min_child_samples": 100,
        "min_split_gain": 0.05,
        "n_estimators": 45,
        "num_leaves": 63,
        "reg_alpha": 0.05,
        "reg_lambda": 0.5,
        "subsample": 1.0,
        "subsample_freq": 2,
    },
    "positive_weight_multiplier": 0.5,
}

# Streamlit app files are generated in APP_DIR.

# ============================================================
# Feature definitions
# ============================================================

BASELINE_FEATURES = [
    "finaldx_degenerative", "finaldx_radicular", "finaldx_stenosis",
    "finaldx_deformity_instability", "finaldx_other_diagnosis",
    "age", "sex", "race", "ethnicity", "cancer_status",
    "chronic_pulmonary_disease", "congestive_heart_failure",
    "connective_tissue_rheumatic_disease", "diabetes_status",
    "myocardial_infarction", "renal_disease", "institution_type",
    "institution_size", "institution_region", "asa", "bmi", "payer_status",
    "alif_llif", "corpectomy", "discectomy", "foraminotomy",
    "instrumentation", "laminectomy_posterior_decompression",
    "pelvic_fixation", "plf", "tlif_plif", "other_lumbar_procedure",
    "number_operated_levels", "operative_region_extent", "PatTobaccoUse",
]
assert len(BASELINE_FEATURES) == 35

FEATURES = BASELINE_FEATURES + [ODI_BINARY_COL]

CONTINUOUS_FEATURES = ["age", "bmi"]
BINARY_FEATURES = [
    "finaldx_degenerative", "finaldx_radicular", "finaldx_stenosis",
    "finaldx_deformity_instability", "finaldx_other_diagnosis", "sex",
    "ethnicity", "cancer_status", "chronic_pulmonary_disease",
    "congestive_heart_failure", "connective_tissue_rheumatic_disease",
    "myocardial_infarction", "renal_disease", "institution_type",
    "alif_llif", "corpectomy", "discectomy", "foraminotomy",
    "instrumentation", "laminectomy_posterior_decompression",
    "pelvic_fixation", "plf", "tlif_plif", "other_lumbar_procedure",
    "operative_region_extent", ODI_BINARY_COL,
]
ORDINAL_FEATURES = ["diabetes_status", "institution_size", "asa", "number_operated_levels"]
NOMINAL_FEATURES = ["race", "institution_region", "payer_status", "PatTobaccoUse"]

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

BINARY_MAPS = {
    "sex": {"female": 0, "f": 0, "male": 1, "m": 1},
    "ethnicity": {"non-hispanic": 0, "non hispanic": 0, "hispanic": 1},
    "cancer_status": {"no cancer": 0, "no": 0, "none": 0, "any cancer": 1, "yes": 1, "cancer": 1},
    "institution_type": {"hospital": 0, "non-hospital": 1, "non hospital": 1, "nonhospital": 1},
    "operative_region_extent": {
        "lumbar only": 0,
        "extended_region_involvement": 1,
        "extended region involvement": 1,
        "extended": 1,
        "thoracolumbar extension only": 1,
        "sacrum pelvis extension only": 1,
        "thoracolumbar and sacrum pelvis extension": 1,
    },
}

ORDINAL_MAPS = {
    "diabetes_status": {
        "no": 0, "none": 0, "0": 0,
        "without comp": 1, "without complication": 1, "without complications": 1, "1": 1,
        "with comp": 2, "with complication": 2, "with complications": 2, "2": 2,
    },
    "institution_size": {
        "between 1-99 beds": 0, "1-99": 0,
        "between 100-399 beds": 1, "100-399": 1,
        ">= 400 beds": 2, ">=400 beds": 2, ">=400": 2, ">= 400": 2,
    },
    "asa": {"1": 1, "i": 1, "2": 2, "ii": 2, "3": 3, "iii": 3, "4": 4, "iv": 4, ">=4": 4, ">= 4": 4, "5": 4, "v": 4},
    "number_operated_levels": {"0": 0, "1": 1, "2": 2, "3": 3, "4": 4, ">=4": 4, ">= 4": 4, "5": 4, "6": 4, "7": 4, "8": 4, "9": 4, "10": 4},
}

PREFERRED_NOMINAL_LEVELS = {
    "race": ["White", "Black", "Other"],
    "institution_region": ["South", "North East", "West", "Midwest"],
    "payer_status": ["Medicare", "Commercial/Private", "Other", "Medicaid/Public/Government"],
    "PatTobaccoUse": ["Unknown/Not reported/Multiple", "Never", "Former", "Current"],
}

FEATURE_LABELS = {
    "finaldx_degenerative": "Degenerative diagnosis",
    "finaldx_radicular": "Radiculopathy diagnosis",
    "finaldx_stenosis": "Spinal stenosis diagnosis",
    "finaldx_deformity_instability": "Deformity or instability diagnosis",
    "finaldx_other_diagnosis": "Other lumbar diagnosis",
    "age": "Age",
    "sex": "Male sex",
    "race": "Race",
    "ethnicity": "Hispanic ethnicity",
    "cancer_status": "Any cancer",
    "chronic_pulmonary_disease": "Chronic pulmonary disease",
    "congestive_heart_failure": "Congestive heart failure",
    "connective_tissue_rheumatic_disease": "Connective tissue/rheumatic disease",
    "diabetes_status": "Diabetes status",
    "myocardial_infarction": "Myocardial infarction",
    "renal_disease": "Renal disease",
    "institution_type": "Non-hospital institution",
    "institution_size": "Institution size",
    "institution_region": "Institution region",
    "asa": "ASA physical status",
    "bmi": "BMI",
    "payer_status": "Primary payer",
    "alif_llif": "Anterior/lateral lumbar interbody fusion",
    "corpectomy": "Corpectomy",
    "discectomy": "Discectomy",
    "foraminotomy": "Foraminotomy",
    "instrumentation": "Instrumentation",
    "laminectomy_posterior_decompression": "Posterior decompression or laminectomy",
    "pelvic_fixation": "Pelvic fixation",
    "plf": "Posterolateral fusion",
    "tlif_plif": "Posterior/transforaminal lumbar interbody fusion",
    "other_lumbar_procedure": "Other lumbar procedure",
    "number_operated_levels": "Number of operated levels",
    "operative_region_extent": "Operative region extent",
    "PatTobaccoUse": "Tobacco use",
    ODI_BINARY_COL: "Preoperative ODI ≥ 40",
}

# ============================================================
# Helpers
# ============================================================

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def to_binary_target(x: Any) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if sx in {"1", "1.0", "yes", "y", "true", "t"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def safe_filename(x: str) -> str:
    x = str(x).replace("≥", "ge").replace("≤", "le")
    x = re.sub(r"[^A-Za-z0-9_.-]+", "_", x)
    x = re.sub(r"_+", "_", x).strip("_")
    return x[:180] if x else "file"


def json_native(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {str(k): json_native(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_native(v) for v in obj]
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return float(obj)
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    return obj


def find_existing_column(columns: List[str], candidates: List[str], label: str) -> str:
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if c.lower() in lookup:
            return lookup[c.lower()]
    raise ValueError(f"Could not find {label}. Tried: {candidates}")


def find_input_csv() -> str:
    for p in INPUT_CSV_CANDIDATES:
        if os.path.exists(p):
            if p.startswith(FALLBACK_DIR):
                dst = os.path.join(BASE_DIR, os.path.basename(p))
                if not os.path.exists(dst):
                    shutil.copy2(p, dst)
                return dst
            return p
    raise FileNotFoundError(f"Could not find an ODI cohort CSV. Tried: {INPUT_CSV_CANDIDATES}")


def safe_average_precision(y: np.ndarray, p: np.ndarray) -> float:
    return np.nan if len(np.unique(y)) < 2 else float(average_precision_score(y, p))


def safe_roc_auc(y: np.ndarray, p: np.ndarray) -> float:
    return np.nan if len(np.unique(y)) < 2 else float(roc_auc_score(y, p))


def brier_safe(y: np.ndarray, p: np.ndarray) -> float:
    try:
        return float(brier_score_loss(y, p))
    except Exception:
        return np.nan


def expected_calibration_error(y: np.ndarray, p: np.ndarray, n_bins: int = ECE_N_BINS) -> float:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    bins = np.linspace(0.0, 1.0, n_bins + 1)
    out = 0.0
    if len(y) == 0:
        return np.nan
    for i in range(n_bins):
        left = bins[i]
        right = bins[i + 1]
        mask = (p >= left) & ((p <= right) if i == n_bins - 1 else (p < right))
        if np.any(mask):
            out += (np.sum(mask) / len(y)) * abs(float(np.mean(y[mask])) - float(np.mean(p[mask])))
    return float(out)


def positive_sample_weights(y: np.ndarray, multiplier: float) -> np.ndarray:
    y = np.asarray(y).astype(int)
    n_pos = int(np.sum(y == 1))
    n_neg = int(np.sum(y == 0))
    if n_pos == 0:
        raise ValueError("Training data contain no events.")
    pos_weight = (n_neg / max(n_pos, 1)) * float(multiplier)
    return np.where(y == 1, pos_weight, 1.0).astype(float)


def patient_level_split(df: pd.DataFrame, target_col: str, group_col: str) -> pd.DataFrame:
    """Disabled by design for final-submission reproducibility.

    The ASR-ODI grading system must use the locked split assignment generated
    by the primary Step 1 LightGBM analysis. Creating a new split here would
    change calibration-set grade boundaries and held-out test results.
    """
    del df, target_col, group_col
    raise RuntimeError(
        "New split creation is disabled. Provide the upstream locked "
        "split_assignment_ODI.csv from the primary Step 1 LightGBM output. "
        "Do not regenerate train/calibration/test splits inside the grading code."
    )


def _source_folder_has_split(path: str) -> bool:
    if not os.path.isdir(path):
        return False
    for root, _, files in os.walk(path):
        if "split_assignment_ODI.csv" in files:
            return True
    return False


def _archive_has_split(zip_path: str) -> bool:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            return any(os.path.basename(n.replace(chr(92), "/")) == "split_assignment_ODI.csv" for n in zf.namelist())
    except Exception:
        return False


def ensure_source_dir_for_split() -> Optional[str]:
    """Locate/extract an upstream Step 1 LightGBM output containing split_assignment_ODI.csv."""
    folder_candidates = []
    for root in [BASE_DIR, FALLBACK_DIR]:
        for name in SOURCE_FOLDER_CANDIDATES:
            folder_candidates.append(os.path.join(root, name))

    for folder in folder_candidates:
        if _source_folder_has_split(folder):
            return folder

    archive_candidates = []
    for root in [BASE_DIR, FALLBACK_DIR]:
        for name in SOURCE_ARCHIVE_CANDIDATES:
            archive_candidates.append(os.path.join(root, name))
        # Also scan root for any Step 1 LightGBM zip that contains split_assignment_ODI.csv.
        if os.path.isdir(root):
            for fn in os.listdir(root):
                if fn.lower().endswith(".zip") and "step1" in fn.lower() and "lightgbm" in fn.lower():
                    archive_candidates.append(os.path.join(root, fn))

    seen = set()
    for candidate in archive_candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        if os.path.exists(candidate) and _archive_has_split(candidate):
            extract_root = os.path.join(BASE_DIR, "_step1_odi_risk_grade_locked_split_source")
            if os.path.isdir(extract_root):
                shutil.rmtree(extract_root)
            os.makedirs(extract_root, exist_ok=True)
            with zipfile.ZipFile(candidate, "r") as zf:
                zf.extractall(extract_root)
            return extract_root

    return None


def load_locked_split_assignment(df: pd.DataFrame, group_col: str, target_col: str) -> Tuple[pd.DataFrame, str]:
    """Load the locked upstream ODI split assignment and never create a new split.

    This is a deliberate final-submission safeguard: ASR-ODI grade cutoffs are
    derived in the calibration split. If a new split is silently generated here,
    the grade boundaries and all downstream Fig. 4 results can change.
    """
    del df, target_col
    source_dir = ensure_source_dir_for_split()
    matches = []
    if source_dir is not None:
        for root, _, files in os.walk(source_dir):
            for fn in files:
                if fn == "split_assignment_ODI.csv":
                    matches.append(os.path.join(root, fn))

    if len(matches) != 1:
        searched_folders = [os.path.join(root, name) for root in [BASE_DIR, FALLBACK_DIR] for name in SOURCE_FOLDER_CANDIDATES]
        searched_archives = [os.path.join(root, name) for root in [BASE_DIR, FALLBACK_DIR] for name in SOURCE_ARCHIVE_CANDIDATES]
        raise FileNotFoundError(
            "Locked split_assignment_ODI.csv was not found uniquely. "
            "For final-submission reproducibility, this grading code is locked-split only "
            "and will not create a new patient-level split. Run the primary Step 1 LightGBM "
            "analysis first and place its output folder or ZIP in /content or /mnt/data.\n"
            f"Folders checked: {searched_folders}\n"
            f"Archives checked: {searched_archives}\n"
            f"Matching split files found: {matches}"
        )

    split_path = matches[0]
    split = pd.read_csv(split_path)
    split.columns = [str(c).strip() for c in split.columns]
    required = [group_col, "split"]
    missing = [c for c in required if c not in split.columns]
    if missing:
        raise ValueError(
            f"Locked split file exists but is missing required columns {missing}: {split_path}"
        )

    split = split[[group_col, "split"]].drop_duplicates().copy()
    split["split"] = split["split"].astype(str).str.strip().str.lower()
    allowed = {"train", "calibration", "test"}
    bad = sorted(set(split["split"].dropna().unique()) - allowed)
    if bad:
        raise ValueError(
            f"Locked split file contains unexpected split labels {bad}. Expected only {sorted(allowed)}."
        )

    counts = split["split"].value_counts().to_dict()
    for needed in ["train", "calibration", "test"]:
        if counts.get(needed, 0) == 0:
            raise ValueError(
                f"Locked split file has no rows for split='{needed}'. File: {split_path}"
            )

    return split, f"locked_split_assignment_ODI.csv::{split_path}"


def parse_binary_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive", "performed"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative", "not performed"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def parse_ordinal_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]:
        return float(ORDINAL_MAPS[feature][sx])
    try:
        v = float(sx)
        if feature == "asa":
            return float(min(max(int(round(v)), 1), 4))
        if feature == "number_operated_levels":
            return float(min(max(int(round(v)), 0), 4))
        if feature == "diabetes_status":
            return float(min(max(int(round(v)), 0), 2))
        if feature == "institution_size":
            return float(min(max(int(round(v)), 0), 2))
        return float(v)
    except Exception:
        return np.nan


def canonical_nominal(feature: str, x: Any) -> Any:
    x = clean_scalar(x)
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    sl = s.lower()
    if feature == "race":
        if sl == "white":
            return "White"
        if sl == "black":
            return "Black"
        return "Other"
    if feature == "institution_region":
        return {
            "south": "South",
            "north east": "North East",
            "northeast": "North East",
            "north-east": "North East",
            "west": "West",
            "midwest": "Midwest",
            "mid west": "Midwest",
        }.get(sl, s)
    if feature == "payer_status":
        if sl == "medicare":
            return "Medicare"
        if sl in {"commercial/private", "commercial", "private", "commercial private"}:
            return "Commercial/Private"
        if sl in {"medicaid/public/government", "medicaid", "public", "government", "public/government"}:
            return "Medicaid/Public/Government"
        return "Other"
    if feature == "PatTobaccoUse":
        if sl == "never":
            return "Never"
        if sl == "former":
            return "Former"
        if sl == "current":
            return "Current"
        return "Unknown/Not reported/Multiple"
    return s


@dataclass
class Step1ODIPreprocessor:
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str, List[str]]

    def __post_init__(self):
        self.continuous_imputer: Optional[SimpleImputer] = None
        self.discrete_imputer: Optional[SimpleImputer] = None
        self.nominal_imputer: Optional[SimpleImputer] = None
        self.onehot: Optional[OneHotEncoder] = None
        self.output_feature_names_: List[str] = []

    def _make_parts(self, X: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        cont = pd.DataFrame(index=X.index)
        for c in self.continuous_features:
            cont[c] = pd.to_numeric(X[c].map(clean_scalar), errors="coerce")

        discrete = pd.DataFrame(index=X.index)
        for c in self.binary_features:
            discrete[c] = X[c].map(lambda z: parse_binary_value(z, c)).astype(float)
        for c in self.ordinal_features:
            discrete[c] = X[c].map(lambda z: parse_ordinal_value(z, c)).astype(float)

        nominal = pd.DataFrame(index=X.index)
        for c in self.nominal_features:
            nominal[c] = X[c].map(lambda z: canonical_nominal(c, z)).astype("object")
        return cont, discrete, nominal

    def fit(self, X: pd.DataFrame):
        cont, discrete, nominal = self._make_parts(X)
        self.continuous_imputer = SimpleImputer(strategy="median")
        self.discrete_imputer = SimpleImputer(strategy="most_frequent")
        self.nominal_imputer = SimpleImputer(strategy="constant", fill_value="Missing")

        self.continuous_imputer.fit(cont)
        self.discrete_imputer.fit(discrete)
        nominal_imp = pd.DataFrame(self.nominal_imputer.fit_transform(nominal), columns=self.nominal_features)

        categories = []
        for c in self.nominal_features:
            preferred = list(self.preferred_nominal_levels.get(c, []))
            observed = nominal_imp[c].astype(str).unique().tolist()
            categories.append(preferred + sorted([v for v in observed if v not in preferred]))

        try:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse_output=False)
        except TypeError:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse=False)
        self.onehot.fit(nominal_imp.astype(str))

        self.output_feature_names_ = (
            self.continuous_features
            + self.binary_features
            + self.ordinal_features
            + self.onehot.get_feature_names_out(self.nominal_features).tolist()
        )
        return self

    def transform(self, X: pd.DataFrame) -> np.ndarray:
        cont, discrete, nominal = self._make_parts(X)
        cont_imp = self.continuous_imputer.transform(cont)
        discrete_imp = self.discrete_imputer.transform(discrete)
        nominal_imp = pd.DataFrame(self.nominal_imputer.transform(nominal), columns=self.nominal_features)
        nominal_oh = self.onehot.transform(nominal_imp.astype(str))
        return np.concatenate([cont_imp, discrete_imp, nominal_oh], axis=1).astype(float)

    @property
    def output_feature_names(self) -> List[str]:
        return self.output_feature_names_


# ============================================================
# Model development
# ============================================================

def prepare_data() -> Tuple[pd.DataFrame, str, str, str]:
    csv_path = find_input_csv()
    df = pd.read_csv(csv_path, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]

    group_col = find_existing_column(df.columns.tolist(), GROUP_COL_CANDIDATES, "patient group column")
    target_col = find_existing_column(df.columns.tolist(), TARGET_COL_CANDIDATES, "Step 1 target column")
    prom_col = find_existing_column(df.columns.tolist(), PROM_COL_CANDIDATES, "preoperative ODI column")

    if group_col != GROUP_COL:
        df[GROUP_COL] = df[group_col]
        group_col = GROUP_COL

    df[target_col] = df[target_col].map(to_binary_target)
    df[prom_col] = pd.to_numeric(df[prom_col].map(clean_scalar), errors="coerce")
    df[ODI_BINARY_COL] = np.where(df[prom_col].notna(), (df[prom_col] >= ODI_THRESHOLD).astype(float), np.nan)

    required = [group_col, target_col, prom_col] + BASELINE_FEATURES + [ODI_BINARY_COL]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    keep = df[target_col].isin([0.0, 1.0]) & df[ODI_BINARY_COL].notna() & df[group_col].notna()
    df = df.loc[keep].copy().reset_index(drop=True)
    df[target_col] = df[target_col].astype(int)
    df[ODI_BINARY_COL] = df[ODI_BINARY_COL].astype(int)

    return df, group_col, target_col, prom_col


def assign_splits(df: pd.DataFrame, group_col: str, target_col: str) -> Tuple[pd.DataFrame, str]:
    split, source = load_locked_split_assignment(df, group_col, target_col)
    out = df.merge(split[[group_col, "split"]].drop_duplicates(), on=group_col, how="left")
    if out["split"].isna().any():
        raise RuntimeError(
            "Some eligible ODI rows could not be matched to the locked split assignment by PersonKey. "
            "Confirm that PROM_ODI_Cohort.csv and split_assignment_ODI.csv came from the same primary Step 1 run."
        )

    valid = out["split"].isin(["train", "calibration", "test"])
    if not valid.all():
        out = out.loc[valid].copy()
    return out.reset_index(drop=True), source


def fit_locked_odi40_model(df: pd.DataFrame, target_col: str) -> Dict[str, Any]:
    train = df[df["split"].eq("train")].copy()
    cal = df[df["split"].eq("calibration")].copy()
    test = df[df["split"].eq("test")].copy()

    if min(train[target_col].sum(), cal[target_col].sum(), test[target_col].sum()) <= 0:
        raise RuntimeError("Each split must contain at least one reoperation event.")

    pre = Step1ODIPreprocessor(
        continuous_features=CONTINUOUS_FEATURES,
        binary_features=BINARY_FEATURES,
        ordinal_features=ORDINAL_FEATURES,
        nominal_features=NOMINAL_FEATURES,
        preferred_nominal_levels=PREFERRED_NOMINAL_LEVELS,
    )

    X_train = train[FEATURES].copy()
    X_cal = cal[FEATURES].copy()
    X_test = test[FEATURES].copy()
    y_train = train[target_col].astype(int).to_numpy()
    y_cal = cal[target_col].astype(int).to_numpy()
    y_test = test[target_col].astype(int).to_numpy()

    Xenc_train = pre.fit(X_train).transform(X_train)
    Xenc_cal = pre.transform(X_cal)
    Xenc_test = pre.transform(X_test)

    params = dict(LOCKED_ODI_LIGHTGBM_SETTINGS["lightgbm_params"])
    model = LGBMClassifier(
        objective="binary",
        boosting_type="gbdt",
        metric="average_precision",
        random_state=RANDOM_STATE,
        n_jobs=-1,
        verbosity=-1,
        force_col_wise=True,
        **params,
    )
    w = positive_sample_weights(y_train, LOCKED_ODI_LIGHTGBM_SETTINGS["positive_weight_multiplier"])
    model.fit(Xenc_train, y_train, sample_weight=w)

    p_train_raw = model.predict_proba(Xenc_train)[:, 1]
    p_cal_raw = model.predict_proba(Xenc_cal)[:, 1]
    p_test_raw = model.predict_proba(Xenc_test)[:, 1]

    calibrator = IsotonicRegression(out_of_bounds="clip")
    calibrator.fit(p_cal_raw, y_cal)

    p_train = np.clip(calibrator.predict(p_train_raw), 0, 1)
    p_cal = np.clip(calibrator.predict(p_cal_raw), 0, 1)
    p_test = np.clip(calibrator.predict(p_test_raw), 0, 1)

    return {
        "preprocessor": pre,
        "model": model,
        "calibrator": calibrator,
        "train": train,
        "calibration": cal,
        "test": test,
        "y_train": y_train,
        "y_cal": y_cal,
        "y_test": y_test,
        "p_train": p_train,
        "p_cal": p_cal,
        "p_test": p_test,
        "p_train_raw": p_train_raw,
        "p_cal_raw": p_cal_raw,
        "p_test_raw": p_test_raw,
    }


def net_benefit_model(y_true: np.ndarray, p_risk: np.ndarray, pt: float) -> float:
    """Calculate model net benefit at a threshold probability."""
    y_true = np.asarray(y_true).astype(int)
    p_risk = np.asarray(p_risk).astype(float)
    pred_positive = p_risk >= pt
    n = len(y_true)
    if n == 0:
        return np.nan
    tp = np.sum((pred_positive == 1) & (y_true == 1))
    fp = np.sum((pred_positive == 1) & (y_true == 0))
    return float((tp / n) - (fp / n) * (pt / (1.0 - pt)))


def net_benefit_treat_all(y_true: np.ndarray, pt: float) -> float:
    """Calculate treat-all net benefit at a threshold probability."""
    y_true = np.asarray(y_true).astype(int)
    prevalence = float(np.mean(y_true)) if len(y_true) else np.nan
    return float(prevalence - (1.0 - prevalence) * (pt / (1.0 - pt)))


def make_decision_curve(
    y_cal: np.ndarray,
    p_cal_model: np.ndarray,
    threshold_grid: np.ndarray = DCA_THRESHOLD_GRID,
) -> pd.DataFrame:
    """Create a decision curve for post-derivation clinical-utility validation."""
    rows = []
    for pt in threshold_grid:
        nb_model = net_benefit_model(y_cal, p_cal_model, float(pt))
        nb_all = net_benefit_treat_all(y_cal, float(pt))
        nb_none = 0.0
        reference = max(nb_all, nb_none)
        rows.append({
            "threshold_probability": float(pt),
            "net_benefit_model": float(nb_model),
            "net_benefit_treat_all": float(nb_all),
            "net_benefit_treat_none": float(nb_none),
            "net_benefit_reference": float(reference),
            "net_benefit_gain_vs_reference": float(nb_model - reference),
        })
    return pd.DataFrame(rows)


def _longest_true_run(mask: np.ndarray) -> Tuple[int, int]:
    """Return start and end indices, inclusive, of the longest True run."""
    best_start, best_end = -1, -1
    current_start = None
    for i, val in enumerate(mask):
        if bool(val) and current_start is None:
            current_start = i
        if ((not bool(val)) or i == len(mask) - 1) and current_start is not None:
            current_end = i if bool(val) and i == len(mask) - 1 else i - 1
            if best_start == -1 or (current_end - current_start) > (best_end - best_start):
                best_start, best_end = current_start, current_end
            current_start = None
    return best_start, best_end


def define_grade_cutoffs(p_cal: np.ndarray, y_cal: Optional[np.ndarray] = None) -> Dict[str, Any]:
    """Define three ASR-ODI risk-grade thresholds using calibration data only.

    The two grade boundaries are prespecified tertile cut points of calibrated
    predicted 1-year reoperation risk in the calibration split. Outcomes are not
    used to define the grade boundaries, and DCA is reserved for downstream
    clinical-utility validation after the boundaries are locked.
    """
    del y_cal
    p_cal = np.asarray(p_cal, dtype=float)
    p_cal = p_cal[np.isfinite(p_cal)]

    if len(p_cal) < N_RISK_GRADES:
        raise RuntimeError("Not enough calibration patients to define three risk grades.")
    if len(np.unique(p_cal)) < N_RISK_GRADES:
        raise RuntimeError("Not enough unique calibrated risks to define three risk grades.")

    cutoff_0_to_1, cutoff_1_to_2 = [
        float(v) for v in np.quantile(p_cal, RISK_GRADE_QUANTILES)
    ]

    if not cutoff_0_to_1 < cutoff_1_to_2:
        unique_scores = np.unique(np.sort(p_cal))
        if len(unique_scores) < N_RISK_GRADES:
            raise RuntimeError("Not enough unique calibrated risks to define ordered grade cutoffs.")
        cutoff_0_to_1 = float(unique_scores[int(np.floor((len(unique_scores) - 1) / 3.0))])
        cutoff_1_to_2 = float(unique_scores[int(np.floor(2.0 * (len(unique_scores) - 1) / 3.0))])

    if not cutoff_0_to_1 < cutoff_1_to_2:
        raise RuntimeError(
            "Calibration-set predicted-risk tertiles did not produce two ordered grade boundaries. "
            "This can occur when isotonic calibration produces too few unique risk values."
        )

    grade_cal = assign_risk_grade_from_scores(p_cal, cutoff_0_to_1, cutoff_1_to_2)
    grade_counts = {int(g): int(np.sum(grade_cal == g)) for g in range(N_RISK_GRADES)}
    if min(grade_counts.values()) == 0:
        raise RuntimeError(
            f"Locked tertile boundaries produced an empty calibration grade: {grade_counts}."
        )

    calibration_warning = None
    if min(grade_counts.values()) < MIN_PATIENTS_PER_GRADE:
        calibration_warning = (
            f"At least one calibration-set grade contains fewer than {MIN_PATIENTS_PER_GRADE} patients: "
            f"{grade_counts}. Interpret grade-specific calibration estimates cautiously."
        )
        warnings.warn(calibration_warning)

    return {
        "grade_score_source": GRADE_SCORE_SOURCE,
        "threshold_algorithm": "prespecified calibration-set calibrated-risk tertiles",
        "threshold_derivation_split": "calibration",
        "n_risk_grades": int(N_RISK_GRADES),
        "risk_grade_quantiles": [float(v) for v in RISK_GRADE_QUANTILES],
        "cutoff_grade0_to_1_score": cutoff_0_to_1,
        "cutoff_grade1_to_2_score": cutoff_1_to_2,
        "score_rule_grade0": f"{GRADE_SCORE_SOURCE} < {cutoff_0_to_1:.12f}",
        "score_rule_grade1": f"{cutoff_0_to_1:.12f} <= {GRADE_SCORE_SOURCE} < {cutoff_1_to_2:.12f}",
        "score_rule_grade2": f"{GRADE_SCORE_SOURCE} >= {cutoff_1_to_2:.12f}",
        "n_calibration_patients_for_threshold_derivation": int(len(p_cal)),
        "n_unique_calibrated_risks_for_threshold_derivation": int(len(np.unique(p_cal))),
        "calibration_grade_counts_at_locked_cutoffs": grade_counts,
        "calibration_grade_distribution_warning": calibration_warning,
    }


def assign_risk_grade_from_scores(score: np.ndarray, c1: float, c2: float) -> np.ndarray:
    """Assign Grade 0/1/2 from calibrated predicted risk and two locked cutoffs."""
    score = np.asarray(score, dtype=float)
    grade = np.full(len(score), 2, dtype=int)
    grade[score < c1] = 0
    grade[(score >= c1) & (score < c2)] = 1
    grade[score >= c2] = 2
    return grade


def assign_risk_grade(p_calibrated: np.ndarray, p_raw: np.ndarray, cutoffs: Dict[str, Any]) -> np.ndarray:
    """Assign Grade 0/1/2 using locked risk-tertile calibration-set thresholds."""
    del p_raw
    score = np.asarray(p_calibrated, dtype=float)
    c1 = float(cutoffs["cutoff_grade0_to_1_score"])
    c2 = float(cutoffs["cutoff_grade1_to_2_score"])
    if not c1 < c2:
        raise RuntimeError("risk-tertile cutoffs must satisfy cutoff_grade0_to_1 < cutoff_grade1_to_2.")
    return assign_risk_grade_from_scores(score, c1, c2)

def evaluate_split(y: np.ndarray, p: np.ndarray, split_name: str) -> Dict[str, Any]:
    return {
        "split": split_name,
        "n": int(len(y)),
        "events": int(np.sum(y)),
        "prevalence": float(np.mean(y)),
        "AP": safe_average_precision(y, p),
        "ROC_AUC": safe_roc_auc(y, p),
        "Brier_score": brier_safe(y, p),
        "ECE_10_equal_width_bins": expected_calibration_error(y, p, ECE_N_BINS),
    }


def bootstrap_ci(values: np.ndarray, y: np.ndarray, p: np.ndarray, grade: np.ndarray, grade_id: int, n_boot: int = N_BOOTSTRAPS) -> Tuple[float, float]:
    rng = np.random.default_rng(RANDOM_STATE + 1000 + grade_id)
    idx_all = np.arange(len(y))
    vals = []
    for _ in range(n_boot):
        idx = rng.choice(idx_all, size=len(idx_all), replace=True)
        mask = grade[idx] == grade_id
        if not np.any(mask):
            continue
        vals.append(float(np.mean(y[idx][mask])))
    if not vals:
        return np.nan, np.nan
    return float(np.quantile(vals, 0.025)), float(np.quantile(vals, 0.975))


def grade_summary(y: np.ndarray, p: np.ndarray, grade: np.ndarray, split_name: str) -> pd.DataFrame:
    y = np.asarray(y).astype(int)
    p = np.asarray(p).astype(float)
    grade = np.asarray(grade).astype(int)
    prevalence = float(np.mean(y)) if len(y) else np.nan
    total_events = int(np.sum(y))
    rows = []
    for gid in sorted(GRADE_LABELS):
        mask = grade == gid
        n = int(np.sum(mask))
        events = int(np.sum(y[mask])) if n > 0 else 0
        observed = float(events / n) if n > 0 else np.nan
        lo, hi = bootstrap_ci(observed, y, p, grade, gid) if n > 0 else (np.nan, np.nan)
        mean_pred = float(np.mean(p[mask])) if n > 0 else np.nan
        rows.append({
            "split": split_name,
            "grade": gid,
            "grade_label": GRADE_LABELS[gid],
            "n": n,
            "events": events,
            "observed_reoperation_rate": observed,
            "observed_reoperation_rate_lower_95": lo,
            "observed_reoperation_rate_upper_95": hi,
            "mean_calibrated_predicted_risk": mean_pred,
            "median_calibrated_predicted_risk": float(np.median(p[mask])) if n > 0 else np.nan,
            "observed_to_expected_ratio": observed / mean_pred if mean_pred > 0 and np.isfinite(observed) else np.nan,
            "risk_lift_vs_split_prevalence": observed / prevalence if prevalence > 0 and np.isfinite(observed) else np.nan,
            "captured_fraction_of_events": events / total_events if total_events > 0 else np.nan,
        })
    return pd.DataFrame(rows)


def make_prediction_table(df_split: pd.DataFrame, target_col: str, p: np.ndarray, p_raw: np.ndarray, grade: np.ndarray, split_name: str, prom_col: str) -> pd.DataFrame:
    cols = [GROUP_COL, target_col, prom_col, ODI_BINARY_COL, "split"]
    out = df_split[cols].copy()
    out["calibrated_predicted_1y_reoperation_risk"] = p
    out["raw_lightgbm_model_score"] = p_raw
    out["ASR_ODI_risk_grade"] = grade.astype(int)
    out["ASR_ODI_risk_grade_label"] = [GRADE_LABELS[int(g)] for g in grade]
    out["split"] = split_name
    return out



def save_dca_plot(dca_df: pd.DataFrame, path: str, title: str, cutoffs: Optional[Dict[str, Any]] = None):
    """Save decision curve analysis plot for clinical-utility validation.

    DCA is not used to derive the grade boundaries in this script. If locked
    cutoffs are provided, the plot overlays them as vertical dashed lines so
    readers can see where the already-defined grade boundaries fall on the
    clinical threshold-probability axis.
    """
    required = [
        "threshold_probability",
        "net_benefit_model",
        "net_benefit_treat_all",
        "net_benefit_treat_none",
    ]
    missing = [c for c in required if c not in dca_df.columns]
    if missing:
        raise ValueError(f"DCA table is missing required columns for plotting: {missing}")

    fig, ax = plt.subplots(figsize=(6.8, 5.2))
    ax.plot(
        dca_df["threshold_probability"] * 100,
        dca_df["net_benefit_model"],
        linewidth=2.0,
        label="ASR-ODI model",
    )
    ax.plot(
        dca_df["threshold_probability"] * 100,
        dca_df["net_benefit_treat_all"],
        linestyle="--",
        linewidth=1.5,
        label="Treat all",
    )
    ax.plot(
        dca_df["threshold_probability"] * 100,
        dca_df["net_benefit_treat_none"],
        linestyle=":",
        linewidth=1.5,
        label="Treat none",
    )

    if "net_benefit_reference" in dca_df.columns:
        ax.plot(
            dca_df["threshold_probability"] * 100,
            dca_df["net_benefit_reference"],
            linestyle="-.",
            linewidth=1.2,
            label="Best default strategy",
        )

    if cutoffs is not None:
        c1 = float(cutoffs["cutoff_grade0_to_1_score"]) * 100.0
        c2 = float(cutoffs["cutoff_grade1_to_2_score"]) * 100.0
        ymin, ymax = ax.get_ylim()
        ax.axvspan(c1, c2, alpha=0.08, label="Intermediate-grade range")
        ax.axvline(c1, linestyle="--", linewidth=1.4, label="Low/intermediate boundary")
        ax.axvline(c2, linestyle="--", linewidth=1.4, label="Intermediate/high boundary")
        ax.text(c1, ymax, f" Low→Intermediate\n {c1:.1f}%", ha="left", va="top", fontsize=8.5)
        ax.text(c2, ymax, f" Intermediate→High\n {c2:.1f}%", ha="left", va="top", fontsize=8.5)
        ax.set_ylim(ymin, ymax)

    ax.set_xlabel("Threshold probability (%)")
    ax.set_ylabel("Net benefit")
    ax.set_title(title, fontweight="bold")
    ax.grid(alpha=0.25)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_grade_boundary_distribution_plot(p_cal: np.ndarray, cutoffs: Dict[str, Any], path: str, title: str):
    """Save calibration-set predicted-risk distribution with locked grade boundaries."""
    p_cal = np.asarray(p_cal, dtype=float)
    p_cal = p_cal[np.isfinite(p_cal)]
    c1 = float(cutoffs["cutoff_grade0_to_1_score"])
    c2 = float(cutoffs["cutoff_grade1_to_2_score"])

    fig, ax = plt.subplots(figsize=(7.0, 5.2))
    ax.hist(p_cal * 100, bins=30, alpha=0.85, edgecolor="black", linewidth=0.4)
    ax.axvline(c1 * 100, linestyle="--", linewidth=1.6, label=f"Low/intermediate: {c1 * 100:.1f}%")
    ax.axvline(c2 * 100, linestyle="--", linewidth=1.6, label=f"Intermediate/high: {c2 * 100:.1f}%")
    ax.set_xlabel("Calibrated predicted 1-year reoperation risk (%)")
    ax.set_ylabel("Calibration-set patient count")
    ax.set_title(title, fontweight="bold")
    ax.grid(axis="y", alpha=0.25)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_calibration_plot(y: np.ndarray, p: np.ndarray, path: str, title: str):
    prob_true, prob_pred = calibration_curve(y, p, n_bins=10, strategy="quantile")
    fig, ax = plt.subplots(figsize=(6.2, 5.4))
    ax.plot([0, 1], [0, 1], linestyle="--", color="gray", linewidth=1.2)
    ax.plot(prob_pred, prob_true, marker="o", color="black", linewidth=1.8)
    ax.set_xlabel("Mean predicted risk")
    ax.set_ylabel("Observed reoperation rate")
    ax.set_title(title, fontweight="bold")
    ax.grid(alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_grade_bar_plot(summary: pd.DataFrame, path: str, title: str):
    d = summary[summary["split"].eq("test")].copy()
    fig, ax = plt.subplots(figsize=(7.4, 5.2))
    x = np.arange(len(d))
    y = d["observed_reoperation_rate"].to_numpy(dtype=float) * 100
    lo = d["observed_reoperation_rate_lower_95"].to_numpy(dtype=float) * 100
    hi = d["observed_reoperation_rate_upper_95"].to_numpy(dtype=float) * 100
    yerr = np.vstack([np.maximum(y - lo, 0), np.maximum(hi - y, 0)])
    ax.bar(x, y, color="#0F766E", alpha=0.85)
    ax.errorbar(x, y, yerr=yerr, fmt="none", color="black", capsize=4)
    ax.set_xticks(x)
    ax.set_xticklabels([GRADE_SHORT_LABELS[int(g)] for g in d["grade"]], rotation=20, ha="right")
    ax.set_ylabel("Observed 1-year reoperation rate (%)")
    ax.set_title(title, fontweight="bold")
    ax.grid(axis="y", alpha=0.25)
    for i, row in d.reset_index(drop=True).iterrows():
        ax.text(i, y[i] + max(0.15, np.nanmax(y) * 0.03), f"n={int(row['n'])}\ne={int(row['events'])}", ha="center", va="bottom", fontsize=8.5)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


# ============================================================
# Streamlit application
# ============================================================

STREAMLIT_APP_CODE = r'''# -*- coding: utf-8 -*-
"""Streamlit interface for the locked ASR-ODI risk-tertile 3-grade model."""
import os
import sys
import warnings
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

import streamlit as st
import lightgbm as lgb  # required for loading the fitted LightGBM object
import joblib

from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder

warnings.filterwarnings("ignore", message="X does not have valid feature names")

ODI_THRESHOLD = 40.0
ODI_BINARY_COL = "preop_ODI_ge_40"
ARTIFACT_FILE = "ASR_ODI_RiskGrade_ODI40_RiskTertile3Grades_LightGBM_artifact.joblib"
BASELINE_FEATURES = [
    "finaldx_degenerative", "finaldx_radicular", "finaldx_stenosis",
    "finaldx_deformity_instability", "finaldx_other_diagnosis",
    "age", "sex", "race", "ethnicity", "cancer_status",
    "chronic_pulmonary_disease", "congestive_heart_failure",
    "connective_tissue_rheumatic_disease", "diabetes_status",
    "myocardial_infarction", "renal_disease", "institution_type",
    "institution_size", "institution_region", "asa", "bmi", "payer_status",
    "alif_llif", "corpectomy", "discectomy", "foraminotomy",
    "instrumentation", "laminectomy_posterior_decompression",
    "pelvic_fixation", "plf", "tlif_plif", "other_lumbar_procedure",
    "number_operated_levels", "operative_region_extent", "PatTobaccoUse",
]
FEATURES = BASELINE_FEATURES + [ODI_BINARY_COL]
CONTINUOUS_FEATURES = ["age", "bmi"]
BINARY_FEATURES = [
    "finaldx_degenerative", "finaldx_radicular", "finaldx_stenosis",
    "finaldx_deformity_instability", "finaldx_other_diagnosis", "sex",
    "ethnicity", "cancer_status", "chronic_pulmonary_disease",
    "congestive_heart_failure", "connective_tissue_rheumatic_disease",
    "myocardial_infarction", "renal_disease", "institution_type",
    "alif_llif", "corpectomy", "discectomy", "foraminotomy",
    "instrumentation", "laminectomy_posterior_decompression",
    "pelvic_fixation", "plf", "tlif_plif", "other_lumbar_procedure",
    "operative_region_extent", ODI_BINARY_COL,
]
ORDINAL_FEATURES = ["diabetes_status", "institution_size", "asa", "number_operated_levels"]
NOMINAL_FEATURES = ["race", "institution_region", "payer_status", "PatTobaccoUse"]
MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}
BINARY_MAPS = {
    "sex": {"female": 0, "f": 0, "male": 1, "m": 1},
    "ethnicity": {"non-hispanic": 0, "non hispanic": 0, "hispanic": 1},
    "cancer_status": {"no cancer": 0, "no": 0, "none": 0, "any cancer": 1, "yes": 1, "cancer": 1},
    "institution_type": {"hospital": 0, "non-hospital": 1, "non hospital": 1, "nonhospital": 1},
    "operative_region_extent": {
        "lumbar only": 0,
        "extended_region_involvement": 1,
        "extended region involvement": 1,
        "extended": 1,
        "thoracolumbar extension only": 1,
        "sacrum pelvis extension only": 1,
        "thoracolumbar and sacrum pelvis extension": 1,
    },
}
ORDINAL_MAPS = {
    "diabetes_status": {
        "no": 0, "none": 0, "0": 0,
        "without comp": 1, "without complication": 1, "without complications": 1, "1": 1,
        "with comp": 2, "with complication": 2, "with complications": 2, "2": 2,
    },
    "institution_size": {
        "between 1-99 beds": 0, "1-99": 0,
        "between 100-399 beds": 1, "100-399": 1,
        ">= 400 beds": 2, ">=400 beds": 2, ">=400": 2, ">= 400": 2,
    },
    "asa": {"1": 1, "i": 1, "2": 2, "ii": 2, "3": 3, "iii": 3, "4": 4, "iv": 4, ">=4": 4, ">= 4": 4, "5": 4, "v": 4},
    "number_operated_levels": {"0": 0, "1": 1, "2": 2, "3": 3, "4": 4, ">=4": 4, ">= 4": 4, "5": 4, "6": 4, "7": 4, "8": 4, "9": 4, "10": 4},
}
PREFERRED_NOMINAL_LEVELS = {
    "race": ["White", "Black", "Other"],
    "institution_region": ["South", "North East", "West", "Midwest"],
    "payer_status": ["Medicare", "Commercial/Private", "Other", "Medicaid/Public/Government"],
    "PatTobaccoUse": ["Unknown/Not reported/Multiple", "Never", "Former", "Current"],
}
GRADE_LABELS = {0: "Grade 0 / Low risk", 1: "Grade 1 / Intermediate risk", 2: "Grade 2 / High risk"}
GRADE_SHORT_LABELS = {0: "Low risk", 1: "Intermediate risk", 2: "High risk"}
GRADE_COLORS = {0: "#16A34A", 1: "#FACC15", 2: "#DC2626"}
GRADE_TEXT_COLORS = {0: "#FFFFFF", 1: "#111827", 2: "#FFFFFF"}

def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x

def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()

def parse_binary_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in BINARY_MAPS and sx in BINARY_MAPS[feature]:
        return float(BINARY_MAPS[feature][sx])
    if sx in {"1", "1.0", "yes", "y", "true", "t", "present", "positive", "performed"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f", "absent", "negative", "not performed"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan

def parse_ordinal_value(x: Any, feature: str) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if feature in ORDINAL_MAPS and sx in ORDINAL_MAPS[feature]:
        return float(ORDINAL_MAPS[feature][sx])
    try:
        v = float(sx)
        if feature == "asa":
            return float(min(max(int(round(v)), 1), 4))
        if feature == "number_operated_levels":
            return float(min(max(int(round(v)), 0), 4))
        if feature == "diabetes_status":
            return float(min(max(int(round(v)), 0), 2))
        if feature == "institution_size":
            return float(min(max(int(round(v)), 0), 2))
        return float(v)
    except Exception:
        return np.nan

def canonical_nominal(feature: str, x: Any) -> Any:
    x = clean_scalar(x)
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    sl = s.lower()
    if feature == "race":
        if sl == "white":
            return "White"
        if sl == "black":
            return "Black"
        return "Other"
    if feature == "institution_region":
        return {"south": "South", "north east": "North East", "northeast": "North East", "north-east": "North East", "west": "West", "midwest": "Midwest", "mid west": "Midwest"}.get(sl, s)
    if feature == "payer_status":
        if sl == "medicare":
            return "Medicare"
        if sl in {"commercial/private", "commercial", "private", "commercial private"}:
            return "Commercial/Private"
        if sl in {"medicaid/public/government", "medicaid", "public", "government", "public/government"}:
            return "Medicaid/Public/Government"
        return "Other"
    if feature == "PatTobaccoUse":
        if sl == "never":
            return "Never"
        if sl == "former":
            return "Former"
        if sl == "current":
            return "Current"
        return "Unknown/Not reported/Multiple"
    return s

@dataclass
class Step1ODIPreprocessor:
    continuous_features: List[str]
    binary_features: List[str]
    ordinal_features: List[str]
    nominal_features: List[str]
    preferred_nominal_levels: Dict[str, List[str]]

    def __post_init__(self):
        self.continuous_imputer: Optional[SimpleImputer] = None
        self.discrete_imputer: Optional[SimpleImputer] = None
        self.nominal_imputer: Optional[SimpleImputer] = None
        self.onehot: Optional[OneHotEncoder] = None
        self.output_feature_names_: List[str] = []

    def _make_parts(self, X: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        cont = pd.DataFrame(index=X.index)
        for c in self.continuous_features:
            cont[c] = pd.to_numeric(X[c].map(clean_scalar), errors="coerce")
        discrete = pd.DataFrame(index=X.index)
        for c in self.binary_features:
            discrete[c] = X[c].map(lambda z: parse_binary_value(z, c)).astype(float)
        for c in self.ordinal_features:
            discrete[c] = X[c].map(lambda z: parse_ordinal_value(z, c)).astype(float)
        nominal = pd.DataFrame(index=X.index)
        for c in self.nominal_features:
            nominal[c] = X[c].map(lambda z: canonical_nominal(c, z)).astype("object")
        return cont, discrete, nominal

    def fit(self, X: pd.DataFrame):
        cont, discrete, nominal = self._make_parts(X)
        self.continuous_imputer = SimpleImputer(strategy="median")
        self.discrete_imputer = SimpleImputer(strategy="most_frequent")
        self.nominal_imputer = SimpleImputer(strategy="constant", fill_value="Missing")
        self.continuous_imputer.fit(cont)
        self.discrete_imputer.fit(discrete)
        nominal_imp = pd.DataFrame(self.nominal_imputer.fit_transform(nominal), columns=self.nominal_features)
        categories = []
        for c in self.nominal_features:
            preferred = list(self.preferred_nominal_levels.get(c, []))
            observed = nominal_imp[c].astype(str).unique().tolist()
            categories.append(preferred + sorted([v for v in observed if v not in preferred]))
        try:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse_output=False)
        except TypeError:
            self.onehot = OneHotEncoder(categories=categories, handle_unknown="ignore", sparse=False)
        self.onehot.fit(nominal_imp.astype(str))
        self.output_feature_names_ = self.continuous_features + self.binary_features + self.ordinal_features + self.onehot.get_feature_names_out(self.nominal_features).tolist()
        return self

    def transform(self, X: pd.DataFrame) -> np.ndarray:
        cont, discrete, nominal = self._make_parts(X)
        cont_imp = self.continuous_imputer.transform(cont)
        discrete_imp = self.discrete_imputer.transform(discrete)
        nominal_imp = pd.DataFrame(self.nominal_imputer.transform(nominal), columns=self.nominal_features)
        nominal_oh = self.onehot.transform(nominal_imp.astype(str))
        return np.concatenate([cont_imp, discrete_imp, nominal_oh], axis=1).astype(float)

    @property
    def output_feature_names(self) -> List[str]:
        return self.output_feature_names_

def assign_risk_grade(p_calibrated: np.ndarray, p_raw: np.ndarray, cutoffs: Dict[str, Any]) -> np.ndarray:
    del p_raw
    score = np.asarray(p_calibrated, dtype=float)
    c1 = float(cutoffs["cutoff_grade0_to_1_score"])
    c2 = float(cutoffs["cutoff_grade1_to_2_score"])
    grade = np.full(len(score), 2, dtype=int)
    grade[score < c1] = 0
    grade[(score >= c1) & (score < c2)] = 1
    grade[score >= c2] = 2
    return grade

import __main__
sys.modules.setdefault("orig_app", sys.modules[__name__])
__main__.Step1ODIPreprocessor = Step1ODIPreprocessor
__main__.clean_scalar = clean_scalar
__main__.norm_text = norm_text
__main__.parse_binary_value = parse_binary_value
__main__.parse_ordinal_value = parse_ordinal_value
__main__.canonical_nominal = canonical_nominal


def patch_artifact_compatibility(artifact: Dict[str, Any]) -> Dict[str, Any]:
    """Apply narrow compatibility patches for scikit-learn minor-version changes."""
    pre = artifact.get("preprocessor")
    if pre is None:
        return artifact
    for name in ("continuous_imputer", "discrete_imputer", "nominal_imputer"):
        imputer = getattr(pre, name, None)
        if imputer is not None and not hasattr(imputer, "_fill_dtype") and hasattr(imputer, "_fit_dtype"):
            imputer._fill_dtype = imputer._fit_dtype
    return artifact

@st.cache_resource
def load_artifact():
    here = os.path.dirname(os.path.abspath(__file__))
    artifact_path = os.path.join(here, ARTIFACT_FILE)
    if not os.path.exists(artifact_path):
        st.error(f"Model artifact not found: {artifact_path}")
        st.stop()
    return patch_artifact_compatibility(joblib.load(artifact_path))

def binary_checkbox(label: str, default: bool = False) -> int:
    return int(st.checkbox(label, value=default))

def build_patient_input() -> pd.DataFrame:
    st.sidebar.header("Patient inputs")
    patient = {f: 0 for f in BASELINE_FEATURES}
    with st.sidebar.expander("Demographics and PROM", expanded=True):
        patient["age"] = st.number_input("Age", min_value=18.0, max_value=110.0, value=65.0, step=1.0)
        patient["sex"] = st.selectbox("Sex", ["female", "male"], index=1)
        patient["race"] = st.selectbox("Race", ["White", "Black", "Other"], index=0)
        patient["ethnicity"] = st.selectbox("Ethnicity", ["non-hispanic", "hispanic"], index=0)
        patient["bmi"] = st.number_input("BMI", min_value=10.0, max_value=80.0, value=30.0, step=0.1)
        preop_odi = st.number_input("Preoperative ODI", min_value=0.0, max_value=100.0, value=40.0, step=1.0)
    with st.sidebar.expander("Diagnosis", expanded=False):
        patient["finaldx_degenerative"] = binary_checkbox("Degenerative diagnosis")
        patient["finaldx_radicular"] = binary_checkbox("Radiculopathy diagnosis")
        patient["finaldx_stenosis"] = binary_checkbox("Spinal stenosis diagnosis", True)
        patient["finaldx_deformity_instability"] = binary_checkbox("Deformity/instability diagnosis")
        patient["finaldx_other_diagnosis"] = binary_checkbox("Other lumbar diagnosis")
    with st.sidebar.expander("Comorbidities", expanded=False):
        patient["cancer_status"] = "yes" if st.checkbox("Any cancer") else "no"
        patient["chronic_pulmonary_disease"] = binary_checkbox("Chronic pulmonary disease")
        patient["congestive_heart_failure"] = binary_checkbox("Congestive heart failure")
        patient["connective_tissue_rheumatic_disease"] = binary_checkbox("Connective tissue/rheumatic disease")
        patient["diabetes_status"] = st.selectbox("Diabetes status", ["0", "1", "2"], index=0)
        patient["myocardial_infarction"] = binary_checkbox("Myocardial infarction")
        patient["renal_disease"] = binary_checkbox("Renal disease")
        patient["asa"] = st.selectbox("ASA physical status", ["1", "2", "3", "4"], index=2)
        patient["PatTobaccoUse"] = st.selectbox("Tobacco use", ["Unknown/Not reported/Multiple", "Never", "Former", "Current"], index=1)
    with st.sidebar.expander("Institution and payer", expanded=False):
        patient["institution_type"] = st.selectbox("Institution type", ["hospital", "non-hospital"], index=0)
        patient["institution_size"] = st.selectbox("Institution size", ["between 1-99 beds", "between 100-399 beds", ">= 400 beds"], index=2)
        patient["institution_region"] = st.selectbox("Institution region", ["South", "North East", "West", "Midwest"], index=0)
        patient["payer_status"] = st.selectbox("Primary payer", ["Medicare", "Commercial/Private", "Other", "Medicaid/Public/Government"], index=0)
    with st.sidebar.expander("Procedure", expanded=False):
        patient["alif_llif"] = binary_checkbox("ALIF/LLIF")
        patient["corpectomy"] = binary_checkbox("Corpectomy")
        patient["discectomy"] = binary_checkbox("Discectomy")
        patient["foraminotomy"] = binary_checkbox("Foraminotomy")
        patient["instrumentation"] = binary_checkbox("Instrumentation")
        patient["laminectomy_posterior_decompression"] = binary_checkbox("Posterior decompression/laminectomy", True)
        patient["pelvic_fixation"] = binary_checkbox("Pelvic fixation")
        patient["plf"] = binary_checkbox("PLF")
        patient["tlif_plif"] = binary_checkbox("TLIF/PLIF")
        patient["other_lumbar_procedure"] = binary_checkbox("Other lumbar procedure")
        patient["number_operated_levels"] = st.selectbox("Number operated levels", ["0", "1", "2", "3", "4"], index=2)
        patient["operative_region_extent"] = st.selectbox("Operative region extent", ["lumbar only", "extended_region_involvement"], index=0)
    patient["preoperative_ODI_value"] = preop_odi
    patient[ODI_BINARY_COL] = int(preop_odi >= ODI_THRESHOLD)
    return pd.DataFrame([patient])

def score_patient(artifact: Dict[str, Any], df: pd.DataFrame) -> Tuple[float, float, int]:
    features = artifact.get("features", FEATURES)
    X = df[features].copy()
    Xenc = artifact["preprocessor"].transform(X)
    raw = float(artifact["model"].predict_proba(Xenc)[:, 1][0])
    risk = float(np.clip(artifact["calibrator"].predict([raw])[0], 0, 1))
    grade = int(assign_risk_grade(np.array([risk]), np.array([raw]), artifact["grade_cutoffs"])[0])
    return risk, raw, grade

def render_grade_card(grade: int) -> None:
    color = GRADE_COLORS[grade]
    text_color = GRADE_TEXT_COLORS[grade]
    st.markdown(
        f"""
        <div style="background-color:{color}; color:{text_color}; padding:32px; border-radius:20px; text-align:center;">
            <div style="font-size:20px; font-weight:700; letter-spacing:0.02em;">ASR-ODI Risk Grade</div>
            <div style="font-size:48px; font-weight:800; margin-top:10px;">{GRADE_SHORT_LABELS[grade]}</div>
            <div style="font-size:18px; margin-top:8px;">{GRADE_LABELS[grade]}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_risk_card(risk: float) -> None:
    st.markdown(
        f"""
        <div style="border:1px solid #E5E7EB; padding:28px; border-radius:20px; text-align:center; background-color:#FFFFFF;">
            <div style="font-size:18px; font-weight:700; color:#374151;">Predicted risk</div>
            <div style="font-size:52px; font-weight:800; color:#111827; margin-top:10px;">{risk * 100:.1f}%</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def main():
    st.set_page_config(page_title="ASR-ODI Risk Grade", page_icon="🟢", layout="wide")
    artifact = load_artifact()
    st.title("ASR-ODI Risk Grade Calculator")

    patient_df = build_patient_input()
    risk, raw, grade = score_patient(artifact, patient_df)

    left, right = st.columns([0.9, 1.1], gap="large")
    with left:
        render_risk_card(risk)
    with right:
        render_grade_card(grade)

    st.caption("Research-use risk stratification tool; not a stand-alone treatment decision rule.")

if __name__ == "__main__":
    main()
'''


def create_streamlit_app(artifact_path: str, app_dir: str) -> Dict[str, str]:
    """Create a self-contained Streamlit app folder for patient-level scoring."""
    os.makedirs(app_dir, exist_ok=True)
    app_path = os.path.join(app_dir, "ASR_ODI_RiskGrade_App.py")
    artifact_dst = os.path.join(app_dir, os.path.basename(artifact_path))
    shutil.copy2(artifact_path, artifact_dst)
    with open(app_path, "w", encoding="utf-8") as f:
        f.write(STREAMLIT_APP_CODE)
    requirements_path = os.path.join(app_dir, "requirements.txt")
    with open(requirements_path, "w", encoding="utf-8") as f:
        f.write("streamlit>=1.35,<2\nlightgbm>=4,<5\nscikit-learn>=1.6,<1.9\npandas>=2,<3\nnumpy>=1.24,<3\njoblib>=1.3,<2\n")
    readme_path = os.path.join(app_dir, "README_ASR_ODI_RiskGrade_App.txt")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(
            "ASR-ODI Risk Grade Streamlit App\n"
            "=================================\n\n"
            "Run locally or in Colab with:\n"
            "  pip install -r requirements.txt\n"
            "  streamlit run ASR_ODI_RiskGrade_App.py\n\n"
            "The app loads the locked LightGBM artifact from the same folder and displays the predicted risk and ASR-ODI risk-tertile 3-grade category.\n"
            "The app interface intentionally does not display model-threshold details; these are retained in the accompanying cutoffs table and source code.\n"
            "This application is a research-use risk-stratification aid and is not a stand-alone treatment decision rule.\n"
        )
    return {"app_path": app_path, "artifact_path": artifact_dst, "requirements_path": requirements_path, "readme_path": readme_path}

# ============================================================
# Archive
# ============================================================

def make_zip(src_dir: str, zip_path: str):
    tmp_zip = zip_path + ".tmp"
    for p in [zip_path, tmp_zip]:
        if os.path.exists(p):
            os.remove(p)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(src_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(src_dir))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 100)
    print("Step 1 ODI ASR reoperation risk grade with ODI cutoff 40: calibration-risk-tertile three-grade version with DCA validation")
    print("=" * 100)

    df, group_col, target_col, prom_col = prepare_data()
    df, split_source = assign_splits(df, group_col, target_col)

    print(f"Rows after eligibility filter: {len(df):,}")
    print(f"Events: {int(df[target_col].sum()):,} ({df[target_col].mean():.3%})")
    print(f"Preoperative ODI column: {prom_col}")
    print(f"Split source: {split_source}")
    if not str(split_source).startswith("locked_split_assignment_ODI.csv"):
        raise RuntimeError(
            "Internal safeguard failed: split_source is not the locked upstream split assignment."
        )
    print(df["split"].value_counts().to_string())

    fit = fit_locked_odi40_model(df, target_col)

    cutoffs = define_grade_cutoffs(fit["p_cal"])
    grade_train = assign_risk_grade(fit["p_train"], fit["p_train_raw"], cutoffs)
    grade_cal = assign_risk_grade(fit["p_cal"], fit["p_cal_raw"], cutoffs)
    grade_test = assign_risk_grade(fit["p_test"], fit["p_test_raw"], cutoffs)

    metrics = pd.DataFrame([
        evaluate_split(fit["y_train"], fit["p_train"], "train"),
        evaluate_split(fit["y_cal"], fit["p_cal"], "calibration"),
        evaluate_split(fit["y_test"], fit["p_test"], "test"),
    ])

    grades = pd.concat([
        grade_summary(fit["y_train"], fit["p_train"], grade_train, "train"),
        grade_summary(fit["y_cal"], fit["p_cal"], grade_cal, "calibration"),
        grade_summary(fit["y_test"], fit["p_test"], grade_test, "test"),
    ], ignore_index=True)

    preds = pd.concat([
        make_prediction_table(fit["train"], target_col, fit["p_train"], fit["p_train_raw"], grade_train, "train", prom_col),
        make_prediction_table(fit["calibration"], target_col, fit["p_cal"], fit["p_cal_raw"], grade_cal, "calibration", prom_col),
        make_prediction_table(fit["test"], target_col, fit["p_test"], fit["p_test_raw"], grade_test, "test", prom_col),
    ], ignore_index=True)

    dca_cal_df = make_decision_curve(y_cal=fit["y_cal"], p_cal_model=fit["p_cal"])
    dca_test_df = make_decision_curve(y_cal=fit["y_test"], p_cal_model=fit["p_test"])
    cutoffs_df = pd.DataFrame([
        {"cutoff_name": k, "value": json.dumps(json_native(v)) if isinstance(v, (dict, list)) else v, "defined_in": GRADE_CUTOFF_MODE}
        for k, v in cutoffs.items()
    ])

    metrics.to_csv(os.path.join(TABLE_DIR, "Step1_ODI40_model_performance_by_split.csv"), index=False)
    grades.to_csv(os.path.join(TABLE_DIR, "Step1_ODI40_risk_grade_summary_by_split.csv"), index=False)
    preds.to_csv(os.path.join(TABLE_DIR, "Step1_ODI40_predictions_with_risk_grades.csv"), index=False)
    cutoffs_df.to_csv(os.path.join(TABLE_DIR, "Step1_ODI40_risk_grade_cutoffs.csv"), index=False)
    dca_cal_df.to_csv(os.path.join(TABLE_DIR, "Step1_ODI40_calibration_DCA_validation_curve.csv"), index=False)
    dca_test_df.to_csv(os.path.join(TABLE_DIR, "Step1_ODI40_test_DCA_validation_curve.csv"), index=False)

    artifact = {
        "analysis": "Step 1 ASR-ODI Reoperation Risk Grade: calibration-risk-tertile three-grade version with DCA validation",
        "cohort": "ODI",
        "model_description": "LightGBM with preoperative ODI dichotomized at 40",
        "preprocessor": fit["preprocessor"],
        "model_object": "lightgbm.LGBMClassifier",
        "model": fit["model"],
        "calibrator": fit["calibrator"],
        "features": FEATURES,
        "baseline_features": BASELINE_FEATURES,
        "ODI_threshold": ODI_THRESHOLD,
        "ODI_binary_feature": ODI_BINARY_COL,
        "grade_cutoff_mode": GRADE_CUTOFF_MODE,
        "manual_cutoffs_prespecified": False,
        "dca_used_for_grade_derivation": False,
        "dca_role": "post-derivation clinical-utility validation only",
        "n_risk_grades": len(GRADE_LABELS),
        "grade_cutoffs": cutoffs,
        "grade_labels": GRADE_LABELS,
        "target_col": target_col,
        "prom_col": prom_col,
        "group_col": group_col,
        "split_source": split_source,
        "locked_split_required": REQUIRE_LOCKED_SPLIT_ASSIGNMENT,
        "new_split_creation_allowed": False,
        "locked_lightgbm_settings": LOCKED_ODI_LIGHTGBM_SETTINGS,
    }

    artifact_path = os.path.join(ARTIFACT_DIR, "ASR_ODI_RiskGrade_ODI40_RiskTertile3Grades_LightGBM_artifact.joblib")
    joblib.dump(artifact, artifact_path)

    with open(os.path.join(TABLE_DIR, "analysis_settings.json"), "w") as f:
        json.dump(json_native({k: v for k, v in artifact.items() if k not in {"preprocessor", "model", "calibrator"}}), f, indent=2, sort_keys=True)

    calibration_dca_plot = os.path.join(PLOT_DIR, "Step1_ODI40_calibration_DCA_validation_plot_with_locked_grade_boundaries.png")
    test_dca_plot = os.path.join(PLOT_DIR, "Step1_ODI40_test_DCA_validation_plot_with_locked_grade_boundaries.png")
    grade_boundary_plot = os.path.join(PLOT_DIR, "Step1_ODI40_calibration_predicted_risk_distribution_with_locked_tertile_boundaries.png")
    calibration_plot = os.path.join(PLOT_DIR, "Step1_ODI40_test_calibration_plot.png")
    grade_plot = os.path.join(PLOT_DIR, "Step1_ODI40_test_observed_reoperation_by_risk_tertile_3grade.png")

    save_dca_plot(
        dca_cal_df,
        calibration_dca_plot,
        "Calibration-set DCA validation with locked ASR-ODI grade boundaries",
        cutoffs=cutoffs,
    )
    save_dca_plot(
        dca_test_df,
        test_dca_plot,
        "Held-out test-set DCA validation with locked ASR-ODI grade boundaries",
        cutoffs=cutoffs,
    )
    save_grade_boundary_distribution_plot(
        fit["p_cal"],
        cutoffs,
        grade_boundary_plot,
        "Calibration-set predicted-risk distribution and locked ASR-ODI grade boundaries",
    )
    save_calibration_plot(fit["y_test"], fit["p_test"], calibration_plot, "Step 1 ODI≥40 model: test-set calibration")
    save_grade_bar_plot(grades, grade_plot, "ASR-ODI risk-tertile 3-grade system: observed reoperation rate")

    app_files = create_streamlit_app(artifact_path, APP_DIR)

    summary_xlsx = os.path.join(OUTPUT_DIR, "Step1_ODI40_RiskGrade_RiskTertile3Grades_DCAValidation_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        metrics.to_excel(writer, sheet_name="model_performance", index=False)
        grades.to_excel(writer, sheet_name="risk_grade_summary", index=False)
        cutoffs_df.to_excel(writer, sheet_name="risk_grade_cutoffs", index=False)
        dca_cal_df.to_excel(writer, sheet_name="calibration_DCA", index=False)
        dca_test_df.to_excel(writer, sheet_name="test_DCA", index=False)
        preds.to_excel(writer, sheet_name="predictions", index=False)

    zip_path = os.path.join(BASE_DIR, "Step1_ODI40_RiskGrade_RiskTertile3Grades_DCAValidation_StreamlitApp.zip")
    make_zip(OUTPUT_DIR, zip_path)

    print("\nDONE")
    print("Output folder:", OUTPUT_DIR)
    print("Model artifact:", artifact_path)
    print("Streamlit app:", app_files["app_path"])
    print("App artifact copy:", app_files["artifact_path"])
    print("Summary Excel:", summary_xlsx)
    print("ZIP:", zip_path)
    print("\nHeld-out test-set performance:")
    print(metrics[metrics["split"].eq("test")].to_string(index=False))
    print("\nHeld-out test-set grade summary:")
    print(grades[grades["split"].eq("test")].to_string(index=False))

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(
                f'<p><b>Step 1 ODI40 calibration-risk-tertile three-grade risk-grade app outputs with DCA validation are ready.</b></p>'
                f'<p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p>'
                f'<p>Path: <code>{zip_path}</code></p>'
            ))
        except Exception as e:
            print("Download link display skipped:", repr(e))

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as e:
            print("Automatic download skipped:", repr(e))


if __name__ == "__main__":
    main()

# %% [markdown] Cell 17
# #**ODI_ASR_Grade_SurvivalValidation_KM_Cox**

# %% Cell 18
# -*- coding: utf-8 -*-
"""
Step 1 ASR-ODI risk-tertile 3-grade survival validation: Kaplan-Meier, log-rank, Cox, and forest plot
======================================================================================================

Purpose
-------
Validate the locked ASR-ODI risk-tertile 3-grade system using time-to-event analyses
in the held-out test set. This script does not refit, recalibrate, retune, or modify
any model, probability calibration, ODI threshold, or grade boundary. It only evaluates
already-assigned locked low/intermediate/high risk grades.

Expected upstream input
-----------------------
The prediction table generated by:
  Step1_ODI40_RiskGrade_RiskTertile3Grades_DCAValidation_StreamlitApp_final.py
"""

# ============================================================
# Imports
# ============================================================

import os
import sys
import json
import zipfile
import shutil
import warnings
import subprocess
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    from lifelines import KaplanMeierFitter, CoxPHFitter
    from lifelines.statistics import multivariate_logrank_test, logrank_test, proportional_hazard_test
    from lifelines.plotting import add_at_risk_counts
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "lifelines"])
    from lifelines import KaplanMeierFitter, CoxPHFitter
    from lifelines.statistics import multivariate_logrank_test, logrank_test, proportional_hazard_test
    from lifelines.plotting import add_at_risk_counts

try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl  # noqa: F401

import matplotlib.pyplot as plt

warnings.filterwarnings("ignore")

# ============================================================
# Configuration
# ============================================================

BASE_DIR = "/content"
FALLBACK_DIR = "/mnt/data"

OUTPUT_DIR = os.path.join(BASE_DIR, "Step1_ODI_RiskTertile3Grade_KM_Cox_Validation")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")
AUDIT_DIR = os.path.join(OUTPUT_DIR, "audit")
for _d in [OUTPUT_DIR, TABLE_DIR, PLOT_DIR, AUDIT_DIR]:
    os.makedirs(_d, exist_ok=True)

# Prefer the current risk-tertile/DCA-validation grading outputs. Older names are kept only as fallbacks.
PREDICTION_CSV_CANDIDATES = [
    os.path.join(BASE_DIR, "Step1_ODI_RiskGrade_ODI40_StreamlitApp_RiskTertile3Grades_DCAValidation", "tables", "Step1_ODI40_predictions_with_risk_grades.csv"),
    os.path.join(FALLBACK_DIR, "Step1_ODI_RiskGrade_ODI40_StreamlitApp_RiskTertile3Grades_DCAValidation", "tables", "Step1_ODI40_predictions_with_risk_grades.csv"),
    os.path.join(BASE_DIR, "Step1_ODI40_predictions_with_risk_grades.csv"),
    os.path.join(FALLBACK_DIR, "Step1_ODI40_predictions_with_risk_grades.csv"),
    os.path.join(BASE_DIR, "Step1_ODI40_predictions_with_risk_grades(1).csv"),
    os.path.join(FALLBACK_DIR, "Step1_ODI40_predictions_with_risk_grades(1).csv"),
    os.path.join(BASE_DIR, "Step1_ODI40_predictions_with_risk_grades(2).csv"),
    os.path.join(FALLBACK_DIR, "Step1_ODI40_predictions_with_risk_grades(2).csv"),
]

INPUT_CSV_CANDIDATES = [
    os.path.join(BASE_DIR, "PROM_ODI_Cohort.csv"),
    os.path.join(BASE_DIR, "PROM_ODI_Cohort(1).csv"),
    os.path.join(FALLBACK_DIR, "PROM_ODI_Cohort.csv"),
    os.path.join(FALLBACK_DIR, "PROM_ODI_Cohort(1).csv"),
]

GROUP_COL_CANDIDATES = ["PersonKey", "PatientKey"]
TARGET_COL_CANDIDATES = [
    "final_reop_step1", "final_reop_1yr", "reop_1yr", "reoperation_1yr",
    "one_year_reoperation", "reop365", "reop_365", "final_reop",
    "reoperation", "reop",
]
PROM_COL_CANDIDATES = [
    "ODI_preop_value", "preop_odi_within_90d", "odi_preop_within_90d",
    "odi_preop_anytime_value", "preop_odi_anytime", "preop_ODI",
    "Preop_ODI", "preoperative_ODI", "Preoperative_ODI", "ODI", "odi",
    "baseline_ODI",
]
REOP_TIME_CANDIDATES = [
    "reoptime", "reop_time", "reoperation_time", "time_to_reoperation",
    "time_to_reoperation_days", "days_to_reoperation", "reoperation_days",
    "days_from_index_to_reoperation", "reop_days", "reoperation_time_days",
]
DEATH_DAYS_CANDIDATES = [
    "days_to_death_from_index_surgery", "days_to_death", "death_time_days",
]

TARGET_WINDOW_DAYS = 365.0
TEST_SPLIT_LABEL = "test"
SURVIVAL_TIME_COL = "time_to_reoperation_or_censor_days"
SURVIVAL_EVENT_COL = "event_reoperation_within_365d"

GRADE_COL = "ASR_ODI_risk_grade"
GRADE_LABEL_COL = "ASR_ODI_risk_grade_label"
PREDICTED_RISK_COL = "calibrated_predicted_1y_reoperation_risk"
SPLIT_COL = "split"

GRADE_SHORT_LABELS = {
    0: "Low risk",
    1: "Intermediate risk",
    2: "High risk",
}
GRADE_LONG_LABELS = {
    0: "Grade 0 / Low risk",
    1: "Grade 1 / Intermediate risk",
    2: "Grade 2 / High risk",
}
GRADE_COLORS = {
    0: "#16A34A",  # green
    1: "#FACC15",  # yellow/amber
    2: "#DC2626",  # red
}

COX_PENALIZER_FALLBACKS = [0.0, 1e-8, 1e-7, 1e-6, 1e-5, 1e-4, 1e-3, 1e-2]
TIME_POINTS_FOR_AT_RISK = [0, 90, 180, 270, 365]
AUTO_DOWNLOAD_ZIP = True
CREATE_COLAB_DOWNLOAD_LINK = True
ZIP_COMPRESSION_LEVEL = 1

MISSING_STRINGS = {"", " ", "na", "n/a", "nan", "none", "null", ".", "missing", "<na>"}

# ============================================================
# Utility functions
# ============================================================


def clean_scalar(x: Any) -> Any:
    if pd.isna(x):
        return np.nan
    if isinstance(x, str):
        s = x.strip().replace("≥", ">=")
        return np.nan if s.lower() in MISSING_STRINGS else s
    return x


def norm_text(x: Any) -> Optional[str]:
    x = clean_scalar(x)
    if pd.isna(x):
        return None
    return str(x).strip().replace("≥", ">=").lower()


def to_binary_target(x: Any) -> float:
    sx = norm_text(x)
    if sx is None:
        return np.nan
    if sx in {"1", "1.0", "yes", "y", "true", "t"}:
        return 1.0
    if sx in {"0", "0.0", "no", "n", "false", "f"}:
        return 0.0
    try:
        v = float(sx)
        return float(v) if v in (0.0, 1.0) else np.nan
    except Exception:
        return np.nan


def find_file(candidates: List[str], label: str) -> str:
    for p in candidates:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(f"Could not find {label}. Tried: {candidates}")


def find_existing_column(columns: List[str], candidates: List[str], label: str) -> str:
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if c.lower() in lookup:
            return lookup[c.lower()]
    raise ValueError(f"Could not find {label}. Tried: {candidates}")


def optional_existing_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    lookup = {str(c).lower(): str(c) for c in columns}
    for c in candidates:
        if c in columns:
            return c
        if c.lower() in lookup:
            return lookup[c.lower()]
    return None


def make_match_count(df: pd.DataFrame, key_cols: List[str]) -> pd.Series:
    return df.groupby(key_cols, dropna=False).cumcount()


def holm_adjust(p_values: np.ndarray) -> np.ndarray:
    p = np.asarray(p_values, dtype=float)
    out = np.full_like(p, np.nan, dtype=float)
    finite = np.where(np.isfinite(p))[0]
    if len(finite) == 0:
        return out
    order = finite[np.argsort(p[finite])]
    m = len(order)
    adjusted_sorted = np.empty(m, dtype=float)
    running = 0.0
    for rank, idx in enumerate(order):
        value = min((m - rank) * p[idx], 1.0)
        running = max(running, value)
        adjusted_sorted[rank] = running
    for rank, idx in enumerate(order):
        out[idx] = adjusted_sorted[rank]
    return out


def binomial_ci_wilson(events: int, n: int, z: float = 1.959963984540054) -> Tuple[float, float]:
    if n <= 0:
        return np.nan, np.nan
    p = events / n
    denom = 1.0 + z**2 / n
    center = (p + z**2 / (2 * n)) / denom
    half = z * np.sqrt((p * (1 - p) / n) + (z**2 / (4 * n**2))) / denom
    return max(0.0, center - half), min(1.0, center + half)


def format_p_value(p: float) -> str:
    if not np.isfinite(p):
        return "NA"
    if p < 1e-300:
        return "<1e-300"
    if p < 0.001:
        return f"{p:.2e}"
    return f"{p:.3f}"


def json_native(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {str(k): json_native(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_native(v) for v in obj]
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return float(obj)
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    return obj

# ============================================================
# Loading and validation dataset construction
# ============================================================


def load_predictions() -> Tuple[pd.DataFrame, str, Dict[str, str]]:
    path = find_file(PREDICTION_CSV_CANDIDATES, "risk-tertile 3-grade prediction table")
    pred = pd.read_csv(path, low_memory=False)
    pred.columns = [str(c).strip() for c in pred.columns]

    group_col = find_existing_column(pred.columns.tolist(), GROUP_COL_CANDIDATES, "prediction-table patient group column")
    target_col = find_existing_column(pred.columns.tolist(), TARGET_COL_CANDIDATES, "prediction-table 1-year reoperation target column")
    prom_col = find_existing_column(pred.columns.tolist(), PROM_COL_CANDIDATES, "prediction-table preoperative ODI column")

    required = [group_col, target_col, prom_col, SPLIT_COL, PREDICTED_RISK_COL, GRADE_COL]
    missing = [c for c in required if c not in pred.columns]
    if missing:
        raise ValueError(f"Prediction table is missing required columns: {missing}")

    pred = pred.copy()
    pred[group_col] = pred[group_col]
    pred[target_col] = pred[target_col].map(to_binary_target)
    pred[prom_col] = pd.to_numeric(pred[prom_col].map(clean_scalar), errors="coerce")
    pred[PREDICTED_RISK_COL] = pd.to_numeric(pred[PREDICTED_RISK_COL], errors="coerce")
    pred[GRADE_COL] = pd.to_numeric(pred[GRADE_COL], errors="coerce").astype("Int64")
    pred[SPLIT_COL] = pred[SPLIT_COL].astype(str).str.lower().str.strip()

    valid = pred[target_col].isin([0.0, 1.0]) & pred[prom_col].notna() & pred[GRADE_COL].notna() & pred[PREDICTED_RISK_COL].notna()
    pred = pred.loc[valid].copy()
    pred[target_col] = pred[target_col].astype(int)
    pred[GRADE_COL] = pred[GRADE_COL].astype(int)

    found_grades = sorted(pred[GRADE_COL].dropna().astype(int).unique().tolist())
    if found_grades != [0, 1, 2]:
        raise RuntimeError(
            f"This script requires locked 3-grade assignments coded as 0, 1, and 2. Found: {found_grades}. "
            "Check that you are using the current risk-tertile 3-grade prediction table."
        )

    if GRADE_LABEL_COL not in pred.columns:
        pred[GRADE_LABEL_COL] = pred[GRADE_COL].map(GRADE_LONG_LABELS)
    pred[GRADE_LABEL_COL] = pred[GRADE_COL].map(GRADE_LONG_LABELS)
    pred["_prediction_row_order"] = np.arange(len(pred))

    cols = {"group_col": group_col, "target_col": target_col, "prom_col": prom_col}
    return pred, path, cols


def load_cohort() -> Tuple[pd.DataFrame, str, Dict[str, str]]:
    path = find_file(INPUT_CSV_CANDIDATES, "PROM ODI cohort CSV")
    cohort = pd.read_csv(path, low_memory=False)
    cohort.columns = [str(c).strip() for c in cohort.columns]

    group_col = find_existing_column(cohort.columns.tolist(), GROUP_COL_CANDIDATES, "cohort patient group column")
    target_col = find_existing_column(cohort.columns.tolist(), TARGET_COL_CANDIDATES, "cohort 1-year reoperation target column")
    prom_col = find_existing_column(cohort.columns.tolist(), PROM_COL_CANDIDATES, "cohort preoperative ODI column")
    reop_time_col = find_existing_column(cohort.columns.tolist(), REOP_TIME_CANDIDATES, "cohort time-to-reoperation column")
    death_days_col = optional_existing_column(cohort.columns.tolist(), DEATH_DAYS_CANDIDATES)

    cols = {
        "group_col": group_col,
        "target_col": target_col,
        "prom_col": prom_col,
        "reop_time_col": reop_time_col,
        "death_days_col": death_days_col if death_days_col is not None else "",
    }
    return cohort, path, cols


def build_validation_dataset(
    pred: pd.DataFrame,
    cohort: pd.DataFrame,
    pred_cols: Dict[str, str],
    cohort_cols: Dict[str, str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    pred_group_col = pred_cols["group_col"]
    pred_target_col = pred_cols["target_col"]
    pred_prom_col = pred_cols["prom_col"]
    cohort_group_col = cohort_cols["group_col"]
    cohort_target_col = cohort_cols["target_col"]
    cohort_prom_col = cohort_cols["prom_col"]
    reop_time_col = cohort_cols["reop_time_col"]
    death_days_col = cohort_cols.get("death_days_col", "") or None

    work = cohort.copy()
    work[cohort_target_col] = work[cohort_target_col].map(to_binary_target)
    work[cohort_prom_col] = pd.to_numeric(work[cohort_prom_col].map(clean_scalar), errors="coerce")
    work[reop_time_col] = pd.to_numeric(work[reop_time_col], errors="coerce")

    keep = work[cohort_group_col].notna() & work[cohort_target_col].isin([0.0, 1.0]) & work[cohort_prom_col].notna()
    work = work.loc[keep].copy().reset_index(drop=True)
    work[cohort_target_col] = work[cohort_target_col].astype(int)

    # Strict merge keys ensure that the survival validation uses exactly the same eligible rows
    # as the prediction table, without creating or changing any grade assignment.
    work["_merge_group"] = work[cohort_group_col].astype(str)
    work["_merge_target"] = work[cohort_target_col].astype(int)
    work["_merge_prom"] = work[cohort_prom_col].round(8)

    pred2 = pred.copy()
    pred2["_merge_group"] = pred2[pred_group_col].astype(str)
    pred2["_merge_target"] = pred2[pred_target_col].astype(int)
    pred2["_merge_prom"] = pred2[pred_prom_col].round(8)

    key_cols = ["_merge_group", "_merge_target", "_merge_prom"]
    work["_within_key_count"] = make_match_count(work, key_cols)
    pred2["_within_key_count"] = make_match_count(pred2, key_cols)
    merge_cols = key_cols + ["_within_key_count"]

    pred_keep = [
        "_prediction_row_order", SPLIT_COL, PREDICTED_RISK_COL, GRADE_COL, GRADE_LABEL_COL,
        "raw_lightgbm_model_score" if "raw_lightgbm_model_score" in pred2.columns else None,
    ]
    pred_keep = [c for c in pred_keep if c is not None]

    merged = work.merge(pred2[merge_cols + pred_keep], on=merge_cols, how="inner", validate="one_to_one")

    target = merged[cohort_target_col].astype(int)
    reop_time = pd.to_numeric(merged[reop_time_col], errors="coerce")
    censor = pd.Series(TARGET_WINDOW_DAYS, index=merged.index, dtype=float)

    if death_days_col is not None and death_days_col in merged.columns:
        death_days = pd.to_numeric(merged[death_days_col], errors="coerce")
        death_days = death_days.where(death_days.gt(0), np.nan)
        censor = pd.concat([censor, death_days], axis=1).min(axis=1, skipna=True)

    # Event occurs if a reoperation is observed within 365 days before censoring.
    event = target.eq(1) & reop_time.notna() & reop_time.gt(0) & reop_time.le(TARGET_WINDOW_DAYS) & reop_time.le(censor)
    duration = censor.copy()
    duration[event] = reop_time[event]

    missing_event_time = target.eq(1) & reop_time.isna()
    nonpositive_duration = duration.isna() | ~np.isfinite(duration) | duration.le(0)

    merged[SURVIVAL_TIME_COL] = duration.astype(float)
    merged[SURVIVAL_EVENT_COL] = event.astype(int)
    merged[GRADE_COL] = merged[GRADE_COL].astype(int)
    merged["grade_short_label"] = merged[GRADE_COL].map(GRADE_SHORT_LABELS)

    before_survival_filter = len(merged)
    analysis = merged.loc[(~missing_event_time) & (~nonpositive_duration)].copy().reset_index(drop=True)

    audit = pd.DataFrame([
        {"audit_item": "prediction_file_rows_after_basic_validity_filter", "value": int(len(pred2)), "note": "Rows in prediction table with valid target, ODI, predicted risk, and grade."},
        {"audit_item": "eligible_cohort_rows", "value": int(len(work)), "note": "Rows after target/PROM/group eligibility filters in cohort."},
        {"audit_item": "merged_rows_before_survival_filter", "value": int(before_survival_filter), "note": "Rows matched between prediction table and cohort."},
        {"audit_item": "unmatched_prediction_rows", "value": int(len(pred2) - before_survival_filter), "note": "Should be 0 if cohort and prediction table are aligned."},
        {"audit_item": "excluded_events_with_missing_reop_time", "value": int(missing_event_time.sum()), "note": "Event-positive rows without usable reoperation time."},
        {"audit_item": "excluded_nonpositive_or_missing_duration", "value": int(nonpositive_duration.sum()), "note": "Rows without valid survival duration."},
        {"audit_item": "rows_after_survival_filter", "value": int(len(analysis)), "note": "Rows available for KM/Cox validation."},
        {"audit_item": "events_after_survival_filter", "value": int(analysis[SURVIVAL_EVENT_COL].sum()), "note": "Reoperations observed by 365 days."},
        {"audit_item": "event_rate_after_survival_filter", "value": float(analysis[SURVIVAL_EVENT_COL].mean()), "note": "Overall event rate in analysis dataset."},
        {"audit_item": "prediction_group_col_used", "value": pred_group_col, "note": ""},
        {"audit_item": "prediction_target_col_used", "value": pred_target_col, "note": ""},
        {"audit_item": "prediction_prom_col_used", "value": pred_prom_col, "note": ""},
        {"audit_item": "cohort_group_col_used", "value": cohort_group_col, "note": ""},
        {"audit_item": "cohort_target_col_used", "value": cohort_target_col, "note": ""},
        {"audit_item": "cohort_prom_col_used", "value": cohort_prom_col, "note": ""},
        {"audit_item": "reop_time_col_used", "value": reop_time_col, "note": ""},
        {"audit_item": "death_days_col_used", "value": death_days_col if death_days_col else "none", "note": ""},
    ])

    if len(pred2) != before_survival_filter:
        raise RuntimeError(
            f"Prediction-to-cohort matching incomplete: matched {before_survival_filter:,} of {len(pred2):,} prediction rows. "
            "Confirm that the prediction table and PROM_ODI_Cohort file came from the same analysis run."
        )

    return analysis, audit

# ============================================================
# Summary tables and tests
# ============================================================


def grade_survival_summary(df: pd.DataFrame, split_name: str) -> pd.DataFrame:
    total_events = int(df[SURVIVAL_EVENT_COL].sum())
    prevalence = float(df[SURVIVAL_EVENT_COL].mean()) if len(df) else np.nan
    rows = []
    for grade in [0, 1, 2]:
        d = df[df[GRADE_COL].eq(grade)]
        n = int(len(d))
        events = int(d[SURVIVAL_EVENT_COL].sum()) if n else 0
        rate = events / n if n else np.nan
        lo, hi = binomial_ci_wilson(events, n)
        mean_pred = float(d[PREDICTED_RISK_COL].mean()) if n else np.nan
        rows.append({
            "split": split_name,
            "grade": grade,
            "grade_label": GRADE_LONG_LABELS[grade],
            "grade_short_label": GRADE_SHORT_LABELS[grade],
            "n": n,
            "events": events,
            "observed_reoperation_rate": rate,
            "observed_reoperation_rate_lower_95": lo,
            "observed_reoperation_rate_upper_95": hi,
            "mean_predicted_risk": mean_pred,
            "median_predicted_risk": float(d[PREDICTED_RISK_COL].median()) if n else np.nan,
            "observed_to_expected_ratio": rate / mean_pred if mean_pred > 0 and np.isfinite(rate) else np.nan,
            "risk_lift_vs_split_prevalence": rate / prevalence if prevalence > 0 and np.isfinite(rate) else np.nan,
            "captured_fraction_of_events": events / total_events if total_events else np.nan,
            "median_followup_or_event_time_days": float(d[SURVIVAL_TIME_COL].median()) if n else np.nan,
        })
    return pd.DataFrame(rows)


def at_risk_table(df: pd.DataFrame, times: List[int]) -> pd.DataFrame:
    rows = []
    for grade in [0, 1, 2]:
        d = df[df[GRADE_COL].eq(grade)]
        for t in times:
            rows.append({
                "grade": grade,
                "grade_label": GRADE_LONG_LABELS[grade],
                "time_days": int(t),
                "n_at_risk": int(np.sum(d[SURVIVAL_TIME_COL] >= t)),
                "cumulative_events_by_time": int(np.sum(d[SURVIVAL_EVENT_COL].eq(1) & (d[SURVIVAL_TIME_COL] <= t))),
            })
    return pd.DataFrame(rows)


def run_logrank_tests(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    global_lr = multivariate_logrank_test(
        event_durations=df[SURVIVAL_TIME_COL],
        groups=df[GRADE_COL].astype(int),
        event_observed=df[SURVIVAL_EVENT_COL].astype(int),
    )
    global_df = pd.DataFrame([{
        "test": "global_logrank_across_3_locked_grades",
        "test_statistic": float(global_lr.test_statistic),
        "p": float(global_lr.p_value),
        "n": int(len(df)),
        "events": int(df[SURVIVAL_EVENT_COL].sum()),
    }])

    pairs = []
    for low, high in [(0, 1), (0, 2), (1, 2)]:
        d0 = df[df[GRADE_COL].eq(low)]
        d1 = df[df[GRADE_COL].eq(high)]
        lr = logrank_test(
            d1[SURVIVAL_TIME_COL],
            d0[SURVIVAL_TIME_COL],
            event_observed_A=d1[SURVIVAL_EVENT_COL],
            event_observed_B=d0[SURVIVAL_EVENT_COL],
        )
        pairs.append({
            "comparison": f"{GRADE_LONG_LABELS[high]} vs {GRADE_LONG_LABELS[low]}",
            "higher_grade": high,
            "reference_grade": low,
            "test_statistic": float(lr.test_statistic),
            "p_raw": float(lr.p_value),
            "n_higher_grade": int(len(d1)),
            "events_higher_grade": int(d1[SURVIVAL_EVENT_COL].sum()),
            "n_reference_grade": int(len(d0)),
            "events_reference_grade": int(d0[SURVIVAL_EVENT_COL].sum()),
        })
    pairwise = pd.DataFrame(pairs)
    pairwise["p_holm"] = holm_adjust(pairwise["p_raw"].to_numpy(dtype=float))
    return global_df, pairwise

# ============================================================
# Cox models
# ============================================================


def _extract_cox_row(cph: CoxPHFitter, term: str, model_name: str, comparison: str, penalizer: float, n: int, events: int) -> Dict[str, Any]:
    s = cph.summary.loc[term]
    return {
        "cox_model": model_name,
        "comparison": comparison,
        "term": term,
        "coef": float(s["coef"]),
        "se(coef)": float(s["se(coef)"]),
        "z": float(s["z"]),
        "p": float(s["p"]),
        "HR": float(s["exp(coef)"]),
        "HR_lower_95": float(s["exp(coef) lower 95%"]),
        "HR_upper_95": float(s["exp(coef) upper 95%"]),
        "n": int(n),
        "events": int(events),
        "method": "unadjusted Cox proportional hazards with robust variance",
        "penalizer": float(penalizer),
        "converged": True,
    }


def _fit_single_cox(work: pd.DataFrame, term: str, model_name: str, comparison: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    cols = [SURVIVAL_TIME_COL, SURVIVAL_EVENT_COL, term]
    d = work[cols].copy()
    d[SURVIVAL_TIME_COL] = pd.to_numeric(d[SURVIVAL_TIME_COL], errors="coerce")
    d[SURVIVAL_EVENT_COL] = pd.to_numeric(d[SURVIVAL_EVENT_COL], errors="coerce").astype(int)
    d[term] = pd.to_numeric(d[term], errors="coerce")
    d = d.replace([np.inf, -np.inf], np.nan).dropna()
    d = d[d[SURVIVAL_TIME_COL] > 0].copy()

    if len(d) == 0 or d[SURVIVAL_EVENT_COL].sum() == 0:
        raise RuntimeError(f"No usable observations/events for Cox model: {model_name}")
    if d[term].nunique() < 2:
        raise RuntimeError(f"No covariate variation for Cox model: {model_name}")

    last_error = None
    for penalizer in COX_PENALIZER_FALLBACKS:
        try:
            cph = CoxPHFitter(penalizer=penalizer)
            cph.fit(
                d,
                duration_col=SURVIVAL_TIME_COL,
                event_col=SURVIVAL_EVENT_COL,
                robust=True,
                show_progress=False,
            )
            row = _extract_cox_row(
                cph=cph,
                term=term,
                model_name=model_name,
                comparison=comparison,
                penalizer=penalizer,
                n=len(d),
                events=int(d[SURVIVAL_EVENT_COL].sum()),
            )
            ph_p = np.nan
            ph_stat = np.nan
            ph_status = "not_evaluated"
            try:
                ph = proportional_hazard_test(cph, d, time_transform="rank")
                ph_p = float(ph.summary.loc[term, "p"])
                ph_stat = float(ph.summary.loc[term, "test_statistic"])
                ph_status = "success"
            except Exception as exc:
                ph_status = f"failed: {repr(exc)}"
            ph_row = {
                "cox_model": model_name,
                "comparison": comparison,
                "term": term,
                "ph_test": "Schoenfeld residual proportional-hazards test; rank time transform",
                "test_statistic": ph_stat,
                "p": ph_p,
                "status": ph_status,
            }
            audit = {
                "cox_model": model_name,
                "comparison": comparison,
                "status": "success",
                "penalizer": float(penalizer),
                "method": row["method"],
                "note": "Penalizer is 0 for the ordinary Cox model; nonzero only if needed for numerical convergence.",
            }
            return row, {**ph_row, **{"_audit": audit}}
        except Exception as exc:
            last_error = exc
            continue

    raise RuntimeError(f"Cox model failed for {model_name}. Last error: {repr(last_error)}")


def fit_ordinal_cox(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    work = df[[SURVIVAL_TIME_COL, SURVIVAL_EVENT_COL, GRADE_COL]].copy()
    work["ASR_ODI_grade_ordinal"] = work[GRADE_COL].astype(float)
    row, ph_bundle = _fit_single_cox(
        work=work[[SURVIVAL_TIME_COL, SURVIVAL_EVENT_COL, "ASR_ODI_grade_ordinal"]],
        term="ASR_ODI_grade_ordinal",
        model_name="ordinal_grade_trend",
        comparison="Per one-grade increase in ASR-ODI risk grade",
    )
    audit = ph_bundle.pop("_audit")
    return pd.DataFrame([row]), pd.DataFrame([ph_bundle]), pd.DataFrame([audit])


def fit_pairwise_cox(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    ph_rows = []
    audits = []
    for ref, comp in [(0, 1), (0, 2), (1, 2)]:
        sub = df[df[GRADE_COL].isin([ref, comp])].copy()
        term = "higher_grade_vs_reference"
        sub[term] = (sub[GRADE_COL] == comp).astype(float)
        model_name = f"pairwise_grade_{comp}_vs_grade_{ref}"
        label = f"{GRADE_LONG_LABELS[comp]} vs {GRADE_LONG_LABELS[ref]}"
        row, ph_bundle = _fit_single_cox(
            work=sub[[SURVIVAL_TIME_COL, SURVIVAL_EVENT_COL, term]],
            term=term,
            model_name=model_name,
            comparison=label,
        )
        row.update({
            "higher_grade": int(comp),
            "reference_grade": int(ref),
            "n_higher_grade": int((sub[GRADE_COL] == comp).sum()),
            "events_higher_grade": int(sub.loc[sub[GRADE_COL] == comp, SURVIVAL_EVENT_COL].sum()),
            "n_reference_grade": int((sub[GRADE_COL] == ref).sum()),
            "events_reference_grade": int(sub.loc[sub[GRADE_COL] == ref, SURVIVAL_EVENT_COL].sum()),
        })
        audit = ph_bundle.pop("_audit")
        rows.append(row)
        ph_rows.append(ph_bundle)
        audits.append(audit)
    return pd.DataFrame(rows), pd.DataFrame(ph_rows), pd.DataFrame(audits)

# ============================================================
# Plots
# ============================================================


def save_km_plot(df: pd.DataFrame, path: str) -> None:
    fig, ax = plt.subplots(figsize=(8.4, 6.2))
    kmfs = []
    for grade in [0, 1, 2]:
        d = df[df[GRADE_COL].eq(grade)]
        kmf = KaplanMeierFitter(label=GRADE_SHORT_LABELS[grade])
        kmf.fit(durations=d[SURVIVAL_TIME_COL], event_observed=d[SURVIVAL_EVENT_COL])
        kmf.plot_survival_function(
            ax=ax,
            ci_show=True,
            linewidth=2.2,
            color=GRADE_COLORS[grade],
        )
        kmfs.append(kmf)
    ax.set_title("ASR-ODI locked 3-grade validation: reoperation-free survival", fontweight="bold")
    ax.set_xlabel("Days after index surgery")
    ax.set_ylabel("Reoperation-free survival probability")
    ax.set_xlim(0, TARGET_WINDOW_DAYS)
    ax.set_ylim(0, 1.01)
    ax.grid(alpha=0.25)
    add_at_risk_counts(*kmfs, ax=ax, rows_to_show=["At risk"])
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_cumulative_risk_plot(df: pd.DataFrame, path: str) -> None:
    fig, ax = plt.subplots(figsize=(8.0, 5.6))
    for grade in [0, 1, 2]:
        d = df[df[GRADE_COL].eq(grade)]
        kmf = KaplanMeierFitter(label=GRADE_SHORT_LABELS[grade])
        kmf.fit(durations=d[SURVIVAL_TIME_COL], event_observed=d[SURVIVAL_EVENT_COL])
        sf = kmf.survival_function_.copy()
        ax.step(
            sf.index,
            (1.0 - sf.iloc[:, 0]) * 100.0,
            where="post",
            linewidth=2.2,
            label=GRADE_SHORT_LABELS[grade],
            color=GRADE_COLORS[grade],
        )
    ax.set_title("ASR-ODI locked 3-grade validation: cumulative reoperation risk", fontweight="bold")
    ax.set_xlabel("Days after index surgery")
    ax.set_ylabel("Cumulative reoperation risk (%)")
    ax.set_xlim(0, TARGET_WINDOW_DAYS)
    ax.grid(alpha=0.25)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_forest_plot(pairwise_df: pd.DataFrame, path: str) -> None:
    d = pairwise_df.copy()
    d = d.sort_values(["reference_grade", "higher_grade"]).reset_index(drop=True)
    labels = d["comparison"].tolist()
    hr = d["HR"].to_numpy(dtype=float)
    lo = d["HR_lower_95"].to_numpy(dtype=float)
    hi = d["HR_upper_95"].to_numpy(dtype=float)
    p_col = "p_holm" if "p_holm" in d.columns else "p"
    pvals = d[p_col].to_numpy(dtype=float)
    y = np.arange(len(d))[::-1]

    finite = np.isfinite(hr) & np.isfinite(lo) & np.isfinite(hi) & (hr > 0) & (lo > 0) & (hi > 0)
    if not finite.all():
        raise RuntimeError("Non-finite HR/CI detected in pairwise Cox output; inspect the Cox audit table.")

    xmin = max(0.1, float(np.min(lo) / 1.6))
    xmax = max(float(np.max(hi) * 1.8), 2.0)

    fig, ax = plt.subplots(figsize=(10.2, max(4.2, 1.1 * len(d) + 2.0)))
    for yi, h, l, u in zip(y, hr, lo, hi):
        ax.errorbar(
            [h],
            [yi],
            xerr=np.array([[h - l], [u - h]]),
            fmt="o",
            capsize=5,
            markersize=7,
            linewidth=1.8,
        )

    ax.axvline(1.0, linestyle="--", linewidth=1.2)
    ax.set_xscale("log")
    ax.set_xlim(xmin, xmax)
    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.set_xlabel("Hazard ratio for 1-year reoperation (log scale)")
    ax.set_title("ASR-ODI locked 3-grade validation: pairwise Cox hazard ratios", fontweight="bold")
    ax.grid(axis="x", alpha=0.25)

    text_x = xmax * 1.05
    for yi, h, l, u, p in zip(y, hr, lo, hi, pvals):
        ax.text(
            text_x,
            yi,
            f"HR {h:.2f} ({l:.2f}-{u:.2f}), Holm p={format_p_value(p)}" if p_col == "p_holm" else f"HR {h:.2f} ({l:.2f}-{u:.2f}), p={format_p_value(p)}",
            va="center",
            ha="left",
            fontsize=9.5,
        )
    ax.set_xlim(xmin, xmax * 2.4)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)

# ============================================================
# Output helpers
# ============================================================


def make_zip(src_dir: str, zip_path: str) -> None:
    tmp_zip = zip_path + ".tmp"
    for p in [zip_path, tmp_zip]:
        if os.path.exists(p):
            os.remove(p)
    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=ZIP_COMPRESSION_LEVEL) as zf:
        for root, _, files in os.walk(src_dir):
            for fn in files:
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, os.path.dirname(src_dir))
                zf.write(full, rel)
    os.replace(tmp_zip, zip_path)

# ============================================================
# Main
# ============================================================


def main() -> None:
    print("=" * 100)
    print("Step 1 ASR-ODI risk-tertile 3-grade survival validation: KM + log-rank + Cox + forest plot")
    print("=" * 100)

    pred, pred_path, pred_cols = load_predictions()
    cohort, cohort_path, cohort_cols = load_cohort()
    validation_all, merge_audit = build_validation_dataset(pred, cohort, pred_cols, cohort_cols)
    validation = validation_all[validation_all[SPLIT_COL].eq(TEST_SPLIT_LABEL)].copy().reset_index(drop=True)

    if len(validation) == 0:
        raise RuntimeError("No held-out test-set rows found after survival dataset construction.")
    if int(validation[SURVIVAL_EVENT_COL].sum()) == 0:
        raise RuntimeError("No reoperation events found in the held-out test set.")
    if sorted(validation[GRADE_COL].dropna().astype(int).unique().tolist()) != [0, 1, 2]:
        raise RuntimeError("The held-out test set must contain all three locked grades for KM/Cox validation.")

    split_summaries = []
    for split in sorted(validation_all[SPLIT_COL].dropna().unique().tolist()):
        split_summaries.append(grade_survival_summary(validation_all[validation_all[SPLIT_COL].eq(split)].copy(), split))
    grade_summary_by_split = pd.concat(split_summaries, ignore_index=True)
    grade_summary_test = grade_survival_summary(validation, TEST_SPLIT_LABEL)
    at_risk = at_risk_table(validation, TIME_POINTS_FOR_AT_RISK)

    global_lr, pairwise_lr = run_logrank_tests(validation)
    cox_ordinal, ph_ordinal, audit_ord = fit_ordinal_cox(validation)
    cox_pairwise, ph_pairwise, audit_pairwise = fit_pairwise_cox(validation)

    # Pairwise Cox models involve three prespecified grade comparisons:
    # intermediate vs low, high vs low, and high vs intermediate.
    # Raw Cox p values are retained, and Holm-adjusted p values are added for
    # multiplicity-aware reporting of the pairwise grade-separation tests.
    cox_pairwise["p_raw"] = pd.to_numeric(cox_pairwise["p"], errors="coerce")
    cox_pairwise["p_holm"] = holm_adjust(cox_pairwise["p_raw"].to_numpy(dtype=float))
    cox_pairwise["multiplicity_family"] = "pairwise Cox comparisons across three locked ASR-ODI grades"

    cox_all = pd.concat([cox_pairwise, cox_ordinal], ignore_index=True, sort=False)
    ph_all = pd.concat([ph_pairwise, ph_ordinal], ignore_index=True, sort=False)
    cox_audit = pd.concat([audit_pairwise, audit_ord], ignore_index=True, sort=False)

    # Save tables.
    validation.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_survival_validation_dataset_test.csv"), index=False)
    grade_summary_by_split.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_survival_summary_by_split.csv"), index=False)
    grade_summary_test.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_survival_summary_test.csv"), index=False)
    at_risk.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_at_risk_table_test.csv"), index=False)
    global_lr.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_global_logrank_test.csv"), index=False)
    pairwise_lr.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_pairwise_logrank_tests_Holm.csv"), index=False)
    cox_all.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_Cox_models.csv"), index=False)
    cox_pairwise.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_Cox_pairwise_forest_table.csv"), index=False)
    ph_all.to_csv(os.path.join(TABLE_DIR, "ASR_ODI_RiskTertile3Grade_Cox_PH_tests_Schoenfeld.csv"), index=False)
    merge_audit.to_csv(os.path.join(AUDIT_DIR, "ASR_ODI_RiskTertile3Grade_survival_merge_audit.csv"), index=False)
    cox_audit.to_csv(os.path.join(AUDIT_DIR, "ASR_ODI_RiskTertile3Grade_Cox_convergence_audit.csv"), index=False)

    # Save plots.
    km_path = os.path.join(PLOT_DIR, "ASR_ODI_RiskTertile3Grade_Kaplan_Meier_reoperation_free_survival.png")
    risk_path = os.path.join(PLOT_DIR, "ASR_ODI_RiskTertile3Grade_cumulative_reoperation_risk.png")
    forest_path = os.path.join(PLOT_DIR, "ASR_ODI_RiskTertile3Grade_Cox_pairwise_forest_plot.png")
    save_km_plot(validation, km_path)
    save_cumulative_risk_plot(validation, risk_path)
    save_forest_plot(cox_pairwise, forest_path)

    # Save analysis settings/audit.
    settings = {
        "prediction_table_used": pred_path,
        "cohort_table_used": cohort_path,
        "prediction_columns_used": pred_cols,
        "cohort_columns_used": cohort_cols,
        "target_window_days": TARGET_WINDOW_DAYS,
        "test_split_label": TEST_SPLIT_LABEL,
        "n_validation_test": int(len(validation)),
        "events_validation_test": int(validation[SURVIVAL_EVENT_COL].sum()),
        "grade_system": "locked ASR-ODI risk-tertile 3-grade system",
        "grade_boundary_handling": "grade boundaries were already locked upstream and are not modified in this script",
        "cox_models": "unadjusted Cox proportional-hazards validation models because grade is a model-derived baseline risk summary; pairwise Cox p values are reported both raw and Holm-adjusted across the three pairwise grade comparisons",
        "ph_assumption": "Schoenfeld residual tests are reported for each Cox validation model when available",
    }
    with open(os.path.join(AUDIT_DIR, "ASR_ODI_RiskTertile3Grade_survival_validation_settings.json"), "w") as f:
        json.dump(json_native(settings), f, indent=2, sort_keys=True)

    # Summary workbook.
    summary_xlsx = os.path.join(OUTPUT_DIR, "ASR_ODI_RiskTertile3Grade_KM_Cox_Validation_summary.xlsx")
    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        grade_summary_test.to_excel(writer, sheet_name="grade_summary_test", index=False)
        global_lr.to_excel(writer, sheet_name="global_logrank", index=False)
        pairwise_lr.to_excel(writer, sheet_name="pairwise_logrank_Holm", index=False)
        cox_all.to_excel(writer, sheet_name="cox_models", index=False)
        cox_pairwise.to_excel(writer, sheet_name="cox_pairwise_forest", index=False)
        ph_all.to_excel(writer, sheet_name="cox_PH_tests", index=False)
        at_risk.to_excel(writer, sheet_name="at_risk_table", index=False)
        merge_audit.to_excel(writer, sheet_name="merge_audit", index=False)
        cox_audit.to_excel(writer, sheet_name="cox_audit", index=False)

    zip_path = os.path.join(BASE_DIR, "ASR_ODI_RiskTertile3Grade_KM_Cox_Validation.zip")
    make_zip(OUTPUT_DIR, zip_path)

    print("\nDONE")
    print("Prediction table:", pred_path)
    print("PROM ODI cohort:", cohort_path)
    print(f"Held-out validation set: n={len(validation):,}, events={int(validation[SURVIVAL_EVENT_COL].sum()):,}")
    print("\nHeld-out test-set grade survival summary:")
    print(grade_summary_test.to_string(index=False))
    print("\nGlobal log-rank test:")
    print(global_lr.to_string(index=False))
    print("\nPairwise log-rank tests with Holm correction:")
    print(pairwise_lr.to_string(index=False))
    print("\nCox pairwise forest-table estimates with Holm correction:")
    print(cox_pairwise[["comparison", "HR", "HR_lower_95", "HR_upper_95", "p_raw", "p_holm", "penalizer", "n", "events", "method"]].to_string(index=False))
    print("\nOrdinal Cox trend:")
    print(cox_ordinal[["comparison", "HR", "HR_lower_95", "HR_upper_95", "p", "penalizer", "n", "events", "method"]].to_string(index=False))
    print("\nOutput folder:", OUTPUT_DIR)
    print("Summary Excel:", summary_xlsx)
    print("KM plot:", km_path)
    print("Cumulative risk plot:", risk_path)
    print("Forest plot:", forest_path)
    print("ZIP:", zip_path)

    if CREATE_COLAB_DOWNLOAD_LINK:
        try:
            from IPython.display import HTML, display
            display(HTML(
                f'<p><b>ASR-ODI risk-tertile 3-grade KM/Cox validation outputs are ready.</b></p>'
                f'<p><a href="/files{zip_path}" download>Click here to download the ZIP archive</a></p>'
                f'<p>Path: <code>{zip_path}</code></p>'
            ))
        except Exception as exc:
            print("Download link display skipped:", repr(exc))

    if AUTO_DOWNLOAD_ZIP:
        try:
            from google.colab import files
            files.download(zip_path)
        except Exception as exc:
            print("Automatic download skipped:", repr(exc))


if __name__ == "__main__":
    main()

# %% Cell 19

